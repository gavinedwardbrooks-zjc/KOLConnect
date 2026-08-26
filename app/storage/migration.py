from __future__ import annotations

"""Read-only Excel to staged SQLite migration for PRE-M8 Batch 1."""

from dataclasses import dataclass
from datetime import date, datetime, timezone
import hashlib
import json
import os
from pathlib import Path
import shutil
from typing import Any, Callable
import uuid

from openpyxl import load_workbook

from runtime_paths import atomic_write_json, load_json_with_backup
from storage.connection import SQLiteConnectionFactory
from storage.errors import (
    SQLiteActivationError,
    SQLiteBackupError,
    SQLiteIntegrityError,
    SQLiteMigrationAmbiguousIdentityError,
    SQLiteMigrationError,
    SQLiteSchemaUnsupportedError,
)
from storage.paths import SQLiteStoragePaths
from storage.schema import (
    CURRENT_SCHEMA_VERSION,
    apply_schema_migrations,
    schema_version,
    validate_schema,
)


MIGRATION_VERSION = "excel-to-sqlite-v1"
MIGRATION_PHASES = (
    "prepared",
    "source_validated",
    "backup_created",
    "schema_created",
    "data_imported",
    "validated",
    "ready_for_activation",
    "activation_authorized",
    "database_activated",
    "authority_activated",
    "completed",
    "failed",
)


BASE_TABLES: tuple[tuple[str, str, str | tuple[str, ...], tuple[str, ...]], ...] = (
    ("Creators", "creators", "creator_id", (
        "creator_id", "name", "platform", "profile_url", "country", "language",
        "content_category", "followers", "insight_level", "status", "created_at",
        "updated_at", "email", "whatsapp", "cooperation_stage", "recent_product",
        "quote", "owner", "last_contact_time", "next_follow_up_time", "note",
        "agency_id", "current_contact_id", "source_contact_id", "bio", "archived_at",
    )),
    ("CreatorAccounts", "creator_accounts", "account_uid", (
        "account_uid", "account_id", "creator_id", "platform", "username", "profile_url",
        "followers", "account_email", "latest_post_date", "last_scrape_time", "data_source",
        "scrape_status", "platform_account_id", "attribution_status", "note", "source_task_id",
        "created_at", "updated_at",
    )),
    ("Videos", "videos", ("creator_id", "video_url"), (
        "creator_id", "video_url", "views", "likes", "comments", "captured_at",
    )),
    ("Insights", "insights", "creator_id", (
        "creator_id", "average_views", "median_views", "stability", "risks", "recommendation",
    )),
    ("CreatorSnapshots", "creator_snapshots", "snapshot_id", (
        "snapshot_id", "creator_id", "platform", "account_uid", "followers", "average_views",
        "median_views", "video_count", "creator_score", "insight_level", "captured_at", "source",
    )),
    ("VideoSnapshots", "video_snapshots", "video_snapshot_id", (
        "video_snapshot_id", "snapshot_id", "creator_id", "video_id", "video_url", "platform",
        "views", "likes", "comments", "captured_at",
    )),
    ("Cooperations", "cooperations", "cooperation_id", (
        "cooperation_id", "creator_id", "campaign", "platform", "contact_date", "price",
        "published_count", "total_views", "average_views", "roi", "result", "note", "created_at",
    )),
    ("Agencies", "agencies", "agency_id", (
        "agency_id", "name", "country", "website", "public_email", "whatsapp",
        "cooperation_stage", "tags", "last_contact_time", "next_follow_up_time", "owner", "note",
        "resource_files", "created_at", "updated_at",
    )),
    ("AgencyContacts", "agency_contacts", "contact_id", (
        "contact_id", "name", "agency_id", "position", "email", "whatsapp", "language", "status",
        "last_contact_time", "next_follow_up_time", "owner", "note", "external_record_id", "source",
        "created_at", "updated_at",
    )),
    ("FollowUpLogs", "follow_up_logs", "follow_up_id", (
        "follow_up_id", "object_type", "object_id", "contact_method", "content", "stage_before",
        "stage_after", "contacted_at", "next_follow_up_time", "owner", "created_at",
    )),
    ("Products", "products", "product_id", (
        "product_id", "name", "company_name", "note", "created_at", "updated_at", "archived_at",
    )),
    ("Campaigns", "campaigns", "campaign_id", (
        "campaign_id", "product_id", "name", "country", "platform", "start_date", "end_date",
        "owner", "status", "budget", "goal", "note", "created_at", "updated_at", "archived_at",
    )),
    ("CampaignCreators", "campaign_creators", "id", (
        "id", "campaign_id", "creator_id", "account_id", "stage", "owner", "creator_quote", "cost",
        "publish_date", "views", "likes", "comments", "roi", "performance_note", "created_at",
        "updated_at", "archived_at",
    )),
    ("_AnalysisData", "analysis_data", "creator_id", (
        "creator_id", "task_id", "account_uid", "status_updated_at", "analysis_json", "source",
    )),
)

SPECIAL_COLUMNS = {
    "Creators": {"tags"},
    "Campaigns": {"platforms"},
    "CampaignCreators": {"account_ids", "planned_publish_dates", "publish_links"},
    "_Metadata": {"schema_version", "last_update_time"},
}
INTEGER_FIELDS = {
    "followers", "views", "likes", "comments", "video_count", "published_count", "total_views",
}
REAL_FIELDS = {
    "quote", "average_views", "median_views", "stability", "creator_score", "price", "roi",
    "budget", "creator_quote", "cost",
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _number(value: object, *, integer: bool, field: str) -> int | float | None:
    text = _safe_text(value)
    if text is None:
        return None
    try:
        number = float(text.replace(",", ""))
    except ValueError as exc:
        raise SQLiteMigrationError(f"Invalid numeric value for {field}.") from exc
    if integer:
        if not number.is_integer():
            raise SQLiteMigrationError(f"Invalid integer value for {field}.")
        return int(number)
    return number


def _value(field: str, value: object) -> object:
    if field in INTEGER_FIELDS:
        return _number(value, integer=True, field=field)
    if field in REAL_FIELDS:
        return _number(value, integer=False, field=field)
    return _safe_text(value)


def _list_value(value: object, *, fallback: object = None, split_plain: bool = False) -> list[str]:
    if value in (None, ""):
        value = fallback
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple)):
        raw = list(value)
    else:
        text = str(value).strip()
        if text.startswith("["):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError as exc:
                raise SQLiteMigrationError("Malformed JSON relation.") from exc
            if not isinstance(parsed, list):
                raise SQLiteMigrationError("Relation value must be a JSON list.")
            raw = parsed
        elif split_plain:
            raw = [part for part in text.replace("\r", "\n").replace(",", "\n").split("\n")]
        else:
            raw = [text]
    result: list[str] = []
    for item in raw:
        normalized = str(item or "").strip()
        if normalized and normalized not in result:
            result.append(normalized)
    return result


def _sheet_rows(workbook, sheet_name: str) -> tuple[list[str], list[dict[str, object]]]:
    if sheet_name not in workbook.sheetnames:
        return [], []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return [], []
    headers = [str(value or "").strip() for value in raw_headers]
    records: list[dict[str, object]] = []
    for values in rows:
        record = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return headers, records


def _known_columns(sheet_name: str) -> set[str]:
    for source, _table, _identity, columns in BASE_TABLES:
        if source == sheet_name:
            return set(columns) | SPECIAL_COLUMNS.get(source, set())
    return SPECIAL_COLUMNS.get(sheet_name, set())


def validate_source_workbook(path: Path) -> dict[str, object]:
    path = Path(path)
    if not path.is_file():
        raise SQLiteMigrationError("Migration source workbook is missing.")
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise SQLiteMigrationError("Migration source workbook is invalid.") from exc
    try:
        if "Creators" not in workbook.sheetnames or "CreatorAccounts" not in workbook.sheetnames:
            raise SQLiteMigrationError("Migration source is missing identity sheets.")
        known_sheets = set(source for source, *_ in BASE_TABLES) | {"_Metadata"}
        for sheet_name in set(workbook.sheetnames) - known_sheets:
            _headers, rows = _sheet_rows(workbook, sheet_name)
            if rows:
                raise SQLiteMigrationError(f"Unsupported non-empty sheet: {sheet_name}.")
        report: dict[str, int] = {}
        for sheet_name in known_sheets:
            headers, rows = _sheet_rows(workbook, sheet_name)
            if not headers and sheet_name not in {"Creators", "CreatorAccounts"}:
                report[sheet_name] = 0
                continue
            unknown = set(headers) - _known_columns(sheet_name)
            unknown_with_data = {
                column
                for column in unknown
                if any(_safe_text(row.get(column)) is not None for row in rows)
            }
            if unknown_with_data:
                raise SQLiteMigrationError(f"Unsupported columns in {sheet_name}.")
            report[sheet_name] = len(rows)
        for required_sheet, required_key in (("Creators", "creator_id"), ("CreatorAccounts", "account_uid")):
            headers, _rows = _sheet_rows(workbook, required_sheet)
            if required_key not in headers:
                raise SQLiteMigrationError(f"Migration source lacks {required_key}.")
        return {"sheets": report, "source_sha256": _sha256(path)}
    finally:
        workbook.close()


@dataclass(frozen=True)
class MigrationResult:
    migration_id: str
    manifest_path: Path
    staged_database_path: Path
    backup_path: Path
    source_sha256_before: str
    source_sha256_after: str
    counts: dict[str, int]
    semantic_digest: str


@dataclass(frozen=True)
class ProductionActivationAuthorization:
    migration_id: str
    source_sha256: str
    confirmed_at: str


class ExcelToSQLiteMigrator:
    def __init__(
        self,
        paths: SQLiteStoragePaths,
        *,
        failure_injector: Callable[[str], None] | None = None,
        production_root_provider: Callable[[], Path] | None = None,
    ) -> None:
        self.paths = paths
        self.failure_injector = failure_injector
        self.production_root_provider = production_root_provider or (
            lambda: SQLiteStoragePaths.for_app_data().app_data_dir
        )

    def _inject(self, phase: str) -> None:
        if self.failure_injector is not None:
            self.failure_injector(phase)

    def _write_manifest(self, path: Path, manifest: dict[str, object], phase: str) -> None:
        if phase not in MIGRATION_PHASES:
            raise SQLiteMigrationError("Migration phase is invalid.")
        manifest["phase"] = phase
        manifest["updated_at"] = _utc_now()
        atomic_write_json(path, manifest)
        self._inject(phase)

    def migrate(self, source_workbook: Path) -> MigrationResult:
        source_workbook = Path(source_workbook)
        self.paths.ensure_migration_directories()
        migration_id = uuid.uuid4().hex
        manifest_path = self.paths.migration_manifest_path(migration_id)
        staged_path = self.paths.staged_database_path(migration_id)
        backup_path = self.paths.migration_backup_dir / f"{source_workbook.stem}.{migration_id}.xlsx"
        manifest: dict[str, object] = {
            "migration_id": migration_id,
            "migration_version": MIGRATION_VERSION,
            "source_name": source_workbook.name,
            "source_sha256": "",
            "target_name": staged_path.name,
            "backup_name": backup_path.name,
            "phase": "prepared",
            "started_at": _utc_now(),
            "updated_at": _utc_now(),
            "failure_code": "",
            "activation_state": "inactive",
        }
        self._write_manifest(manifest_path, manifest, "prepared")
        try:
            source_report = validate_source_workbook(source_workbook)
            before_hash = str(source_report["source_sha256"])
            manifest["source_sha256"] = before_hash
            self._write_manifest(manifest_path, manifest, "source_validated")

            try:
                shutil.copy2(source_workbook, backup_path)
            except OSError as exc:
                raise SQLiteBackupError("Migration source backup failed.") from exc
            if _sha256(backup_path) != before_hash:
                raise SQLiteBackupError("Migration source backup validation failed.")
            self._write_manifest(manifest_path, manifest, "backup_created")

            factory = SQLiteConnectionFactory(staged_path)
            with factory.read_connection() as connection:
                apply_schema_migrations(connection, migration_reference=migration_id)
            self._write_manifest(manifest_path, manifest, "schema_created")

            counts, digest = self._import_and_validate(source_workbook, factory)
            self._write_manifest(manifest_path, manifest, "data_imported")
            with factory.read_connection() as connection:
                validate_schema(connection)
            self._write_manifest(manifest_path, manifest, "validated")

            after_hash = _sha256(source_workbook)
            if after_hash != before_hash:
                raise SQLiteIntegrityError("Migration source workbook changed during migration.")
            manifest["counts"] = counts
            manifest["semantic_digest"] = digest
            self._write_manifest(manifest_path, manifest, "ready_for_activation")
            return MigrationResult(
                migration_id=migration_id,
                manifest_path=manifest_path,
                staged_database_path=staged_path,
                backup_path=backup_path,
                source_sha256_before=before_hash,
                source_sha256_after=after_hash,
                counts=counts,
                semantic_digest=digest,
            )
        except Exception as exc:
            manifest["failure_code"] = getattr(exc, "code", "SQLITE_MIGRATION_FAILED")
            try:
                self._write_manifest(manifest_path, manifest, "failed")
            except Exception:
                pass
            raise

    def _import_and_validate(
        self, source_workbook: Path, factory: SQLiteConnectionFactory
    ) -> tuple[dict[str, int], str]:
        workbook = load_workbook(source_workbook, read_only=True, data_only=False)
        counts: dict[str, int] = {}
        try:
            with factory.write_transaction() as connection:
                source_rows: dict[str, list[dict[str, object]]] = {}
                for source, table, identity, columns in BASE_TABLES:
                    _headers, rows = _sheet_rows(workbook, source)
                    source_rows[source] = rows
                    self._insert_rows(connection, table, identity, columns, rows)
                    counts[table] = len(rows)
                    self._inject(f"entity_imported:{table}")
                self._insert_relations(connection, source_rows)
                counts.update(self._relation_counts(connection))
                self._import_workbook_metadata(connection, workbook)
            with factory.read_connection() as connection:
                validate_schema(connection)
                digest = semantic_digest(connection)
                self._validate_expected_counts(connection, counts)
            source_digest = source_semantic_digest(workbook)
            if digest != source_digest:
                raise SQLiteIntegrityError("Migrated semantic projections differ from source.")
            return counts, digest
        finally:
            workbook.close()

    @staticmethod
    def _insert_rows(connection, table: str, identity: str | tuple[str, ...], columns: tuple[str, ...], rows):
        identity_fields = (identity,) if isinstance(identity, str) else identity
        seen: set[tuple[str, ...]] = set()
        placeholders = ",".join("?" for _ in columns)
        sql = f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})"
        for row in rows:
            identity_value = tuple(str(row.get(field) or "").strip() for field in identity_fields)
            if any(not value for value in identity_value):
                raise SQLiteMigrationAmbiguousIdentityError(f"Missing identity in {table}.")
            if identity_value in seen:
                raise SQLiteMigrationAmbiguousIdentityError(f"Duplicate identity in {table}.")
            seen.add(identity_value)
            try:
                connection.execute(sql, tuple(_value(field, row.get(field)) for field in columns))
            except Exception as exc:
                if "UNIQUE constraint" in str(exc):
                    raise SQLiteMigrationAmbiguousIdentityError(
                        f"Duplicate identity in {table}."
                    ) from exc
                raise SQLiteMigrationError(f"Failed to migrate {table}.") from exc

    @staticmethod
    def _insert_relations(connection, source_rows: dict[str, list[dict[str, object]]]) -> None:
        for creator in source_rows.get("Creators", []):
            creator_id = str(creator.get("creator_id") or "").strip()
            for position, tag in enumerate(_list_value(creator.get("tags"), split_plain=True)):
                connection.execute(
                    "INSERT INTO creator_tags(creator_id, position, tag) VALUES (?, ?, ?)",
                    (creator_id, position, tag),
                )
        for campaign in source_rows.get("Campaigns", []):
            campaign_id = str(campaign.get("campaign_id") or "").strip()
            platforms = _list_value(
                campaign.get("platforms"), fallback=campaign.get("platform"), split_plain=True
            )
            for position, platform in enumerate(platforms):
                connection.execute(
                    "INSERT INTO campaign_platforms(campaign_id, position, platform) VALUES (?, ?, ?)",
                    (campaign_id, position, platform),
                )
        account_uids_by_external_id: dict[str, str] = {}
        for account in source_rows.get("CreatorAccounts", []):
            account_uid = str(account.get("account_uid") or "").strip()
            account_id = str(account.get("account_id") or "").strip()
            if account_uid:
                account_uids_by_external_id[account_uid] = account_uid
            if account_id:
                account_uids_by_external_id[account_id] = account_uid
        for relation in source_rows.get("CampaignCreators", []):
            relation_id = str(relation.get("id") or "").strip()
            account_ids = _list_value(
                relation.get("account_ids"), fallback=relation.get("account_id")
            )
            for position, account_id in enumerate(account_ids):
                account_uid = account_uids_by_external_id.get(account_id)
                if not account_uid:
                    raise SQLiteMigrationError(
                        "Campaign execution account identity is invalid."
                    )
                connection.execute(
                    "INSERT INTO campaign_creator_accounts(campaign_creator_id, position, account_uid) VALUES (?, ?, ?)",
                    (relation_id, position, account_uid),
                )
            dates = _list_value(
                relation.get("planned_publish_dates"), fallback=relation.get("publish_date")
            )
            for position, planned_date in enumerate(dates):
                try:
                    date.fromisoformat(planned_date)
                except ValueError as exc:
                    raise SQLiteMigrationError("Malformed planned publish date.") from exc
                connection.execute(
                    "INSERT INTO campaign_creator_planned_dates(campaign_creator_id, position, planned_date) VALUES (?, ?, ?)",
                    (relation_id, position, planned_date),
                )
            links = _list_value(relation.get("publish_links"), split_plain=True)
            for position, link in enumerate(links):
                connection.execute(
                    "INSERT INTO campaign_creator_publish_links(campaign_creator_id, position, publish_link) VALUES (?, ?, ?)",
                    (relation_id, position, link),
                )

    @staticmethod
    def _relation_counts(connection) -> dict[str, int]:
        names = (
            "creator_tags", "campaign_platforms", "campaign_creator_accounts",
            "campaign_creator_planned_dates", "campaign_creator_publish_links",
        )
        return {
            name: int(connection.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0])
            for name in names
        }

    @staticmethod
    def _import_workbook_metadata(connection, workbook) -> None:
        _headers, rows = _sheet_rows(workbook, "_Metadata")
        if rows:
            row = rows[0]
            for source_key, target_key in (
                ("schema_version", "source_workbook_schema_version"),
                ("last_update_time", "source_workbook_last_update_time"),
            ):
                value = _safe_text(row.get(source_key))
                if value is not None:
                    connection.execute(
                        "INSERT OR REPLACE INTO storage_metadata(key, value) VALUES (?, ?)",
                        (target_key, value),
                    )

    @staticmethod
    def _validate_expected_counts(connection, counts: dict[str, int]) -> None:
        for table, expected in counts.items():
            actual = int(connection.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
            if actual != expected:
                raise SQLiteIntegrityError("Migrated table count mismatch.")

    def activate_synthetic(
        self,
        result: MigrationResult,
        *,
        inject_after_database_activation: bool = False,
    ) -> Path:
        production_root = Path(self.production_root_provider()).resolve()
        if self.paths.app_data_dir.resolve() == production_root:
            raise SQLiteActivationError("Production SQLite activation is disabled in Batch 1.")
        manifest, _source = load_json_with_backup(result.manifest_path)
        if not isinstance(manifest, dict) or manifest.get("phase") != "ready_for_activation":
            raise SQLiteActivationError("Staged migration is not ready for activation.")
        final_path = self.paths.database_path
        if final_path.exists():
            raise SQLiteActivationError("Synthetic target database already exists.")
        os.replace(result.staged_database_path, final_path)
        manifest["activation_state"] = "database_activated"
        self._write_manifest(result.manifest_path, manifest, "database_activated")
        if inject_after_database_activation:
            raise SQLiteActivationError("Injected activation interruption.")
        self._activate_marker(result.migration_id, final_path, result.manifest_path, manifest)
        return final_path

    def activate_production(
        self,
        result: MigrationResult,
        *,
        source_workbook: Path,
        authorization: ProductionActivationAuthorization,
    ) -> Path:
        """Publish a prepared production database after explicit authorization."""
        self._require_production_root()
        if authorization.migration_id != result.migration_id:
            raise SQLiteActivationError("Production activation authorization mismatch.")
        if authorization.source_sha256 != result.source_sha256_before:
            raise SQLiteActivationError("Production activation source authorization mismatch.")
        if resolve_authority(self.paths) != "legacy_excel":
            raise SQLiteActivationError("Production authority is not legacy Excel.")
        manifest = self._ready_manifest(result)
        source_workbook = Path(source_workbook)
        if not source_workbook.is_file() or _sha256(source_workbook) != result.source_sha256_before:
            raise SQLiteActivationError("SQLITE_MIGRATION_SOURCE_CHANGED")
        self._validate_staged_database(result.staged_database_path)
        final_path = self.paths.database_path
        if final_path.exists():
            raise SQLiteActivationError("Production target database already exists.")
        manifest["activation_state"] = "authorized"
        manifest["confirmed_at"] = authorization.confirmed_at
        self._write_manifest(result.manifest_path, manifest, "activation_authorized")
        os.replace(result.staged_database_path, final_path)
        manifest["activation_state"] = "database_activated"
        self._write_manifest(result.manifest_path, manifest, "database_activated")
        self._activate_marker(result.migration_id, final_path, result.manifest_path, manifest)
        if resolve_authority(self.paths) != "sqlite_active":
            raise SQLiteActivationError("Production SQLite authority verification failed.")
        return final_path

    def recover_production_activation(self, migration_id: str) -> Path:
        """Complete only an activation whose database publication commit point passed."""
        self._require_production_root()
        manifest_path = self.paths.migration_manifest_path(migration_id)
        manifest, _source = load_json_with_backup(manifest_path)
        if not isinstance(manifest, dict) or str(manifest.get("migration_id")) != migration_id:
            raise SQLiteActivationError("Interrupted production activation cannot be proven.")
        phase = str(manifest.get("phase") or "")
        if phase == "completed" and resolve_authority(self.paths) == "sqlite_active":
            return self.paths.database_path
        if phase == "activation_authorized":
            staged_path = self.paths.staged_database_path(migration_id)
            if self.paths.database_path.is_file() and not staged_path.exists():
                self._validate_staged_database(self.paths.database_path)
                manifest["activation_state"] = "database_activated"
                self._write_manifest(manifest_path, manifest, "database_activated")
                phase = "database_activated"
            elif staged_path.is_file() and not self.paths.database_path.exists():
                manifest["activation_state"] = "inactive"
                self._write_manifest(manifest_path, manifest, "ready_for_activation")
                raise SQLiteActivationError("Production activation did not cross commit point.")
            else:
                raise SQLiteActivationError("Interrupted production activation is ambiguous.")
        if phase not in {"database_activated", "authority_activated"}:
            raise SQLiteActivationError("Interrupted production activation cannot be proven.")
        self._validate_staged_database(self.paths.database_path)
        if phase == "database_activated":
            self._activate_marker(migration_id, self.paths.database_path, manifest_path, manifest)
        else:
            self._write_manifest(manifest_path, manifest, "completed")
        if resolve_authority(self.paths) != "sqlite_active":
            raise SQLiteActivationError("Recovered production authority is invalid.")
        return self.paths.database_path

    def _require_production_root(self) -> None:
        canonical = Path(self.production_root_provider()).resolve()
        if self.paths.app_data_dir.resolve() != canonical:
            raise SQLiteActivationError("Production activation requires the canonical app data root.")

    def _ready_manifest(self, result: MigrationResult) -> dict[str, object]:
        manifest, _source = load_json_with_backup(result.manifest_path)
        if (
            not isinstance(manifest, dict)
            or manifest.get("phase") != "ready_for_activation"
            or str(manifest.get("migration_id") or "") != result.migration_id
            or str(manifest.get("source_sha256") or "") != result.source_sha256_before
        ):
            raise SQLiteActivationError("Staged production migration is not ready.")
        return manifest

    @staticmethod
    def _validate_staged_database(path: Path) -> None:
        if not Path(path).is_file():
            raise SQLiteActivationError("Staged production database is missing.")
        try:
            factory = SQLiteConnectionFactory(Path(path))
            with factory.read_connection() as connection:
                if schema_version(connection) > CURRENT_SCHEMA_VERSION:
                    raise SQLiteSchemaUnsupportedError("Staged schema is newer than supported.")
                validate_schema(connection)
        except SQLiteSchemaUnsupportedError:
            raise
        except Exception as exc:
            raise SQLiteActivationError("Staged production database is invalid.") from exc

    def recover_synthetic_activation(self, migration_id: str) -> Path:
        manifest_path = self.paths.migration_manifest_path(migration_id)
        manifest, _source = load_json_with_backup(manifest_path)
        if not isinstance(manifest, dict) or manifest.get("phase") != "database_activated":
            raise SQLiteActivationError("Interrupted activation cannot be proven.")
        final_path = self.paths.database_path
        factory = SQLiteConnectionFactory(final_path)
        with factory.read_connection() as connection:
            validate_schema(connection)
        self._activate_marker(migration_id, final_path, manifest_path, manifest)
        return final_path

    def _activate_marker(self, migration_id: str, final_path: Path, manifest_path: Path, manifest):
        marker = {
            "authority": "sqlite",
            "database_name": final_path.name,
            "migration_id": migration_id,
            "schema_version": CURRENT_SCHEMA_VERSION,
            "activated_at": _utc_now(),
        }
        atomic_write_json(self.paths.authority_marker_path, marker)
        self._inject("authority_marker_written")
        manifest["activation_state"] = "authority_activated"
        self._write_manifest(manifest_path, manifest, "authority_activated")
        self._write_manifest(manifest_path, manifest, "completed")


def resolve_authority(paths: SQLiteStoragePaths) -> str:
    marker, _source = load_json_with_backup(paths.authority_marker_path)
    if marker is None:
        return "legacy_excel"
    if not isinstance(marker, dict):
        return "migration_error"
    authority = str(marker.get("authority") or "")
    if authority == "legacy_excel":
        return "legacy_excel"
    if authority != "sqlite":
        return "migration_error"
    if int(marker.get("schema_version") or 0) > CURRENT_SCHEMA_VERSION:
        return "unsupported_schema"
    if not paths.database_path.is_file():
        return "migration_error"
    try:
        factory = SQLiteConnectionFactory(paths.database_path)
        with factory.read_connection() as connection:
            if schema_version(connection) > CURRENT_SCHEMA_VERSION:
                return "unsupported_schema"
            validate_schema(connection)
    except SQLiteSchemaUnsupportedError:
        return "unsupported_schema"
    except Exception:
        return "migration_error"
    return "sqlite_active"


def semantic_projection(connection) -> dict[str, object]:
    creators = []
    for row in connection.execute("SELECT * FROM creators ORDER BY creator_id"):
        creator = dict(row)
        creator["tags"] = [
            item[0]
            for item in connection.execute(
                "SELECT tag FROM creator_tags WHERE creator_id=? ORDER BY position",
                (row["creator_id"],),
            )
        ]
        creator["accounts"] = [
            dict(item)
            for item in connection.execute(
                "SELECT * FROM creator_accounts WHERE creator_id=? ORDER BY account_uid",
                (row["creator_id"],),
            )
        ]
        creators.append(creator)
    campaigns = []
    for row in connection.execute("SELECT * FROM campaigns ORDER BY campaign_id"):
        campaign = dict(row)
        campaign["platforms"] = [
            item[0]
            for item in connection.execute(
                "SELECT platform FROM campaign_platforms WHERE campaign_id=? ORDER BY position",
                (row["campaign_id"],),
            )
        ]
        campaign["members"] = []
        for member in connection.execute(
            "SELECT * FROM campaign_creators WHERE campaign_id=? ORDER BY id", (row["campaign_id"],)
        ):
            value = dict(member)
            value["account_ids"] = [
                item[0]
                for item in connection.execute(
                    "SELECT account_uid FROM campaign_creator_accounts WHERE campaign_creator_id=? ORDER BY position",
                    (member["id"],),
                )
            ]
            value["planned_publish_dates"] = [
                item[0]
                for item in connection.execute(
                    "SELECT planned_date FROM campaign_creator_planned_dates WHERE campaign_creator_id=? ORDER BY position",
                    (member["id"],),
                )
            ]
            value["publish_links"] = [
                item[0]
                for item in connection.execute(
                    "SELECT publish_link FROM campaign_creator_publish_links WHERE campaign_creator_id=? ORDER BY position",
                    (member["id"],),
                )
            ]
            campaign["members"].append(value)
        campaigns.append(campaign)
    return {"creators": creators, "campaigns": campaigns}


def source_semantic_projection(workbook) -> dict[str, object]:
    _headers, creator_rows = _sheet_rows(workbook, "Creators")
    _headers, account_rows = _sheet_rows(workbook, "CreatorAccounts")
    _headers, campaign_rows = _sheet_rows(workbook, "Campaigns")
    _headers, member_rows = _sheet_rows(workbook, "CampaignCreators")
    account_uids_by_external_id: dict[str, str] = {}
    for account in account_rows:
        account_uid = str(account.get("account_uid") or "").strip()
        account_id = str(account.get("account_id") or "").strip()
        if account_uid:
            account_uids_by_external_id[account_uid] = account_uid
        if account_id and account_uid:
            account_uids_by_external_id[account_id] = account_uid
    creators = []
    for row in sorted(creator_rows, key=lambda item: str(item.get("creator_id") or "")):
        creator = {
            field: _value(field, row.get(field))
            for source, table, _identity, columns in BASE_TABLES
            if source == "Creators" and table == "creators"
            for field in columns
        }
        creator["tags"] = _list_value(row.get("tags"), split_plain=True)
        creator_id = str(row.get("creator_id") or "").strip()
        creator["accounts"] = [
            {
                field: _value(field, account.get(field))
                for source, table, _identity, columns in BASE_TABLES
                if source == "CreatorAccounts" and table == "creator_accounts"
                for field in columns
            }
            for account in sorted(
                (item for item in account_rows if str(item.get("creator_id") or "").strip() == creator_id),
                key=lambda item: str(item.get("account_uid") or ""),
            )
        ]
        creators.append(creator)
    campaigns = []
    for row in sorted(campaign_rows, key=lambda item: str(item.get("campaign_id") or "")):
        campaign = {
            field: _value(field, row.get(field))
            for source, table, _identity, columns in BASE_TABLES
            if source == "Campaigns" and table == "campaigns"
            for field in columns
        }
        campaign["platforms"] = _list_value(
            row.get("platforms"), fallback=row.get("platform"), split_plain=True
        )
        campaign_id = str(row.get("campaign_id") or "").strip()
        campaign["members"] = []
        for member in sorted(
            (item for item in member_rows if str(item.get("campaign_id") or "").strip() == campaign_id),
            key=lambda item: str(item.get("id") or ""),
        ):
            value = {
                field: _value(field, member.get(field))
                for source, table, _identity, columns in BASE_TABLES
                if source == "CampaignCreators" and table == "campaign_creators"
                for field in columns
            }
            value["account_ids"] = [
                account_uids_by_external_id.get(account_id, account_id)
                for account_id in _list_value(
                    member.get("account_ids"), fallback=member.get("account_id")
                )
            ]
            value["planned_publish_dates"] = _list_value(
                member.get("planned_publish_dates"), fallback=member.get("publish_date")
            )
            value["publish_links"] = _list_value(member.get("publish_links"), split_plain=True)
            campaign["members"].append(value)
        campaigns.append(campaign)
    return {"creators": creators, "campaigns": campaigns}


def _projection_digest(projection: dict[str, object]) -> str:
    serialized = json.dumps(projection, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def semantic_digest(connection) -> str:
    return _projection_digest(semantic_projection(connection))


def source_semantic_digest(workbook) -> str:
    return _projection_digest(source_semantic_projection(workbook))

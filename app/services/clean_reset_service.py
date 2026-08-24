from __future__ import annotations

"""Controlled reset of local business data while preserving app configuration."""

import hashlib
import json
import os
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from creator_repository import _WORKBOOK_SHEETS
from excel_workbook_store import ExcelWorkbookStore, WorkbookReadError
from local_storage_lock import shared_storage_lock
from runtime_paths import atomic_write_json, json_backup_path


BUSINESS_SHEETS = tuple(name for name in _WORKBOOK_SHEETS if name != "_Metadata")
PRESERVED_SHEETS = ("_Metadata",)


class CleanResetError(RuntimeError):
    pass


class CleanResetService:
    def __init__(
        self,
        workbook_path: Path,
        *,
        settings_path: Path,
        data_protection_path: Path,
        mail_messages_path: Path,
        tasks_dir: Path,
        cache_invalidators: tuple[Callable[[], None], ...] = (),
        now_provider: Callable[[], datetime] | None = None,
    ) -> None:
        self.workbook_path = Path(workbook_path)
        self.settings_path = Path(settings_path)
        self.data_protection_path = Path(data_protection_path)
        self.mail_messages_path = Path(mail_messages_path)
        self.tasks_dir = Path(tasks_dir)
        self.cache_invalidators = cache_invalidators
        self.now_provider = now_provider or (lambda: datetime.now(timezone.utc))

    def preview(self) -> dict:
        store = ExcelWorkbookStore(self.workbook_path)
        try:
            with store.read_only_workbook() as workbook:
                actual_sheets = set(workbook.sheetnames)
                expected_sheets = set(_WORKBOOK_SHEETS)
                unknown = sorted(actual_sheets - expected_sheets)
                missing = sorted(expected_sheets - actual_sheets)
                header_errors = self._header_errors(workbook)
                rows = {
                    name: self._data_row_count(workbook[name])
                    for name in BUSINESS_SHEETS
                    if name in workbook.sheetnames
                }
                preserved = {
                    name: self._data_row_count(workbook[name])
                    for name in PRESERVED_SHEETS
                    if name in workbook.sheetnames
                }
        except WorkbookReadError as exc:
            raise CleanResetError(str(exc)) from exc

        review_items = [
            *[f"UNKNOWN_SHEET:{name}" for name in unknown],
            *[f"MISSING_SHEET:{name}" for name in missing],
            *header_errors,
        ]
        return {
            "status": "blocked" if review_items else "success",
            "workbook": self.workbook_path.name,
            "backup_required": True,
            "clear_sheets": rows,
            "preserve_sheets": preserved,
            "external_business_data": {
                "data_protection_entries": self._json_entry_count(self.data_protection_path),
                "mail_messages": self._mail_message_count(),
                "task_files": self._task_file_count(),
            },
            "preserved_configuration": {
                "app_settings": True,
                "chrome_profiles": True,
                "email_accounts": True,
                "feishu": True,
                "schema": not missing and not header_errors,
            },
            "review_items": review_items,
            "summary": {
                "creators": rows.get("Creators", 0),
                "accounts": rows.get("CreatorAccounts", 0),
                "campaigns": rows.get("Campaigns", 0),
                "snapshots": rows.get("CreatorSnapshots", 0)
                + rows.get("VideoSnapshots", 0),
                "videos": rows.get("Videos", 0),
            },
        }

    def execute(self, *, confirm: object) -> dict:
        if confirm is not True:
            raise ValueError("CLEAN_RESET_CONFIRMATION_REQUIRED")

        with shared_storage_lock():
            preview = self.preview()
            if preview["status"] != "success":
                raise CleanResetError("CLEAN_RESET_SCHEMA_REVIEW_REQUIRED")

            timestamp = self.now_provider().astimezone(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
            backup_dir = self.workbook_path.parent / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)
            workbook_backup = self._unique_path(
                backup_dir / f"{self.workbook_path.stem}_before_clean_reset_{timestamp}{self.workbook_path.suffix}"
            )
            support_backup_dir = self._unique_path(backup_dir / f"clean_reset_{timestamp}")
            support_backup_dir.mkdir(parents=True, exist_ok=False)

            settings_digest = self._file_digest(self.settings_path)
            json_backups: dict[Path, Path | None] = {}
            tasks_backup: Path | None = None
            workbook_saved = False
            try:
                self._copy_and_validate_workbook(workbook_backup)
                for source in (self.data_protection_path, self.mail_messages_path):
                    destination = support_backup_dir / source.name
                    if source.is_file():
                        shutil.copy2(source, destination)
                        json_backups[source] = destination
                    else:
                        json_backups[source] = None

                if self._task_file_count():
                    tasks_backup = support_backup_dir / "tasks"
                    self.tasks_dir.replace(tasks_backup)
                    self.tasks_dir.mkdir(parents=True, exist_ok=True)

                workbook = load_workbook(self.workbook_path)
                try:
                    for name in BUSINESS_SHEETS:
                        sheet = workbook[name]
                        if sheet.max_row > 1:
                            sheet.delete_rows(2, sheet.max_row - 1)
                    ExcelWorkbookStore(self.workbook_path).save(workbook)
                    workbook_saved = True
                finally:
                    workbook.close()

                atomic_write_json(self.data_protection_path, {})
                atomic_write_json(
                    self.mail_messages_path,
                    {"version": 1, "updated_at": "", "accounts": {}, "messages": []},
                )
                self._refresh_json_backup(self.data_protection_path)
                self._refresh_json_backup(self.mail_messages_path)
                shutil.copy2(self.workbook_path, self.workbook_path.with_suffix(".xlsx.bak"))

                verified = self.preview()
                remaining = sum(verified["clear_sheets"].values())
                if verified["status"] != "success" or remaining:
                    raise CleanResetError("CLEAN_RESET_VERIFICATION_FAILED")
                if self._file_digest(self.settings_path) != settings_digest:
                    raise CleanResetError("CLEAN_RESET_SETTINGS_CHANGED")
            except Exception:
                if workbook_saved or workbook_backup.is_file():
                    ExcelWorkbookStore(self.workbook_path).restore_transaction_backup(workbook_backup)
                for target, backup in json_backups.items():
                    if backup and backup.is_file():
                        self._restore_file_backup(target, backup)
                    elif target.exists():
                        target.unlink()
                if tasks_backup and tasks_backup.exists():
                    if self.tasks_dir.exists():
                        self.tasks_dir.rmdir()
                    tasks_backup.replace(self.tasks_dir)
                raise

            for invalidate in self.cache_invalidators:
                invalidate()
            return {
                "status": "success",
                "backup": {
                    "path": str(workbook_backup),
                    "filename": workbook_backup.name,
                    "size": workbook_backup.stat().st_size,
                    "sha256": self._file_digest(workbook_backup),
                },
                "cleared": preview["clear_sheets"],
                "cleared_external": preview["external_business_data"],
                "preserved_configuration": preview["preserved_configuration"],
                "after": verified["clear_sheets"],
            }

    @staticmethod
    def _data_row_count(sheet) -> int:
        return sum(
            1
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(value not in (None, "") for value in row)
        )

    @staticmethod
    def _header_errors(workbook) -> list[str]:
        errors: list[str] = []
        for name, expected in _WORKBOOK_SHEETS.items():
            if name not in workbook.sheetnames:
                continue
            actual = [str(cell.value or "").strip() for cell in workbook[name][1]]
            if actual[: len(expected)] != list(expected):
                errors.append(f"HEADER_MISMATCH:{name}")
        return errors

    @staticmethod
    def _file_digest(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest() if path.is_file() else ""

    @staticmethod
    def _json_entry_count(path: Path) -> int:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return 0
        return len(data) if isinstance(data, dict) else 0

    def _mail_message_count(self) -> int:
        try:
            data = json.loads(self.mail_messages_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return 0
        messages = data.get("messages") if isinstance(data, dict) else None
        return len(messages) if isinstance(messages, list) else 0

    def _task_file_count(self) -> int:
        if not self.tasks_dir.is_dir():
            return 0
        return sum(path.is_file() for path in self.tasks_dir.rglob("*"))

    @staticmethod
    def _unique_path(path: Path) -> Path:
        if not path.exists():
            return path
        for index in range(1, 1000):
            candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
            if not candidate.exists():
                return candidate
        raise CleanResetError("CLEAN_RESET_BACKUP_NAME_EXHAUSTED")

    def _copy_and_validate_workbook(self, destination: Path) -> None:
        if not self.workbook_path.is_file():
            raise CleanResetError("CLEAN_RESET_WORKBOOK_MISSING")
        shutil.copy2(self.workbook_path, destination)
        try:
            load_workbook(destination, read_only=True).close()
        except Exception as exc:
            destination.unlink(missing_ok=True)
            raise CleanResetError("CLEAN_RESET_BACKUP_INVALID") from exc

    @staticmethod
    def _restore_file_backup(target: Path, backup: Path) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        fd, raw_temp_path = tempfile.mkstemp(
            prefix=f"{target.name}.restore.", suffix=".tmp", dir=target.parent
        )
        os.close(fd)
        temp_path = Path(raw_temp_path)
        try:
            shutil.copy2(backup, temp_path)
            os.replace(temp_path, target)
        finally:
            temp_path.unlink(missing_ok=True)

    @staticmethod
    def _refresh_json_backup(path: Path) -> None:
        if path.is_file():
            shutil.copy2(path, json_backup_path(path))

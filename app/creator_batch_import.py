from __future__ import annotations

"""Pure XLSX parsing and response-safe errors for Creator batch import."""

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook


TEMPLATE_HEADERS = (
    "platform",
    "profile_url",
    "name",
    "country",
    "language",
    "content_category",
    "email",
    "whatsapp",
    "agency_id",
    "bio",
)
REQUIRED_HEADERS = frozenset({"platform", "profile_url"})


@dataclass(frozen=True)
class ParsedCreatorImportRow:
    excel_row: int
    values: dict[str, str]


class CreatorBatchImportError(ValueError):
    def __init__(
        self,
        code: str,
        *,
        summary: dict[str, int] | None = None,
        rows: list[dict[str, Any]] | None = None,
    ) -> None:
        super().__init__(code)
        self.code = code
        self.summary = dict(summary or {})
        self.rows = [dict(row) for row in rows or []]

    def to_response(self) -> dict[str, Any]:
        response: dict[str, Any] = {"ok": False, "error": self.code}
        if self.summary:
            response["summary"] = dict(self.summary)
        if self.rows:
            response["rows"] = [
                {
                    key: row[key]
                    for key in ("row", "status", "code", "field")
                    if key in row
                }
                for row in self.rows
            ]
        return response


def build_creator_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"
    sheet.append(list(TEMPLATE_HEADERS))
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def parse_creator_import_workbook(payload: bytes) -> list[ParsedCreatorImportRow]:
    if not isinstance(payload, (bytes, bytearray)) or not payload:
        raise CreatorBatchImportError("INVALID_FILE")
    try:
        workbook = load_workbook(io.BytesIO(bytes(payload)), read_only=True, data_only=True)
    except Exception as exc:
        raise CreatorBatchImportError("INVALID_FILE") from exc

    try:
        sheet = workbook.active
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value or "").strip() for value in raw_headers]
        missing = REQUIRED_HEADERS - set(headers)
        if missing:
            raise CreatorBatchImportError("MISSING_REQUIRED_COLUMN")

        recognized_columns = {
            index: header
            for index, header in enumerate(headers)
            if header in TEMPLATE_HEADERS
        }
        rows: list[ParsedCreatorImportRow] = []
        for excel_row, raw_row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            values = {
                header: _cell_text(raw_row[index] if index < len(raw_row) else None)
                for index, header in recognized_columns.items()
            }
            if not any(values.values()):
                continue
            rows.append(ParsedCreatorImportRow(excel_row=excel_row, values=values))
        if not rows:
            raise CreatorBatchImportError("EMPTY_IMPORT")
        return rows
    finally:
        workbook.close()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

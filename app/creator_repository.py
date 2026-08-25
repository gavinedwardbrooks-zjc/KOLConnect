from __future__ import annotations

"""Creator Library storage contract backed by a cloud-syncable Excel workbook.

The public methods are intentionally storage-neutral. A future Supabase or
PostgreSQL adapter can keep this contract without changing the HTTP API or UI.
"""

import json
import hashlib
import os  # Compatibility patch point for existing atomic-save tests.
import re
import time
import uuid
from functools import wraps
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from domain.normalization import normalize_country, normalize_followers, normalize_tags

from openpyxl.styles import Font

from app_logging import log_event
from campaign_creator_repository import CAMPAIGN_CREATORS_HEADERS
from campaign_repository import CAMPAIGNS_HEADERS, migrate_legacy_campaign_archives
from ports.agency_port import AgencyPort
from repositories.agency_repository import AGENCIES_HEADERS, AGENCY_CONTACTS_HEADERS
from excel_workbook_store import (
    ExcelWorkbookStore,
    WORKBOOK_LOCK,
    WorkbookReadError,
    WorkbookSaveError,
)
from product_repository import PRODUCTS_HEADERS
from runtime_paths import load_json_with_backup
from local_storage_lock import shared_storage_lock


CREATOR_LIBRARY_STATUSES = {
    "discovered",
    "contacted",
    "negotiating",
    "cooperating",
    "completed",
    "rejected",
}

CREATOR_LIBRARY_PAGE_SIZES = frozenset({12, 24, 25, 48, 50, 100})
CREATOR_LIBRARY_SORT_FIELDS = frozenset({
    "created_at", "updated_at", "creator_name", "followers", "platform",
})
CREATOR_LIBRARY_FILTER_FIELDS = frozenset({
    "search", "platform", "country", "language", "content_category",
    "agency_id", "tag", "ai_tag", "followers_min", "followers_max",
    "insight_level", "status",
})
_PLATFORM_SORT_ORDER = {"instagram": 0, "tiktok": 1, "youtube": 2}

_TASK_ID_PATTERN = re.compile(r"^task_[0-9]{8}T[0-9]{6}Z_[0-9a-f]{8}$")
CREATOR_LIBRARY_SCHEMA_VERSION = "2.0-product-campaign-phase2-api"
_CREATORS_HEADERS = [
    "creator_id", "name", "platform", "profile_url", "country", "language",
    "content_category", "followers", "insight_level", "status", "created_at", "tags", "updated_at",
    "email", "whatsapp", "cooperation_stage", "recent_product", "quote", "owner",
    "last_contact_time", "next_follow_up_time", "note", "agency_id",
    "current_contact_id", "source_contact_id",
    "bio", "archived_at",
]
_CREATOR_ACCOUNTS_HEADERS = [
    "account_id", "creator_id", "account_uid", "platform", "username", "profile_url",
    "followers", "account_email", "latest_post_date", "last_scrape_time", "data_source",
    "scrape_status", "platform_account_id", "attribution_status", "note", "source_task_id",
    "created_at", "updated_at",
]
_VIDEOS_HEADERS = ["creator_id", "video_url", "views", "likes", "comments", "captured_at"]
_INSIGHTS_HEADERS = ["creator_id", "average_views", "median_views", "stability", "risks", "recommendation"]
_CREATOR_SNAPSHOTS_HEADERS = [
    "snapshot_id", "creator_id", "platform", "account_uid", "followers", "average_views",
    "median_views", "video_count", "creator_score", "insight_level", "captured_at", "source",
]
_VIDEO_SNAPSHOTS_HEADERS = [
    "video_snapshot_id", "snapshot_id", "creator_id", "video_id", "video_url", "platform",
    "views", "likes", "comments", "captured_at",
]
_COOPERATIONS_HEADERS = [
    "cooperation_id", "creator_id", "campaign", "platform", "contact_date", "price",
    "published_count", "total_views", "average_views", "roi", "result", "note", "created_at",
]
# Legacy cooperation records remain read/write compatible until the UI and
# Dashboard move to CampaignCreators. Phase 1 does not migrate or delete them.
_AGENCIES_HEADERS = AGENCIES_HEADERS
_AGENCY_CONTACTS_HEADERS = AGENCY_CONTACTS_HEADERS
_FOLLOW_UP_LOGS_HEADERS = [
    "follow_up_id", "object_type", "object_id", "contact_method", "content",
    "stage_before", "stage_after", "contacted_at", "next_follow_up_time", "owner",
    "created_at",
]
# Hidden technical metadata preserves the full snapshot and review-task handoff.
_ANALYSIS_METADATA_HEADERS = ["creator_id", "task_id", "account_uid", "status_updated_at", "analysis_json", "source"]
_WORKBOOK_METADATA_HEADERS = ["schema_version", "last_update_time"]
_WORKBOOK_SHEETS = {
    "Creators": _CREATORS_HEADERS,
    "CreatorAccounts": _CREATOR_ACCOUNTS_HEADERS,
    "Videos": _VIDEOS_HEADERS,
    "Insights": _INSIGHTS_HEADERS,
    "CreatorSnapshots": _CREATOR_SNAPSHOTS_HEADERS,
    "VideoSnapshots": _VIDEO_SNAPSHOTS_HEADERS,
    "Cooperations": _COOPERATIONS_HEADERS,
    "Agencies": _AGENCIES_HEADERS,
    "AgencyContacts": _AGENCY_CONTACTS_HEADERS,
    "FollowUpLogs": _FOLLOW_UP_LOGS_HEADERS,
    "Products": PRODUCTS_HEADERS,
    "Campaigns": CAMPAIGNS_HEADERS,
    "CampaignCreators": CAMPAIGN_CREATORS_HEADERS,
    "_AnalysisData": _ANALYSIS_METADATA_HEADERS,
    "_Metadata": _WORKBOOK_METADATA_HEADERS,
}
_WORKBOOK_LOCK = WORKBOOK_LOCK
_CAMPAIGN_LIFECYCLE_LOGGED_PATHS: set[str] = set()


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _synchronized(method):
    """Serialize workbook access within the local desktop process."""
    @wraps(method)
    def wrapped(self, *args, **kwargs):
        with _WORKBOOK_LOCK:
            return method(self, *args, **kwargs)
    return wrapped


def _mutation_synchronized(method):
    """Serialize one workbook mutation across threads and app processes."""
    @wraps(method)
    def wrapped(self, *args, **kwargs):
        with shared_storage_lock(), _WORKBOOK_LOCK:
            return method(self, *args, **kwargs)
    return wrapped


class CreatorRepository:
    """Excel adapter for Creator Library data, suitable for WPS/OneDrive cloud folders."""

    def __init__(
        self,
        workbook_path: Path | ExcelWorkbookStore,
        legacy_analysis_dir: Path | None = None,
        legacy_library_file: Path | None = None,
    ) -> None:
        self.store = (
            workbook_path
            if isinstance(workbook_path, ExcelWorkbookStore)
            else ExcelWorkbookStore(workbook_path)
        )
        self.workbook_path = self.store.workbook_path
        self.legacy_analysis_dir = legacy_analysis_dir
        self.legacy_library_file = legacy_library_file
        self.last_migration_report: dict[str, Any] | None = None
        self.last_campaign_lifecycle_report: dict[str, Any] | None = None
        self.store.register_migration(self._apply_schema_migrations)
        self.store.register_before_save(self._prepare_workbook_save)

    @_mutation_synchronized
    def saveCreator(self, analysis: dict[str, Any]) -> dict[str, Any]:
        """Save the latest creator view and append an immutable analysis snapshot."""
        self._validate_analysis(analysis)
        workbook = self._load_workbook()
        requested_creator_id = str(analysis["analysis_id"])
        account_uid = str(analysis.get("account_uid") or "").strip()
        preferred_creator_id = str(analysis.get("creator_id") or "").strip()
        creator_id = (
            self._creator_id_for_account_uid(workbook, account_uid)
            or (preferred_creator_id if self._creator_row(workbook["Creators"], preferred_creator_id) else "")
            or requested_creator_id
        )
        existing = self._creator_row(workbook["Creators"], creator_id)
        is_new_creator = not bool(existing)
        status = str(existing.get("status") or "discovered") if existing else "discovered"
        existing_metadata = self._metadata_row(workbook["_AnalysisData"], creator_id)
        existing_analysis = self._decode_analysis(existing_metadata.get("analysis_json"))
        existing_crm = (
            existing_analysis.get("_crm")
            if isinstance(existing_analysis.get("_crm"), dict)
            else {}
        )
        status_updated_at = existing_metadata.get("status_updated_at", "")
        creator_values = self._creator_values(analysis, status, creator_id)
        if existing:
            # Extension imports should never erase optional data already curated in Excel.
            for field in (
                "country", "language", "tags", "email", "whatsapp", "cooperation_stage",
                "recent_product", "quote", "owner", "last_contact_time",
                "next_follow_up_time", "note", "agency_id", "current_contact_id",
                "source_contact_id", "bio", "archived_at",
            ):
                existing_value = existing.get(field) or existing_crm.get(field)
                if not creator_values.get(field) and existing_value:
                    creator_values[field] = existing_value
            creator_values["created_at"] = existing.get("created_at") or creator_values["created_at"]
            if not self._account_row_by_uid(workbook["CreatorAccounts"], account_uid):
                for field in ("platform", "profile_url", "followers"):
                    creator_values[field] = existing.get(field) or creator_values.get(field)
        self._upsert_row(workbook["Creators"], "creator_id", creator_id, creator_values)
        account = self._upsert_account_from_analysis(workbook, analysis, creator_id)
        self._replace_video_rows(workbook["Videos"], creator_id, analysis.get("videos"))
        self._upsert_row(
            workbook["Insights"],
            "creator_id",
            creator_id,
            self._insight_values(analysis, creator_id),
        )
        stored_analysis = dict(analysis)
        if existing_crm:
            stored_analysis["_crm"] = dict(existing_crm)
        self._upsert_row(
            workbook["_AnalysisData"],
            "creator_id",
            creator_id,
            {
                "creator_id": creator_id,
                "task_id": str(analysis.get("task_id") or ""),
                "account_uid": str(analysis.get("account_uid") or ""),
                "source": str(analysis.get("source") or ""),
                "status_updated_at": status_updated_at,
                "analysis_json": json.dumps(stored_analysis, ensure_ascii=False),
            },
        )
        snapshot = self.createSnapshot(analysis, creator_id, workbook)
        self._save_workbook(workbook)
        return {
            **analysis,
            "creator_id": creator_id,
            "account_id": account["account_id"],
            "snapshot_id": snapshot["snapshot_id"],
            "is_new_creator": is_new_creator,
        }

    @_mutation_synchronized
    def createSnapshot(self, analysis: dict[str, Any], creator_id: str, workbook=None) -> dict[str, Any]:
        """Append one time-stamped analysis without replacing earlier snapshots."""
        should_save = workbook is None
        workbook = workbook or self._load_workbook()
        snapshot_id = f"snapshot_{analysis['task_id']}"
        creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        metrics = analysis.get("video_analysis") if isinstance(analysis.get("video_analysis"), dict) else {}
        insight = analysis.get("creator_insight") if isinstance(analysis.get("creator_insight"), dict) else {}
        captured_at = str(analysis.get("imported_at") or _utc_now())
        snapshot = {
            "snapshot_id": snapshot_id,
            "creator_id": creator_id,
            "platform": str(creator.get("platform") or ""),
            "account_uid": str(analysis.get("account_uid") or ""),
            "followers": str(creator.get("followers") or ""),
            "average_views": metrics.get("average_views", ""),
            "median_views": metrics.get("median_views", ""),
            "video_count": len(analysis.get("videos") if isinstance(analysis.get("videos"), list) else []),
            "creator_score": insight.get("creator_score", insight.get("rule_score", "")),
            "insight_level": str(insight.get("level") or insight.get("grade") or "insufficient"),
            "captured_at": captured_at,
            "source": str(analysis.get("source") or ""),
        }
        self._upsert_row(workbook["CreatorSnapshots"], "snapshot_id", snapshot_id, snapshot)
        self._replace_video_snapshot_rows(workbook["VideoSnapshots"], snapshot_id, creator_id, analysis.get("videos"), captured_at)
        if should_save:
            self._save_workbook(workbook)
        return snapshot

    @_synchronized
    def getCreatorSnapshots(self, creator_id: str) -> list[dict[str, Any]]:
        workbook = self._load_workbook()
        return self._snapshots_from_workbook(workbook, creator_id)

    @_synchronized
    def getCreatorAccounts(self, creator_id: str = "") -> list[dict[str, Any]]:
        """Return normalized social accounts, optionally scoped to one creator."""
        workbook = self._load_workbook()
        creator_id = str(creator_id or "").strip()
        accounts = self._rows(workbook["CreatorAccounts"])
        if creator_id:
            accounts = [
                row for row in accounts
                if str(row.get("creator_id") or "") == creator_id
            ]
        accounts.sort(key=lambda row: (str(row.get("created_at") or ""), str(row.get("account_id") or "")))
        return accounts

    @_synchronized
    def getCreatorInventoryRows(self) -> dict[str, list[dict[str, Any]]]:
        """Return detached authoritative rows for identity-sensitive read operations."""
        workbook = self._load_workbook()
        return {
            "creators": self._rows(workbook["Creators"]),
            "accounts": self._rows(workbook["CreatorAccounts"]),
            "insights": self._rows(workbook["Insights"]),
            "snapshots": self._rows(workbook["CreatorSnapshots"]),
        }

    def getCreatorAccountIdentityRows(self) -> dict[str, list[dict[str, Any]]]:
        """Read only the authoritative identity sheets without migration or save hooks."""
        with shared_storage_lock(), self.store.read_only_workbook() as workbook:
            return {
                "creators": self._rows(workbook["Creators"]),
                "accounts": self._rows(workbook["CreatorAccounts"]),
            }

    @_synchronized
    def getExistingCreatorAccountUids(self, account_uids: set[str]) -> set[str]:
        """Return only requested identities that already belong to a Creator."""
        requested = {str(value or "").strip() for value in account_uids if str(value or "").strip()}
        if not requested:
            return set()
        workbook = self._load_workbook()
        creator_ids = {
            str(row.get("creator_id") or "")
            for row in self._rows(workbook["Creators"])
            if str(row.get("creator_id") or "")
        }
        return {
            str(row.get("account_uid") or "")
            for row in self._rows(workbook["CreatorAccounts"])
            if str(row.get("account_uid") or "") in requested
            and str(row.get("creator_id") or "") in creator_ids
        }

    @_mutation_synchronized
    def createCreatorsBatch(self, records: list[dict[str, Any]]) -> dict[str, int]:
        """Create normalized Creator rows in one workbook save without overwrites."""
        if not records:
            return {"created": 0, "skipped_existing": 0}
        workbook = self._load_workbook()
        creators_sheet = workbook["Creators"]
        accounts_sheet = workbook["CreatorAccounts"]
        creator_ids = {
            str(row.get("creator_id") or "")
            for row in self._rows(creators_sheet)
            if str(row.get("creator_id") or "")
        }
        accounts_by_uid = {
            str(row.get("account_uid") or ""): row
            for row in self._rows(accounts_sheet)
            if str(row.get("account_uid") or "")
        }
        now = _utc_now()
        created = 0
        skipped_existing = 0
        for record in records:
            account_uid = str(record.get("account_uid") or "").strip()
            if not account_uid:
                raise ValueError("Creator account identity is required.")
            existing_account = accounts_by_uid.get(account_uid, {})
            existing_creator_id = str(existing_account.get("creator_id") or "")
            if existing_creator_id and existing_creator_id in creator_ids:
                skipped_existing += 1
                continue

            creator_id = f"creator_{hashlib.sha256(account_uid.encode('utf-8')).hexdigest()[:16]}"
            if creator_id in creator_ids:
                skipped_existing += 1
                continue
            creator_values = {
                "creator_id": creator_id,
                "name": str(record.get("name") or ""),
                "platform": str(record.get("platform") or ""),
                "profile_url": str(record.get("profile_url") or ""),
                "country": str(record.get("country") or ""),
                "language": str(record.get("language") or ""),
                "content_category": str(record.get("content_category") or ""),
                "followers": "",
                "insight_level": "insufficient",
                "status": "discovered",
                "created_at": now,
                "tags": "",
                "updated_at": now,
                "email": str(record.get("email") or ""),
                "whatsapp": str(record.get("whatsapp") or ""),
                "agency_id": str(record.get("agency_id") or ""),
                "bio": str(record.get("bio") or ""),
                "archived_at": "",
            }
            self._append_row(creators_sheet, creator_values)
            account_values = {
                "account_id": self._account_id(account_uid),
                "creator_id": creator_id,
                "account_uid": account_uid,
                "platform": str(record.get("platform") or ""),
                "username": self._username_from_profile_url(
                    str(record.get("platform") or ""), str(record.get("profile_url") or "")
                ),
                "profile_url": str(record.get("profile_url") or ""),
                "followers": "",
                "account_email": str(record.get("email") or ""),
                "latest_post_date": "",
                "last_scrape_time": "",
                "data_source": "excel_batch_import",
                "scrape_status": "",
                "platform_account_id": "",
                "attribution_status": "confirmed",
                "note": "",
                "source_task_id": "",
                "created_at": now,
                "updated_at": now,
            }
            if existing_account:
                # Preserve the legacy repair path for an orphaned account row.
                self._upsert_row(accounts_sheet, "account_uid", account_uid, account_values)
            else:
                self._append_row(accounts_sheet, account_values)
            creator_ids.add(creator_id)
            accounts_by_uid[account_uid] = account_values
            created += 1
        if created:
            self._save_workbook(workbook)
        return {"created": created, "skipped_existing": skipped_existing}

    @_synchronized
    def getCreatorsByAgency(self, agency_id: str) -> list[dict[str, Any]]:
        workbook = self._load_workbook()
        agency_id = str(agency_id or "").strip()
        return [
            row
            for row in self._rows(workbook["Creators"])
            if str(row.get("agency_id") or "") == agency_id
        ]

    @_synchronized
    def getCreatorCountsByAgency(self) -> dict[str, int]:
        workbook = self._load_workbook()
        counts: dict[str, int] = {}
        for creator in self._rows(workbook["Creators"]):
            agency_id = str(creator.get("agency_id") or "")
            if agency_id:
                counts[agency_id] = counts.get(agency_id, 0) + 1
        return counts

    @_mutation_synchronized
    def importTaskResults(
        self,
        task_id: str,
        records: list[dict[str, Any]],
        *,
        source: str,
        imported_at: str = "",
    ) -> dict[str, Any]:
        """Idempotently add valid task results to the local creator library."""
        task_id = str(task_id or "").strip()
        if not _TASK_ID_PATTERN.fullmatch(task_id):
            raise ValueError("任务 ID 无效。")
        workbook = self._load_workbook()
        imported_at = str(imported_at or _utc_now())
        seen_uids: set[str] = set()
        creator_ids: list[str] = []
        account_ids: list[str] = []
        summary = {
            "input_records": len(records),
            "created_creators": 0,
            "created_accounts": 0,
            "updated_accounts": 0,
            "duplicate_records": 0,
            "skipped_failed": 0,
            "skipped_invalid": 0,
        }

        for record in records:
            if not isinstance(record, dict):
                summary["skipped_invalid"] += 1
                continue
            scrape_status = str(record.get("scrape_status") or "").strip()
            if scrape_status not in {"success", "partial_success"}:
                summary["skipped_failed"] += 1
                continue
            account_uid = str(record.get("account_uid") or "").strip()
            platform = str(record.get("platform") or "").strip()
            profile_url = str(record.get("profile_url") or "").strip()
            if not account_uid or platform not in {"TikTok", "Instagram", "YouTube"} or not profile_url:
                summary["skipped_invalid"] += 1
                continue
            if account_uid in seen_uids:
                summary["duplicate_records"] += 1
                continue
            seen_uids.add(account_uid)

            existing_account = self._account_row_by_uid(workbook["CreatorAccounts"], account_uid)
            creator_id = str(existing_account.get("creator_id") or "") if existing_account else ""
            if not creator_id or not self._creator_row(workbook["Creators"], creator_id):
                creator_id = f"creator_{hashlib.sha256(account_uid.encode('utf-8')).hexdigest()[:16]}"
            existing_creator = self._creator_row(workbook["Creators"], creator_id)
            existing_metadata = self._metadata_row(workbook["_AnalysisData"], creator_id)
            existing_analysis = self._decode_analysis(existing_metadata.get("analysis_json"))
            existing_analysis_creator = (
                existing_analysis.get("creator")
                if isinstance(existing_analysis.get("creator"), dict)
                else {}
            )
            existing_crm = (
                existing_analysis.get("_crm")
                if isinstance(existing_analysis.get("_crm"), dict)
                else {}
            )
            is_new_creator = not bool(existing_creator)
            creator_name = str(record.get("creator_name") or "").strip()
            incoming_email = str(record.get("email") or "").strip()
            creator_email = str(existing_creator.get("email") or incoming_email)
            if record.get("email_recheck") and incoming_email:
                creator_email = incoming_email
            creator_values = {
                "creator_id": creator_id,
                "name": creator_name or str(existing_creator.get("name") or ""),
                "platform": str(existing_creator.get("platform") or platform),
                "profile_url": str(existing_creator.get("profile_url") or profile_url),
                "country": str(existing_creator.get("country") or record.get("country") or ""),
                "language": str(existing_creator.get("language") or record.get("language") or ""),
                "content_category": str(existing_creator.get("content_category") or record.get("content_category") or ""),
                "followers": str(record.get("followers") or existing_creator.get("followers") or ""),
                "insight_level": str(existing_creator.get("insight_level") or "insufficient"),
                "status": str(existing_creator.get("status") or "discovered"),
                "created_at": str(existing_creator.get("created_at") or imported_at),
                "tags": str(existing_creator.get("tags") or record.get("tags") or ""),
                "updated_at": imported_at,
                "email": creator_email,
                "whatsapp": str(existing_creator.get("whatsapp") or record.get("whatsapp") or ""),
                "note": str(existing_creator.get("note") or record.get("note") or ""),
                "agency_id": str(existing_creator.get("agency_id") or ""),
                "current_contact_id": str(existing_creator.get("current_contact_id") or ""),
                "source_contact_id": str(
                    existing_creator.get("source_contact_id")
                    or (record.get("source_contact_id") if is_new_creator else "")
                    or ""
                ),
                "bio": str(
                    existing_creator.get("bio")
                    or record.get("bio")
                    or existing_crm.get("bio")
                    or existing_analysis_creator.get("bio")
                    or ""
                ),
                "archived_at": str(
                    existing_creator.get("archived_at")
                    or existing_crm.get("archived_at")
                    or ""
                ),
            }
            self._upsert_row(workbook["Creators"], "creator_id", creator_id, creator_values)
            if is_new_creator:
                summary["created_creators"] += 1

            account_values = self._task_account_values(record, creator_id, task_id, source, imported_at)
            if existing_account:
                account_values["account_id"] = str(existing_account.get("account_id") or account_values["account_id"])
                account_values["created_at"] = str(existing_account.get("created_at") or account_values["created_at"])
                if not account_values.get("account_email"):
                    account_values["account_email"] = str(existing_account.get("account_email") or "")
                summary["updated_accounts"] += 1
            else:
                summary["created_accounts"] += 1
            self._upsert_row(
                workbook["CreatorAccounts"],
                "account_uid",
                account_uid,
                account_values,
            )

            analysis = self._task_analysis(record, creator_id, task_id, source, imported_at)
            keep_existing_analysis = (
                str(existing_metadata.get("source") or "") == "chrome_extension"
                and bool(self._decode_analysis(existing_metadata.get("analysis_json")))
            )
            if not keep_existing_analysis:
                self._upsert_row(
                    workbook["_AnalysisData"],
                    "creator_id",
                    creator_id,
                    {
                        "creator_id": creator_id,
                        "task_id": task_id,
                        "account_uid": account_uid,
                        "source": source,
                        "status_updated_at": str(existing_metadata.get("status_updated_at") or ""),
                        "analysis_json": json.dumps(analysis, ensure_ascii=False),
                    },
                )
            self._upsert_task_snapshot(
                workbook,
                creator_id,
                task_id,
                account_uid,
                record,
                source,
                imported_at,
            )
            creator_ids.append(creator_id)
            account_ids.append(str(account_values["account_id"]))

        if creator_ids or account_ids:
            self._save_workbook(workbook)
        return {
            **summary,
            "creator_ids": list(dict.fromkeys(creator_ids)),
            "account_ids": list(dict.fromkeys(account_ids)),
        }

    @_mutation_synchronized
    def updateCreatorRelations(
        self,
        creator_id: str,
        payload: dict[str, Any],
        *,
        agency_port: AgencyPort | None = None,
    ) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise ValueError("达人关系数据无效。")
        allowed = {"agency_id", "current_contact_id", "source_contact_id"}
        unknown = set(payload) - allowed
        if unknown:
            raise ValueError(f"不允许修改关系字段：{', '.join(sorted(unknown))}")
        workbook = self._load_workbook()
        creator_id = str(creator_id or "").strip()
        creator = self._creator_row(workbook["Creators"], creator_id)
        if not creator:
            raise ValueError("未找到达人分析记录。")
        agency_id = str(payload.get("agency_id") if "agency_id" in payload else creator.get("agency_id") or "").strip()
        if agency_id:
            try:
                if agency_port is None:
                    if not self._row_by_key(workbook["Agencies"], "agency_id", agency_id):
                        raise ValueError
                else:
                    agency_port.get_agency(agency_id)
            except ValueError as exc:
                raise ValueError("关联的 Agency 不存在。") from exc
        relations = {"agency_id": agency_id}
        for field in ("current_contact_id", "source_contact_id"):
            contact_id = str(payload.get(field) if field in payload else creator.get(field) or "").strip()
            if contact_id:
                try:
                    if agency_port is None:
                        if not self._row_by_key(
                            workbook["AgencyContacts"], "contact_id", contact_id
                        ):
                            raise ValueError
                    else:
                        agency_port.get_contact(contact_id)
                except ValueError as exc:
                    raise ValueError("关联的 Agency 联系人不存在。") from exc
            relations[field] = contact_id
        updated = {**creator, **relations, "updated_at": _utc_now()}
        self._upsert_row(workbook["Creators"], "creator_id", creator_id, updated)
        self._save_workbook(workbook)
        return {"creator_id": creator_id, **relations}

    @_mutation_synchronized
    def updateCreator(
        self,
        creator_id: str,
        payload: dict[str, Any],
        *,
        agency_port: AgencyPort | None = None,
    ) -> dict[str, Any]:
        """Update CRM-maintained Creator fields without changing identity or schema."""
        if not isinstance(payload, dict) or not payload:
            raise ValueError("缺少可更新的达人资料。")
        allowed = {
            "creator_name", "profile_url", "followers", "country", "language",
            "content_category",
            "bio", "agency_id", "archived_at",
        }
        unknown = set(payload) - allowed
        if unknown:
            raise ValueError(f"不允许修改达人字段：{', '.join(sorted(unknown))}")

        workbook = self._load_workbook()
        creator_id = str(creator_id or "").strip()
        creator = self._creator_row(workbook["Creators"], creator_id)
        if not creator:
            raise ValueError("未找到达人记录。")

        metadata = self._metadata_row(workbook["_AnalysisData"], creator_id)
        analysis = self._decode_analysis(metadata.get("analysis_json"))
        if not analysis:
            analysis = self._rebuild_analysis(workbook, creator, metadata)
        analysis_creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        analysis_creator = dict(analysis_creator)
        crm = analysis.get("_crm") if isinstance(analysis.get("_crm"), dict) else {}
        crm = dict(crm)
        updated = dict(creator)
        now = _utc_now()

        if "creator_name" in payload:
            name = str(payload.get("creator_name") or "").strip()
            if not name:
                raise ValueError("达人名称不能为空。")
            updated["name"] = name
            analysis_creator["creator_name"] = name
            crm["creator_name"] = name

        old_profile_url = str(creator.get("profile_url") or "").strip()
        if "profile_url" in payload:
            profile_url = str(payload.get("profile_url") or "").strip()
            parsed_url = urlparse(profile_url)
            if not profile_url or parsed_url.scheme not in {"http", "https"} or not parsed_url.netloc:
                raise ValueError("主页链接必须是有效的 HTTP 或 HTTPS 地址。")
            updated["profile_url"] = profile_url
            analysis_creator["profile_url"] = profile_url
            crm["profile_url"] = profile_url

        if "followers" in payload:
            followers = str(payload.get("followers") or "").strip()
            updated["followers"] = followers
            analysis_creator["followers"] = followers
            crm["followers"] = followers

        if "country" in payload:
            country = str(payload.get("country") or "").strip()
            updated["country"] = country
            analysis_creator["country"] = country
            crm["country"] = country

        if "language" in payload:
            language = str(payload.get("language") or "").strip()
            updated["language"] = language
            analysis_creator["language"] = language
            crm["language"] = language

        if "content_category" in payload:
            content_category = str(payload.get("content_category") or "").strip()
            updated["content_category"] = content_category
            analysis["content_category"] = content_category
            crm["content_category"] = content_category

        if "bio" in payload:
            bio = str(payload.get("bio") or "").strip()
            updated["bio"] = bio
            crm["bio"] = bio
            analysis_creator["bio"] = bio

        if "agency_id" in payload:
            agency_id = str(payload.get("agency_id") or "").strip()
            if agency_id:
                try:
                    if agency_port is None:
                        if not self._row_by_key(
                            workbook["Agencies"], "agency_id", agency_id
                        ):
                            raise ValueError
                    else:
                        agency_port.get_agency(agency_id)
                except ValueError as exc:
                    raise ValueError("关联的 Agency 不存在。") from exc
            updated["agency_id"] = agency_id
            crm["agency_id"] = agency_id

        if "archived_at" in payload:
            archived_at = payload.get("archived_at")
            if archived_at is not None:
                archived_at = str(archived_at or "").strip()
                try:
                    datetime.fromisoformat(archived_at.replace("Z", "+00:00"))
                except ValueError as exc:
                    raise ValueError("归档时间必须是有效的 ISO 时间。") from exc
            # openpyxl's cell(row, column, None) does not clear an existing
            # value, so restoration is persisted as an explicit empty cell.
            updated["archived_at"] = archived_at or ""
            crm["archived_at"] = archived_at

        updated["updated_at"] = now
        analysis["creator"] = analysis_creator
        analysis["_crm"] = crm
        self._upsert_row(workbook["Creators"], "creator_id", creator_id, updated)

        if any(field in payload for field in ("profile_url", "followers")):
            accounts = [
                row for row in self._rows(workbook["CreatorAccounts"])
                if str(row.get("creator_id") or "") == creator_id
            ]
            metadata_uid = str(metadata.get("account_uid") or "")
            primary_account = next(
                (row for row in accounts if str(row.get("account_uid") or "") == metadata_uid),
                next(
                    (row for row in accounts if str(row.get("profile_url") or "") == old_profile_url),
                    accounts[0] if accounts else {},
                ),
            )
            account_id = str(primary_account.get("account_id") or "")
            if account_id:
                account_values = {**primary_account, "updated_at": now}
                if "profile_url" in payload:
                    account_values["profile_url"] = updated["profile_url"]
                if "followers" in payload:
                    account_values["followers"] = updated["followers"]
                self._upsert_row(workbook["CreatorAccounts"], "account_id", account_id, account_values)

        self._upsert_row(
            workbook["_AnalysisData"],
            "creator_id",
            creator_id,
            {
                **metadata,
                "creator_id": creator_id,
                "analysis_json": json.dumps(analysis, ensure_ascii=False),
            },
        )
        self._save_workbook(workbook)
        return {
            "creator_id": creator_id,
            "creator_name": str(updated.get("name") or ""),
            "profile_url": str(updated.get("profile_url") or ""),
            "followers": str(updated.get("followers") or ""),
            "content_category": str(updated.get("content_category") or ""),
            "bio": str(updated.get("bio") or crm.get("bio") or analysis_creator.get("bio") or ""),
            "agency_id": str(updated.get("agency_id") or ""),
            "archived_at": updated.get("archived_at") or None,
            "updated_at": now,
        }

    @_mutation_synchronized
    def set_creator_archived(self, creator_id: str, archived: bool) -> dict[str, Any]:
        """Idempotently archive or restore a Creator without changing related data."""
        creator_id = str(creator_id or "").strip()
        current = self.getCreatorDetail(creator_id)["record"]
        current_archived_at = str(current.get("archived_at") or "").strip()
        archived_at = (current_archived_at or _utc_now()) if archived else None
        return self.updateCreator(creator_id, {"archived_at": archived_at})

    @_synchronized
    def getCreatorTrend(self, creator_id: str) -> dict[str, Any]:
        """Return latest/previous snapshots, field deltas, and data freshness."""
        workbook = self._load_workbook()
        if not self._creator_row(workbook["Creators"], str(creator_id or "")):
            raise ValueError("未找到达人分析记录。")
        snapshots = self._snapshots_from_workbook(workbook, creator_id)
        return {
            "creator_id": str(creator_id or ""),
            "snapshots": snapshots,
            **self._trend_from_snapshots(snapshots),
        }

    @_synchronized
    def getCreators(self, include_archived: bool = False) -> list[dict[str, Any]]:
        """Return concise Creator Library records from the Excel workbook."""
        request_started = time.perf_counter()
        workbook = self._load_workbook()
        load_duration_ms = round((time.perf_counter() - request_started) * 1000, 2)
        records, index = self._creator_records_from_workbook(workbook)
        if not include_archived:
            records = [record for record in records if not record.get("archived_at")]
        response_duration_ms = round((time.perf_counter() - request_started) * 1000, 2)
        log_event(
            "CreatorLibrary",
            "列表加载完成"
            f" | creators_count={len(records)}"
            f" | accounts_count={index['accounts_count']}"
            f" | snapshots_count={index['snapshots_count']}"
            f" | load_duration_ms={load_duration_ms}"
            f" | index_duration_ms={index['index_duration_ms']}"
            f" | response_duration_ms={response_duration_ms}",
        )
        return records

    @_synchronized
    def getCreatorsPage(
        self,
        page: int = 1,
        page_size: int = 24,
        sort: str = "created_at",
        order: str = "desc",
        include_archived: bool = False,
        filters: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Return one stable, server-sorted Creator Library page."""
        self._validate_creator_page_query(page, page_size, sort, order)
        snapshot = self.getCreatorLibrarySnapshot()
        return self.getCreatorsPageFromSnapshot(
            snapshot,
            page=page,
            page_size=page_size,
            sort=sort,
            order=order,
            include_archived=include_archived,
            filters=filters,
        )

    def getCreatorLibrarySnapshot(self) -> dict[str, Any]:
        """Build the existing Creator Library read model for cache storage."""
        workbook = self._load_workbook()
        records, index = self._creator_records_from_workbook(workbook)
        active_records = [
            record for record in records if not record.get("archived_at")
        ]
        creator_counts: dict[str, int] = {}
        for record in records:
            agency_id = str(record.get("agency_id") or "")
            if agency_id:
                creator_counts[agency_id] = creator_counts.get(agency_id, 0) + 1
        contact_counts: dict[str, int] = {}
        for contact in self._rows(workbook["AgencyContacts"]):
            agency_id = str(contact.get("agency_id") or "")
            if agency_id:
                contact_counts[agency_id] = contact_counts.get(agency_id, 0) + 1
        agency_options = [
            {
                **agency,
                "creator_count": creator_counts.get(
                    str(agency.get("agency_id") or ""), 0
                ),
                "contact_count": contact_counts.get(
                    str(agency.get("agency_id") or ""), 0
                ),
            }
            for agency in self._rows(workbook["Agencies"])
        ]
        agency_options.sort(
            key=lambda agency: str(agency.get("name") or "").casefold()
        )
        return {
            "creators": records,
            "creator_id_index": {
                str(record.get("creator_id") or ""): record
                for record in records
                if str(record.get("creator_id") or "")
            },
            "filter_options": {
                "active": self._creator_filter_options(active_records),
                "all": self._creator_filter_options(records),
            },
            "agency_options": agency_options,
            "accounts_count": index["accounts_count"],
            "snapshots_count": index["snapshots_count"],
        }

    def getCreatorsPageFromSnapshot(
        self,
        snapshot: dict[str, Any],
        page: int = 1,
        page_size: int = 24,
        sort: str = "created_at",
        order: str = "desc",
        include_archived: bool = False,
        filters: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Apply the existing query contract to a detached cached snapshot."""
        self._validate_creator_page_query(page, page_size, sort, order)

        request_started = time.perf_counter()
        records = list(snapshot["creators"])
        if not include_archived:
            records = [record for record in records if not record.get("archived_at")]
        filter_options = dict(
            snapshot["filter_options"]["all" if include_archived else "active"]
        )
        records = self._filter_creator_records(records, filters or {})
        records = self._sort_creator_records(records, sort, order)
        total = len(records)
        pages = (total + page_size - 1) // page_size
        start = (page - 1) * page_size
        creators = records[start:start + page_size]
        log_event(
            "CreatorLibrary",
            "分页列表加载完成"
            f" | total={total} | page={page} | page_size={page_size}"
            f" | sort={sort} | order={order}"
            f" | accounts_count={snapshot['accounts_count']}"
            f" | response_duration_ms={round((time.perf_counter() - request_started) * 1000, 2)}",
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "creators": creators,
            "filter_options": filter_options,
        }

    @staticmethod
    def _validate_creator_page_query(
        page: int,
        page_size: int,
        sort: str,
        order: str,
    ) -> None:
        if page < 1:
            raise ValueError("page 必须大于或等于 1。")
        if page_size not in CREATOR_LIBRARY_PAGE_SIZES:
            raise ValueError("page_size 仅支持 12、24、25、48、50、100。")
        if sort not in CREATOR_LIBRARY_SORT_FIELDS:
            raise ValueError("不支持的达人排序字段。")
        if order not in {"asc", "desc"}:
            raise ValueError("order 仅支持 asc 或 desc。")

    @classmethod
    def _filter_creator_records(
        cls,
        records: list[dict[str, Any]],
        filters: dict[str, Any],
    ) -> list[dict[str, Any]]:
        unknown = set(filters) - CREATOR_LIBRARY_FILTER_FIELDS
        if unknown:
            raise ValueError(f"不支持的达人筛选字段：{', '.join(sorted(unknown))}")
        normalized = {
            key: str(value or "").strip()
            for key, value in filters.items()
            if str(value or "").strip()
        }
        if not normalized:
            return records

        country_filter = normalize_country(normalized.get("country")) if normalized.get("country") else None
        followers_min = normalize_followers(normalized.get("followers_min"))
        followers_max = normalize_followers(normalized.get("followers_max"))

        exact_fields = ("platform", "language", "content_category", "agency_id", "insight_level")

        def matches(record: dict[str, Any]) -> bool:
            for field in exact_fields:
                expected = normalized.get(field)
                if expected and str(record.get(field) or "").strip().casefold() != expected.casefold():
                    return False
            if normalized.get("country"):
                record_country = normalize_country(record.get("country"))
                if country_filter is not None:
                    if record_country != country_filter:
                        return False
                elif str(record.get("country") or "").strip().casefold() != normalized["country"].casefold():
                    return False
            followers = normalize_followers(record.get("followers"))
            if followers_min is not None and (followers is None or followers < followers_min):
                return False
            if followers_max is not None and (followers is None or followers > followers_max):
                return False
            status = normalized.get("status")
            if status == "archived":
                if not record.get("archived_at"):
                    return False
            elif status and (record.get("archived_at") or str(record.get("status") or "") != status):
                return False
            tag = normalized.get("tag")
            if tag:
                tags = {
                    item.strip().casefold()
                    for item in re.split(r"[,，]", str(record.get("tags") or ""))
                    if item.strip()
                }
                if tag.casefold() not in tags:
                    return False
            ai_tag = normalized.get("ai_tag")
            if ai_tag:
                ai_tags = {
                    item.casefold()
                    for item in cls._derived_ai_tags(record)
                }
                if ai_tag.casefold() not in ai_tags:
                    return False
            search = normalized.get("search")
            if search:
                searchable = " ".join(str(record.get(field) or "") for field in (
                    "creator_name", "platform", "profile_url", "country", "language",
                    "content_category", "tags", "agency_name",
                )).casefold()
                if search.casefold() not in searchable:
                    return False
            return True

        return [record for record in records if matches(record)]

    @staticmethod
    def _creator_filter_options(records: list[dict[str, Any]]) -> dict[str, list[str]]:
        def unique(field: str) -> list[str]:
            return sorted({
                str(record.get(field) or "").strip()
                for record in records
                if str(record.get(field) or "").strip()
            }, key=str.casefold)

        tags = sorted({
            tag.strip()
            for record in records
            for tag in re.split(r"[,，]", str(record.get("tags") or ""))
            if tag.strip()
        }, key=str.casefold)
        return {
            "platform": unique("platform"),
            "country": unique("country"),
            "language": unique("language"),
            "content_category": unique("content_category"),
            "agency_id": unique("agency_id"),
            "tag": tags,
            "ai_tag": sorted({
                tag
                for record in records
                for tag in CreatorRepository._derived_ai_tags(record)
            }, key=str.casefold),
        }

    @staticmethod
    def _derived_ai_tags(record: dict[str, Any]) -> list[str]:
        values = []
        category = str(record.get("content_category") or "").strip()
        platform = str(record.get("platform") or "").strip()
        if category:
            values.append(f"category:{category}")
        if platform:
            values.append(f"platform:{platform}")
        return normalize_tags(values)

    @classmethod
    def _sort_creator_records(
        cls,
        records: list[dict[str, Any]],
        sort: str,
        order: str,
    ) -> list[dict[str, Any]]:
        """Sort populated values first and use Creator ID as a stable tie-breaker."""
        def sort_value(record: dict[str, Any]):
            if sort == "followers":
                return cls._metric_sort_value(record.get("followers"))
            if sort == "platform":
                platform = str(record.get("platform") or "").strip().lower()
                return _PLATFORM_SORT_ORDER.get(platform)
            if sort == "creator_name":
                name = str(record.get("creator_name") or "").strip()
                return cls._localized_name_sort_key(name) if name else None
            field = "analysis_time" if sort == "created_at" else "data_updated_at"
            value = str(record.get(field) or "").strip()
            return value or None

        populated = []
        missing = []
        for record in records:
            value = sort_value(record)
            (populated if value is not None else missing).append((value, record))
        populated.sort(
            key=lambda item: (item[0], str(item[1].get("creator_id") or "")),
            reverse=order == "desc",
        )
        missing.sort(key=lambda item: str(item[1].get("creator_id") or ""))
        return [record for _value, record in populated + missing]

    @staticmethod
    def _metric_sort_value(value: Any) -> float | None:
        number = normalize_followers(value)
        if number is not None:
            return float(number)
        # Keep historical Chinese shorthand sorting compatibility outside the
        # canonical import/storage contract.
        raw = str(value or "").strip().replace(",", "").replace(" ", "")
        match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)(万|亿)", raw)
        return float(match.group(1)) * {"万": 1e4, "亿": 1e8}[match.group(2)] if match else None

    @staticmethod
    def _localized_name_sort_key(value: str) -> bytes:
        """Return a deterministic pinyin-oriented key without using system locale."""
        normalized = value.casefold()
        return normalized.encode("gb18030", errors="replace")

    def _creator_records_from_workbook(self, workbook) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        """Build request-scoped indexes and assemble Creator summaries in linear time."""
        index_started = time.perf_counter()
        creators = self._rows(workbook["Creators"])
        insights = {
            str(row["creator_id"]): row
            for row in self._rows(workbook["Insights"])
            if row.get("creator_id")
        }
        metadata = {
            str(row["creator_id"]): row
            for row in self._rows(workbook["_AnalysisData"])
            if row.get("creator_id")
        }
        agency_names = {
            str(row.get("agency_id") or ""): str(row.get("name") or "")
            for row in self._rows(workbook["Agencies"])
            if str(row.get("agency_id") or "")
        }
        accounts_by_creator: dict[str, list[dict[str, Any]]] = {}
        accounts = self._rows(workbook["CreatorAccounts"])
        for account in accounts:
            account_creator_id = str(account.get("creator_id") or "")
            if account_creator_id:
                accounts_by_creator.setdefault(account_creator_id, []).append(account)

        indexed_snapshots_by_creator: dict[str, list[tuple[int, dict[str, Any]]]] = {}
        all_snapshots = self._rows(workbook["CreatorSnapshots"])
        for snapshot_index, snapshot in enumerate(all_snapshots):
            snapshot_creator_id = str(snapshot.get("creator_id") or "")
            if snapshot_creator_id:
                indexed_snapshots_by_creator.setdefault(snapshot_creator_id, []).append(
                    (snapshot_index, snapshot)
                )
        snapshots_by_creator = {
            creator_id: [
                row
                for _index, row in sorted(
                    snapshots,
                    key=lambda item: (str(item[1].get("captured_at") or ""), item[0]),
                    reverse=True,
                )
            ]
            for creator_id, snapshots in indexed_snapshots_by_creator.items()
        }

        records = []
        creator_by_id: dict[str, dict[str, Any]] = {}
        for creator in creators:
            creator_id = str(creator.get("creator_id") or "")
            if not creator_id:
                continue
            creator_by_id[creator_id] = creator
            insight = insights.get(creator_id, {})
            meta = metadata.get(creator_id, {})
            snapshots = snapshots_by_creator.get(creator_id, [])
            snapshot = snapshots[0] if snapshots else {}
            trend = self._trend_from_snapshots(snapshots)
            creator_accounts = accounts_by_creator.get(creator_id, [])
            analysis = self._decode_analysis(meta.get("analysis_json"))
            analysis_creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
            crm = analysis.get("_crm") if isinstance(analysis.get("_crm"), dict) else {}
            metadata_uid = str(meta.get("account_uid") or "")
            primary_account = next(
                (
                    account for account in creator_accounts
                    if str(account.get("account_uid") or "") == metadata_uid
                ),
                creator_accounts[0] if creator_accounts else {},
            )
            agency_id = str(crm.get("agency_id") if "agency_id" in crm else creator.get("agency_id") or "")
            records.append({
                "analysis_id": creator_id,
                "creator_id": creator_id,
                "task_id": str(meta.get("task_id") or ""),
                "account_uid": str(primary_account.get("account_uid") or metadata_uid),
                "creator_name": str(crm.get("creator_name") or creator.get("name") or ""),
                "platform": str(primary_account.get("platform") or creator.get("platform") or ""),
                "profile_url": str(crm.get("profile_url") or primary_account.get("profile_url") or creator.get("profile_url") or ""),
                "followers": str(crm.get("followers") if "followers" in crm else creator.get("followers") or primary_account.get("followers") or snapshot.get("followers") or ""),
                "content_category": str(crm.get("content_category") if "content_category" in crm else creator.get("content_category") or ""),
                "bio": str(
                    creator.get("bio")
                    or crm.get("bio")
                    or analysis_creator.get("bio")
                    or ""
                ),
                "country": str(creator.get("country") or ""),
                "language": str(creator.get("language") or ""),
                "tags": str(creator.get("tags") or ""),
                "insight_level": str(snapshot.get("insight_level") or creator.get("insight_level") or "insufficient"),
                "average_views": snapshot.get("average_views", insight.get("average_views")),
                "median_views": snapshot.get("median_views", insight.get("median_views")),
                "analysis_time": str(creator.get("created_at") or ""),
                "created_at": str(creator.get("created_at") or ""),
                "last_analysis_time": str(snapshot.get("captured_at") or creator.get("created_at") or ""),
                "data_updated_at": str(creator.get("updated_at") or creator.get("created_at") or ""),
                "updated_at": str(creator.get("updated_at") or creator.get("created_at") or ""),
                "source": str(snapshot.get("source") or meta.get("source") or "excel"),
                "status": self._status_value(creator.get("status")),
                "status_updated_at": str(meta.get("status_updated_at") or ""),
                "trend": trend,
                "account_count": len(creator_accounts),
                "agency_id": agency_id,
                "agency_name": agency_names.get(agency_id, ""),
                "current_contact_id": str(creator.get("current_contact_id") or ""),
                "source_contact_id": str(creator.get("source_contact_id") or ""),
                "archived_at": creator.get("archived_at") or crm.get("archived_at"),
            })
        records.sort(key=lambda item: item["analysis_time"], reverse=True)
        return records, {
            "creator_by_id": creator_by_id,
            "metadata_by_creator": metadata,
            "accounts_by_creator": accounts_by_creator,
            "snapshots_by_creator": snapshots_by_creator,
            "accounts_count": len(accounts),
            "snapshots_count": len(all_snapshots),
            "index_duration_ms": round((time.perf_counter() - index_started) * 1000, 2),
        }

    @_synchronized
    def getCreatorDetail(self, analysis_id: str) -> dict[str, Any]:
        """Return one full analysis snapshot reconstructed from the workbook."""
        creator_id = str(analysis_id or "").strip()
        workbook = self._load_workbook()
        records, index = self._creator_records_from_workbook(workbook)
        creator = index["creator_by_id"].get(creator_id, {})
        if not creator:
            raise ValueError("未找到达人分析记录。")
        metadata = index["metadata_by_creator"].get(creator_id, {})
        analysis = self._decode_analysis(metadata.get("analysis_json"))
        if not analysis:
            analysis = self._rebuild_analysis(workbook, creator, metadata)
        record = next((item for item in records if item["analysis_id"] == creator_id), None)
        if not record:
            raise ValueError("达人分析记录不可用。")
        analysis_creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        analysis["creator"] = {
            **analysis_creator,
            "creator_name": record.get("creator_name") or analysis_creator.get("creator_name") or "",
            "platform": record.get("platform") or analysis_creator.get("platform") or "",
            "profile_url": record.get("profile_url") or analysis_creator.get("profile_url") or "",
            "followers": record.get("followers") or analysis_creator.get("followers") or "",
            "bio": record.get("bio") if "bio" in record else analysis_creator.get("bio") or "",
        }
        analysis["content_category"] = record.get("content_category") or analysis.get("content_category") or ""
        snapshots = index["snapshots_by_creator"].get(creator_id, [])
        history_times = [str(item.get("captured_at") or "") for item in snapshots]
        if not history_times and creator.get("created_at"):
            history_times = [str(creator["created_at"])]
        cooperations = self.getCreatorCooperations(creator_id, workbook)
        accounts = index["accounts_by_creator"].get(creator_id, [])
        return {
            "record": record,
            "analysis": analysis,
            "accounts": accounts,
            "snapshots": snapshots,
            "trend": self._trend_from_snapshots(snapshots),
            "history_analysis_times": sorted(history_times, reverse=True),
            "cooperations": cooperations,
            "cooperation_statistics": self._cooperation_statistics(cooperations),
        }

    @_synchronized
    def getCreatorSummarySourceData(self, creator_id: str) -> dict[str, Any]:
        """Read only the safe factual inputs required by CreatorSummaryService."""
        creator_id = str(creator_id or "").strip()
        workbook = self._load_workbook()
        creator = self._creator_row(workbook["Creators"], creator_id)
        if not creator:
            raise ValueError("未找到达人分析记录。")
        insight = self._creator_row(workbook["Insights"], creator_id)
        snapshots = self._snapshots_from_workbook(workbook, creator_id)
        campaign_creator_count = sum(
            str(row.get("creator_id") or "") == creator_id
            for row in self._rows(workbook["CampaignCreators"])
        )
        return {
            "creator": {
                field: creator.get(field)
                for field in (
                    "creator_id", "name", "platform", "profile_url", "followers",
                    "country", "language", "content_category", "bio",
                    "insight_level", "archived_at",
                )
            },
            "insight": {
                field: insight.get(field)
                for field in ("average_views", "median_views", "stability")
            },
            "snapshots": [
                {
                    field: snapshot.get(field)
                    for field in (
                        "average_views", "median_views", "video_count",
                        "creator_score", "insight_level", "captured_at",
                    )
                }
                for snapshot in snapshots
            ],
            "campaign_creator_count": campaign_creator_count,
        }

    @_synchronized
    def getCreatorIntelligenceSourceData(self, creator_id: str) -> dict[str, Any]:
        """Return factual, detached inputs for deterministic intelligence."""
        creator_id = str(creator_id or "").strip()
        workbook = self._load_workbook()
        creator = self._creator_row(workbook["Creators"], creator_id)
        if not creator:
            raise ValueError("未找到达人分析记录。")
        metadata = self._metadata_row(workbook["_AnalysisData"], creator_id)
        analysis = self._decode_analysis(metadata.get("analysis_json"))
        return {
            "creator": dict(creator),
            "accounts": [
                dict(row)
                for row in self._rows(workbook["CreatorAccounts"])
                if str(row.get("creator_id") or "") == creator_id
            ],
            "snapshots": [
                dict(row)
                for row in self._rows(workbook["CreatorSnapshots"])
                if str(row.get("creator_id") or "") == creator_id
            ],
            "videos": [
                dict(row)
                for row in (analysis.get("videos") or [])
                if isinstance(row, dict)
            ],
            "campaign_creators": [
                dict(row)
                for row in self._rows(workbook["CampaignCreators"])
                if str(row.get("creator_id") or "") == creator_id
                and not str(row.get("archived_at") or "").strip()
            ],
        }

    @_synchronized
    def getCreatorCooperations(self, creator_id: str, workbook=None) -> list[dict[str, Any]]:
        """Return one creator's cooperation history, latest contact first."""
        workbook = workbook or self._load_workbook()
        rows = [
            row for row in self._rows(workbook["Cooperations"])
            if str(row.get("creator_id") or "") == str(creator_id or "")
        ]
        rows.sort(key=lambda row: (str(row.get("contact_date") or ""), str(row.get("created_at") or "")), reverse=True)
        return rows

    @_synchronized
    def getCooperations(self) -> list[dict[str, Any]]:
        """Return all cooperation records in one workbook read for Dashboard requests."""
        workbook = self._load_workbook()
        rows = self._rows(workbook["Cooperations"])
        rows.sort(
            key=lambda row: (
                str(row.get("contact_date") or ""),
                str(row.get("created_at") or ""),
            ),
            reverse=True,
        )
        return rows

    @_mutation_synchronized
    def saveCooperation(self, creator_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        """Reject writes to the preserved, read-only Legacy Cooperation store."""
        raise PermissionError("请使用 Campaign 创建新的合作。")

    @_mutation_synchronized
    def updateCreatorStatus(self, analysis_id: str, status: object) -> dict[str, str]:
        """Update only the local Creator Library status in the Excel workbook."""
        creator_id = str(analysis_id or "").strip()
        status_value = self._status_value(status)
        workbook = self._load_workbook()
        creator = self._creator_row(workbook["Creators"], creator_id)
        if not creator:
            raise ValueError("未找到达人分析记录。")
        self._upsert_row(
            workbook["Creators"],
            "creator_id",
            creator_id,
            {**creator, "status": status_value, "updated_at": _utc_now()},
        )
        metadata = self._metadata_row(workbook["_AnalysisData"], creator_id)
        self._upsert_row(
            workbook["_AnalysisData"],
            "creator_id",
            creator_id,
            {**metadata, "creator_id": creator_id, "status_updated_at": _utc_now()},
        )
        self._save_workbook(workbook)
        return {"analysis_id": creator_id, "status": status_value, "updated_at": _utc_now()}

    def _load_workbook(self):
        try:
            return self.store.open()
        except WorkbookReadError as exc:
            detail = exc.__cause__ or exc
            log_event("Excel", f"读取失败: {self.workbook_path} | {detail}")
            raise RuntimeError(f"无法读取达人库 Excel 文件：{detail}") from exc
        except WorkbookSaveError as exc:
            log_event("Excel", f"保存失败，文件可能被占用: {self.workbook_path} | {exc}")
            raise RuntimeError(
                "无法保存达人库 Excel 文件。请先关闭 WPS 或 Excel 中打开的该文件。"
            ) from exc

    def _apply_schema_migrations(self, workbook, created: bool) -> bool:
        if created:
            self._initialize_new_workbook(workbook)
            self._migrate_legacy_json(workbook)
            self.last_migration_report = self._migrate_phase1_workbook(
                workbook,
                from_schema="new",
                backup_path="",
                created_sheets={"CreatorAccounts", "Agencies", "AgencyContacts", "FollowUpLogs"},
            )
            log_event("Excel", f"已创建达人库文件: {self.workbook_path}")
            return True

        log_event("Excel", f"打开成功: {self.workbook_path}")
        from_schema = self._schema_version(workbook)
        existing_sheets = set(workbook.sheetnames)
        requires_migration = self._requires_phase1_migration(workbook, from_schema)
        backup_path = self._create_migration_backup() if requires_migration else None
        try:
            changed = self._ensure_sheets(workbook)
            lifecycle_changed, review_required = migrate_legacy_campaign_archives(workbook)
            changed = lifecycle_changed or changed
            self.last_campaign_lifecycle_report = {
                "manual_review_required": review_required,
                "count": len(review_required),
            }
            lifecycle_report_key = str(self.workbook_path.resolve())
            if review_required and lifecycle_report_key not in _CAMPAIGN_LIFECYCLE_LOGGED_PATHS:
                log_event(
                    "Migration",
                    json.dumps(
                        {
                            "type": "campaign_archive_lifecycle",
                            "manual_review_required": review_required,
                            "count": len(review_required),
                        },
                        ensure_ascii=False,
                    ),
                )
                _CAMPAIGN_LIFECYCLE_LOGGED_PATHS.add(lifecycle_report_key)
            if requires_migration:
                self.last_migration_report = self._migrate_phase1_workbook(
                    workbook,
                    from_schema=from_schema,
                    backup_path=str(backup_path or ""),
                    created_sheets={
                        name
                        for name in ("CreatorAccounts", "Agencies", "AgencyContacts", "FollowUpLogs")
                        if name not in existing_sheets
                    },
                )
                changed = True
        except Exception as exc:
            log_event(
                "Migration",
                f"迁移失败，原工作簿保持不变 | backup={backup_path or '--'} | {exc}",
            )
            raise
        return changed

    def _new_workbook(self):
        workbook = self.store.new_workbook()
        self._initialize_new_workbook(workbook)
        return workbook

    def _initialize_new_workbook(self, workbook) -> None:
        creators = workbook.active
        creators.title = "Creators"
        self._set_headers(creators, _CREATORS_HEADERS)
        self._set_headers(workbook.create_sheet("CreatorAccounts"), _CREATOR_ACCOUNTS_HEADERS)
        self._set_headers(workbook.create_sheet("Videos"), _VIDEOS_HEADERS)
        self._set_headers(workbook.create_sheet("Insights"), _INSIGHTS_HEADERS)
        self._set_headers(workbook.create_sheet("CreatorSnapshots"), _CREATOR_SNAPSHOTS_HEADERS)
        self._set_headers(workbook.create_sheet("VideoSnapshots"), _VIDEO_SNAPSHOTS_HEADERS)
        self._set_headers(workbook.create_sheet("Cooperations"), _COOPERATIONS_HEADERS)
        self._set_headers(workbook.create_sheet("Agencies"), _AGENCIES_HEADERS)
        self._set_headers(workbook.create_sheet("AgencyContacts"), _AGENCY_CONTACTS_HEADERS)
        self._set_headers(workbook.create_sheet("FollowUpLogs"), _FOLLOW_UP_LOGS_HEADERS)
        self._set_headers(workbook.create_sheet("Products"), PRODUCTS_HEADERS)
        self._set_headers(workbook.create_sheet("Campaigns"), CAMPAIGNS_HEADERS)
        self._set_headers(workbook.create_sheet("CampaignCreators"), CAMPAIGN_CREATORS_HEADERS)
        metadata = workbook.create_sheet("_AnalysisData")
        self._set_headers(metadata, _ANALYSIS_METADATA_HEADERS)
        metadata.sheet_state = "hidden"
        workbook_metadata = workbook.create_sheet("_Metadata")
        self._set_headers(workbook_metadata, _WORKBOOK_METADATA_HEADERS)
        workbook_metadata.sheet_state = "hidden"

    def _ensure_sheets(self, workbook) -> bool:
        changed = False
        for name, headers in _WORKBOOK_SHEETS.items():
            if name not in workbook.sheetnames:
                self._set_headers(workbook.create_sheet(name), headers)
                changed = True
            else:
                changed = self._ensure_headers(workbook[name], headers) or changed
        workbook["_AnalysisData"].sheet_state = "hidden"
        workbook["_Metadata"].sheet_state = "hidden"
        return changed

    @staticmethod
    def _set_headers(sheet, headers: list[str]) -> bool:
        existing = [cell.value for cell in sheet[1]] if sheet.max_row else []
        changed = existing != headers
        if existing != headers:
            for column, header in enumerate(headers, start=1):
                sheet.cell(1, column, header)
                sheet.cell(1, column).font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + min(len(headers), 26))}{max(sheet.max_row, 1)}"
        return changed

    @staticmethod
    def _ensure_headers(sheet, headers: list[str]) -> bool:
        """Append missing columns without renaming or reordering existing data."""
        existing = [str(cell.value or "") for cell in sheet[1]] if sheet.max_row else []
        changed = False
        for header in headers:
            if header in existing:
                continue
            column = len(existing) + 1
            sheet.cell(1, column, header)
            sheet.cell(1, column).font = Font(bold=True)
            existing.append(header)
            changed = True
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + min(len(existing), 26))}{max(sheet.max_row, 1)}"
        return changed

    def _schema_version(self, workbook) -> str:
        if "_Metadata" not in workbook.sheetnames:
            return ""
        versions = [
            str(row.get("schema_version") or "").strip()
            for row in self._rows(workbook["_Metadata"])
            if str(row.get("schema_version") or "").strip()
        ]
        return versions[-1] if versions else ""

    def _requires_phase1_migration(self, workbook, schema_version: str) -> bool:
        required = {
            # M1-C2 columns are appended lazily by _ensure_sheets and must not
            # retrigger the older Phase 1 entity migration.
            "Creators": _CREATORS_HEADERS[:-2],
            "CreatorAccounts": _CREATOR_ACCOUNTS_HEADERS,
            "Agencies": _AGENCIES_HEADERS,
            "AgencyContacts": _AGENCY_CONTACTS_HEADERS,
            "FollowUpLogs": _FOLLOW_UP_LOGS_HEADERS,
        }
        for sheet_name, headers in required.items():
            if sheet_name not in workbook.sheetnames:
                return True
            existing = {str(cell.value or "") for cell in workbook[sheet_name][1]}
            if not set(headers).issubset(existing):
                return True
        if schema_version == CREATOR_LIBRARY_SCHEMA_VERSION:
            return False
        # Do not downgrade a future v2+ workbook that already satisfies this contract.
        return not schema_version.startswith("2.")

    def _create_migration_backup(self) -> Path:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
        backup_path = self.store.create_backup(f".pre_v2_{timestamp}")
        log_event("Migration", f"迁移前备份已创建: {backup_path}")
        return backup_path

    def _migrate_phase1_workbook(
        self,
        workbook,
        *,
        from_schema: str,
        backup_path: str,
        created_sheets: set[str],
    ) -> dict[str, Any]:
        creators = self._rows(workbook["Creators"])
        metadata_by_creator = {
            str(row.get("creator_id") or ""): row
            for row in self._rows(workbook["_AnalysisData"])
            if str(row.get("creator_id") or "")
        }
        snapshots_by_creator: dict[str, list[dict[str, Any]]] = {}
        for snapshot in self._rows(workbook["CreatorSnapshots"]):
            creator_id = str(snapshot.get("creator_id") or "")
            if creator_id:
                snapshots_by_creator.setdefault(creator_id, []).append(snapshot)

        created_accounts = 0
        duplicates = 0
        unresolved = 0
        for creator in creators:
            creator_id = str(creator.get("creator_id") or "").strip()
            if not creator_id:
                unresolved += 1
                continue
            metadata = metadata_by_creator.get(creator_id, {})
            analysis = self._decode_analysis(metadata.get("analysis_json"))
            analysis_creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
            snapshots = snapshots_by_creator.get(creator_id, [])
            snapshots.sort(key=lambda item: str(item.get("captured_at") or ""), reverse=True)
            latest_snapshot = snapshots[0] if snapshots else {}
            platform = str(creator.get("platform") or analysis_creator.get("platform") or latest_snapshot.get("platform") or "").strip()
            profile_url = str(creator.get("profile_url") or analysis_creator.get("profile_url") or "").strip()
            account_uid = str(metadata.get("account_uid") or latest_snapshot.get("account_uid") or "").strip()
            if not account_uid and platform and profile_url:
                account_uid = self._build_account_uid(platform, profile_url)
            if not account_uid or not platform or not profile_url:
                unresolved += 1
                continue

            existing_account = self._account_row_by_uid(workbook["CreatorAccounts"], account_uid)
            if existing_account:
                if str(existing_account.get("creator_id") or "") != creator_id:
                    duplicates += 1
                continue
            captured_at = str(
                latest_snapshot.get("captured_at")
                or creator.get("updated_at")
                or creator.get("created_at")
                or _utc_now()
            )
            self._upsert_row(
                workbook["CreatorAccounts"],
                "account_uid",
                account_uid,
                {
                    "account_id": self._account_id(account_uid),
                    "creator_id": creator_id,
                    "account_uid": account_uid,
                    "platform": platform,
                    "username": self._username_from_profile_url(platform, profile_url),
                    "profile_url": profile_url,
                    "followers": str(latest_snapshot.get("followers") or creator.get("followers") or ""),
                    "account_email": str(analysis_creator.get("email") or ""),
                    "latest_post_date": str(analysis_creator.get("latest_post_date") or ""),
                    "last_scrape_time": captured_at,
                    "data_source": str(metadata.get("source") or latest_snapshot.get("source") or "legacy_excel"),
                    "scrape_status": "",
                    "platform_account_id": "",
                    "attribution_status": "confirmed",
                    "note": "",
                    "source_task_id": str(metadata.get("task_id") or analysis.get("task_id") or ""),
                    "created_at": str(creator.get("created_at") or captured_at),
                    "updated_at": captured_at,
                },
            )
            created_accounts += 1

        report = {
            "from_schema": from_schema or "unknown",
            "to_schema": CREATOR_LIBRARY_SCHEMA_VERSION,
            "legacy_creator_rows": len(creators),
            "creators_created": 0,
            "creators_preserved": len(creators),
            "accounts_created": created_accounts,
            "duplicate_accounts": duplicates,
            "unresolved_accounts": unresolved,
            "agencies_sheet_created": "Agencies" in created_sheets,
            "agency_contacts_sheet_created": "AgencyContacts" in created_sheets,
            "follow_up_logs_sheet_created": "FollowUpLogs" in created_sheets,
            "backup_path": backup_path,
            "result": "success",
            "migrated_at": _utc_now(),
        }
        log_event("Migration", json.dumps(report, ensure_ascii=False))
        return report

    def _save_workbook(self, workbook) -> None:
        try:
            self.store.save(workbook)
            log_event("Excel", f"保存成功: {self.workbook_path}")
        except WorkbookSaveError as exc:
            log_event("Excel", f"保存失败，文件可能被占用: {self.workbook_path} | {exc}")
            raise RuntimeError(
                "无法保存达人库 Excel 文件。请先关闭 WPS 或 Excel 中打开的该文件。"
            ) from exc

    def _prepare_workbook_save(self, workbook) -> None:
        self._upsert_row(
            workbook["_Metadata"],
            "schema_version",
            CREATOR_LIBRARY_SCHEMA_VERSION,
            {"schema_version": CREATOR_LIBRARY_SCHEMA_VERSION, "last_update_time": _utc_now()},
        )

    @staticmethod
    def _rows(sheet) -> list[dict[str, Any]]:
        headers = [str(cell.value or "") for cell in sheet[1]]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return rows

    def _creator_row(self, sheet, creator_id: str) -> dict[str, Any]:
        return next((row for row in self._rows(sheet) if str(row.get("creator_id") or "") == creator_id), {})

    def _row_by_key(self, sheet, key: str, value: str) -> dict[str, Any]:
        return next(
            (row for row in self._rows(sheet) if str(row.get(key) or "") == str(value or "")),
            {},
        )

    def _account_row_by_uid(self, sheet, account_uid: str) -> dict[str, Any]:
        return self._row_by_key(sheet, "account_uid", str(account_uid or "").strip())

    def _metadata_row(self, sheet, creator_id: str) -> dict[str, Any]:
        return self._creator_row(sheet, creator_id)

    @staticmethod
    def _account_id(account_uid: str) -> str:
        return f"account_{hashlib.sha256(account_uid.encode('utf-8')).hexdigest()[:16]}"

    @staticmethod
    def _build_account_uid(platform: str, profile_url: str) -> str:
        try:
            import scraper as scraper_module
        except ImportError:
            from . import scraper as scraper_module
        return scraper_module.build_creator_uid({"platform": platform, "url": profile_url})

    @staticmethod
    def _username_from_profile_url(platform: str, profile_url: str) -> str:
        parsed = urlparse(str(profile_url or ""))
        parts = [part for part in parsed.path.split("/") if part]
        if not parts:
            return ""
        if platform == "TikTok":
            return parts[0].lstrip("@")
        if platform == "Instagram":
            return parts[0]
        if platform == "YouTube":
            if parts[0].startswith("@"):
                return parts[0].lstrip("@")
            if len(parts) >= 2 and parts[0].lower() in {"channel", "c", "user"}:
                return parts[1]
        return ""

    def _upsert_account_from_analysis(
        self,
        workbook,
        analysis: dict[str, Any],
        creator_id: str,
    ) -> dict[str, Any]:
        creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        platform = str(creator.get("platform") or "").strip()
        profile_url = str(creator.get("profile_url") or "").strip()
        account_uid = str(analysis.get("account_uid") or "").strip()
        if not account_uid and platform and profile_url:
            account_uid = self._build_account_uid(platform, profile_url)
        if not account_uid:
            return {"account_id": ""}
        existing = self._account_row_by_uid(workbook["CreatorAccounts"], account_uid)
        now = str(analysis.get("imported_at") or _utc_now())
        values = {
            **existing,
            "account_id": str(existing.get("account_id") or self._account_id(account_uid)),
            "creator_id": creator_id,
            "account_uid": account_uid,
            "platform": platform or str(existing.get("platform") or ""),
            "username": str(
                creator.get("username")
                or existing.get("username")
                or self._username_from_profile_url(platform, profile_url)
                or ""
            ),
            "profile_url": profile_url or str(existing.get("profile_url") or ""),
            "followers": str(creator.get("followers") or existing.get("followers") or ""),
            "account_email": str(creator.get("account_email") or creator.get("email") or existing.get("account_email") or ""),
            "latest_post_date": str(creator.get("latest_post_date") or existing.get("latest_post_date") or ""),
            "last_scrape_time": now,
            "data_source": str(analysis.get("source") or existing.get("data_source") or ""),
            "scrape_status": str(analysis.get("scrape_status") or existing.get("scrape_status") or ""),
            "platform_account_id": str(creator.get("platform_account_id") or existing.get("platform_account_id") or ""),
            "attribution_status": str(existing.get("attribution_status") or "confirmed"),
            "note": str(existing.get("note") or ""),
            "source_task_id": str(analysis.get("task_id") or existing.get("source_task_id") or ""),
            "created_at": str(existing.get("created_at") or now),
            "updated_at": now,
        }
        self._upsert_row(workbook["CreatorAccounts"], "account_uid", account_uid, values)
        return values

    def _task_account_values(
        self,
        record: dict[str, Any],
        creator_id: str,
        task_id: str,
        source: str,
        imported_at: str,
    ) -> dict[str, Any]:
        account_uid = str(record.get("account_uid") or "").strip()
        platform = str(record.get("platform") or "").strip()
        profile_url = str(record.get("profile_url") or "").strip()
        return {
            "account_id": self._account_id(account_uid),
            "creator_id": creator_id,
            "account_uid": account_uid,
            "platform": platform,
            "username": str(record.get("username") or self._username_from_profile_url(platform, profile_url)),
            "profile_url": profile_url,
            "followers": str(record.get("followers") or ""),
            "account_email": str(record.get("email") or ""),
            "latest_post_date": str(record.get("latest_post_date") or ""),
            "last_scrape_time": str(record.get("last_scrape_time") or imported_at),
            "data_source": str(record.get("data_source") or source),
            "scrape_status": str(record.get("scrape_status") or ""),
            "platform_account_id": str(record.get("platform_account_id") or ""),
            "attribution_status": str(record.get("attribution_status") or "confirmed"),
            "note": str(record.get("note") or ""),
            "source_task_id": task_id,
            "created_at": imported_at,
            "updated_at": imported_at,
        }

    @staticmethod
    def _task_analysis(
        record: dict[str, Any],
        creator_id: str,
        task_id: str,
        source: str,
        imported_at: str,
    ) -> dict[str, Any]:
        return {
            "schema_version": "2.0-phase1",
            "analysis_id": creator_id,
            "task_id": task_id,
            "account_uid": str(record.get("account_uid") or ""),
            "imported_at": imported_at,
            "source": source,
            "creator": {
                "creator_name": str(record.get("creator_name") or ""),
                "platform": str(record.get("platform") or ""),
                "profile_url": str(record.get("profile_url") or ""),
                "followers": str(record.get("followers") or ""),
                "bio": str(record.get("bio") or ""),
                "country": str(record.get("country") or ""),
                "language": str(record.get("language") or ""),
            },
            "content_category": str(record.get("content_category") or ""),
            "video_analysis": {},
            "videos": [],
            "creator_insight": {"level": "insufficient", "risks": [], "recommendation": ""},
        }

    def _upsert_task_snapshot(
        self,
        workbook,
        creator_id: str,
        task_id: str,
        account_uid: str,
        record: dict[str, Any],
        source: str,
        imported_at: str,
    ) -> None:
        suffix = hashlib.sha256(account_uid.encode("utf-8")).hexdigest()[:12]
        snapshot_id = f"snapshot_{task_id}_{suffix}"
        self._upsert_row(
            workbook["CreatorSnapshots"],
            "snapshot_id",
            snapshot_id,
            {
                "snapshot_id": snapshot_id,
                "creator_id": creator_id,
                "platform": str(record.get("platform") or ""),
                "account_uid": account_uid,
                "followers": str(record.get("followers") or ""),
                "average_views": record.get("average_views", ""),
                "median_views": record.get("median_views", ""),
                "video_count": record.get("video_count", 0),
                "creator_score": record.get("creator_score", ""),
                "insight_level": str(record.get("insight_level") or "insufficient"),
                "captured_at": str(record.get("last_scrape_time") or imported_at),
                "source": source,
            },
        )

    def _snapshots_from_workbook(self, workbook, creator_id: str) -> list[dict[str, Any]]:
        snapshots = [
            row for row in self._rows(workbook["CreatorSnapshots"])
            if str(row.get("creator_id") or "") == str(creator_id or "")
        ]
        # Captures can occur within the same second; later appended rows are newer in that tie.
        indexed = enumerate(snapshots)
        return [row for _index, row in sorted(indexed, key=lambda item: (str(item[1].get("captured_at") or ""), item[0]), reverse=True)]

    @staticmethod
    def _metric_number(value: object) -> float | None:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        raw = str(value).strip().lower().replace(",", "")
        match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)\s*([kmb])?", raw)
        if not match:
            return None
        multiplier = {"k": 1_000, "m": 1_000_000, "b": 1_000_000_000}.get(match.group(2) or "", 1)
        return float(match.group(1)) * multiplier

    @classmethod
    def _metric_change(cls, latest: object, previous: object) -> dict[str, Any]:
        latest_number = cls._metric_number(latest)
        previous_number = cls._metric_number(previous)
        if latest_number is None or previous_number is None:
            return {"status": "unavailable", "direction": "", "delta": None}
        delta = latest_number - previous_number
        direction = "growth" if delta > 0 else "decline" if delta < 0 else "stable"
        return {"status": "available", "direction": direction, "delta": delta}

    @staticmethod
    def _freshness(captured_at: object) -> dict[str, Any]:
        raw = str(captured_at or "").strip()
        try:
            captured = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            if captured.tzinfo is None:
                captured = captured.replace(tzinfo=timezone.utc)
        except ValueError:
            return {"status": "unknown", "days": None}
        days = max(0, (datetime.now(timezone.utc) - captured).days)
        status = "fresh" if days <= 7 else "update_recommended" if days <= 30 else "stale"
        return {"status": status, "days": days}

    @classmethod
    def _trend_from_snapshots(cls, snapshots: list[dict[str, Any]]) -> dict[str, Any]:
        latest = snapshots[0] if snapshots else {}
        previous = snapshots[1] if len(snapshots) > 1 else {}
        if not previous:
            changes = {
                key: {"status": "no_history", "direction": "", "delta": None}
                for key in ("followers", "median_views", "creator_score")
            }
        else:
            changes = {
                "followers": cls._metric_change(latest.get("followers"), previous.get("followers")),
                "median_views": cls._metric_change(latest.get("median_views"), previous.get("median_views")),
                "creator_score": cls._metric_change(latest.get("creator_score"), previous.get("creator_score")),
            }
        return {
            "latest": latest,
            "previous": previous,
            "changes": changes,
            "freshness": cls._freshness(latest.get("captured_at")),
        }

    def _creator_id_for_account_uid(self, workbook, account_uid: str) -> str:
        account_uid = str(account_uid or "").strip()
        if not account_uid:
            return ""
        if "CreatorAccounts" in workbook.sheetnames:
            account = self._account_row_by_uid(workbook["CreatorAccounts"], account_uid)
            candidate = str(account.get("creator_id") or "")
            if candidate and self._creator_row(workbook["Creators"], candidate):
                return candidate
        for snapshot in self._rows(workbook["CreatorSnapshots"]):
            if str(snapshot.get("account_uid") or "") == account_uid:
                candidate = str(snapshot.get("creator_id") or "")
                if self._creator_row(workbook["Creators"], candidate):
                    return candidate
        for metadata in self._rows(workbook["_AnalysisData"]):
            if str(metadata.get("account_uid") or "") == account_uid:
                candidate = str(metadata.get("creator_id") or "")
                if self._creator_row(workbook["Creators"], candidate):
                    return candidate
        return ""

    @staticmethod
    def _upsert_row(sheet, key: str, key_value: str, values: dict[str, Any]) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        key_index = headers.index(key) + 1
        row_index = next((row for row in range(2, sheet.max_row + 1) if str(sheet.cell(row, key_index).value or "") == key_value), sheet.max_row + 1)
        for column, header in enumerate(headers, start=1):
            if header in values:
                sheet.cell(row_index, column, values[header])

    @staticmethod
    def _append_row(sheet, values: dict[str, Any]) -> None:
        """Append a known-new record without scanning or touching existing rows."""
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _replace_video_rows(self, sheet, creator_id: str, videos: object) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        existing = self._rows(sheet)
        retained = [row for row in existing if str(row.get("creator_id") or "") != creator_id]
        sheet.delete_rows(2, max(sheet.max_row - 1, 0))
        for row in retained:
            sheet.append([row.get(header, "") for header in headers])
        for video in videos if isinstance(videos, list) else []:
            if not isinstance(video, dict):
                continue
            sheet.append([
                creator_id,
                str(video.get("video_url") or ""),
                video.get("views", ""),
                video.get("likes", ""),
                video.get("comments", ""),
                str(video.get("captured_at") or ""),
            ])

    def _replace_video_snapshot_rows(self, sheet, snapshot_id: str, creator_id: str, videos: object, captured_at: str) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        retained = [row for row in self._rows(sheet) if str(row.get("snapshot_id") or "") != snapshot_id]
        sheet.delete_rows(2, max(sheet.max_row - 1, 0))
        for row in retained:
            sheet.append([row.get(header, "") for header in headers])
        for index, video in enumerate(videos if isinstance(videos, list) else []):
            if not isinstance(video, dict):
                continue
            video_url = str(video.get("video_url") or "")
            video_id = str(video.get("video_id") or video.get("video_key") or "").strip()
            if not video_id:
                video_id = hashlib.sha256(video_url.encode("utf-8")).hexdigest()[:16] if video_url else str(index + 1)
            sheet.append([
                f"{snapshot_id}:{video_id}", snapshot_id, creator_id, video_id, video_url,
                str(video.get("platform") or ""), video.get("views", ""), video.get("likes", ""),
                video.get("comments", ""), str(video.get("captured_at") or captured_at),
            ])

    @staticmethod
    def _optional_number(value: object, label: str, integer: bool = False) -> int | float | str:
        raw = str(value if value is not None else "").strip()
        if not raw:
            return ""
        try:
            number = float(raw.replace(",", ""))
        except ValueError as exc:
            raise ValueError(f"{label}必须是数字。") from exc
        if integer:
            if not number.is_integer() or number < 0:
                raise ValueError(f"{label}必须是非负整数。")
            return int(number)
        return number

    @staticmethod
    def _cooperation_statistics(cooperations: list[dict[str, Any]]) -> dict[str, Any]:
        def values(field: str) -> list[float]:
            result = []
            for cooperation in cooperations:
                value = cooperation.get(field)
                if isinstance(value, (int, float)):
                    result.append(float(value))
                elif isinstance(value, str) and value.strip():
                    try:
                        result.append(float(value.replace(",", "")))
                    except ValueError:
                        continue
            return result

        prices = values("price")
        average_views = values("average_views")
        rois = values("roi")
        return {
            "cooperation_count": len(cooperations),
            "total_spend": sum(prices) if prices else 0,
            "average_views": sum(average_views) / len(average_views) if average_views else 0,
            "average_roi": sum(rois) / len(rois) if rois else 0,
        }

    @staticmethod
    def _creator_values(analysis: dict[str, Any], status: str, creator_id: str) -> dict[str, Any]:
        creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        insight = analysis.get("creator_insight") if isinstance(analysis.get("creator_insight"), dict) else {}
        return {
            "creator_id": creator_id,
            "name": str(creator.get("creator_name") or ""),
            "platform": str(creator.get("platform") or ""),
            "profile_url": str(creator.get("profile_url") or ""),
            "country": str(creator.get("country") or ""),
            "language": str(creator.get("language") or ""),
            "content_category": str(analysis.get("content_category") or ""),
            "tags": CreatorRepository._tags_value(analysis),
            "followers": str(creator.get("followers") or ""),
            "insight_level": str(insight.get("level") or insight.get("grade") or "insufficient"),
            "status": status,
            "created_at": str(analysis.get("imported_at") or _utc_now()),
            "updated_at": _utc_now(),
            "email": str(creator.get("email") or ""),
            "whatsapp": str(creator.get("whatsapp") or ""),
            "cooperation_stage": str(creator.get("cooperation_stage") or ""),
            "recent_product": str(creator.get("recent_product") or ""),
            "quote": str(creator.get("quote") or ""),
            "owner": str(creator.get("owner") or ""),
            "last_contact_time": str(creator.get("last_contact_time") or ""),
            "next_follow_up_time": str(creator.get("next_follow_up_time") or ""),
            "note": str(creator.get("note") or ""),
            "agency_id": str(creator.get("agency_id") or ""),
            "current_contact_id": str(creator.get("current_contact_id") or ""),
            "source_contact_id": str(creator.get("source_contact_id") or ""),
            "bio": str(creator.get("bio") or ""),
            "archived_at": str(creator.get("archived_at") or ""),
        }

    @staticmethod
    def _insight_values(analysis: dict[str, Any], creator_id: str) -> dict[str, Any]:
        metrics = analysis.get("video_analysis") if isinstance(analysis.get("video_analysis"), dict) else {}
        insight = analysis.get("creator_insight") if isinstance(analysis.get("creator_insight"), dict) else {}
        return {
            "creator_id": creator_id,
            "average_views": metrics.get("average_views", ""),
            "median_views": metrics.get("median_views", ""),
            "stability": metrics.get("view_stability", ""),
            "risks": json.dumps(insight.get("risks") if isinstance(insight.get("risks"), list) else [], ensure_ascii=False),
            "recommendation": str(insight.get("recommendation") or ""),
        }

    def _rebuild_analysis(self, workbook, creator: dict[str, Any], metadata: dict[str, Any]) -> dict[str, Any]:
        creator_id = str(creator.get("creator_id") or "")
        insight = self._creator_row(workbook["Insights"], creator_id)
        videos = [row for row in self._rows(workbook["Videos"]) if str(row.get("creator_id") or "") == creator_id]
        risks = self._decode_json_list(insight.get("risks"))
        return {
            "schema_version": "1.0",
            "analysis_id": creator_id,
            "task_id": str(metadata.get("task_id") or ""),
            "account_uid": str(metadata.get("account_uid") or ""),
            "imported_at": str(creator.get("created_at") or ""),
            "source": str(metadata.get("source") or "excel"),
            "creator": {
                "creator_name": str(creator.get("name") or ""),
                "platform": str(creator.get("platform") or ""),
                "profile_url": str(creator.get("profile_url") or ""),
                "followers": str(creator.get("followers") or ""),
                "bio": str(creator.get("bio") or ""),
                "country": str(creator.get("country") or ""),
                "language": str(creator.get("language") or ""),
            },
            "content_category": str(creator.get("content_category") or ""),
            "tags": str(creator.get("tags") or ""),
            "video_analysis": {
                "sample_size": len(videos),
                "average_views": insight.get("average_views"),
                "median_views": insight.get("median_views"),
                "view_stability": insight.get("stability"),
            },
            "videos": videos,
            "creator_insight": {
                "level": str(creator.get("insight_level") or "insufficient"),
                "risks": risks,
                "recommendation": str(insight.get("recommendation") or ""),
                "strengths": [],
            },
        }

    def _migrate_legacy_json(self, workbook) -> None:
        if not self.legacy_analysis_dir or not self.legacy_analysis_dir.exists():
            return
        statuses = self._legacy_statuses()
        for path in self.legacy_analysis_dir.glob("analysis_task_*.json"):
            analysis, _source_path = load_json_with_backup(path)
            if not isinstance(analysis, dict):
                continue
            try:
                self._validate_analysis(analysis)
            except ValueError:
                continue
            creator_id = str(analysis["analysis_id"])
            state = statuses.get(creator_id, {})
            self._upsert_row(
                workbook["Creators"],
                "creator_id",
                creator_id,
                self._creator_values(analysis, self._status_value(state.get("status")), creator_id),
            )
            self._replace_video_rows(workbook["Videos"], creator_id, analysis.get("videos"))
            self._upsert_row(
                workbook["Insights"],
                "creator_id",
                creator_id,
                self._insight_values(analysis, creator_id),
            )
            self._upsert_row(workbook["_AnalysisData"], "creator_id", creator_id, {
                "creator_id": creator_id,
                "task_id": str(analysis.get("task_id") or ""),
                "account_uid": str(analysis.get("account_uid") or ""),
                "source": str(analysis.get("source") or "legacy_json"),
                "status_updated_at": str(state.get("updated_at") or ""),
                "analysis_json": json.dumps(analysis, ensure_ascii=False),
            })

    def _legacy_statuses(self) -> dict[str, Any]:
        if not self.legacy_library_file:
            return {}
        data, _source_path = load_json_with_backup(self.legacy_library_file)
        return data.get("records", {}) if isinstance(data, dict) and isinstance(data.get("records"), dict) else {}

    @staticmethod
    def _validate_analysis(analysis: dict[str, Any]) -> None:
        task_id = str(analysis.get("task_id") or "").strip()
        analysis_id = str(analysis.get("analysis_id") or "").strip()
        if not _TASK_ID_PATTERN.fullmatch(task_id) or analysis_id != f"analysis_{task_id}":
            raise ValueError("达人分析标识无效。")

    @staticmethod
    def _status_value(value: object) -> str:
        status = str(value or "discovered").strip()
        if status not in CREATOR_LIBRARY_STATUSES:
            raise ValueError("达人状态无效。")
        return status

    @staticmethod
    def _tags_value(analysis: dict[str, Any]) -> str:
        raw_tags = analysis.get("tags")
        if isinstance(raw_tags, list):
            return ", ".join(str(tag).strip() for tag in raw_tags if str(tag).strip())
        creator = analysis.get("creator") if isinstance(analysis.get("creator"), dict) else {}
        return str(creator.get("tags") or raw_tags or "").strip()

    @staticmethod
    def _decode_analysis(value: object) -> dict[str, Any]:
        if not isinstance(value, str) or not value.strip():
            return {}
        try:
            data = json.loads(value)
            return data if isinstance(data, dict) else {}
        except json.JSONDecodeError:
            return {}

    @staticmethod
    def _decode_json_list(value: object) -> list[Any]:
        if not isinstance(value, str):
            return []
        try:
            data = json.loads(value)
            return data if isinstance(data, list) else []
        except json.JSONDecodeError:
            return []


def _ensure_default_workbook_schema(workbook, created: bool) -> bool:
    """Domain-owned schema callback used by path-compatible repositories."""
    changed = False
    for name, headers in _WORKBOOK_SHEETS.items():
        if name not in workbook.sheetnames:
            if created and name == "Creators" and workbook.sheetnames == ["Sheet"]:
                sheet = workbook.active
                sheet.title = name
            else:
                sheet = workbook.create_sheet(name)
            CreatorRepository._set_headers(sheet, headers)
            changed = True
        else:
            changed = CreatorRepository._ensure_headers(workbook[name], headers) or changed
    workbook["_AnalysisData"].sheet_state = "hidden"
    workbook["_Metadata"].sheet_state = "hidden"
    if changed:
        CreatorRepository._upsert_row(
            workbook["_Metadata"],
            "schema_version",
            CREATOR_LIBRARY_SCHEMA_VERSION,
            {"schema_version": CREATOR_LIBRARY_SCHEMA_VERSION, "last_update_time": _utc_now()},
        )
    return changed


ExcelWorkbookStore.register_default_migration(_ensure_default_workbook_schema)

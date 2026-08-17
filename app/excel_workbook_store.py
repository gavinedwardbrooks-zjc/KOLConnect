from __future__ import annotations

"""Storage-only lifecycle management for the local Excel workbook."""

import os
import shutil
import time
from contextlib import contextmanager, nullcontext
from pathlib import Path
from typing import Any, Callable, Iterator

from openpyxl import Workbook, load_workbook
from local_storage_lock import (
    LOCAL_STORAGE_MUTATION_LOCK,
    shared_storage_lock,
    shared_storage_lock_held,
)


MigrationCallback = Callable[[Any, bool], bool | None]
BeforeSaveCallback = Callable[[Any], None]
WORKBOOK_LOCK = LOCAL_STORAGE_MUTATION_LOCK


class WorkbookReadError(RuntimeError):
    pass


class WorkbookSaveError(RuntimeError):
    pass


class ExcelWorkbookStore:
    """Own workbook I/O without knowing any domain sheets or repositories."""

    _default_migration_callbacks: list[MigrationCallback] = []

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = Path(workbook_path)
        self._migration_callbacks: list[MigrationCallback] = []
        self._before_save_callbacks: list[BeforeSaveCallback] = []
        self._scoped_workbook: Any | None = None
        self._scope_depth = 0
        self._scope_write_requested = False
        self._scope_failed = False
        self._scope_defer_writes = False
        self._scope_lock_held = False
        self._scope_shared_lock_context = None

    def register_migration(self, callback: MigrationCallback) -> None:
        if callback not in self._migration_callbacks:
            self._migration_callbacks.append(callback)

    @classmethod
    def register_default_migration(cls, callback: MigrationCallback) -> None:
        """Register schema behavior from a domain module without importing it here."""
        if callback not in cls._default_migration_callbacks:
            cls._default_migration_callbacks.append(callback)

    def register_before_save(self, callback: BeforeSaveCallback) -> None:
        if callback not in self._before_save_callbacks:
            self._before_save_callbacks.append(callback)

    @staticmethod
    def new_workbook():
        return Workbook()

    def open(self):
        if self._scope_depth > 0:
            if self._scoped_workbook is None:
                lock_context = (
                    nullcontext()
                    if shared_storage_lock_held()
                    else shared_storage_lock()
                )
                with lock_context:
                    WORKBOOK_LOCK.acquire()
                    self._scope_lock_held = True
                    try:
                        self._scoped_workbook = self._open_now()
                    except Exception:
                        self._scope_lock_held = False
                        WORKBOOK_LOCK.release()
                        raise
            return self._scoped_workbook
        with shared_storage_lock():
            return self._open_now()

    @contextmanager
    def read_only_workbook(self) -> Iterator[Any]:
        """Load an existing workbook without migrations, creation, or save hooks."""
        with WORKBOOK_LOCK:
            if not self.workbook_path.is_file():
                raise WorkbookReadError("达人库 Excel 文件不存在。")
            try:
                workbook = load_workbook(
                    self.workbook_path,
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:
                raise WorkbookReadError(f"无法读取 Excel 文件：{exc}") from exc
            try:
                yield workbook
            finally:
                workbook.close()

    def _open_now(self):
        created = not self.workbook_path.exists()
        if created:
            workbook = self.new_workbook()
        else:
            try:
                workbook = load_workbook(self.workbook_path)
            except Exception as exc:
                raise WorkbookReadError(f"无法读取 Excel 文件：{exc}") from exc

        try:
            changed = created
            callbacks = [*self._migration_callbacks, *self._default_migration_callbacks]
            for callback in callbacks:
                changed = bool(callback(workbook, created)) or changed
            if changed:
                self.save(workbook)
            return workbook
        except Exception:
            workbook.close()
            raise

    def save(self, workbook) -> None:
        if workbook is self._scoped_workbook and self._scope_defer_writes:
            self._scope_write_requested = True
            return
        with shared_storage_lock():
            with WORKBOOK_LOCK:
                self._save_now(workbook)

    def _save_now(self, workbook) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = self.workbook_path.with_suffix(".tmp.xlsx")
        backup_path = self.workbook_path.with_suffix(".xlsx.bak")
        try:
            for callback in self._before_save_callbacks:
                callback(workbook)
            workbook.save(temp_path)
            load_workbook(temp_path, read_only=True).close()
            if self.workbook_path.exists():
                shutil.copy2(self.workbook_path, backup_path)
            for attempt in range(3):
                try:
                    os.replace(temp_path, self.workbook_path)
                    break
                except PermissionError:
                    if attempt == 2:
                        raise
                    time.sleep(0.1 * (attempt + 1))
        except PermissionError as exc:
            raise WorkbookSaveError(
                "无法保存 Excel 文件。请先关闭 WPS 或 Excel 中打开的该文件。"
            ) from exc
        finally:
            temp_path.unlink(missing_ok=True)

    def create_backup(self, suffix: str) -> Path:
        backup_path = self.workbook_path.with_name(
            f"{self.workbook_path.stem}{suffix}{self.workbook_path.suffix}"
        )
        with shared_storage_lock():
            with WORKBOOK_LOCK:
                shutil.copy2(self.workbook_path, backup_path)
        return backup_path

    def create_transaction_backup(self, backup_path: Path) -> Path:
        """Create and validate one exact backup for a recoverable local transaction."""
        backup_path = Path(backup_path)
        with shared_storage_lock(), WORKBOOK_LOCK:
            if not self.workbook_path.is_file():
                raise WorkbookReadError("达人库 Excel 文件不存在。")
            backup_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(self.workbook_path, backup_path)
            try:
                load_workbook(backup_path, read_only=True).close()
            except Exception as exc:
                backup_path.unlink(missing_ok=True)
                raise WorkbookReadError("事务工作簿备份验证失败。") from exc
        return backup_path

    def restore_transaction_backup(self, backup_path: Path) -> None:
        """Validate and atomically restore an explicitly selected transaction backup."""
        backup_path = Path(backup_path)
        with shared_storage_lock(), WORKBOOK_LOCK:
            if not backup_path.is_file():
                raise WorkbookReadError("事务工作簿备份不存在。")
            try:
                load_workbook(backup_path, read_only=True).close()
            except Exception as exc:
                raise WorkbookReadError("事务工作簿备份无效。") from exc
            self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
            restore_path = self.workbook_path.with_suffix(".restore.tmp.xlsx")
            try:
                shutil.copy2(backup_path, restore_path)
                load_workbook(restore_path, read_only=True).close()
                os.replace(restore_path, self.workbook_path)
            except Exception as exc:
                raise WorkbookSaveError("事务工作簿恢复失败。") from exc
            finally:
                restore_path.unlink(missing_ok=True)

    @contextmanager
    def workbook(self, *, write: bool = False) -> Iterator[Any]:
        """Open one operation, or reuse an explicitly active request scope."""
        if self._scope_depth > 0:
            workbook = self.open()
            try:
                yield workbook
                if write:
                    self.save(workbook)
            except Exception:
                self._scope_failed = True
                raise
            return

        if write:
            with shared_storage_lock(), WORKBOOK_LOCK:
                workbook = self.open()
                try:
                    yield workbook
                    self.save(workbook)
                finally:
                    workbook.close()
            return

        with shared_storage_lock():
            WORKBOOK_LOCK.acquire()
            try:
                workbook = self.open()
            except Exception:
                WORKBOOK_LOCK.release()
                raise
        try:
            yield workbook
        finally:
            workbook.close()
            WORKBOOK_LOCK.release()

    @contextmanager
    def scope(
        self,
        *,
        write: bool = False,
        defer_writes: bool = False,
    ) -> Iterator[ExcelWorkbookStore]:
        """Allow request-scoped repositories to share one workbook explicitly."""
        if self._scope_depth == 0:
            self._scope_write_requested = write
            self._scope_failed = False
            self._scope_defer_writes = defer_writes
            if defer_writes:
                self._scope_shared_lock_context = shared_storage_lock()
                self._scope_shared_lock_context.__enter__()
        else:
            self._scope_write_requested = self._scope_write_requested or write
            self._scope_defer_writes = self._scope_defer_writes or defer_writes
        self._scope_depth += 1
        try:
            yield self
        except Exception:
            self._scope_failed = True
            raise
        finally:
            self._scope_depth -= 1
            if self._scope_depth == 0:
                workbook = self._scoped_workbook
                try:
                    if (
                        workbook is not None
                        and self._scope_write_requested
                        and not self._scope_failed
                        and self._scope_defer_writes
                    ):
                        self._save_now(workbook)
                finally:
                    if workbook is not None:
                        workbook.close()
                    self._scoped_workbook = None
                    self._scope_write_requested = False
                    self._scope_failed = False
                    self._scope_defer_writes = False
                    if self._scope_lock_held:
                        self._scope_lock_held = False
                        WORKBOOK_LOCK.release()
                    if self._scope_shared_lock_context is not None:
                        context = self._scope_shared_lock_context
                        self._scope_shared_lock_context = None
                        context.__exit__(None, None, None)

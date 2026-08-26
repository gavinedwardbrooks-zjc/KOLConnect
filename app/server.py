from __future__ import annotations

"""KOL联系助手本地服务。

职责保持简单：
1. 为桌面前端提供状态与设置接口。
2. 启动、停止并读取 scraper.py 抓取任务。
3. 保存本地配置，打开结果文件并触发飞书同步。
"""

import csv
import io
import json
import imaplib
import mail_sync as mail_sync_module
import os
import re
import smtplib
import subprocess
import sys
import task_manager
import threading
import time
import webbrowser
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import scraper as scraper_module
from adapters.task_manager_adapter import TaskManagerAdapter
from campaign_creator_repository import CampaignCreatorRepository
from campaign_repository import CampaignRepository
from creator_repository import CreatorRepository
from dashboard_repository import DashboardRepository
from dashboard_service import DashboardService
from product_repository import ProductRepository
from ports.creator_delete_impact_port import CreatorDeleteImpactPort
from ports.creator_port import (
    CreatorAnalysisSnapshot,
    CreatorImportResult,
    CreatorImportSummary,
    CreatorPort,
    EmailRecheckCandidateScan,
    ExternalAgencyContact,
    ExternalAgencyContactCommand,
    FourTableSyncCommand,
    FourTableSyncResult,
    ImportTaskResultsCommand,
    ManualTaskPreparationCommand,
    ManualTaskProtectionCommand,
    PreparedManualTask,
    PreparedFourTableSync,
    PreparedTaskResultUpdate,
    TaskResultImportCommand,
    TaskResultUpdateCommand,
)
from ports.task_port import (
    CreatorImportLinkage,
    ManualTaskCreateCommand,
    ManualTaskInitializationCommand,
    ManualReviewTaskCommand,
    TaskPort,
)
from repositories.task_repository import TaskRepository
from repositories.agency_repository import AgencyRepository
from repository_factory import RepositoryFactory, get_active_repository_factory
from services.agency_service import AgencyService
from services.analytics_service import AnalyticsService
from services.workbook_backup_service import WorkbookBackupService
from services.creator_delete_impact_service import CreatorDeleteImpactService
from services.creator_hard_delete_service import CreatorHardDeleteService
from services.creator_merge_service import CreatorMergeService
from services.creator_library_cache import CreatorLibraryCache
from services.campaign_creator_service import CampaignCreatorService
from services.dashboard_response_cache import DashboardResponseCache
from services.creator_service import CreatorService
from services.creator_summary_service import CreatorSummaryService
from services.creator_intelligence_service import (
    CreatorIntelligenceService,
    CreatorIntelligenceSummaryFacade,
)
from services.clean_reset_service import CleanResetService
from services.assistant_service import AssistantService
from services.assistant_provider import DeterministicAssistantProvider
from services.task_service import TaskService
from services.mail_auth_service import classify_imap_error
from services.risk_service import RiskService
from app_logging import log_error, log_event
from api_contract import (
    error_payload,
    get_trace_id,
    new_trace_id,
    set_trace_id,
    success_payload,
)
from openpyxl import load_workbook
from runtime_paths import (
    get_app_data_dir,
    get_logs_dir,
    get_resource_dir,
    atomic_write_json,
    json_backup_path,
    load_json_with_backup,
    scraper_worker_command,
)
from local_storage_lock import shared_storage_lock
from staged_delete_transaction import recover_pending_delete_transactions
from feishu_client import FeishuClient
from local_request_security import (
    MUTATING_METHODS,
    allowed_host_header,
    allowed_mutation_origin,
    browser_shutdown_allowed,
)
from version import APP_DISPLAY_VERSION
from http_handlers import (
    analytics_handler,
    assistant_handler,
    campaign_handler,
    clean_reset_handler,
    creator_handler,
    dashboard_handler,
    feishu_chat_handler,
    feishu_delete_handler,
    feishu_sync_handler,
    settings_handler,
    storage_migration_handler,
    task_handler,
    risk_handler,
)
from services.feishu_sync_service import FeishuSyncService
from services.feishu_chat_transport import FeishuChatTransport
from services.feishu_delete_intent_service import (
    FeishuDeleteIntentStore,
    FeishuDeleteReconciliationService,
)
from services.production_migration_service import ProductionMigrationService
from storage.paths import SQLiteStoragePaths


APP_DIR = get_resource_dir()
DATA_DIR = get_app_data_dir()
LOGS_DIR = get_logs_dir()
STATIC_DIR = APP_DIR / "webapp"
STATE_FILE = DATA_DIR / "settings.json"
TASKS_DIR = DATA_DIR / "tasks"
DATA_PROTECTION_FILE = DATA_DIR / "data_protection.json"
CREATOR_ANALYSIS_DIR = DATA_DIR / "creator_analysis"
CREATOR_LIBRARY_FILE = DATA_DIR / "creator_library.json"
DEFAULT_CREATOR_LIBRARY_WORKBOOK = DATA_DIR / "Creator_Library.xlsx"
RUN_LOG_FILE = LOGS_DIR / "kolconnect.log"
DIAGNOSTICS_FILE = DATA_DIR / "system_diagnostics.json"
HOST = "127.0.0.1"
PORT = 8765
SENSITIVE_MASK = "********"
TASK_HEARTBEAT_SECONDS = 240
TASK_INTERRUPTION_TIMEOUT_SECONDS = 15 * 60

REVIEW_FIELD_WHATSAPP = "WhatsApp"
REVIEW_FIELD_NOTE = "备注"
REVIEW_FIELD_DATA_STATUS = "数据状态"
REVIEW_FIELD_MODIFIED_AT = "最后修改时间"
REVIEW_CSV_FIELDS = [
    REVIEW_FIELD_WHATSAPP,
    REVIEW_FIELD_NOTE,
    REVIEW_FIELD_DATA_STATUS,
    REVIEW_FIELD_MODIFIED_AT,
]
REVIEW_EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
REVIEW_WHATSAPP_PATTERN = re.compile(r"^[0-9+()\- ]+$")
TASK_PLATFORM_OPTIONS = {"全部", "TikTok", "Instagram", "YouTube"}
PLATFORM_LABELS = {"tiktok": "TikTok", "instagram": "Instagram", "youtube": "YouTube"}
RETRYABLE_SCRAPE_STATUSES = {"missing_data", "failed", "login_required", "platform_error"}
LEGACY_COOPERATION_READ_ONLY_MESSAGE = "请使用 Campaign 创建新的合作。"
LEGACY_COOPERATION_PATH_PATTERN = re.compile(r"/api/creator-library/[^/]+/cooperations")
BLOCKING_SCRAPE_STATUSES = {"missing_data", "failed", "login_required", "platform_error"}
INSTAGRAM_ERROR_STATUSES = {"failed", "login_required", "platform_error"}
AGENCY_CONTACT_FIELD_NAME = "联系人姓名"
AGENCY_CONTACT_FIELD_WHATSAPP = "WhatsApp"
AGENCY_CONTACT_FIELD_AGENCY = "所属 Agency"
CHROME_USER_DATA = Path.home() / "AppData" / "Local" / "Google" / "Chrome" / "User Data"
KOLCONNECT_CHROME_USER_DATA = DATA_DIR / "ChromeProfile"
AUTOMATION_PROFILE_NAME = "KOLConnect Automation"
AUTOMATION_PROFILE_DIRECTORY = "Default"
CHROME_EXE_CANDIDATES = [
    Path(os.environ.get("PROGRAMFILES", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
    Path(os.environ.get("PROGRAMFILES(X86)", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
    Path(os.environ.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
]

DEFAULT_STATE = {
    "ui": {"language": "zh", "debug_mode": False},
    "profiles": {"selected": AUTOMATION_PROFILE_NAME},
    "accounts": {"entries": []},
    "feishu": {
        "app_id": "",
        "app_secret": "",
        "app_token": "",
        "creator_table_id": "",
        "account_table_id": "",
        "agency_table_id": "",
        "contact_table_id": "",
        "chat_enabled": False,
    },
    "mail": {
        "accounts": [],
        "template_subject": "",
        "template_body": "",
    },
    "creator_library": {
        "workbook_path": str(DEFAULT_CREATOR_LIBRARY_WORKBOOK),
    },
}

MAIL_PROVIDER_PRESETS = {
    "gmail": {
        "imap_host": "imap.gmail.com",
        "imap_port": "993",
        "smtp_host": "smtp.gmail.com",
        "smtp_port": "587",
    },
    "netease": {
        "imap_host": "imap.163.com",
        "imap_port": "993",
        "smtp_host": "smtp.163.com",
        "smtp_port": "465",
    },
    "outlook": {
        "imap_host": "outlook.office365.com",
        "imap_port": "993",
        "smtp_host": "smtp.office365.com",
        "smtp_port": "587",
    },
}

HANDLERS = [
    assistant_handler,
    analytics_handler,
    dashboard_handler,
    risk_handler,
    campaign_handler,
    feishu_chat_handler,
    feishu_delete_handler,
    feishu_sync_handler,
    clean_reset_handler,
    storage_migration_handler,
    settings_handler,
    creator_handler,
    task_handler,
]
CREATOR_LIBRARY_CACHE = CreatorLibraryCache()
DASHBOARD_RESPONSE_CACHE = DashboardResponseCache(
    build_event_logger=lambda message: log_event("DashboardCache", message)
)
ASSISTANT_SERVICE = None
FEISHU_CHAT_TRANSPORT = None
PRODUCTION_MIGRATION_SERVICE = None


def get_mail_provider_preset(provider: str) -> dict[str, str]:
    return MAIL_PROVIDER_PRESETS.get(provider, {})


def normalize_mail_account(raw: dict | None) -> dict:
    raw = raw or {}
    provider = str(raw.get("provider") or "custom").strip().lower() or "custom"
    preset = get_mail_provider_preset(provider)
    return {
        "name": str(raw.get("name") or "").strip(),
        "provider": provider if provider in {"aliyun", "netease", "gmail", "outlook", "custom"} else "custom",
        "email": str(raw.get("email") or "").strip(),
        "sender_name": str(raw.get("sender_name") or "").strip(),
        "imap_host": str(raw.get("imap_host") or preset.get("imap_host") or "").strip(),
        "imap_port": str(raw.get("imap_port") or preset.get("imap_port") or "").strip(),
        "smtp_host": str(raw.get("smtp_host") or preset.get("smtp_host") or "").strip(),
        "smtp_port": str(raw.get("smtp_port") or preset.get("smtp_port") or "").strip(),
        "username": str(raw.get("username") or "").strip(),
        "password": str(raw.get("password") or ""),
        "enabled": bool(raw.get("enabled")),
    }


def normalize_mail_state(raw_mail: dict | None) -> dict:
    raw_mail = raw_mail or {}
    if isinstance(raw_mail.get("accounts"), list):
        return {
            "accounts": [
                normalize_mail_account(item)
                for item in raw_mail["accounts"]
                if isinstance(item, dict) and str(item.get("name") or item.get("email") or item.get("username") or "").strip()
            ],
            "template_subject": str(raw_mail.get("template_subject") or ""),
            "template_body": str(raw_mail.get("template_body") or ""),
        }
    return clone_default_state()["mail"]


def is_sensitive_mask(value: object) -> bool:
    return str(value or "").strip() == SENSITIVE_MASK


def mail_account_identity(account: dict) -> tuple[str, str, str]:
    """Use stable account values to retain a masked password during saves."""
    return (
        str(account.get("email") or "").strip().lower(),
        str(account.get("username") or "").strip().lower(),
        str(account.get("name") or "").strip().lower(),
    )


def merge_masked_mail_passwords(raw_mail: dict | None, existing_mail: dict | None) -> dict:
    """Replace API mask placeholders with the already saved account password."""
    merged = dict(raw_mail) if isinstance(raw_mail, dict) else {}
    raw_accounts = merged.get("accounts")
    if not isinstance(raw_accounts, list):
        return merged

    existing_accounts = existing_mail.get("accounts") if isinstance(existing_mail, dict) else []
    existing_by_identity = {
        mail_account_identity(account): account
        for account in existing_accounts
        if isinstance(account, dict) and any(mail_account_identity(account))
    }
    merged_accounts = []
    for raw_account in raw_accounts:
        if not isinstance(raw_account, dict):
            continue
        account = dict(raw_account)
        if is_sensitive_mask(account.get("password")):
            saved_account = existing_by_identity.get(mail_account_identity(account))
            account["password"] = str(saved_account.get("password") or "") if saved_account else ""
        merged_accounts.append(account)
    merged["accounts"] = merged_accounts
    return merged


def state_for_client(state: dict) -> dict:
    """Return settings needed by the UI without exposing saved secrets."""
    client_state = json.loads(json.dumps(state))
    feishu = client_state.get("feishu") if isinstance(client_state.get("feishu"), dict) else {}
    if feishu.get("app_secret"):
        feishu["app_secret"] = SENSITIVE_MASK
    if feishu.get("app_token"):
        feishu["app_token"] = SENSITIVE_MASK
    mail = client_state.get("mail") if isinstance(client_state.get("mail"), dict) else {}
    accounts = mail.get("accounts") if isinstance(mail.get("accounts"), list) else []
    for account in accounts:
        if isinstance(account, dict) and account.get("password"):
            account["password"] = SENSITIVE_MASK
    return client_state


def parse_port(value: str, default: int) -> int:
    try:
        port = int(str(value or "").strip())
        return port if port > 0 else default
    except Exception:
        return default


def test_imap_login(account: dict) -> None:
    host = str(account.get("imap_host") or "").strip()
    username = str(account.get("username") or "").strip()
    password = str(account.get("password") or "")
    port = parse_port(account.get("imap_port") or "", 993)
    if not host or not username or not password:
        raise RuntimeError("请完整填写 IMAP Host、IMAP Port、用户名和密码/授权码。")
    client = None
    try:
        if port == 993:
            client = imaplib.IMAP4_SSL(host, port, timeout=10)
        else:
            client = imaplib.IMAP4(host, port, timeout=10)
            if hasattr(client, "starttls"):
                client.starttls()
        client.login(username, password)
    except Exception as exc:
        raise classify_imap_error(exc) from exc
    finally:
        if client is not None:
            try:
                client.logout()
            except Exception:
                pass


def test_smtp_login(account: dict) -> None:
    host = str(account.get("smtp_host") or "").strip()
    username = str(account.get("username") or "").strip()
    password = str(account.get("password") or "")
    port = parse_port(account.get("smtp_port") or "", 587)
    if not host or not username or not password:
        raise RuntimeError("请完整填写 SMTP Host、SMTP Port、用户名和密码/授权码。")
    client = None
    try:
        if port == 465:
            client = smtplib.SMTP_SSL(host, port, timeout=10)
            client.ehlo()
        else:
            client = smtplib.SMTP(host, port, timeout=10)
            client.ehlo()
            if client.has_extn("starttls"):
                client.starttls()
                client.ehlo()
        client.login(username, password)
    except Exception as exc:
        raise RuntimeError(f"SMTP 登录失败：{exc}") from exc
    finally:
        if client is not None:
            try:
                client.quit()
            except Exception:
                pass

def clone_default_state() -> dict:
    return json.loads(json.dumps(DEFAULT_STATE))


def get_profiles() -> list[str]:
    profiles: list[str] = [AUTOMATION_PROFILE_NAME]
    if not CHROME_USER_DATA.exists():
        return profiles + ["Default"]

    for child in CHROME_USER_DATA.iterdir():
        if child.is_dir() and (child.name == "Default" or child.name.startswith("Profile ")):
            profiles.append(child.name)
    chrome_profiles = sorted({name for name in profiles if name != AUTOMATION_PROFILE_NAME}, key=lambda name: (name != "Default", name))
    return [AUTOMATION_PROFILE_NAME, *chrome_profiles] or [AUTOMATION_PROFILE_NAME, "Default"]


def resolve_chrome_launch_config(profile: str | None) -> tuple[Path, str]:
    selected = (profile or "").strip() or AUTOMATION_PROFILE_NAME
    if selected == AUTOMATION_PROFILE_NAME:
        KOLCONNECT_CHROME_USER_DATA.mkdir(parents=True, exist_ok=True)
        return KOLCONNECT_CHROME_USER_DATA, AUTOMATION_PROFILE_DIRECTORY
    return CHROME_USER_DATA, selected


def normalize_state(raw: dict | None) -> dict:
    """Normalize persisted application state into the current structure."""
    state = clone_default_state()
    raw = raw or {}

    if isinstance(raw.get("ui"), dict):
        state["ui"].update(raw["ui"])
    if state["ui"].get("language") not in {"zh", "en"}:
        state["ui"]["language"] = "zh"
    state["ui"]["debug_mode"] = bool(state["ui"].get("debug_mode"))

    if isinstance(raw.get("profiles"), dict):
        state["profiles"].update(raw["profiles"])
    if not state["profiles"].get("selected"):
        state["profiles"]["selected"] = AUTOMATION_PROFILE_NAME
    elif state["profiles"]["selected"] == "Default":
        state["profiles"]["selected"] = AUTOMATION_PROFILE_NAME

    if isinstance(raw.get("accounts"), dict) and isinstance(raw["accounts"].get("entries"), list):
        state["accounts"]["entries"] = [
            {
                "profile": str(item.get("profile") or "").strip(),
                "alias": str(item.get("alias") or "").strip(),
                "usage": str(item.get("usage") or "通用").strip() or "通用",
            }
            for item in raw["accounts"]["entries"]
            if isinstance(item, dict) and str(item.get("profile") or "").strip()
        ]

    if isinstance(raw.get("feishu"), dict):
        feishu = raw["feishu"]
        state["feishu"]["app_id"] = str(feishu.get("app_id") or "").strip()
        state["feishu"]["app_secret"] = str(feishu.get("app_secret") or "").strip()
        for key in (
            "app_token",
            "creator_table_id",
            "account_table_id",
            "agency_table_id",
            "contact_table_id",
        ):
            state["feishu"][key] = str(feishu.get(key) or "").strip()
        state["feishu"]["chat_enabled"] = bool(feishu.get("chat_enabled"))

    if isinstance(raw.get("mail"), dict):
        state["mail"] = normalize_mail_state(raw["mail"])

    if isinstance(raw.get("creator_library"), dict):
        state["creator_library"]["workbook_path"] = normalize_creator_library_workbook_path(
            raw["creator_library"].get("workbook_path")
        )

    return state


def normalize_creator_library_workbook_path(value: object) -> str:
    raw_path = str(value or "").strip()
    if not raw_path:
        return str(DEFAULT_CREATOR_LIBRARY_WORKBOOK)
    path = Path(os.path.expandvars(os.path.expanduser(raw_path)))
    if path.suffix.lower() != ".xlsx":
        raise ValueError("达人库文件必须是 .xlsx 格式。")
    return str(path.resolve())


def load_state() -> dict:
    data, source_path = load_json_with_backup(STATE_FILE)
    if data is None:
        if STATE_FILE.exists() or json_backup_path(STATE_FILE).exists():
            RUN_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
            with RUN_LOG_FILE.open("a", encoding="utf-8") as handle:
                handle.write("设置文件损坏且无法恢复，已使用默认配置。\n")
        return normalize_state(None)
    if source_path == json_backup_path(STATE_FILE):
        RUN_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with RUN_LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write("设置文件损坏，已从 settings.json.bak 恢复。\n")
    return normalize_state(data if isinstance(data, dict) else None)


def save_state(state: dict) -> None:
    atomic_write_json(STATE_FILE, state)


STATE = load_state()


def _load_diagnostics() -> dict:
    data, _source_path = load_json_with_backup(DIAGNOSTICS_FILE)
    return data if isinstance(data, dict) else {}


DIAGNOSTICS = _load_diagnostics()
DIAGNOSTICS_LOCK = threading.Lock()


def _save_diagnostics() -> None:
    with shared_storage_lock(), DIAGNOSTICS_LOCK:
        atomic_write_json(DIAGNOSTICS_FILE, DIAGNOSTICS)


def _record_diagnostic(key: str, value: dict) -> None:
    with shared_storage_lock(), DIAGNOSTICS_LOCK:
        DIAGNOSTICS[key] = value
        atomic_write_json(DIAGNOSTICS_FILE, DIAGNOSTICS)


def _record_last_error(message: str) -> None:
    _record_diagnostic("last_error", {"message": message, "time": _utc_now()})


def _diagnostic_timestamp_is_stale(value: object, days: int = 7) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    try:
        timestamp = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return True
    if timestamp.tzinfo is None:
        timestamp = timestamp.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - timestamp.astimezone(timezone.utc)).total_seconds() > days * 86400


def _friendly_error_message(exc: BaseException | str) -> str:
    raw = str(exc or "").strip()
    lowered = raw.lower()
    if isinstance(exc, PermissionError) or "permissionerror" in lowered or "permission denied" in lowered:
        return "Excel 文件正在被其他程序使用，请关闭 WPS 或 Excel 后重试。"
    if "excel" in lowered:
        return "Excel 文件处理失败，请确认文件可读取且未被 WPS 或 Excel 占用。"
    if "connectionrefused" in lowered or "failed to establish a new connection" in lowered:
        return "服务连接失败，请确认 KOLConnect 正在运行。"
    if "traceback" in lowered:
        return "操作失败，请查看系统日志中的详细原因。"
    return raw or "操作失败，请查看系统日志中的详细原因。"


def get_system_health() -> dict:
    """Run read-only local checks used by the settings diagnostics panel."""
    workbook_path = Path(STATE.get("creator_library", {}).get("workbook_path") or DEFAULT_CREATOR_LIBRARY_WORKBOOK)
    required_sheets = {
        "Creators", "CreatorAccounts", "Videos", "Insights", "Cooperations",
        "CreatorSnapshots", "VideoSnapshots", "Agencies", "AgencyContacts",
        "FollowUpLogs", "_Metadata",
    }
    checks: list[dict] = [
        {"key": "data_directory", "label": "数据目录", "status": "ok", "message": str(DATA_DIR)},
        {"key": "api", "label": "API 服务", "status": "ok", "message": f"{HOST}:{PORT}"},
    ]
    excel_status = "warning"
    excel_message = "文件尚未创建，首次导入达人时会自动创建。"
    sheets_message = "等待创建 Excel 文件后检查。"
    if workbook_path.exists():
        try:
            workbook = load_workbook(workbook_path, read_only=True)
            sheet_names = set(workbook.sheetnames)
            workbook.close()
            excel_status = "ok"
            excel_message = str(workbook_path)
            missing_sheets = sorted(required_sheets - sheet_names)
            if missing_sheets:
                sheets_message = f"缺少工作表：{', '.join(missing_sheets)}"
                sheets_status = "error"
            else:
                sheets_message = "工作表结构完整。"
                sheets_status = "ok"
        except Exception as exc:
            excel_status = "error"
            excel_message = _friendly_error_message(exc)
            sheets_status = "error"
            sheets_message = "无法读取 Excel 工作簿。"
    else:
        sheets_status = "warning"
    checks.extend(
        [
            {"key": "excel", "label": "Excel 文件", "status": excel_status, "message": excel_message},
            {"key": "excel_structure", "label": "Excel 结构", "status": sheets_status, "message": sheets_message},
        ]
    )
    last_import = DIAGNOSTICS.get("last_extension_import") if isinstance(DIAGNOSTICS.get("last_extension_import"), dict) else {}
    import_time = str(last_import.get("time") or "")
    checks.append(
        {
            "key": "extension", "label": "Chrome 插件", "status": "warning" if _diagnostic_timestamp_is_stale(import_time) else "ok",
            "message": "最近 7 天无导入记录。" if _diagnostic_timestamp_is_stale(import_time) else f"最近导入：{import_time}",
        }
    )
    feishu = STATE.get("feishu") if isinstance(STATE.get("feishu"), dict) else {}
    required_feishu = ("app_id", "app_secret", "app_token")
    missing_feishu = [key for key in required_feishu if not str(feishu.get(key) or "").strip()]
    checks.append(
        {
            "key": "feishu", "label": "飞书配置", "status": "warning" if missing_feishu else "ok",
            "message": f"缺少：{', '.join(missing_feishu)}" if missing_feishu else "配置已填写，未执行真实连接测试。",
        }
    )
    overall = "error" if any(item["status"] == "error" for item in checks) else "warning" if any(item["status"] == "warning" for item in checks) else "ok"
    last_error = DIAGNOSTICS.get("last_error") if isinstance(DIAGNOSTICS.get("last_error"), dict) else {}
    return {
        "status": overall,
        "checks": checks,
        "debug": {
            "version": APP_DISPLAY_VERSION,
            "api_status": "正常",
            "excel_path": str(workbook_path),
            "excel_status": excel_status,
            "last_extension_import": last_import,
            "last_error": last_error,
        },
    }


def find_chrome_exe() -> Path | None:
    for candidate in CHROME_EXE_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


def open_chrome_profile(profile: str) -> None:
    chrome_exe = find_chrome_exe()
    if not chrome_exe:
        raise RuntimeError("未找到 Chrome，请先安装 Google Chrome。")
    user_data_dir, profile_directory = resolve_chrome_launch_config(profile)
    subprocess.Popen(
        [
            str(chrome_exe),
            f"--user-data-dir={user_data_dir}",
            f"--profile-directory={profile_directory}",
        ],
        cwd=str(DATA_DIR),
    )


def get_four_table_feishu_config() -> dict:
    """Return named four-table settings without falling back to a legacy table."""
    return {
        "app_id": STATE["feishu"].get("app_id", ""),
        "app_secret": STATE["feishu"].get("app_secret", ""),
        "app_token": STATE["feishu"].get("app_token", ""),
        "creator_table_id": STATE["feishu"].get("creator_table_id", ""),
        "account_table_id": STATE["feishu"].get("account_table_id", ""),
        "agency_table_id": STATE["feishu"].get("agency_table_id", ""),
        "contact_table_id": STATE["feishu"].get("contact_table_id", ""),
    }


def _feishu_relation_labels(value) -> list[str]:
    """Read relation display text without using it as an identifier."""
    items = value if isinstance(value, list) else [value]
    labels: list[str] = []
    for item in items:
        if isinstance(item, dict):
            label = str(item.get("text") or item.get("name") or "").strip()
        else:
            label = str(item or "").strip()
        if label and label not in labels:
            labels.append(label)
    return labels


def get_agency_contact_options() -> dict:
    """Read contact choices for manual tasks; no records are changed here."""
    config = get_four_table_feishu_config()
    contact_table_id = str(config.get("contact_table_id") or "").strip()
    if not contact_table_id:
        return {"configured": False, "contacts": []}

    required = ("app_id", "app_secret", "app_token", "contact_table_id")
    missing = [key for key in required if not str(config.get(key) or "").strip()]
    if missing:
        raise RuntimeError(f"Agency联系人表飞书配置不完整：缺少 {', '.join(missing)}。")

    access_token = scraper_module._four_table_access_token(config)
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    contacts: list[dict] = []
    for item in scraper_module._four_table_list_records(contact_table_id, config, headers):
        fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
        record_id = str(item.get("record_id") or "").strip()
        if not record_id:
            continue
        contacts.append(
            {
                "record_id": record_id,
                "name": str(fields.get(AGENCY_CONTACT_FIELD_NAME) or "").strip(),
                "whatsapp": str(fields.get(AGENCY_CONTACT_FIELD_WHATSAPP) or "").strip(),
                "agencies": _feishu_relation_labels(fields.get(AGENCY_CONTACT_FIELD_AGENCY)),
            }
        )
    contacts.sort(key=lambda item: (not bool(item["name"]), item["name"].casefold(), item["record_id"]))
    return {"configured": True, "contacts": contacts}


def _resolve_source_contact(source_contact_record_id: object) -> dict | None:
    """Validate a selected contact by stable Feishu record ID before saving a manual task."""
    record_id = str(source_contact_record_id or "").strip()
    if not record_id:
        return None
    options = get_agency_contact_options()
    if not options["configured"]:
        raise ValueError("未配置 Agency联系人表，无法选择来源联系人。")
    for contact in options["contacts"]:
        if contact["record_id"] == record_id:
            return contact
    raise ValueError("来源联系人不存在或已被删除，请重新选择。")


@dataclass
class ScrapeJob:
    running: bool = False
    process: subprocess.Popen[str] | None = None
    logs: list[str] = field(default_factory=list)
    task_id: str = ""
    results_file: Path | None = None
    pause_requested: bool = False
    stop_requested: bool = False
    lock: threading.Lock = field(default_factory=threading.Lock)

    def append(self, text: str) -> None:
        with self.lock:
            self.logs.append(text)
            self.logs = self.logs[-1000:]
            log_event("Scraper", text.strip())

    def snapshot(self) -> dict:
        with self.lock:
            status = "idle"
            if self.running:
                status = "stopping" if self.stop_requested else "paused" if self.pause_requested else "running"
                if self.task_id:
                    try:
                        task = get_task_service().get_runtime_task_snapshot(self.task_id)
                        persisted_status = task.status
                        if persisted_status in {"running", "finalizing", "paused", "stopping"}:
                            status = persisted_status
                    except ValueError:
                        pass
            return {
                "running": self.running,
                "status": status,
                "pause_requested": self.pause_requested,
                "stop_requested": self.stop_requested,
                "logs": "".join(self.logs),
                "has_results": bool(self.results_file and self.results_file.exists()),
                "task_id": self.task_id,
            }


SCRAPE_JOB = ScrapeJob()


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _task_timestamp_is_stale(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    try:
        timestamp = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return True
    if timestamp.tzinfo is None:
        timestamp = timestamp.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - timestamp.astimezone(timezone.utc)).total_seconds() > TASK_INTERRUPTION_TIMEOUT_SECONDS


def detect_interrupted_tasks() -> int:
    """Mark running tasks with no live worker or a stale heartbeat as interrupted."""
    interrupted = 0
    recovered_stopping = 0
    for task in get_task_service().list_recovery_candidates():
        status = task.status
        task_id = task.task_id
        active_worker = (
            SCRAPE_JOB.running
            and SCRAPE_JOB.task_id == task_id
            and SCRAPE_JOB.process is not None
            and SCRAPE_JOB.process.poll() is None
        )
        if status == "stopping" and not active_worker:
            get_task_service().recover_stopping_task(task_id, finished_at=task.finished_at or _utc_now())
            recovered_stopping += 1
            continue
        if status != "running":
            continue
        heartbeat = task.heartbeat_time or task.started_at
        if active_worker:
            continue
        reason = (
            "任务心跳超时，可能由于程序关闭、电脑异常退出或进程异常结束"
            if _task_timestamp_is_stale(heartbeat)
            else "任务执行进程不存在，可能由于程序关闭、Chrome 关闭或进程异常结束"
        )
        get_task_service().mark_task_interrupted(task_id, interrupted_at=_utc_now(), reason=reason)
        interrupted += 1
    if interrupted:
        SCRAPE_JOB.append(f"检测到任务异常中断：{interrupted} 个。\n")
    if recovered_stopping:
        SCRAPE_JOB.append(f"已恢复无活动进程的停止中任务：{recovered_stopping} 个。\n")
    SCRAPE_JOB.append("任务状态检查完成。\n")
    return interrupted + recovered_stopping


def _task_next_pending_item(task_paths: dict[str, Path]) -> str:
    try:
        links = [line.strip() for line in task_paths["links"].read_text(encoding="utf-8").splitlines() if line.strip()]
    except OSError:
        return ""
    completed = set(scraper_module.load_progress(str(task_paths["progress"])))
    return next((url for url in links if url not in completed), "")


def _instagram_error_count(progress_file: Path) -> int:
    """Return the current consecutive Instagram acquisition failures from task-local progress."""
    if not progress_file.exists():
        return 0
    consecutive = 0
    try:
        with progress_file.open(encoding="utf-8-sig", newline="", errors="ignore") as handle:
            for row in csv.DictReader(handle):
                if str(row.get(scraper_module.FIELD_PLATFORM) or "").strip() != "Instagram":
                    continue
                status = str(row.get(scraper_module.FIELD_SCRAPE_STATUS) or "success").strip()
                if status in INSTAGRAM_ERROR_STATUSES:
                    consecutive += 1
                else:
                    consecutive = 0
    except OSError:
        return 0
    return consecutive


def _monitor_scrape_task(task_id: str, task_paths: dict[str, Path], stop_event: threading.Event) -> None:
    """Update task liveness and progress without touching scraper output files."""
    last_completed = len(scraper_module.load_progress(str(task_paths["progress"])))
    last_heartbeat = time.monotonic()
    while not stop_event.wait(2):
        if not SCRAPE_JOB.running or SCRAPE_JOB.task_id != task_id:
            return
        try:
            task = get_task_service().get_runtime_task_snapshot(task_id)
            changes: dict[str, object] = {}
            completed = len(scraper_module.load_progress(str(task_paths["progress"])))
            if completed > last_completed:
                changes["last_progress_time"] = _utc_now()
                changes["current_item"] = _task_next_pending_item(task_paths)
                changes["completed_count"] = completed
                changes["last_successful_index"] = completed
                last_completed = completed
            instagram_errors = _instagram_error_count(task_paths["progress"])
            if instagram_errors >= 5:
                changes.update(
                    instagram_error_count=instagram_errors,
                    instagram_status="login_required",
                    instagram_message="Instagram登录状态异常，请重新登录后继续。",
                )
            if task.status in {"running", "stopping"} and time.monotonic() - last_heartbeat >= TASK_HEARTBEAT_SECONDS:
                changes["heartbeat_time"] = _utc_now()
                last_heartbeat = time.monotonic()
                SCRAPE_JOB.append("任务心跳更新。\n")
            if changes:
                get_task_service().persist_runtime_progress(
                    task_id,
                    RuntimeProgressUpdate(
                        completed_count=int(changes.get("completed_count") or 0),
                        last_successful_index=int(changes.get("last_successful_index") or 0),
                        current_item=str(changes.get("current_item") or ""),
                        last_progress_time=str(changes.get("last_progress_time") or ""),
                        heartbeat_time=str(changes.get("heartbeat_time") or ""),
                        instagram_error_count=int(changes.get("instagram_error_count") or 0),
                        instagram_status=str(changes.get("instagram_status") or ""),
                        instagram_message=str(changes.get("instagram_message") or ""),
                    ),
                )
        except Exception as exc:
            SCRAPE_JOB.append(f"任务监控更新失败：{exc}\n")


def build_accounts_payload() -> list[dict]:
    """构造账号管理页面使用的 Chrome Profile 列表。"""
    profile_names = get_profiles()
    by_profile = {item["profile"]: item for item in STATE["accounts"]["entries"]}
    rows: list[dict] = []
    for profile in profile_names:
        saved = by_profile.get(profile, {})
        rows.append(
            {
                "profile": profile,
                "alias": saved.get("alias", ""),
                "usage": saved.get("usage", "通用"),
                "is_default": profile == STATE["profiles"].get("selected"),
            }
        )
    return rows


def start_scrape(payload: dict) -> dict:
    if SCRAPE_JOB.running:
        raise RuntimeError("已有任务正在运行。")

    task_id = str(payload.get("taskId") or "").strip()
    if not task_id:
        raise RuntimeError("请选择任务。")
    runtime_task = get_task_service().get_runtime_task_snapshot(task_id)
    runtime_documents = get_task_service().get_runtime_documents(task_id)
    task_paths = {
        "links": Path(runtime_documents.links_file),
        "progress": Path(runtime_documents.progress_file),
        "results": Path(runtime_documents.results_file),
    }
    links_file = runtime_documents.links_file
    progress_file = runtime_documents.progress_file
    output_file = runtime_documents.results_file

    profile = (payload.get("profile") or "").strip() or STATE["profiles"].get("selected", AUTOMATION_PROFILE_NAME)
    chrome_user_data_dir, chrome_profile_directory = resolve_chrome_launch_config(profile)

    command = [
        *scraper_worker_command(),
        "--file",
        links_file,
        "--progress-file",
        progress_file,
        "--output",
        output_file,
        "--task-file",
        runtime_documents.metadata_file,
        "--chrome-dir",
        str(chrome_user_data_dir),
        "--chrome-profile",
        chrome_profile_directory,
    ]
    command.append("--no-feishu")
    startup_logs: list[str] = ["同步模式：手动四表同步\n"]

    STATE["profiles"]["selected"] = profile
    save_state(STATE)
    completed_before_start = len(scraper_module.load_progress(str(task_paths["progress"])))
    get_task_service().start_runtime_task(task_id, profile=profile, started_at=_utc_now(), heartbeat_interval=TASK_HEARTBEAT_SECONDS, completed_count=completed_before_start, current_item="", last_progress_time=runtime_task.heartbeat_time)

    SCRAPE_JOB.logs = startup_logs
    SCRAPE_JOB.running = True
    SCRAPE_JOB.task_id = task_id
    SCRAPE_JOB.results_file = Path(output_file)
    SCRAPE_JOB.pause_requested = False
    SCRAPE_JOB.stop_requested = False
    SCRAPE_JOB.append("任务心跳已启动。\n")

    def worker() -> None:
        return_code: int | None = None
        error_message = ""
        monitor_stop_event = threading.Event()
        monitor_thread: threading.Thread | None = None
        try:
            SCRAPE_JOB.process = subprocess.Popen(
                command,
                cwd=str(DATA_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            get_task_service().mark_runtime_worker_running(task_id)
            monitor_thread = threading.Thread(
                target=_monitor_scrape_task,
                args=(task_id, task_paths, monitor_stop_event),
                daemon=True,
            )
            monitor_thread.start()
            assert SCRAPE_JOB.process.stdout is not None
            for line in SCRAPE_JOB.process.stdout:
                SCRAPE_JOB.append(line)
            return_code = SCRAPE_JOB.process.wait()
        except Exception as exc:
            error_message = str(exc)
            SCRAPE_JOB.append(f"\n启动抓取失败：{exc}\n")
        finally:
            monitor_stop_event.set()
            if monitor_thread:
                monitor_thread.join(timeout=3)
            try:
                completed_count = len(scraper_module.load_progress(str(task_paths["progress"])))
                instagram_errors = _instagram_error_count(task_paths["progress"])
                sync_summary: dict = {}
                sync_errors: list[str] = []
                final_task = get_task_service().get_task_metadata(task_id)
                stop_requested = bool(final_task.get("stop_requested")) or SCRAPE_JOB.stop_requested
                retry_urls = [
                    str(url or "").strip()
                    for url in final_task.get("retry_requested_urls", [])
                    if str(url or "").strip()
                ]
                retry_history = list(final_task.get("retry_history") or [])
                retry_changes: dict[str, object] = {}
                if retry_urls:
                    latest_rows = scraper_module.load_progress(str(task_paths["progress"]))
                    retry_success = sum(
                        1
                        for url in retry_urls
                        if str(latest_rows.get(url, {}).get("scrape_status") or "") == "success"
                    )
                    remaining_retry_urls = [
                        url for url in retry_urls
                        if str(latest_rows.get(url, {}).get("scrape_status") or "") != "success"
                    ]
                    retry_history.append(
                        {
                            "time": _utc_now(),
                            "count": len(retry_urls),
                            "success": retry_success,
                            "failed": len(remaining_retry_urls),
                            "round": int(final_task.get("retry_round") or 0),
                        }
                    )
                    retry_changes = {
                        "retry_history": retry_history,
                        "retry_requested_urls": remaining_retry_urls,
                    }
                last_error = error_message or ("" if return_code == 0 else f"抓取进程退出码：{return_code}")
                common_changes = {
                    "completed_count": completed_count,
                    "sync_summary": sync_summary,
                    "sync_errors": sync_errors,
                    "pause_requested": False,
                    "stop_requested": False,
                    "heartbeat_time": _utc_now(),
                    "current_item": "",
                    "last_successful_index": completed_count,
                    "browser_status": "closed",
                    "worker_status": "stopped",
                    "instagram_error_count": instagram_errors,
                    "instagram_status": "login_required" if instagram_errors >= 5 else "",
                    "instagram_message": "Instagram登录状态异常，请重新登录后继续。" if instagram_errors >= 5 else "",
                    "has_system_supplement": True
                    if str(final_task.get("task_type") or "") == "manual" and return_code == 0
                    else final_task.get("has_system_supplement", False),
                    **retry_changes,
                }

                if stop_requested:
                    status = "stopped"
                    sync_status = "not_started"
                elif return_code != 0:
                    status = "failed"
                    sync_status = "not_started"
                else:
                    status = "finalizing"
                    sync_status = "not_requested"

                final_state_persisted = False
                if status == "finalizing":
                    get_task_service().mark_runtime_finalizing(task_id, metadata_changes={"finished_at": "", "last_error": "", "sync_status": sync_status, **common_changes})
                    try:
                        finalization = import_task_results_to_creator_library(
                            task_id,
                            allowed_task_statuses={"finalizing"},
                        )
                    except Exception as library_error:
                        status = "failed"
                        sync_status = "not_started"
                        last_error = f"Creator Library 入库失败：{library_error}"
                        log_error(
                            "CreatorLibrary",
                            f"任务完成后导入达人库失败 | task_id={task_id}",
                            library_error,
                        )
                        get_task_service().mark_runtime_finalizing(task_id, metadata_changes={"creator_library_import_error": str(library_error)})
                    else:
                        if finalization.get("status") == "failed":
                            status = "failed"
                            sync_status = str(
                                finalization.get("sync_status") or "not_started"
                            )
                            last_error = str(finalization.get("last_error") or "")
                        else:
                            status = "completed"
                        try:
                            persisted_task = get_task_service().get_task_metadata(task_id)
                            final_state_persisted = (
                                str(persisted_task.get("status") or "") == status
                            )
                        except ValueError:
                            final_state_persisted = False

                if not final_state_persisted:
                    if status == "completed":
                        get_task_service().complete_runtime_task(task_id, finished_at=_utc_now(), metadata_changes={"sync_status": sync_status, **common_changes})
                    else:
                        get_task_service().fail_runtime_task(task_id, finished_at=_utc_now(), error=last_error, metadata_changes={"sync_status": sync_status, **common_changes})
                if status == "completed":
                    SCRAPE_JOB.append("任务完成。\n")
                elif status == "failed":
                    SCRAPE_JOB.append(f"任务失败：{last_error or '结果处理失败'}\n")
            except Exception as task_error:
                SCRAPE_JOB.append(f"\n任务状态保存失败：{task_error}\n")
                try:
                    get_task_service().fail_runtime_task(task_id, finished_at=_utc_now(), error=str(task_error), metadata_changes={"pause_requested": False, "stop_requested": False, "browser_status": "closed", "worker_status": "stopped"})
                except Exception as recovery_error:
                    SCRAPE_JOB.append(f"任务失败状态保存失败：{recovery_error}\n")
            finally:
                with SCRAPE_JOB.lock:
                    SCRAPE_JOB.running = False
                    SCRAPE_JOB.process = None
                    SCRAPE_JOB.pause_requested = False
                    SCRAPE_JOB.stop_requested = False

    threading.Thread(target=worker, daemon=True).start()
    return {"task_id": task_id}


def _active_scrape_task() -> str:
    if not SCRAPE_JOB.running or not SCRAPE_JOB.task_id:
        raise RuntimeError("当前没有正在运行的抓取任务。")
    return SCRAPE_JOB.task_id


def _active_scrape_is_finalizing(task_id: str) -> bool:
    try:
        task = get_task_service().get_runtime_task_snapshot(task_id)
    except ValueError:
        task = {}
    if task.status == "finalizing":
        return True
    process = SCRAPE_JOB.process
    return bool(process is not None and process.poll() is not None)


def _reject_finalizing_control(task_id: str) -> None:
    if _active_scrape_is_finalizing(task_id):
        raise RuntimeError("任务已经完成抓取，正在处理中。")


def pause_scrape() -> dict:
    task_id = _active_scrape_task()
    _reject_finalizing_control(task_id)
    if SCRAPE_JOB.stop_requested:
        raise RuntimeError("任务正在停止，不能暂停。")
    if not SCRAPE_JOB.pause_requested:
        SCRAPE_JOB.pause_requested = True
        get_task_service().mark_runtime_paused(task_id)
        SCRAPE_JOB.append("任务已暂停，等待继续。\n")
    return {"task_id": task_id, "status": "paused"}


def resume_scrape() -> dict:
    task_id = _active_scrape_task()
    _reject_finalizing_control(task_id)
    if SCRAPE_JOB.stop_requested:
        raise RuntimeError("任务正在停止，不能继续。")
    if SCRAPE_JOB.pause_requested:
        SCRAPE_JOB.pause_requested = False
        get_task_service().mark_runtime_resumed(task_id)
        SCRAPE_JOB.append("任务恢复运行。\n")
    return {"task_id": task_id, "status": "running"}


def request_stop_scrape() -> dict:
    task_id = _active_scrape_task()
    _reject_finalizing_control(task_id)
    if not SCRAPE_JOB.stop_requested:
        SCRAPE_JOB.stop_requested = True
        SCRAPE_JOB.pause_requested = False
        get_task_service().request_runtime_stop(task_id)
        SCRAPE_JOB.append("收到停止请求，正在保存当前进度。\n")
    return {"task_id": task_id, "status": "stopping"}


def resume_task(task_id: str) -> dict:
    """Resume an in-memory pause or relaunch a persisted paused/interrupted task."""
    plan = get_task_service().resume_task(
        task_id,
        runtime_running=SCRAPE_JOB.running,
        runtime_task_id=SCRAPE_JOB.task_id,
    )
    if plan.runtime_action == "resume":
        return resume_scrape()
    profile = plan.profile
    user_data_dir, profile_directory = resolve_chrome_launch_config(profile)
    if profile and profile != AUTOMATION_PROFILE_NAME and not (user_data_dir / profile_directory).is_dir():
        raise RuntimeError("无法恢复任务：原 Chrome Profile 不存在，请在账号管理中选择有效 Profile 后重新开始任务。")
    SCRAPE_JOB.append("恢复任务：将重新启动 Chrome 和抓取进程，并从已保存进度继续。\n")
    return start_scrape({"taskId": task_id, "profile": profile})


def stop_task(task_id: str) -> dict:
    """Stop active work gracefully; persist stopped state when no worker remains."""
    plan = get_task_service().stop_task(
        task_id,
        runtime_active=SCRAPE_JOB.running and SCRAPE_JOB.task_id == task_id,
    )
    if plan.runtime_action == "stop":
        return request_stop_scrape()
    return dict(plan.response)


def _read_task_csv(path: Path) -> tuple[list[str], list[dict]]:
    if not path.exists():
        raise ValueError(f"未找到任务文件：{path.name}")
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"任务文件格式无效：{path.name}")
        return list(reader.fieldnames), [dict(row) for row in reader]


def _review_value(row: dict, field: str) -> str:
    if field == REVIEW_FIELD_DATA_STATUS:
        return str(row.get(field) or "待检查")
    return str(row.get(field) or "")


def _account_uid_for_row(row: dict) -> str:
    return scraper_module.build_creator_uid(scraper_module.row_to_result(row))


def _review_record(row: dict) -> dict:
    result = scraper_module.row_to_result(row)
    return {
        "account_uid": _account_uid_for_row(row),
        scraper_module.FIELD_NAME: str(row.get(scraper_module.FIELD_NAME) or ""),
        scraper_module.FIELD_PLATFORM: str(row.get(scraper_module.FIELD_PLATFORM) or ""),
        scraper_module.FIELD_URL: str(row.get(scraper_module.FIELD_URL) or ""),
        scraper_module.FIELD_EMAIL: str(row.get(scraper_module.FIELD_EMAIL) or ""),
        scraper_module.FIELD_EMAIL_SOURCE: str(row.get(scraper_module.FIELD_EMAIL_SOURCE) or ""),
        scraper_module.FIELD_EXTERNAL_LINK: str(row.get(scraper_module.FIELD_EXTERNAL_LINK) or ""),
        scraper_module.FIELD_EXTERNAL_SOURCE: str(row.get(scraper_module.FIELD_EXTERNAL_SOURCE) or ""),
        scraper_module.FIELD_LATEST_DATE: str(row.get(scraper_module.FIELD_LATEST_DATE) or ""),
        scraper_module.FIELD_FOLLOWER_COUNT: str(row.get(scraper_module.FIELD_FOLLOWER_COUNT) or ""),
        scraper_module.FIELD_STATUS: str(row.get(scraper_module.FIELD_STATUS) or ""),
        scraper_module.FIELD_SCRAPE_STATUS: str(result.get("scrape_status") or "success"),
        scraper_module.FIELD_STATUS_REASON: str(result.get("status_reason") or ""),
        scraper_module.FIELD_LAST_SCRAPE_TIME: str(row.get(scraper_module.FIELD_LAST_SCRAPE_TIME) or ""),
        scraper_module.FIELD_RETRY_COUNT: str(row.get(scraper_module.FIELD_RETRY_COUNT) or "0"),
        REVIEW_FIELD_WHATSAPP: _review_value(row, REVIEW_FIELD_WHATSAPP),
        REVIEW_FIELD_NOTE: _review_value(row, REVIEW_FIELD_NOTE),
        REVIEW_FIELD_DATA_STATUS: _review_value(row, REVIEW_FIELD_DATA_STATUS),
        REVIEW_FIELD_MODIFIED_AT: _review_value(row, REVIEW_FIELD_MODIFIED_AT),
    }


def _task_progress(task_id: str, fallback_total: int = 0) -> dict:
    """Calculate task-local progress from its own input and latest progress rows."""
    try:
        documents = get_task_service().get_task_summary_documents(task_id)
    except ValueError:
        return {"total_links": fallback_total, "completed_links": 0, "failed_links": 0, "pending_links": fallback_total, "progress": 0}

    try:
        links = list(documents.links)
    except OSError:
        links = []
    total_links = len(links) or fallback_total
    task_urls = set(links)
    latest_status_by_url: dict[str, str] = {}
    for row in documents.progress_rows:
        url = str(row.get(scraper_module.FIELD_URL) or "").strip()
        if url and (not task_urls or url in task_urls):
            latest_status_by_url[url] = str(row.get(scraper_module.FIELD_STATUS) or "").strip()

    completed = sum(1 for status in latest_status_by_url.values() if status == "完成")
    failed = sum(1 for status in latest_status_by_url.values() if status == "失败")
    completed = min(completed, total_links)
    failed = min(failed, max(0, total_links - completed))
    pending = max(0, total_links - completed - failed)
    progress = round(((completed + failed) / total_links) * 100, 1) if total_links else 0
    return {
        "total_links": total_links,
        "completed_links": completed,
        "failed_links": failed,
        "pending_links": pending,
        "progress": progress,
    }


def _read_task_links(path: Path) -> list[str]:
    if not path.exists():
        raise ValueError("未找到任务链接文件。")
    return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _write_task_links(path: Path, links: list[str]) -> None:
    with shared_storage_lock():
        temp_path = path.with_suffix(f"{path.suffix}.tmp")
        try:
            temp_path.write_text("\n".join(links) + ("\n" if links else ""), encoding="utf-8")
            temp_path.replace(path)
        finally:
            temp_path.unlink(missing_ok=True)


def _task_platform_summary_from_links(links: list[str]) -> dict[str, int]:
    summary = {"TikTok": 0, "Instagram": 0, "YouTube": 0}
    for link in links:
        platform = str(scraper_module.normalize_link_record(link).get("platform") or "")
        if platform in summary:
            summary[platform] += 1
    return summary


def _platform_display(platforms: list[str]) -> str:
    normalized = task_manager.normalize_platforms(platforms)
    if normalized == list(task_manager.PLATFORM_KEYS):
        return "全部"
    return "、".join(PLATFORM_LABELS[key] for key in normalized)


def prepare_task_links(raw_links: list[str], selected_platforms: object = None) -> dict:
    """Normalize once, then keep only links selected for this local task."""
    platforms = task_manager.normalize_platforms(selected_platforms)

    platform_summary = {"TikTok": 0, "Instagram": 0, "YouTube": 0}
    selected_links: list[str] = []
    filtered_links: list[dict] = []
    invalid_links: list[str] = []
    seen_links: set[str] = set()
    seen_invalid: set[str] = set()

    for raw_link in raw_links:
        record = scraper_module.normalize_link_record(raw_link)
        if not record.get("valid"):
            value = str(record.get("input") or "").strip()
            if value and value not in seen_invalid:
                seen_invalid.add(value)
                invalid_links.append(value)
            continue

        normalized_url = str(record.get("normalized_url") or "").strip()
        platform = {
            "tiktok": "TikTok",
            "instagram": "Instagram",
            "youtube": "YouTube",
        }.get(str(record.get("platform") or "").strip().lower(), "")
        if not normalized_url or platform not in platform_summary or normalized_url in seen_links:
            continue
        seen_links.add(normalized_url)
        platform_summary[platform] += 1
        if platform.lower() in platforms:
            selected_links.append(normalized_url)
        else:
            filtered_links.append(
                {
                    "url": normalized_url,
                    "platform": platform,
                    "reason": "非目标平台",
                }
            )

    return {
        "target_platform": _platform_display(platforms),
        "platforms": platforms,
        "platform_summary": platform_summary,
        "normalized_links": selected_links,
        "filtered_links": filtered_links,
        "invalid_links": invalid_links,
    }


def _feishu_link_value(value) -> str:
    """Read the URL field returned by Feishu without guessing a profile URL."""
    if isinstance(value, dict):
        return str(value.get("link") or value.get("url") or value.get("text") or "").strip()
    if isinstance(value, list):
        for item in value:
            link = _feishu_link_value(item)
            if link:
                return link
        return ""
    return str(value or "").strip()


def _account_email_is_empty(value) -> bool:
    return not str(value or "").strip() or str(value or "").strip() == scraper_module.NO_EMAIL


def load_data_protection() -> dict:
    data, _source = load_json_with_backup(DATA_PROTECTION_FILE)
    if not isinstance(data, dict):
        return {}
    return {
        str(account_uid): fields
        for account_uid, fields in data.items()
        if isinstance(fields, dict) and str(account_uid).strip()
    }


def _save_data_protection(data: dict) -> None:
    atomic_write_json(DATA_PROTECTION_FILE, data)


def _merge_data_protection(
    protection: dict, account_uid: str, values: dict[str, str], source: str, task_id: str, updated_at: str
) -> bool:
    return CreatorService.merge_data_protection(
        protection, account_uid, values, source, task_id, updated_at
    )




def _email_recheck_summary(task_id: str) -> dict:
    try:
        documents = get_task_service().get_task_summary_documents(task_id)
        if not documents.results_available:
            return {"email_found_count": 0, "email_failed_count": 0}
        rows = documents.result_rows
    except (OSError, ValueError):
        return {"email_found_count": 0, "email_failed_count": 0}
    found = sum(
        1
        for row in rows
        if not _account_email_is_empty(row.get(scraper_module.FIELD_EMAIL))
    )
    return {"email_found_count": found, "email_failed_count": max(0, len(rows) - found)}


def create_manual_task(
    payload: dict,
    *,
    defer_library_import: bool = True,
    task_port: TaskPort | None = None,
) -> dict:
    """Keep legacy callers stable while routing the manual API through TaskService."""
    if task_port is not None or not defer_library_import:
        return _create_manual_task_legacy(
            payload,
            defer_library_import=defer_library_import,
            task_port=task_port,
        )
    return get_task_service().create_manual_task(
        payload, defer_library_import=defer_library_import
    )


def _create_manual_task_legacy(
    payload: dict,
    *,
    defer_library_import: bool = False,
    task_port: TaskPort | None = None,
) -> dict:
    """Create a one-record task that enters the same review and four-table flow."""
    profile_url = str(payload.get("profile_url") or "").strip()
    if not profile_url:
        raise ValueError("主页链接不能为空。")

    normalized = scraper_module.normalize_link_record(profile_url)
    if not normalized.get("valid"):
        raise ValueError(str(normalized.get("reason") or "主页链接无效。"))

    platform_by_key = {
        "tiktok": "TikTok",
        "instagram": "Instagram",
        "youtube": "YouTube",
    }
    normalized_platform = platform_by_key.get(str(normalized.get("platform") or "").lower(), "")
    selected_platform = str(payload.get("platform") or "").strip()
    if selected_platform not in {"TikTok", "Instagram", "YouTube"}:
        raise ValueError("请选择平台。")
    if selected_platform != normalized_platform:
        raise ValueError(f"主页链接属于 {normalized_platform}，请确认所选平台。")

    name = str(payload.get("name") or "").strip()
    email = str(payload.get("email") or "").strip()
    whatsapp = str(payload.get("whatsapp") or "").strip()
    note = str(payload.get("note") or "")
    source_contact = _resolve_source_contact(payload.get("source_contact_record_id"))
    follower_count = _normalize_follower_count(payload.get("follower_count"))
    if email:
        _validate_review_updates({scraper_module.FIELD_NAME: name or "未命名"}, {scraper_module.FIELD_EMAIL: email})
    if whatsapp:
        _validate_review_updates(
            {scraper_module.FIELD_NAME: name or "未命名"},
            {REVIEW_FIELD_WHATSAPP: whatsapp},
        )
    if follower_count:
        _validate_review_updates(
            {scraper_module.FIELD_NAME: name or "未命名"},
            {scraper_module.FIELD_FOLLOWER_COUNT: follower_count},
        )

    url = str(normalized.get("normalized_url") or "").strip()
    platform_summary = {
        "TikTok": int(selected_platform == "TikTok"),
        "Instagram": int(selected_platform == "Instagram"),
        "YouTube": int(selected_platform == "YouTube"),
    }
    active_task_port = task_port or get_task_port()
    task = active_task_port.create_manual_task(
        ManualTaskCreateCommand(
            normalized_url=url,
            task_name=str(payload.get("task_name") or ""),
            platform=selected_platform,
            platform_summary=platform_summary,
            defer_library_import=defer_library_import,
        )
    ).task.to_response()
    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    result = scraper_module.build_result(
        url=url,
        platform=selected_platform,
        name=name,
        emails=[] if not email else [item.strip() for item in email.split(",") if item.strip()],
        email_source="人工录入" if email else "",
        follower_count=follower_count,
        status="手动录入",
        whatsapp=whatsapp,
        note=note,
        data_status="待检查",
        last_modified_at=now,
    )
    account_uid = scraper_module.build_creator_uid(result)
    result_row = scraper_module.result_to_row(result)
    progress_row = dict(result_row)
    # Keep the manual URL pending so supplemental crawling can reuse the existing crawler.
    progress_row[scraper_module.FIELD_STATUS] = "待补充抓取"
    manual_values = {
        field: value
        for field, value in {
            scraper_module.FIELD_NAME: name,
            scraper_module.FIELD_EMAIL: email,
            scraper_module.FIELD_FOLLOWER_COUNT: follower_count,
            REVIEW_FIELD_WHATSAPP: whatsapp,
            REVIEW_FIELD_NOTE: note,
        }.items()
        if value
    }
    modifications = []
    if manual_values:
        modifications.append(
            {
                "account_uid": account_uid,
                "modified_fields": {
                    field: {"old": "", "new": value} for field, value in manual_values.items()
                },
                "status": "pending_sync",
                "time": now,
                "source": "manual_task",
            }
        )
    local_source_contact_id = ""
    if source_contact:
        try:
            local_contact = get_agency_service().upsert_external_contact(
                source_contact["record_id"],
                name=source_contact["name"],
                whatsapp=source_contact["whatsapp"],
            )
            local_source_contact_id = str(local_contact.get("contact_id") or "")
        except (OSError, RuntimeError, ValueError) as exc:
            log_error(
                "CreatorLibrary",
                f"来源联系人本地兼容保存失败 | record_id={source_contact['record_id']}",
                exc,
            )
    task.update(
        {
            "status": "manual_created",
            "completed_count": 0,
            "modified_count": len(modifications),
            "last_modified_time": now if manual_values else "",
            "source_contact_record_id": source_contact["record_id"] if source_contact else "",
            "local_source_contact_id": local_source_contact_id,
        }
    )
    if source_contact:
        task["source_contact_name"] = source_contact["name"]
    initialized = active_task_port.initialize_manual_task(
        task["id"],
        ManualTaskInitializationCommand(
            creator_name=name,
            platform=selected_platform,
            profile_url=url,
            follower_count=follower_count,
            email=email,
            whatsapp=whatsapp,
            note=note,
            source_contact_record_id=source_contact["record_id"] if source_contact else "",
            source_contact_name=source_contact["name"] if source_contact else "",
            local_source_contact_id=local_source_contact_id,
        ),
    )
    task = initialized.task.to_response()
    account_uid = initialized.account_uid
    if manual_values:
        with shared_storage_lock():
            protection = load_data_protection()
            if _merge_data_protection(protection, account_uid, manual_values, "人工录入", task["id"], now):
                _save_data_protection(protection)
    library_import = None
    if not defer_library_import:
        try:
            library_import = import_task_results_to_creator_library(
                task["id"],
                allowed_task_statuses={"manual_created"},
            )
            task = active_task_port.get_task(task["id"]).to_response()
        except (OSError, RuntimeError, ValueError) as exc:
            log_error("CreatorLibrary", f"人工任务进入达人库失败 | task_id={task['id']}", exc)
            library_import = {"status": "failed", "error": str(exc)}
    return {"task": task, "account_uid": account_uid, "creator_library_import": library_import}


def get_creator_repository() -> CreatorRepository:
    """Create the active local adapter; swap this factory for a cloud adapter later."""
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.creator()


def get_agency_repository() -> AgencyRepository:
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.agency()


def get_agency_service() -> AgencyService:
    return AgencyService(
        get_agency_repository,
        get_creator_repository,
        lambda: CREATOR_LIBRARY_CACHE,
    )


def get_creator_delete_impact_port() -> CreatorDeleteImpactPort:
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.creator_delete_impact()


def get_creator_delete_impact_service() -> CreatorDeleteImpactService:
    return CreatorDeleteImpactService(get_creator_delete_impact_port)


def get_creator_hard_delete_repository():
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.creator_hard_delete()


def get_creator_hard_delete_service() -> CreatorHardDeleteService:
    return CreatorHardDeleteService(
        get_creator_delete_impact_service,
        get_creator_hard_delete_repository,
        lambda: DATA_DIR,
        lambda exc: log_error("CreatorDelete", "永久删除失败", exc),
        creator_library_cache_invalidator=CREATOR_LIBRARY_CACHE.invalidate,
        dashboard_response_cache_invalidator=DASHBOARD_RESPONSE_CACHE.invalidate,
        feishu_delete_intent_store=FeishuDeleteIntentStore(DATA_DIR),
    )


def get_creator_merge_repository():
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.creator_merge(mail_messages_path=mail_sync_module.MAIL_MESSAGES_FILE)


def get_creator_merge_service() -> CreatorMergeService:
    return CreatorMergeService(
        get_creator_merge_repository,
        cache_invalidators=(
            CREATOR_LIBRARY_CACHE.invalidate,
            DASHBOARD_RESPONSE_CACHE.invalidate,
        ),
    )


def get_campaign_creator_service() -> CampaignCreatorService:
    return CampaignCreatorService(
        get_campaign_creator_repository,
        DASHBOARD_RESPONSE_CACHE.invalidate,
    )


def get_creator_service() -> CreatorService:
    """Create a stateless facade whose provider resolves the active request repository."""
    return CreatorService(
        get_creator_repository,
        get_task_port,
        load_data_protection,
        _save_data_protection,
        _resolve_source_contact,
        get_four_table_feishu_config,
        get_agency_repository,
        lambda: CREATOR_LIBRARY_CACHE,
        DASHBOARD_RESPONSE_CACHE.invalidate,
    )


def get_creator_summary_service() -> CreatorIntelligenceSummaryFacade:
    return CreatorIntelligenceSummaryFacade(
        CreatorSummaryService(get_creator_repository),
        CreatorIntelligenceService(get_creator_repository),
    )


def get_feishu_sync_service() -> FeishuSyncService:
    return FeishuSyncService(
        get_creator_repository(),
        lambda: FeishuClient(get_four_table_feishu_config()),
    )


def get_feishu_delete_reconciliation_service() -> FeishuDeleteReconciliationService:
    return FeishuDeleteReconciliationService(
        FeishuDeleteIntentStore(DATA_DIR),
        lambda: FeishuClient(get_four_table_feishu_config()),
    )


def _recover_feishu_delete_intents_on_startup() -> None:
    """Recover local delete state, then perform one bounded lifecycle pass."""
    try:
        recover_pending_delete_transactions(DATA_DIR)
        get_feishu_delete_reconciliation_service().reconcile(max_intents=10)
    except Exception as exc:
        log_error("FeishuDelete", "飞书删除意图启动恢复失败", exc)


def get_task_port() -> TaskPort:
    """Build a stateless adapter for one task operation."""
    return TaskManagerAdapter(
        lambda: TASKS_DIR,
        scrape_status_provider=SCRAPE_JOB.snapshot,
        heartbeat_interval=TASK_HEARTBEAT_SECONDS,
    )


class _CreatorAnalysisPortAdapter:
    """Expose the read-only Creator capability required by TaskService."""

    def __init__(self, creator_service_provider=None) -> None:
        self._creator_service_provider = creator_service_provider or get_creator_service

    def _service(self) -> CreatorService:
        return self._creator_service_provider()

    def get_creator_analysis(self, creator_id: str) -> CreatorAnalysisSnapshot:
        detail = self._service().get_creator_detail(creator_id)
        return CreatorAnalysisSnapshot(
            creator_id=creator_id,
            analysis=detail["analysis"],
        )

    def get_email_recheck_candidates(self) -> EmailRecheckCandidateScan:
        return self._service().get_email_recheck_candidates()

    def prepare_four_table_sync(
        self, command: FourTableSyncCommand
    ) -> PreparedFourTableSync:
        return self._service().prepare_four_table_sync(command)

    def execute_four_table_sync(
        self, prepared: PreparedFourTableSync
    ) -> FourTableSyncResult:
        return self._service().execute_four_table_sync(prepared)

    def prepare_manual_task(
        self, command: ManualTaskPreparationCommand
    ) -> PreparedManualTask:
        return self._service().prepare_manual_task(command)

    def commit_manual_task_protection(
        self, command: ManualTaskProtectionCommand
    ) -> None:
        self._service().commit_manual_task_protection(command)

    def upsert_external_agency_contact(
        self, contact: ExternalAgencyContactCommand
    ) -> ExternalAgencyContact:
        return self._service().upsert_external_agency_contact(contact)

    def prepare_task_result_update(
        self, command: TaskResultUpdateCommand
    ) -> PreparedTaskResultUpdate:
        return self._service().prepare_task_result_update(command)

    def commit_task_result_protection(
        self, task_id: str, update: PreparedTaskResultUpdate
    ) -> None:
        self._service().commit_task_result_protection(task_id, update)

    def import_task_results(
        self, command: ImportTaskResultsCommand | TaskResultImportCommand
    ) -> CreatorImportSummary | CreatorImportResult:
        result = self._service().import_task_results(command)
        if not isinstance(result, (CreatorImportSummary, CreatorImportResult)):
            raise RuntimeError("Creator 导入结果无效。")
        return result


def get_creator_port() -> CreatorPort:
    return _CreatorAnalysisPortAdapter()


def get_task_service() -> TaskService:
    """Create a facade whose dependencies resolve within the active operation."""
    return TaskService(
        get_task_port,
        get_creator_port,
        lambda: TaskRepository(TASKS_DIR),
        lambda task_id, exc: log_error(
            "CreatorLibrary", f"审核结果更新达人库失败 | task_id={task_id}", exc
        ),
        lambda record_id, exc: log_error(
            "CreatorLibrary",
            f"来源联系人本地兼容保存失败 | record_id={record_id}",
            exc,
        ),
        lambda task_id, exc: log_error(
            "CreatorLibrary",
            f"任务完成后导入达人库失败 | task_id={task_id}",
            exc,
        ),
        lambda task_id, exc: log_error(
            "Feishu", f"同步失败 | task_id={task_id}", exc
        ),
    )


# Recovery now depends on the composed TaskService boundary.
detect_interrupted_tasks()


@contextmanager
def background_task_service_scope():
    """Create and close repositories for one non-HTTP finalizing operation."""
    factory = _new_repository_factory()
    with factory.store.scope(defer_writes=False):
        creator_service = CreatorService(
            factory.creator,
            get_task_port,
            load_data_protection,
            _save_data_protection,
            _resolve_source_contact,
            get_four_table_feishu_config,
            factory.agency,
            lambda: CREATOR_LIBRARY_CACHE,
            DASHBOARD_RESPONSE_CACHE.invalidate,
        )
        creator_port = _CreatorAnalysisPortAdapter(lambda: creator_service)
        yield TaskService(
            get_task_port,
            lambda: creator_port,
            lambda: TaskRepository(TASKS_DIR),
            lambda task_id, exc: log_error(
                "CreatorLibrary",
                f"审核结果更新达人库失败 | task_id={task_id}",
                exc,
            ),
            lambda record_id, exc: log_error(
                "CreatorLibrary",
                f"来源联系人本地兼容保存失败 | record_id={record_id}",
                exc,
            ),
            lambda task_id, exc: log_error(
                "CreatorLibrary",
                f"任务完成后导入达人库失败 | task_id={task_id}",
                exc,
            ),
            lambda task_id, exc: log_error(
                "Feishu", f"同步失败 | task_id={task_id}", exc
            ),
        )


def _creator_library_workbook_path() -> Path:
    return Path(STATE.get("creator_library", {}).get("workbook_path") or DEFAULT_CREATOR_LIBRARY_WORKBOOK)


def _new_repository_factory() -> RepositoryFactory:
    return RepositoryFactory.for_runtime(
        _creator_library_workbook_path(),
        legacy_analysis_dir=CREATOR_ANALYSIS_DIR,
        legacy_library_file=CREATOR_LIBRARY_FILE,
        tasks_dir=TASKS_DIR,
        data_protection_file=DATA_PROTECTION_FILE,
    )


def get_product_repository() -> ProductRepository:
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.product()


def get_campaign_repository() -> CampaignRepository:
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.campaign()


def get_campaign_creator_repository() -> CampaignCreatorRepository:
    factory = get_active_repository_factory() or _new_repository_factory()
    return factory.campaign_creator()


def get_risk_service() -> RiskService:
    factory = get_active_repository_factory() or _new_repository_factory()
    return RiskService(factory.risk())


def get_analytics_service() -> AnalyticsService:
    factory = get_active_repository_factory() or _new_repository_factory()
    return AnalyticsService(factory.creator(), factory.campaign_creator())


def get_workbook_backup_service() -> WorkbookBackupService:
    return WorkbookBackupService(
        _creator_library_workbook_path,
        store_provider=lambda: (
            get_active_repository_factory() or _new_repository_factory()
        ).store,
    )


def get_clean_reset_service() -> CleanResetService:
    return CleanResetService(
        _creator_library_workbook_path(),
        settings_path=STATE_FILE,
        data_protection_path=DATA_PROTECTION_FILE,
        mail_messages_path=mail_sync_module.MAIL_MESSAGES_FILE,
        tasks_dir=TASKS_DIR,
        store_provider=lambda: (
            get_active_repository_factory() or _new_repository_factory()
        ).store,
        cache_invalidators=(
            CREATOR_LIBRARY_CACHE.invalidate,
            DASHBOARD_RESPONSE_CACHE.invalidate,
        ),
    )


def get_production_migration_service() -> ProductionMigrationService:
    global PRODUCTION_MIGRATION_SERVICE
    if PRODUCTION_MIGRATION_SERVICE is None:
        paths = SQLiteStoragePaths.for_app_data(DATA_DIR)
        PRODUCTION_MIGRATION_SERVICE = ProductionMigrationService(
            paths,
            _creator_library_workbook_path,
            production_root_provider=lambda: DATA_DIR,
        )
    return PRODUCTION_MIGRATION_SERVICE


def _assistant_search_creators(arguments: dict) -> list[dict]:
    filters = {
        key: arguments.get(key)
        for key in (
            "country", "platform", "language", "content_category", "search",
            "followers_min", "followers_max", "ai_tag",
        )
        if str(arguments.get(key) or "").strip()
    }
    include_archived = bool(arguments.get("include_archived"))
    service = get_creator_service()
    records: list[dict] = []
    page = 1
    while True:
        result = service.get_creator_library(
            include_archived=include_archived,
            page=page,
            page_size=100,
            filters=filters,
        )
        records.extend(dict(item) for item in result.get("creators", []))
        if page >= int(result.get("pages") or 0):
            break
        page += 1
    return records


def _assistant_campaign_detail(campaign_id: str) -> dict:
    campaign = get_campaign_repository().getCampaign(campaign_id)
    relations = get_campaign_creator_repository().getCampaignCreators(
        campaign_id=campaign_id
    )
    return {"campaign": campaign, "campaign_creators": relations}


def _assistant_create_capture_task(arguments: dict) -> dict:
    url = str(arguments.get("url") or "").strip()
    prepared = prepare_task_links([url], arguments.get("platform"))
    if not prepared["normalized_links"]:
        raise ValueError("没有符合目标平台的有效链接。")
    task = get_task_service().create_scrape_task(
        normalized_links=prepared["normalized_links"],
        invalid_links=prepared["invalid_links"],
        input_count=1,
        name=arguments.get("name"),
        target_platform=prepared["target_platform"],
        platforms=prepared["platforms"],
        platform_summary=prepared["platform_summary"],
        filtered_links=prepared["filtered_links"],
    )
    return {"task": task, "invalid_links": prepared["invalid_links"]}


def _assistant_daily_summary() -> dict:
    dashboard = get_dashboard_data()
    task_data = get_task_service().get_tasks()
    tasks = list(task_data.get("tasks") or [])
    statuses = {"pending": 0, "running": 0, "failed": 0}
    for task in tasks:
        status = str(task.get("status") or "")
        if status in statuses:
            statuses[status] += 1
    overview = dict(dashboard.get("overview") or {})
    return {
        "creator_total": overview.get("total_creators"),
        "new_creators_7d": overview.get("new_creators_7d"),
        "active_campaign_count": len(
            [
                campaign
                for campaign in get_campaign_repository().getCampaigns()
                if str(campaign.get("status") or "") in {"sourcing", "running"}
            ]
        ),
        "tasks": statuses,
    }


def get_assistant_service() -> AssistantService:
    global ASSISTANT_SERVICE
    if ASSISTANT_SERVICE is None:
        ASSISTANT_SERVICE = AssistantService(
            DeterministicAssistantProvider(),
            {
                "search_creators": _assistant_search_creators,
                "get_creator_detail": lambda creator_id: get_creator_service().get_creator_detail(creator_id),
                "list_campaigns": lambda arguments: get_campaign_repository().getCampaigns(
                    status=str(arguments.get("status") or "")
                ),
                "get_campaign_detail": _assistant_campaign_detail,
                "get_task_status": lambda task_id: get_task_service().get_task_details(task_id),
                "feishu_sync_dry_run": lambda: get_feishu_sync_service().dry_run(),
                "daily_summary": _assistant_daily_summary,
                "create_capture_task": _assistant_create_capture_task,
                "feishu_full_sync": lambda: get_feishu_sync_service().full_sync(confirm=True),
            },
            event_logger=lambda message: log_event("Assistant", message),
        )
    return ASSISTANT_SERVICE


def get_feishu_chat_transport() -> FeishuChatTransport:
    global FEISHU_CHAT_TRANSPORT
    if FEISHU_CHAT_TRANSPORT is None:
        FEISHU_CHAT_TRANSPORT = FeishuChatTransport(
            get_four_table_feishu_config,
            get_assistant_service,
            trace_id_provider=new_trace_id,
            event_logger=lambda message: log_event("FeishuChat", message),
            error_logger=lambda message, exc: log_error("FeishuChat", message, exc),
        )
    return FEISHU_CHAT_TRANSPORT


def import_task_results_to_creator_library(
    task_id: str,
    *,
    allowed_task_statuses: set[str] | None = None,
) -> dict:
    """Compatibility entry for scoped task-result imports and finalizing."""
    allowed_statuses = allowed_task_statuses or {"completed"}
    with background_task_service_scope() as service:
        if allowed_statuses == {"finalizing"}:
            finalized = service.finalize_background_task(task_id)
            if finalized.import_result.get("status") == "success":
                log_event(
                    "CreatorLibrary",
                    f"任务结果已导入 | task_id={task_id} | creators={len(finalized.import_result.get('creator_ids', []))} | accounts={len(finalized.import_result.get('account_ids', []))}",
                )
            return {
                "status": finalized.status,
                "sync_status": finalized.sync_status,
                "last_error": finalized.last_error,
            }
        result = service.import_task_results_to_creator_library(
            task_id, allowed_task_statuses=allowed_statuses
        )
    if result.get("status") == "success":
        log_event(
            "CreatorLibrary",
            f"任务结果已导入 | task_id={task_id} | creators={len(result.get('creator_ids', []))} | accounts={len(result.get('account_ids', []))}",
        )
    return result


def get_dashboard_data() -> dict:
    """Return read-only operational dashboard data from the Creator Repository."""
    creator_repository = get_creator_repository()

    def build_response() -> dict:
        factory = get_active_repository_factory()
        repository = (
            factory.dashboard(creator_repository)
            if factory
            else DashboardRepository(creator_repository)
        )
        service = DashboardService(repository)
        return {
            "overview": service.getOverview(),
            "creator_health": service.getCreatorHealth(),
            "health_summary": service.getHealthSummary(),
            "cooperation_performance": service.getCooperationPerformance(),
            "action_items": service.getActionItems(),
            "platform_distribution": service.getPlatformDistribution(),
            "creator_status_distribution": service.getCreatorStatusDistribution(),
            "creator_growth_trend": service.getCreatorGrowthTrend(),
        }

    return DASHBOARD_RESPONSE_CACHE.get_response(
        creator_repository.store, build_response
    )


def _extension_analysis_payload(payload: dict, task: dict, account_uid: str) -> dict:
    creator = payload.get("creator") if isinstance(payload.get("creator"), dict) else {}
    videos = payload.get("videos") if isinstance(payload.get("videos"), list) else []
    video_analysis = payload.get("video_analysis") if isinstance(payload.get("video_analysis"), dict) else {}
    creator_insight = payload.get("creator_insight") if isinstance(payload.get("creator_insight"), dict) else {}
    # The extension is capped at 20 videos; keep the same bound when persisting its snapshot.
    videos = [item for item in videos[:20] if isinstance(item, dict)]
    return {
        "schema_version": "1.0",
        "analysis_id": f"analysis_{task['id']}",
        "task_id": task["id"],
        "account_uid": account_uid,
        "imported_at": _utc_now(),
        "source": "chrome_extension",
        "creator": {
            "creator_name": str(creator.get("creator_name") or "").strip(),
            "platform": str(creator.get("platform") or "").strip(),
            "profile_url": str(creator.get("profile_url") or "").strip(),
            "followers": str(creator.get("followers") or "").strip(),
            "bio": str(creator.get("bio") or "").strip(),
            "email": str(creator.get("email") or payload.get("email") or "").strip(),
            "whatsapp": str(creator.get("whatsapp") or payload.get("whatsapp") or "").strip(),
            "country": str(creator.get("country") or payload.get("country") or "").strip(),
            "language": str(creator.get("language") or payload.get("language") or "").strip(),
            "language_source": str(creator.get("language_source") or "").strip(),
            "agency_id": str(creator.get("agency_id") or payload.get("agency_id") or "").strip(),
        },
        "content_category": str(
            payload.get("content_category") or creator.get("content_category") or ""
        ).strip(),
        "video_analysis": video_analysis,
        "videos": videos,
        "creator_insight": creator_insight,
    }


def import_extension_capture(payload: dict) -> dict:
    """Create one reviewable manual task and persist its extension-only analysis snapshot."""
    creator = payload.get("creator") if isinstance(payload.get("creator"), dict) else {}
    profile_url = str(creator.get("profile_url") or "").strip()
    if not profile_url:
        raise ValueError("缺少达人主页链接。")

    normalized = scraper_module.normalize_link_record(profile_url)
    if not normalized.get("valid"):
        raise ValueError(str(normalized.get("reason") or "主页链接无效。"))
    normalized_url = str(normalized.get("normalized_url") or "").strip()
    email = str(creator.get("email") or payload.get("email") or "").strip()
    whatsapp = str(creator.get("whatsapp") or payload.get("whatsapp") or "").strip()
    country = str(creator.get("country") or payload.get("country") or "").strip()
    language = str(creator.get("language") or payload.get("language") or "").strip()
    agency_id = str(creator.get("agency_id") or payload.get("agency_id") or "").strip()
    content_category = str(
        payload.get("content_category") or creator.get("content_category") or ""
    ).strip()
    normalized_payload = {
        **payload,
        "creator": {
            **creator,
            "email": email,
            "whatsapp": whatsapp,
            "country": country,
            "language": language,
            "agency_id": agency_id,
        },
        "content_category": content_category,
    }
    manual_result = create_manual_task(
        {
            "task_name": payload.get("task_name"),
            "name": creator.get("creator_name"),
            "platform": creator.get("platform"),
            "profile_url": normalized_url,
            "follower_count": creator.get("followers"),
            "email": email,
            "whatsapp": whatsapp,
            "note": payload.get("note"),
        },
        defer_library_import=True,
        task_port=get_task_port(),
    )
    task = manual_result["task"]
    analysis = _extension_analysis_payload(normalized_payload, task, manual_result["account_uid"])
    saved_analysis = get_creator_service().import_creator_from_extension(
        analysis,
        compensation_task_id=task["id"],
    )
    task = get_task_port().attach_creator_import(
        task["id"],
        CreatorImportLinkage(
            creator_id=saved_analysis["creator_id"],
            snapshot_id=saved_analysis["snapshot_id"],
            imported_at=analysis["imported_at"],
            country=country,
            language=language,
            content_category=content_category,
        ),
    ).to_response()
    return {
        "duplicate": False,
        "is_new_creator": saved_analysis["is_new_creator"],
        "task": task,
        "account_uid": manual_result["account_uid"],
        "analysis_id": saved_analysis["creator_id"],
        "account_id": saved_analysis["account_id"],
        "snapshot_id": saved_analysis["snapshot_id"],
    }


def save_creator_library_cooperation(analysis_id: str, payload: dict) -> dict:
    return get_creator_repository().saveCooperation(analysis_id, payload)


def _normalize_follower_count(value: object) -> str:
    raw = str(value or "").strip()
    normalized = scraper_module.normalize_follower_count(raw)
    if raw and not normalized:
        raise ValueError("粉丝数格式错误，请填写如 10K、1.2M 或 100000。")
    return normalized


def _validate_review_updates(row: dict, fields: dict) -> dict[str, str]:
    return get_creator_service().validate_task_result_updates(row, fields)


def _csv_content(fieldnames: list[str], rows: list[dict]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def retry_failed_task_results(task_id: str, account_uids: list[object] | None = None) -> dict:
    """Queue retryable records inside the existing task without duplicating task files."""
    if SCRAPE_JOB.running:
        raise RuntimeError("已有任务正在运行，暂不能重新抓取。")
    return get_task_service().retry_failed_results(task_id, account_uids)




def _assert_task_sync_lifecycle(task: dict) -> None:
    task_status = str(task.get("status") or "").strip()
    if task_status == "running":
        raise RuntimeError("任务抓取中，请稍候")
    if task_status == "finalizing":
        raise RuntimeError("任务入库收尾中，请稍候")


def sync_task_results_to_four_tables(task_id: str) -> dict:
    """Compatibility wrapper for the TaskService four-table sync workflow."""
    log_event("Feishu", f"同步开始 | task_id={task_id}")
    result = get_task_service().sync_four_tables(task_id)
    if result["sync_status"] in {"success", "failed"}:
        log_event(
            "Feishu",
            f"同步{result['sync_status']} | task_id={task_id} | errors={len(result['sync_errors'])}",
        )
    return result


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        return

    def _json(self, data: dict, status: int = 200) -> None:
        payload = dict(data)
        payload.setdefault("trace_id", self._request_trace_id())
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Trace-ID", payload["trace_id"])
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)
        if self.path.startswith("/api/"):
            outcome = "success" if status < 400 else "failed"
            error = payload.get("error")
            detail = str(error.get("code") or "") if isinstance(error, dict) else str(error or "")
            log_event("API", f"{self.command} {urlparse(self.path).path} | {outcome} | status={status}{f' | {detail}' if detail else ''}")

    def _binary(self, data: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Trace-ID", self._request_trace_id())
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _ok(self, **extra) -> None:
        data = {"ok": True}
        data.update(extra)
        self._json(data)

    def _error(self, message: str, status: int = 400) -> None:
        friendly_message = _friendly_error_message(message)
        _record_last_error(friendly_message)
        log_error("API", f"{self.command} {urlparse(self.path).path} | status={status} | {friendly_message}")
        self._json({"error": friendly_message}, status=status)

    def _request_trace_id(self) -> str:
        trace_id = get_trace_id()
        if not trace_id:
            trace_id = new_trace_id()
            set_trace_id(trace_id)
        return trace_id

    def _api_success(self, data=None, *, legacy: dict | None = None, status: int = 200) -> None:
        self._json(success_payload(data, legacy=legacy), status=status)

    def _api_error(
        self,
        code: str,
        message: str,
        *,
        status: int = 400,
        details=None,
        legacy: dict | None = None,
    ) -> None:
        friendly_message = _friendly_error_message(message)
        _record_last_error(friendly_message)
        log_error(
            "API",
            f"{self.command} {urlparse(self.path).path} | status={status} | code={code} | {friendly_message}",
        )
        self._json(
            error_payload(code, friendly_message, details=details, legacy=legacy),
            status=status,
        )

    def _repository_error(self, exc: Exception) -> None:
        message = str(exc)
        if isinstance(exc, RuntimeError):
            return self._error(message, status=500)
        if "不存在" in message:
            return self._error(message, status=404)
        if "已加入" in message or "已归档" in message or "不能归档" in message:
            return self._error(message, status=409)
        return self._error(message, status=400)

    def _save_state_and_ok(self) -> None:
        save_state(STATE)
        self._ok()

    def _allow_local_request(self) -> bool:
        set_trace_id(new_trace_id())
        port = int(self.server.server_port)
        if not allowed_host_header(self.headers.get("Host"), port):
            self._json({"error": "LOCAL_REQUEST_REJECTED"}, status=403)
            return False
        path = urlparse(self.path).path
        if self.command in MUTATING_METHODS and not allowed_mutation_origin(
            self.headers.get("Origin"), path, port
        ):
            self._json({"error": "LOCAL_REQUEST_REJECTED"}, status=403)
            return False
        return True

    def _handle_runtime_shutdown(self, path: str) -> bool:
        """Stop only a Browser Mode server after its local response is delivered."""
        if path != "/api/runtime/shutdown":
            return False
        if not browser_shutdown_allowed(path, os.environ.get("KOLCONNECT_BROWSER")):
            self._json({"error": "LOCAL_REQUEST_REJECTED"}, status=403)
            return True
        self._ok(shutting_down=True)
        request_runtime_shutdown()
        return True

    def _normalize_save_state_and_ok(self) -> None:
        global STATE
        STATE = normalize_state(STATE)
        save_state(STATE)
        self._ok()

    def _read_body(self) -> bytes:
        length = int(self.headers.get("Content-Length", "0"))
        return self.rfile.read(length) if length else b""

    def _read_json(self) -> dict:
        raw = self._read_body()
        return json.loads((raw or b"{}").decode("utf-8"))

    def _request_context(self, parsed, query: dict) -> dict:
        body_mode = None
        body_loaded = False
        body = b""
        payload = None

        def read_body(mode: str) -> bytes:
            nonlocal body_mode, body_loaded, body
            if body_mode is not None and body_mode != mode:
                raise RuntimeError("请求正文不能同时按 JSON 和原始数据读取。")
            body_mode = mode
            if not body_loaded:
                body = self._read_body()
                body_loaded = True
            return body

        def get_payload() -> dict:
            nonlocal payload
            if payload is None:
                payload = json.loads((read_body("json") or b"{}").decode("utf-8"))
            return payload

        def get_raw_body() -> bytes:
            return read_body("raw")

        return {
            "method": self.command,
            "path": parsed.path,
            "query": query,
            "trace_id": self._request_trace_id(),
            "payload": None,
            "get_payload": get_payload,
            "get_raw_body": get_raw_body,
        }

    def _handler_context(self) -> dict:
        def get_state() -> dict:
            return STATE

        def save_current_state() -> None:
            save_state(STATE)

        def normalize_and_save_state() -> None:
            global STATE
            STATE = normalize_state(STATE)
            save_state(STATE)

        return {
            "request": {"trace_id": self._request_trace_id()},
            "state": {
                "get": get_state,
                "save": save_current_state,
                "normalize_and_save": normalize_and_save_state,
            },
            "scrape_job": SCRAPE_JOB,
            "repositories": {
                "agency": get_agency_repository,
                "creator": get_creator_repository,
                "product": get_product_repository,
                "campaign": get_campaign_repository,
                "campaign_creator": get_campaign_creator_repository,
            },
            "ports": {"agency": get_agency_repository, "task": get_task_port},
            "services": {
                "agency": get_agency_service(),
                "assistant": get_assistant_service(),
                "analytics": get_analytics_service(),
                "workbook_backup": get_workbook_backup_service(),
                "clean_reset": get_clean_reset_service(),
                "storage_migration": get_production_migration_service(),
                "creator": get_creator_service(),
                "creator_summary": get_creator_summary_service(),
                "creator_delete_impact": get_creator_delete_impact_service(),
                "creator_hard_delete": get_creator_hard_delete_service(),
                "creator_merge": get_creator_merge_service(),
                "feishu_sync": get_feishu_sync_service(),
                "feishu_delete_reconciliation": get_feishu_delete_reconciliation_service(),
                "feishu_chat": get_feishu_chat_transport(),
                "campaign_creator": get_campaign_creator_service(),
                "task": get_task_service(),
                "risk": get_risk_service(),
                "build_accounts_payload": build_accounts_payload,
                "get_dashboard_data": get_dashboard_data,
                "invalidate_dashboard_response_cache": DASHBOARD_RESPONSE_CACHE.invalidate,
                "get_agency_contact_options": get_agency_contact_options,
                "get_four_table_feishu_config": get_four_table_feishu_config,
                "get_profiles": get_profiles,
                "get_system_health": get_system_health,
                "is_sensitive_mask": is_sensitive_mask,
                "import_extension_capture": import_extension_capture,
                "merge_masked_mail_passwords": merge_masked_mail_passwords,
                "normalize_creator_library_workbook_path": normalize_creator_library_workbook_path,
                "normalize_mail_account": normalize_mail_account,
                "normalize_mail_state": normalize_mail_state,
                "open_chrome_profile": open_chrome_profile,
                "record_diagnostic": _record_diagnostic,
                "create_manual_task": create_manual_task,
                "pause_scrape": pause_scrape,
                "prepare_task_links": prepare_task_links,
                "request_stop_scrape": request_stop_scrape,
                "resume_scrape": resume_scrape,
                "resume_task": resume_task,
                "retry_failed_task_results": retry_failed_task_results,
                "state_for_client": state_for_client,
                "test_imap_login": test_imap_login,
                "test_smtp_login": test_smtp_login,
                "start_scrape": start_scrape,
                "stop_task": stop_task,
                "utc_now": _utc_now,
            },
            "task_manager": task_manager,
            "modules": {"mail_sync": mail_sync_module, "scraper": scraper_module},
            "paths": {"tasks": TASKS_DIR, "static": STATIC_DIR, "data": DATA_DIR},
            "config": {
                "automation_profile_name": AUTOMATION_PROFILE_NAME,
                "legacy_cooperation_pattern": LEGACY_COOPERATION_PATH_PATTERN,
                "legacy_cooperation_read_only_message": LEGACY_COOPERATION_READ_ONLY_MESSAGE,
            },
            "logging": {"event": log_event, "error": log_error},
        }

    @staticmethod
    def _repository_request_scope():
        """Resolve the active workbook path once for this HTTP request."""
        return _new_repository_factory().request_scope()

    def _dispatch(self, request: dict) -> bool:
        context = self._handler_context()
        for endpoint_handler in HANDLERS:
            if endpoint_handler.handle(self, request, context):
                return True
        return False

    def _serve_file(self, file_path: Path) -> None:
        try:
            static_root = STATIC_DIR.resolve()
            resolved_path = file_path.resolve()
            resolved_path.relative_to(static_root)
        except (OSError, ValueError):
            self.send_error(404)
            return

        if not resolved_path.exists() or not resolved_path.is_file():
            self.send_error(404)
            return
        suffix = resolved_path.suffix.lower()
        mime = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "application/javascript; charset=utf-8",
            ".png": "image/png",
            ".ico": "image/x-icon",
        }.get(suffix, "application/octet-stream")
        data = resolved_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        if not self._allow_local_request():
            return
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)

        with self._repository_request_scope():
            if self._dispatch(self._request_context(parsed, query)):
                return

            if parsed.path in {"", "/"}:
                return self._serve_file(STATIC_DIR / "index.html")

            return self._serve_file(STATIC_DIR / parsed.path.lstrip("/"))

    def do_POST(self) -> None:
        if not self._allow_local_request():
            return
        parsed = urlparse(self.path)
        if self._handle_runtime_shutdown(parsed.path):
            return
        request = self._request_context(parsed, parse_qs(parsed.query))
        try:
            with self._repository_request_scope():
                if self._dispatch(request):
                    return
                request["get_payload"]()
                return self._error("接口不存在。", status=404)
        except json.JSONDecodeError:
            return self._error("请求数据不是有效 JSON。")
        except Exception as exc:
            log_error("API", f"未处理异常: {self.command} {parsed.path}", exc)
            return self._error(_friendly_error_message(exc), status=500)

    def do_PATCH(self) -> None:
        if not self._allow_local_request():
            return
        parsed = urlparse(self.path)
        request = self._request_context(parsed, parse_qs(parsed.query))
        try:
            with self._repository_request_scope():
                if self._dispatch(request):
                    return
                payload = request["get_payload"]()

                return self._error("接口不存在。", status=404)
        except json.JSONDecodeError:
            return self._error("请求数据不是有效 JSON。")
        except Exception as exc:
            log_error("API", f"未处理异常: {self.command} {parsed.path}", exc)
            return self._error(_friendly_error_message(exc), status=500)

    def do_PUT(self) -> None:
        if not self._allow_local_request():
            return
        parsed = urlparse(self.path)
        with self._repository_request_scope():
            if self._dispatch(self._request_context(parsed, parse_qs(parsed.query))):
                return
            return self._error("接口不存在。", status=404)

    def do_DELETE(self) -> None:
        if not self._allow_local_request():
            return
        parsed = urlparse(self.path)
        with self._repository_request_scope():
            if self._dispatch(self._request_context(parsed, parse_qs(parsed.query))):
                return
            return self._error("接口不存在。", status=404)


_RUNTIME_SERVER_LOCK = threading.RLock()
_RUNTIME_SERVER = None
_RUNTIME_SHUTDOWN_THREAD: threading.Thread | None = None


def request_runtime_shutdown() -> bool:
    """Idempotently stop the active local server through one lifecycle path."""
    global _RUNTIME_SHUTDOWN_THREAD
    with _RUNTIME_SERVER_LOCK:
        server = _RUNTIME_SERVER
        if server is None:
            return False
        if _RUNTIME_SHUTDOWN_THREAD and _RUNTIME_SHUTDOWN_THREAD.is_alive():
            return True
        _RUNTIME_SHUTDOWN_THREAD = threading.Thread(
            target=server.shutdown,
            name="kolconnect-runtime-shutdown",
            daemon=True,
        )
        _RUNTIME_SHUTDOWN_THREAD.start()
        return True


def run() -> None:
    global _RUNTIME_SERVER, _RUNTIME_SHUTDOWN_THREAD
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    with _RUNTIME_SERVER_LOCK:
        _RUNTIME_SERVER = server
        _RUNTIME_SHUTDOWN_THREAD = None
    chat_transport = get_feishu_chat_transport()
    workbook_path = STATE.get("creator_library", {}).get("workbook_path") or DEFAULT_CREATOR_LIBRARY_WORKBOOK
    log_event(
        "KOLConnect Start",
        f"version={APP_DISPLAY_VERSION} | platform={sys.platform} | data_path={DATA_DIR} | excel_path={workbook_path}",
    )
    if (
        os.environ.get("KOLCONNECT_DESKTOP") != "1"
        and os.environ.get("KOLCONNECT_BROWSER") != "1"
    ):
        webbrowser.open(f"http://{HOST}:{PORT}/?v={int(time.time())}")
    if STATE.get("feishu", {}).get("chat_enabled"):
        chat_transport.start()
    threading.Thread(
        target=_recover_feishu_delete_intents_on_startup,
        name="kolconnect-feishu-delete-recovery",
        daemon=True,
    ).start()
    try:
        server.serve_forever()
    finally:
        chat_transport.close()
        server.server_close()
        with _RUNTIME_SERVER_LOCK:
            if _RUNTIME_SERVER is server:
                _RUNTIME_SERVER = None
                _RUNTIME_SHUTDOWN_THREAD = None


if __name__ == "__main__":
    run()


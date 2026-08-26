from __future__ import annotations

"""PRE-M8 C12 synthetic legacy-to-SQLite acceptance harness."""

import argparse
import hashlib
import json
from pathlib import Path
import sys
from uuid import uuid4

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
TESTS = ROOT / "tests"
for path in (APP, TESTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from repository_factory import RepositoryFactory
from storage.authority import resolve_runtime_authority
from storage.migration import ExcelToSQLiteMigrator
from storage.paths import SQLiteStoragePaths
from test_pre_m8_excel_sqlite_migration import build_fixture


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def workbook_counts(path: Path) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            sheet.title: max(0, sheet.max_row - 1)
            for sheet in workbook.worksheets
            if not sheet.title.startswith("_")
        }
    finally:
        workbook.close()


def runtime_projection(factory: RepositoryFactory) -> dict:
    creators = factory.creator().getCreators(include_archived=True)
    accounts = factory.creator().getCreatorAccounts()
    campaigns = factory.campaign().getCampaigns(include_archived=True)
    relations = factory.campaign_creator().getCampaignCreators(include_archived=True)
    return {
        "creator_ids": sorted(str(row.get("creator_id") or "") for row in creators),
        "account_ownership": sorted(
            (str(row.get("account_uid") or ""), str(row.get("creator_id") or ""))
            for row in accounts
        ),
        "campaign_ids": sorted(str(row.get("campaign_id") or "") for row in campaigns),
        "campaign_memberships": sorted(
            (
                str(row.get("id") or ""),
                str(row.get("campaign_id") or ""),
                str(row.get("creator_id") or ""),
                tuple(row.get("account_ids") or ()),
                tuple(row.get("planned_publish_dates") or ()),
            )
            for row in relations
        ),
        "archived_creator_ids": sorted(
            str(row.get("creator_id") or "")
            for row in creators
            if str(row.get("archived_at") or "").strip()
        ),
    }


def run_acceptance(root: Path) -> dict:
    root = Path(root).resolve()
    root.mkdir(parents=True, exist_ok=False)
    logs = root / "logs"
    logs.mkdir()
    import app_logging
    import local_storage_lock

    app_logging.get_logs_dir = lambda: logs
    local_storage_lock.get_shared_storage_lock_path = (
        lambda: root / "locks" / "shared_storage.lock"
    )

    workbook = root / "legacy_creator_library.xlsx"
    build_fixture(workbook, creator_count=100, snapshot_count=1000)
    paths = SQLiteStoragePaths.for_app_data(root / "appdata")
    source_before = sha256(workbook)
    authority_before = resolve_runtime_authority(paths, workbook)
    source_counts = workbook_counts(workbook)

    migrator = ExcelToSQLiteMigrator(paths)
    migration = migrator.migrate(workbook)
    source_after_migration = sha256(workbook)
    migrator.activate_synthetic(migration)
    authority_after = resolve_runtime_authority(paths, workbook)
    factory = RepositoryFactory.for_runtime(workbook, storage_paths=paths)
    projection = runtime_projection(factory)

    workbook_object = load_workbook(workbook)
    workbook_object["Creators"][2][1].value = "Externally changed legacy name"
    workbook_object.save(workbook)
    workbook_object.close()
    externally_changed_hash = sha256(workbook)
    runtime_name_after_excel_edit = factory.creator().getCreatorDetail("creator_0000")[
        "record"
    ]["creator_name"]

    factory.creator().updateCreator("creator_0000", {"country": "Portugal"})
    source_after_sqlite_write = sha256(workbook)
    revision_after_write = factory.store.business_revision()

    backup = root / "state-a.db"
    factory.store.create_transaction_backup(backup)
    factory.creator().updateCreator("creator_0000", {"country": "Japan"})
    factory.store.restore_transaction_backup(backup)
    restarted = RepositoryFactory.for_runtime(workbook, storage_paths=paths)
    restored_country = restarted.creator().getCreatorDetail("creator_0000")["record"]["country"]

    exported = root / "compatibility-export.xlsx"
    restarted.store.export_workbook(exported)
    reimport_paths = SQLiteStoragePaths.for_app_data(root / "reimport_appdata")
    reimport_migration = ExcelToSQLiteMigrator(reimport_paths).migrate(exported)
    ExcelToSQLiteMigrator(reimport_paths).activate_synthetic(reimport_migration)
    reimported = RepositoryFactory.for_runtime(exported, storage_paths=reimport_paths)
    reimport_projection = runtime_projection(reimported)

    result = {
        "sandbox": str(root),
        "authority_before": authority_before.kind,
        "authority_after": authority_after.kind,
        "source_sha256_before": source_before,
        "source_sha256_after_migration": source_after_migration,
        "source_unchanged_by_migration": source_before == source_after_migration,
        "source_counts": source_counts,
        "migration_counts": migration.counts,
        "projection": projection,
        "legacy_excel_edit_ignored": runtime_name_after_excel_edit == "Creator 0",
        "sqlite_write_does_not_touch_excel": externally_changed_hash == source_after_sqlite_write,
        "revision_after_write": revision_after_write,
        "backup_restore_country": restored_country,
        "export_size_bytes": exported.stat().st_size,
        "export_reimport_parity": reimport_projection == projection,
        "restart_authority": resolve_runtime_authority(paths, workbook).kind,
    }
    (root / "acceptance-result.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path)
    args = parser.parse_args()
    root = args.root or ROOT / ".pre_m8_batch3_acceptance" / uuid4().hex
    result = run_acceptance(root)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

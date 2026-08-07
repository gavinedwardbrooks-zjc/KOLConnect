from __future__ import annotations

"""Safely migrate legacy Creator CRM values into first-class Excel columns."""

import argparse
import json
import os
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from creator_repository import CreatorRepository  # noqa: E402


MIGRATION_COLUMNS = ("bio", "archived_at")


def _rows(sheet) -> list[dict[str, Any]]:
    headers = [str(cell.value or "") for cell in sheet[1]]
    return [
        {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]


def _legacy_crm_by_creator(workbook) -> dict[str, dict[str, Any]]:
    if "_AnalysisData" not in workbook.sheetnames:
        return {}
    result: dict[str, dict[str, Any]] = {}
    for metadata in _rows(workbook["_AnalysisData"]):
        creator_id = str(metadata.get("creator_id") or "").strip()
        if not creator_id:
            continue
        analysis = CreatorRepository._decode_analysis(metadata.get("analysis_json"))
        crm = analysis.get("_crm") if isinstance(analysis.get("_crm"), dict) else {}
        result[creator_id] = crm
    return result


def migrate_workbook(workbook_path: Path, *, dry_run: bool = True) -> dict[str, Any]:
    """Preview or apply the CRM-column migration to one explicitly named workbook."""
    workbook_path = Path(workbook_path).expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")

    workbook = load_workbook(workbook_path)
    backup_path: Path | None = None
    try:
        if "Creators" not in workbook.sheetnames:
            raise ValueError("Workbook does not contain a Creators sheet.")

        sheet = workbook["Creators"]
        original_headers = [str(cell.value or "") for cell in sheet[1]]
        column_indexes = {
            header: original_headers.index(header) + 1
            for header in MIGRATION_COLUMNS
            if header in original_headers
        }
        legacy_by_creator = _legacy_crm_by_creator(workbook)
        stats: dict[str, Any] = {
            "workbook": str(workbook_path),
            "mode": "dry-run" if dry_run else "apply",
            "scanned": 0,
            "bio_migrated": 0,
            "archived_at_migrated": 0,
            "skipped_existing": 0,
            "no_legacy_value": 0,
            "conflicts": 0,
            "errors": 0,
            "backup_path": "",
        }

        for row_index in range(2, sheet.max_row + 1):
            row = {
                original_headers[index - 1]: sheet.cell(row_index, index).value
                for index in range(1, len(original_headers) + 1)
                if original_headers[index - 1]
            }
            creator_id = str(row.get("creator_id") or "").strip()
            if not creator_id:
                continue
            stats["scanned"] += 1
            crm = legacy_by_creator.get(creator_id, {})
            has_legacy_value = False
            for field in MIGRATION_COLUMNS:
                legacy_value = crm.get(field)
                if legacy_value is None or str(legacy_value).strip() == "":
                    continue
                has_legacy_value = True
                existing_value = row.get(field)
                if existing_value is not None and str(existing_value).strip() != "":
                    stats["skipped_existing"] += 1
                    if str(existing_value).strip() != str(legacy_value).strip():
                        stats["conflicts"] += 1
                    continue
                stats[f"{field}_migrated"] += 1
            if not has_legacy_value:
                stats["no_legacy_value"] += 1

        if dry_run:
            return stats

        for field in MIGRATION_COLUMNS:
            if field not in column_indexes:
                column_index = sheet.max_column + 1
                sheet.cell(1, column_index, field)
                column_indexes[field] = column_index

        # Populate values after missing columns have acquired physical indexes.
        for row_index in range(2, sheet.max_row + 1):
            creator_id_index = original_headers.index("creator_id") + 1
            creator_id = str(sheet.cell(row_index, creator_id_index).value or "").strip()
            crm = legacy_by_creator.get(creator_id, {})
            for field in MIGRATION_COLUMNS:
                cell = sheet.cell(row_index, column_indexes[field])
                legacy_value = crm.get(field)
                if cell.value in (None, "") and legacy_value is not None and str(legacy_value).strip():
                    cell.value = legacy_value

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(
            f"Creator_Library_before_m1_c2_{timestamp}{workbook_path.suffix}"
        )
        suffix = 1
        while backup_path.exists():
            backup_path = workbook_path.with_name(
                f"Creator_Library_before_m1_c2_{timestamp}_{suffix}{workbook_path.suffix}"
            )
            suffix += 1
        shutil.copy2(workbook_path, backup_path)
        stats["backup_path"] = str(backup_path)

        temp_path = workbook_path.with_name(f".{workbook_path.stem}.m1_c2.tmp.xlsx")
        try:
            workbook.save(temp_path)
            validation = load_workbook(temp_path, read_only=True)
            validation.close()
            os.replace(temp_path, workbook_path)
        finally:
            if temp_path.exists():
                temp_path.unlink()
        return stats
    except Exception:
        if "stats" in locals():
            stats["errors"] += 1
        raise
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path, help="Workbook copy to inspect or migrate")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Preview only (default)")
    mode.add_argument("--apply", action="store_true", help="Back up and update the named workbook")
    args = parser.parse_args()
    result = migrate_workbook(args.workbook, dry_run=not args.apply)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

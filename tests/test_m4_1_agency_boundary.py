from __future__ import annotations

import json
import sys
import tempfile
import threading
import unittest
import urllib.error
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import server
from repository_factory import RepositoryFactory
from services.agency_service import AgencyService


class AgencyBoundaryHttpTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        factory = RepositoryFactory.for_path(self.workbook_path)
        factory.creator().getCreators()
        self._seed_creator()

        self.patchers = [
            mock.patch.object(
                server,
                "_creator_library_workbook_path",
                return_value=self.workbook_path,
            ),
            mock.patch.object(
                server,
                "get_agency_contact_options",
                return_value={
                    "configured": True,
                    "contacts": [
                        {
                            "record_id": "external-contact-1",
                            "name": "External Contact",
                            "whatsapp": "+1 555",
                            "agencies": ["External Agency"],
                        }
                    ],
                },
            ),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in self.patchers:
            patcher.start()

        self.httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.httpd.server_port}"

    def tearDown(self) -> None:
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=5)
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp_dir.cleanup()

    def _seed_creator(self) -> None:
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Creators"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        values = {
            "creator_id": "creator_one",
            "name": "Creator One",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@creator-one",
            "status": "discovered",
            "created_at": "2026-08-01T00:00:00Z",
            "updated_at": "2026-08-01T00:00:00Z",
        }
        sheet.append([values.get(header, "") for header in headers])
        workbook.save(self.workbook_path)
        workbook.close()

    def request(
        self,
        method: str,
        path: str,
        payload: dict | None = None,
    ) -> tuple[int, dict]:
        data = None
        headers = {}
        if payload is not None:
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json"
        request = urllib.request.Request(
            self.base_url + path,
            data=data,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                return response.status, json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            with exc:
                return exc.code, json.loads(exc.read().decode("utf-8"))

    def test_local_agency_http_contracts_and_creator_relations(self) -> None:
        status, created = self.request(
            "POST",
            "/api/local/agencies",
            {"name": "Agency One", "country": "Brazil"},
        )
        self.assertEqual(200, status)
        self.assertTrue(created["ok"])
        agency = created["agency"]
        self.assertRegex(agency["agency_id"], r"^agency_[0-9a-f]{16}$")
        created_at = agency["created_at"]

        status, updated = self.request(
            "POST",
            "/api/local/agencies",
            {"agency_id": agency["agency_id"], "note": "updated"},
        )
        self.assertEqual(200, status)
        self.assertEqual(agency["agency_id"], updated["agency"]["agency_id"])
        self.assertEqual("Agency One", updated["agency"]["name"])
        self.assertEqual(created_at, updated["agency"]["created_at"])

        status, contact_body = self.request(
            "POST",
            "/api/local/agency-contacts",
            {"name": "Contact One", "agency_id": agency["agency_id"]},
        )
        self.assertEqual(200, status)
        contact = contact_body["contact"]
        self.assertRegex(contact["contact_id"], r"^contact_[0-9a-f]{16}$")
        status, updated_contact = self.request(
            "POST",
            "/api/local/agency-contacts",
            {"contact_id": contact["contact_id"], "note": "updated"},
        )
        self.assertEqual(200, status)
        self.assertEqual(
            contact["contact_id"], updated_contact["contact"]["contact_id"]
        )
        self.assertEqual("Contact One", updated_contact["contact"]["name"])

        status, relation = self.request(
            "POST",
            "/api/creator-library/creator_one/relations",
            {
                "agency_id": agency["agency_id"],
                "current_contact_id": contact["contact_id"],
                "source_contact_id": contact["contact_id"],
            },
        )
        self.assertEqual(200, status)
        self.assertEqual(agency["agency_id"], relation["agency_id"])
        self.assertEqual(contact["contact_id"], relation["current_contact_id"])

        status, detail = self.request(
            "GET", f"/api/local/agencies/{agency['agency_id']}"
        )
        self.assertEqual(200, status)
        self.assertEqual({"ok", "agency", "contacts", "creators"}, set(detail))
        self.assertEqual(1, len(detail["contacts"]))
        self.assertEqual(1, len(detail["creators"]))

        status, agencies = self.request("GET", "/api/local/agencies")
        self.assertEqual(200, status)
        self.assertEqual(1, agencies["agencies"][0]["creator_count"])
        self.assertEqual(1, agencies["agencies"][0]["contact_count"])
        status, contacts = self.request("GET", "/api/local/agency-contacts")
        self.assertEqual(200, status)
        self.assertEqual(contact["contact_id"], contacts["contacts"][0]["contact_id"])

        status, error = self.request(
            "POST", "/api/local/agencies", {"name": ""}
        )
        self.assertEqual(400, status)
        self.assertEqual("Agency 名称不能为空。", error["error"])
        status, error = self.request(
            "POST",
            "/api/local/agency-contacts",
            {"name": "Invalid", "agency_id": "missing"},
        )
        self.assertEqual(400, status)
        self.assertEqual("联系人关联的 Agency 不存在。", error["error"])
        status, error = self.request("GET", "/api/local/agencies/missing")
        self.assertEqual(404, status)
        self.assertEqual("未找到 Agency。", error["error"])

    def test_creator_relation_validation_http_contract(self) -> None:
        for creator_id, payload, expected_error in (
            ("creator_one", {"agency_id": "missing"}, "关联的 Agency 不存在。"),
            (
                "creator_one",
                {"current_contact_id": "missing"},
                "关联的 Agency 联系人不存在。",
            ),
            ("missing", {}, "未找到达人分析记录。"),
        ):
            status, body = self.request(
                "POST", f"/api/creator-library/{creator_id}/relations", payload
            )
            self.assertEqual(400, status)
            self.assertEqual(expected_error, body["error"])

    def test_external_agency_contacts_remain_a_separate_compatibility_route(self) -> None:
        status, body = self.request("GET", "/api/agency-contacts")
        self.assertEqual(200, status)
        self.assertEqual({"ok", "configured", "contacts"}, set(body))
        self.assertTrue(body["configured"])
        self.assertEqual("external-contact-1", body["contacts"][0]["record_id"])

        with mock.patch.object(
            server,
            "get_agency_contact_options",
            return_value={"configured": False, "contacts": []},
        ):
            status, empty = self.request("GET", "/api/agency-contacts")
        self.assertEqual(200, status)
        self.assertEqual(
            {"ok": True, "configured": False, "contacts": []}, empty
        )

        with mock.patch.object(
            server,
            "get_agency_contact_options",
            side_effect=RuntimeError("Agency联系人表飞书配置不完整。"),
        ):
            status, error = self.request("GET", "/api/agency-contacts")
        self.assertEqual(400, status)
        self.assertEqual("Agency联系人表飞书配置不完整。", error["error"])


class AgencyBoundaryScopeTests(unittest.TestCase):
    def test_factory_shares_one_store_and_one_scoped_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            factory = RepositoryFactory.for_path(
                Path(temp_dir) / "Creator_Library.xlsx"
            )
            creator_repository = factory.creator()
            agency_repository = factory.agency()
            self.assertIs(factory.store, creator_repository.store)
            self.assertIs(factory.store, agency_repository.store)
            self.assertIs(agency_repository, factory.agency())

            service = AgencyService(factory.agency, factory.creator)
            with mock.patch.object(
                factory.store,
                "_open_now",
                wraps=factory.store._open_now,
            ) as open_now:
                with factory.request_scope():
                    self.assertEqual([], service.get_agencies()["agencies"])
                self.assertEqual(1, open_now.call_count)

    def test_excel_schema_ids_and_creator_relation_columns_are_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            factory = RepositoryFactory.for_path(workbook_path)
            factory.creator().getCreators()
            workbook = load_workbook(workbook_path)
            sheets_before = list(workbook.sheetnames)
            creator_headers_before = [cell.value for cell in workbook["Creators"][1]]
            workbook.close()

            agency = factory.agency().save_agency({"name": "Stable Agency"})
            contact = factory.agency().save_contact(
                {"name": "Stable Contact", "agency_id": agency["agency_id"]}
            )
            updated_agency = factory.agency().save_agency(
                {"agency_id": agency["agency_id"], "note": "same id"}
            )
            updated_contact = factory.agency().save_contact(
                {"contact_id": contact["contact_id"], "note": "same id"}
            )

            workbook = load_workbook(workbook_path)
            self.assertEqual(sheets_before, list(workbook.sheetnames))
            self.assertEqual(
                creator_headers_before,
                [cell.value for cell in workbook["Creators"][1]],
            )
            self.assertIn("agency_id", creator_headers_before)
            self.assertIn("current_contact_id", creator_headers_before)
            self.assertIn("source_contact_id", creator_headers_before)
            workbook.close()
            self.assertEqual(agency["agency_id"], updated_agency["agency_id"])
            self.assertEqual(contact["contact_id"], updated_contact["contact_id"])

    def test_external_contact_upsert_is_stable_and_never_infers_agency(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            factory = RepositoryFactory.for_path(
                Path(temp_dir) / "Creator_Library.xlsx"
            )
            factory.creator().getCreators()
            first = factory.agency().upsert_external_contact(
                "external-record-1", name="External", whatsapp="+1"
            )
            second = factory.agency().upsert_external_contact(
                "external-record-1", name="External Updated", whatsapp="+2"
            )
            self.assertEqual(first["contact_id"], second["contact_id"])
            self.assertEqual("", second["agency_id"])
            self.assertEqual("external-record-1", second["external_record_id"])


if __name__ == "__main__":
    unittest.main()

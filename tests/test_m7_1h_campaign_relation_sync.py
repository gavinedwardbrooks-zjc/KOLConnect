from __future__ import annotations

import copy
import shutil
import sys
import unittest
from contextlib import nullcontext
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))
sys.path.insert(0, str(ROOT / "tests"))

from campaign_creator_repository import CampaignCreatorRepository  # noqa: E402
from campaign_repository import CampaignRepository  # noqa: E402
from creator_repository import _WORKBOOK_SHEETS  # noqa: E402
from feishu_client import FeishuClientError  # noqa: E402
from product_repository import ProductRepository  # noqa: E402
from services.feishu_sync_service import (  # noqa: E402
    ACCOUNT_CREATOR_RELATION_FIELD,
    ACCOUNT_FIELDS,
    CREATOR_ACCOUNT_RELATION_FIELD,
    CREATOR_FIELDS,
    FeishuSyncService,
)
from test_support.runtime_sandbox import test_artifact_path  # noqa: E402


class CampaignMultiValueTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime = test_artifact_path("m7_1h")
        runtime.mkdir(exist_ok=True)
        self.root = runtime / uuid4().hex
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in _WORKBOOK_SHEETS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(list(headers))
        workbook.save(self.workbook_path)
        workbook.close()
        self.lock_patch = patch(
            "excel_workbook_store.shared_storage_lock",
            side_effect=lambda *args, **kwargs: nullcontext(),
        )
        self.lock_patch.start()

    def tearDown(self) -> None:
        self.lock_patch.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    @staticmethod
    def _append(sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _seed(self):
        product = ProductRepository(self.workbook_path).createProduct({"name": "Game"})
        with patch("excel_workbook_store.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()):
            from openpyxl import load_workbook
            workbook = load_workbook(self.workbook_path)
            self._append(workbook["Creators"], {"creator_id": "creator-1", "name": "INSA"})
            for account_id, platform in (("yt", "YouTube"), ("tt", "TikTok"), ("ig", "Instagram")):
                self._append(workbook["CreatorAccounts"], {
                    "account_id": account_id,
                    "creator_id": "creator-1",
                    "account_uid": f"{platform.lower()}|{account_id}",
                    "platform": platform,
                    "profile_url": f"https://example.com/{account_id}",
                })
            workbook.save(self.workbook_path)
            workbook.close()
        return product

    def test_multi_platform_accounts_and_dates_roundtrip_without_duplicate_creator(self):
        product = self._seed()
        campaign = CampaignRepository(self.workbook_path).createCampaign({
            "name": "multi",
            "product_id": product["product_id"],
            "platforms": ["TikTok", "YouTube"],
        })
        self.assertEqual(["TikTok", "YouTube"], campaign["platforms"])
        relation = CampaignCreatorRepository(self.workbook_path).createCampaignCreator({
            "campaign_id": campaign["campaign_id"],
            "creator_id": "creator-1",
            "account_ids": ["yt", "tt", "yt"],
            "planned_publish_dates": ["2026-08-18", "2026-08-14"],
        })
        self.assertEqual(["yt", "tt"], relation["account_ids"])
        self.assertEqual(["2026-08-14", "2026-08-18"], relation["planned_publish_dates"])
        self.assertEqual("yt", relation["account_id"])
        self.assertEqual("2026-08-14", relation["publish_date"])
        rows = CampaignCreatorRepository(self.workbook_path).getCampaignCreators(
            campaign_id=campaign["campaign_id"]
        )
        self.assertEqual(1, len(rows))
        self.assertEqual(2, len(rows[0]["execution_accounts"]))

        updated = CampaignCreatorRepository(self.workbook_path).updateCampaignCreator(
            relation["id"], {"account_id": "tt"}
        )
        self.assertEqual(["tt"], updated["account_ids"])
        self.assertEqual("tt", updated["account_id"])

    def test_platform_eligibility_unrestricted_and_legacy_single_value(self):
        product = self._seed()
        repository = CampaignRepository(self.workbook_path)
        legacy = repository.createCampaign({
            "name": "legacy", "product_id": product["product_id"], "platform": "TikTok"
        })
        self.assertEqual(["TikTok"], repository.getCampaign(legacy["campaign_id"])["platforms"])
        with self.assertRaisesRegex(ValueError, "目标平台"):
            CampaignCreatorRepository(self.workbook_path).createCampaignCreator({
                "campaign_id": legacy["campaign_id"], "creator_id": "creator-1", "account_ids": ["ig"]
            })
        unrestricted = repository.createCampaign({"name": "any", "product_id": product["product_id"]})
        relation = CampaignCreatorRepository(self.workbook_path).createCampaignCreator({
            "campaign_id": unrestricted["campaign_id"],
            "creator_id": "creator-1",
            "account_ids": ["yt", "tt", "ig"],
        })
        self.assertEqual(["yt", "tt", "ig"], relation["account_ids"])

    def test_malformed_legacy_publish_date_remains_readable(self):
        self.assertEqual(
            [],
            CampaignCreatorRepository._date_list("legacy-date", strict=False),
        )
        with self.assertRaisesRegex(ValueError, "YYYY-MM-DD"):
            CampaignCreatorRepository._date_list("legacy-date")


def _schema(specs, relation_name):
    preferred = {"boolean": 7, "url": 15, "number": 2, "datetime": 5, "text": 1}
    target_table_id = (
        "accounts" if relation_name == CREATOR_ACCOUNT_RELATION_FIELD else "creators"
    )
    return [
        {"field_name": spec.remote_name, "type": preferred[spec.kind]} for spec in specs
    ] + [{
        "field_name": relation_name,
        "type": 21,
        "property": {"table_id": target_table_id, "multiple": True},
    }]


class Source:
    def __init__(self, creators, accounts):
        self.data = {"creators": creators, "accounts": accounts, "insights": [], "snapshots": []}

    def getCreatorInventoryRows(self):
        return copy.deepcopy(self.data)


class Client:
    creator_table_id = "creators"
    account_table_id = "accounts"
    batch_size = 100

    def __init__(self):
        self.records = {"creators": [], "accounts": []}
        self.fail_relation_once = False
        self.mutations = []

    def authenticate(self):
        return None

    def list_fields(self, table_id):
        return copy.deepcopy(
            _schema(CREATOR_FIELDS, CREATOR_ACCOUNT_RELATION_FIELD)
            if table_id == "creators"
            else _schema(ACCOUNT_FIELDS, ACCOUNT_CREATOR_RELATION_FIELD)
        )

    def list_records(self, table_id):
        return copy.deepcopy(self.records[table_id])

    def batch_create(self, table_id, payloads):
        created = []
        for payload in payloads:
            record = {
                "record_id": f"{table_id}-{len(self.records[table_id]) + 1}",
                "fields": copy.deepcopy(payload),
            }
            self.records[table_id].append(record)
            created.append(copy.deepcopy(record))
        self.mutations.append(("create", table_id, copy.deepcopy(created)))
        return created

    def batch_update(self, table_id, updates):
        updates = copy.deepcopy(list(updates))
        relation_update = any(
            CREATOR_ACCOUNT_RELATION_FIELD in item["fields"]
            or ACCOUNT_CREATOR_RELATION_FIELD in item["fields"]
            for item in updates
        )
        if relation_update and self.fail_relation_once:
            self.fail_relation_once = False
            raise FeishuClientError("TRANSIENT_REMOTE_ERROR", "temporary")
        for update in updates:
            record = next(row for row in self.records[table_id] if row["record_id"] == update["record_id"])
            record["fields"].update(update["fields"])
        self.mutations.append(("update", table_id, updates))
        return updates


def creator(identity):
    return {"creator_id": identity, "name": identity, "created_at": "2026-08-01T00:00:00Z"}


def account(uid, creator_id):
    return {
        "account_uid": uid,
        "creator_id": creator_id,
        "platform": "TikTok",
        "profile_url": f"https://example.com/{uid}",
    }


class FeishuRelationSyncTests(unittest.TestCase):
    def service(self, source, client):
        return FeishuSyncService(
            source,
            lambda: client,
            now_provider=lambda: datetime(2026, 8, 24, tzinfo=timezone.utc),
        )

    def test_one_creator_three_accounts_relations_and_second_run_idempotency(self):
        source = Source([creator("creator-a")], [account(f"uid-{i}", "creator-a") for i in range(3)])
        client = Client()
        service = self.service(source, client)
        dry = service.dry_run()
        self.assertEqual((1, 3, 3), (
            dry["creator_create_count"], dry["account_create_count"], dry["relation_add_count"]
        ))
        result = service.full_sync(confirm=True)
        self.assertEqual("success", result["status"])
        creator_links = FeishuSyncService._relation_ids(
            client.records["creators"][0]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )
        self.assertEqual(3, len(creator_links))
        self.assertTrue(all(
            FeishuSyncService._relation_ids(row["fields"][ACCOUNT_CREATOR_RELATION_FIELD])
            == [client.records["creators"][0]["record_id"]]
            for row in client.records["accounts"]
        ))
        second = service.dry_run()
        self.assertEqual((0, 0, 0, 3), (
            second["creator_create_count"], second["account_create_count"],
            second["relation_update_count"], second["relation_unchanged_count"],
        ))

    def test_two_creators_group_three_accounts_and_restore_missing_edge(self):
        source = Source(
            [creator("creator-a"), creator("creator-b")],
            [
                account("uid-a1", "creator-a"),
                account("uid-a2", "creator-a"),
                account("uid-b1", "creator-b"),
            ],
        )
        client = Client()
        service = self.service(source, client)
        self.assertEqual("success", service.full_sync(confirm=True)["status"])
        creators = {
            row["fields"]["KOLConnect Creator ID"]: row
            for row in client.records["creators"]
        }
        self.assertEqual(2, len(FeishuSyncService._relation_ids(
            creators["creator-a"]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )))
        self.assertEqual(1, len(FeishuSyncService._relation_ids(
            creators["creator-b"]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )))

        creator_a_links = FeishuSyncService._relation_ids(
            creators["creator-a"]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )
        creators["creator-a"]["fields"][CREATOR_ACCOUNT_RELATION_FIELD] = creator_a_links[:1]
        dry = service.dry_run()
        self.assertEqual(1, dry["relation_add_count"])
        self.assertEqual(1, dry["relation_update_count"])
        self.assertEqual("success", service.full_sync(confirm=True)["status"])
        self.assertEqual(2, len(FeishuSyncService._relation_ids(
            creators["creator-a"]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )))

    def test_relation_removal_and_partial_failure_rerun(self):
        source = Source([creator("creator-a")], [account("uid-1", "creator-a"), account("uid-2", "creator-a")])
        client = Client()
        service = self.service(source, client)
        self.assertEqual("success", service.full_sync(confirm=True)["status"])
        source.data["accounts"] = [account("uid-1", "creator-a")]
        dry = service.dry_run()
        self.assertEqual(1, dry["relation_remove_count"])
        client.fail_relation_once = True
        partial = service.full_sync(confirm=True)
        self.assertEqual("failed", partial["status"])
        rerun = service.full_sync(confirm=True)
        self.assertEqual("success", rerun["status"])
        links = FeishuSyncService._relation_ids(
            client.records["creators"][0]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )
        self.assertEqual([client.records["accounts"][0]["record_id"]], links)
        self.assertEqual([], FeishuSyncService._relation_ids(
            client.records["accounts"][1]["fields"].get(ACCOUNT_CREATOR_RELATION_FIELD)
        ))

    def test_business_create_then_relation_failure_is_partial_and_rerunnable(self):
        source = Source([creator("creator-a")], [account("uid-1", "creator-a")])
        client = Client()
        client.fail_relation_once = True
        service = self.service(source, client)
        first = service.full_sync(confirm=True)
        self.assertEqual("partial", first["status"])
        self.assertEqual("relation_update", first["phase"])
        self.assertEqual((1, 1), (first["creator_created"], first["account_created"]))
        second = service.full_sync(confirm=True)
        self.assertEqual("success", second["status"])
        self.assertEqual(1, len(FeishuSyncService._relation_ids(
            client.records["creators"][0]["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )))

    def test_manual_merge_moves_account_relation_without_duplicate_create(self):
        client = Client()
        before = Source(
            [creator("creator-a"), creator("creator-b")],
            [account("uid-y", "creator-a"), account("uid-t", "creator-b")],
        )
        self.assertEqual("success", self.service(before, client).full_sync(confirm=True)["status"])
        after = Source(
            [creator("creator-a")],
            [account("uid-y", "creator-a"), account("uid-t", "creator-a")],
        )
        service = self.service(after, client)
        dry = service.dry_run()
        self.assertEqual(0, dry["account_create_count"])
        self.assertGreaterEqual(dry["relation_update_count"], 1)
        self.assertEqual("success", service.full_sync(confirm=True)["status"])
        creator_a = next(
            row for row in client.records["creators"]
            if row["fields"].get("KOLConnect Creator ID") == "creator-a"
        )
        self.assertEqual(2, len(FeishuSyncService._relation_ids(
            creator_a["fields"][CREATOR_ACCOUNT_RELATION_FIELD]
        )))
        creator_b = next(
            row for row in client.records["creators"]
            if row["fields"].get("KOLConnect Creator ID") == "creator-b"
        )
        self.assertEqual([], FeishuSyncService._relation_ids(
            creator_b["fields"].get(CREATOR_ACCOUNT_RELATION_FIELD)
        ))


if __name__ == "__main__":
    unittest.main()

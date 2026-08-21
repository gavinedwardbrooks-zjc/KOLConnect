from __future__ import annotations

import json
import sys
import tempfile
import threading
import unittest
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import server
from repositories.agency_repository import AgencyRepository


class CreatorLibraryCrmTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        self.repository = creator_repository.CreatorRepository(self.workbook_path)
        self.repository.getCreators()
        self._seed_data()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _append(self, sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _seed_data(self) -> None:
        workbook = load_workbook(self.workbook_path)
        analysis = {
            "analysis_id": "creator_one",
            "creator": {
                "creator_name": "Creator One",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@creator-one",
                "followers": "10K",
                "bio": "Original bio",
            },
            "content_category": "Gaming",
            "creator_insight": {"level": "good"},
        }
        self._append(workbook["Creators"], {
            "creator_id": "creator_one",
            "name": "Creator One",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@creator-one",
            "followers": "10K",
            "content_category": "Gaming",
            "status": "contacted",
            "created_at": "2026-08-01T00:00:00Z",
            "updated_at": "2026-08-01T00:00:00Z",
        })
        self._append(workbook["CreatorAccounts"], {
            "account_id": "account_one",
            "creator_id": "creator_one",
            "account_uid": "tiktok|creator-one",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@creator-one",
            "followers": "10K",
        })
        self._append(workbook["CreatorSnapshots"], {
            "snapshot_id": "snapshot_one",
            "creator_id": "creator_one",
            "followers": "10K",
            "captured_at": "2026-08-01T00:00:00Z",
        })
        self._append(workbook["Insights"], {"creator_id": "creator_one", "average_views": 5000})
        self._append(workbook["Agencies"], {"agency_id": "agency_one", "name": "Agency One"})
        self._append(workbook["CampaignCreators"], {
            "id": "campaign_creator_one",
            "campaign_id": "campaign_one",
            "creator_id": "creator_one",
            "account_id": "account_one",
            "stage": "executing",
        })
        self._append(workbook["_AnalysisData"], {
            "creator_id": "creator_one",
            "account_uid": "tiktok|creator-one",
            "analysis_json": json.dumps(analysis, ensure_ascii=False),
        })
        workbook.save(self.workbook_path)
        workbook.close()

    def test_edit_and_archive_preserve_identity_history_and_relations(self) -> None:
        workbook = load_workbook(self.workbook_path)
        creator_headers_before = [cell.value for cell in workbook["Creators"][1]]
        workbook.close()

        updated = self.repository.updateCreator("creator_one", {
            "creator_name": "Creator Updated",
            "profile_url": "https://www.tiktok.com/@creator-updated",
            "followers": "25K",
            "content_category": "Lifestyle",
            "bio": "Updated bio",
            "agency_id": "agency_one",
        })
        self.assertEqual("creator_one", updated["creator_id"])
        detail = self.repository.getCreatorDetail("creator_one")
        self.assertEqual("Creator Updated", detail["record"]["creator_name"])
        self.assertEqual("25K", detail["record"]["followers"])
        self.assertEqual("Updated bio", detail["analysis"]["creator"]["bio"])
        self.assertEqual("agency_one", detail["record"]["agency_id"])

        archived_at = "2026-08-03T12:00:00Z"
        self.repository.updateCreator("creator_one", {"archived_at": archived_at})
        self.assertEqual([], self.repository.getCreators())
        archived = self.repository.getCreators(include_archived=True)
        self.assertEqual(1, len(archived))
        self.assertEqual("contacted", archived[0]["status"])
        self.assertEqual(archived_at, archived[0]["archived_at"])

        workbook = load_workbook(self.workbook_path)
        self.assertEqual(creator_headers_before, [cell.value for cell in workbook["Creators"][1]])
        self.assertEqual("creator_one", workbook["Creators"][2][0].value)
        account_headers = [cell.value for cell in workbook["CreatorAccounts"][1]]
        account = dict(zip(account_headers, [cell.value for cell in workbook["CreatorAccounts"][2]]))
        self.assertEqual("account_one", account["account_id"])
        self.assertEqual("tiktok|creator-one", account["account_uid"])
        self.assertEqual("TikTok", account["platform"])
        self.assertEqual(1, workbook["CreatorSnapshots"].max_row - 1)
        self.assertEqual(1, workbook["Insights"].max_row - 1)
        self.assertEqual(1, workbook["CampaignCreators"].max_row - 1)
        workbook.close()

        self.repository.updateCreator("creator_one", {"archived_at": None})
        self.assertEqual(1, len(self.repository.getCreators()))

    def test_invalid_agency_and_identity_fields_are_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "Agency"):
            self.repository.updateCreator("creator_one", {"agency_id": "missing"})
        with self.assertRaisesRegex(ValueError, "不允许修改"):
            self.repository.updateCreator("creator_one", {"creator_id": "changed"})

    def test_creator_patch_and_archived_query_api(self) -> None:
        patchers = [
            mock.patch.object(server, "get_creator_repository", return_value=self.repository),
            mock.patch.object(
                server,
                "get_agency_repository",
                return_value=AgencyRepository(self.repository.store),
            ),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in patchers:
            patcher.start()
        httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        thread.start()
        base_url = f"http://127.0.0.1:{httpd.server_port}"
        try:
            payload = json.dumps({
                "creator_name": "API Updated",
                "profile_url": "https://www.tiktok.com/@api-updated",
                "followers": "30K",
                "content_category": "Education",
                "bio": "API bio",
                "agency_id": "agency_one",
            }).encode("utf-8")
            request = urllib.request.Request(
                f"{base_url}/api/creator-library/creator_one",
                data=payload,
                headers={"Content-Type": "application/json"},
                method="PATCH",
            )
            with urllib.request.urlopen(request, timeout=5) as response:
                body = json.loads(response.read().decode("utf-8"))
            self.assertTrue(body["ok"])
            self.assertEqual("API Updated", body["creator"]["creator_name"])

            archive_request = urllib.request.Request(
                f"{base_url}/api/creator-library/creator_one",
                data=json.dumps({"archived_at": "2026-08-03T12:00:00Z"}).encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="PATCH",
            )
            with urllib.request.urlopen(archive_request, timeout=5):
                pass
            with urllib.request.urlopen(f"{base_url}/api/creator-library", timeout=5) as response:
                active = json.loads(response.read().decode("utf-8"))
            with urllib.request.urlopen(
                f"{base_url}/api/creator-library?include_archived=true", timeout=5
            ) as response:
                history = json.loads(response.read().decode("utf-8"))
            self.assertEqual([], active["records"])
            self.assertEqual(1, len(history["records"]))
        finally:
            httpd.shutdown()
            httpd.server_close()
            thread.join(timeout=5)
            for patcher in reversed(patchers):
                patcher.stop()


if __name__ == "__main__":
    with mock.patch.object(creator_repository, "log_event"):
        unittest.main()

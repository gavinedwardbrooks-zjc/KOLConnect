from __future__ import annotations

import json
import shutil
import sys
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import creator_repository  # noqa: E402
from http_handlers import creator_handler  # noqa: E402
from local_storage_lock import (  # noqa: E402
    SharedStorageLockTimeout,
    shared_storage_lock_held,
)
from repository_factory import RepositoryFactory  # noqa: E402
from services.creator_delete_impact_service import CreatorDeleteImpactService  # noqa: E402
from services.creator_hard_delete_service import (  # noqa: E402
    CreatorHardDeleteError,
    CreatorHardDeleteService,
)
from services.feishu_delete_intent_service import FeishuDeleteIntentStore  # noqa: E402
from staged_delete_transaction import StagedDeleteTransaction  # noqa: E402


def append_row(workbook, sheet_name: str, values: dict) -> None:
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    sheet.append([values.get(header, "") for header in headers])


def rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    result = [
        {
            headers[index]: "" if values[index] is None else values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]
    workbook.close()
    return result


class FakeHandler:
    def __init__(self) -> None:
        self.headers = {"Content-Type": "application/json"}
        self.response = None

    def _json(self, data, status=200):
        self.response = (status, data)


class CreatorHardDeleteTests(unittest.TestCase):
    def setUp(self) -> None:
        self.root = ROOT / f".m4_6_delete_test_{uuid.uuid4().hex}"
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.tasks_dir = self.root / "tasks"
        self.protection_file = self.root / "data_protection.json"
        self.legacy_dir = self.root / "creator_analysis"
        self.legacy_file = self.root / "creator_library.json"
        self.runtime_dir = self.root / "runtime"
        self.lock_patcher = mock.patch(
            "local_storage_lock.get_shared_storage_lock_path",
            return_value=self.root / "locks" / "shared_storage.lock",
        )
        self.lock_patcher.start()
        self.log_patcher = mock.patch.object(creator_repository, "log_event")
        self.log_patcher.start()
        self.factory = RepositoryFactory.for_path(
            self.workbook_path,
            tasks_dir=self.tasks_dir,
            data_protection_file=self.protection_file,
            legacy_analysis_dir=self.legacy_dir,
            legacy_library_file=self.legacy_file,
        )
        self.factory.creator().getCreators()
        self.impact_service = CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        )
        self.errors: list[BaseException] = []
        self.intent_store = FeishuDeleteIntentStore(self.runtime_dir)
        self.service = CreatorHardDeleteService(
            lambda: self.impact_service,
            self.factory.creator_hard_delete,
            lambda: self.runtime_dir,
            self.errors.append,
            lock_timeout=2,
            feishu_delete_intent_store=self.intent_store,
        )

    def tearDown(self) -> None:
        self.log_patcher.stop()
        self.lock_patcher.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def append(self, sheet_name: str, values: dict) -> None:
        workbook = load_workbook(self.workbook_path)
        append_row(workbook, sheet_name, values)
        workbook.save(self.workbook_path)
        workbook.close()

    def seed_creator(
        self,
        creator_id: str,
        *,
        archived: bool = False,
        agency_id: str = "",
    ) -> None:
        self.append("Creators", {
            "creator_id": creator_id,
            "name": creator_id,
            "agency_id": agency_id,
            "email": f"{creator_id}@private.example",
            "bio": "secret bio",
            "archived_at": "2026-08-17T00:00:00Z" if archived else "",
            "updated_at": "2026-08-17T00:00:00Z",
        })

    def preview(self, creator_id: str) -> dict:
        return self.impact_service.get_delete_impact(creator_id)

    def delete(self, creator_id: str, fingerprint: str) -> dict:
        return self.service.delete_creator(
            creator_id,
            confirm=True,
            preview_fingerprint=fingerprint,
        )

    def assert_error(self, code: str, callback) -> CreatorHardDeleteError:
        with self.assertRaises(CreatorHardDeleteError) as caught:
            callback()
        self.assertEqual(code, caught.exception.code)
        return caught.exception

    def test_successful_local_delete_promotes_durable_remote_intent(self) -> None:
        self.seed_creator("creator_remote_intent")
        self.append("CreatorAccounts", {
            "account_id": "account_remote_intent",
            "creator_id": "creator_remote_intent",
            "account_uid": "uid_remote_intent",
            "platform": "TikTok",
        })
        result = self.delete(
            "creator_remote_intent",
            self.preview("creator_remote_intent")["preview_fingerprint"],
        )
        intents = self.intent_store.list()
        self.assertTrue(result["deleted"])
        self.assertEqual(1, len(intents))
        self.assertEqual("pending_remote", intents[0]["status"])
        self.assertEqual("creator_remote_intent", intents[0]["creator_id"])
        self.assertEqual(["uid_remote_intent"], intents[0]["account_uids"])

    def test_failed_local_delete_aborts_remote_intent(self) -> None:
        self.seed_creator("creator_failed_intent")
        preview = self.preview("creator_failed_intent")
        with mock.patch.object(
            self.factory.creator_hard_delete(),
            "delete_workbook_resources",
            side_effect=RuntimeError("synthetic local failure"),
        ):
            self.assert_error(
                "CREATOR_DELETE_FAILED",
                lambda: self.delete(
                    "creator_failed_intent", preview["preview_fingerprint"]
                ),
            )
        intents = self.intent_store.list()
        self.assertEqual(1, len(intents))
        self.assertEqual("aborted", intents[0]["status"])

    def test_request_contract_requires_literal_confirmation_and_fingerprint(self) -> None:
        self.seed_creator("creator_contract")
        before = self.workbook_path.read_bytes()
        for confirm, fingerprint, code in (
            (None, None, "DELETE_CONFIRMATION_REQUIRED"),
            (False, "fingerprint", "DELETE_CONFIRMATION_REQUIRED"),
            ("true", "fingerprint", "DELETE_CONFIRMATION_REQUIRED"),
            (True, None, "DELETE_PREVIEW_REQUIRED"),
            (True, "   ", "DELETE_PREVIEW_REQUIRED"),
        ):
            self.assert_error(
                code,
                lambda confirm=confirm, fingerprint=fingerprint: self.service.delete_creator(
                    "creator_contract",
                    confirm=confirm,
                    preview_fingerprint=fingerprint,
                ),
            )
        self.assertEqual(before, self.workbook_path.read_bytes())
        self.assertFalse((self.runtime_dir / "delete_transactions").exists())

    def test_not_found_and_blocked_are_fail_closed_and_privacy_safe(self) -> None:
        missing = self.assert_error(
            "CREATOR_NOT_FOUND",
            lambda: self.delete("missing", "a" * 64),
        )
        self.assertEqual(404, missing.status)

        self.seed_creator("creator_blocked", archived=True)
        self.append("Cooperations", {
            "cooperation_id": "cooperation_block",
            "creator_id": "creator_blocked",
            "note": "private cooperation note",
        })
        impact = self.preview("creator_blocked")
        before = self.workbook_path.read_bytes()
        blocked = self.assert_error(
            "DELETE_BLOCKED",
            lambda: self.delete("creator_blocked", impact["preview_fingerprint"]),
        )
        self.assertEqual(409, blocked.status)
        response = json.dumps(blocked.to_response(), ensure_ascii=False)
        self.assertIn("COOPERATION_RETENTION_ANONYMIZATION_GAP", response)
        self.assertNotIn("private cooperation note", response)
        self.assertNotIn(str(self.root), response)
        self.assertEqual(before, self.workbook_path.read_bytes())

    def test_safe_delete_removes_exact_resources_and_preserves_roots(self) -> None:
        self.append("Agencies", {"agency_id": "agency_keep", "name": "Keep Agency"})
        self.append("Products", {"product_id": "product_keep", "name": "Keep Product"})
        self.append("Campaigns", {
            "campaign_id": "campaign_keep",
            "product_id": "product_keep",
            "name": "Keep Campaign",
        })
        self.seed_creator("creator_target", agency_id="agency_keep")
        self.seed_creator("creator_other")
        self.append("CreatorAccounts", {
            "account_id": "account_target",
            "creator_id": "creator_target",
            "account_uid": "uid_target",
        })
        self.append("CreatorAccounts", {
            "account_id": "account_other",
            "creator_id": "creator_other",
            "account_uid": "uid_other",
        })
        self.append("Videos", {"creator_id": "creator_target", "video_url": "target-1"})
        self.append("Videos", {"creator_id": "creator_target", "video_url": "target-2"})
        self.append("Videos", {"creator_id": "creator_other", "video_url": "other"})
        self.append("Insights", {"creator_id": "creator_target", "recommendation": "target"})
        self.append("_AnalysisData", {
            "creator_id": "creator_target",
            "analysis_json": json.dumps({"creator_id": "creator_target"}),
        })
        self.append("CreatorSnapshots", {
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_target",
        })
        self.append("VideoSnapshots", {
            "video_snapshot_id": "video_snapshot_target",
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_target",
        })
        self.append("CampaignCreators", {
            "id": "relation_target",
            "campaign_id": "campaign_keep",
            "creator_id": "creator_target",
            "account_id": "account_target",
            "archived_at": "2026-08-16T00:00:00Z",
        })
        self.append("FollowUpLogs", {
            "follow_up_id": "followup_target",
            "object_type": "creator",
            "object_id": "creator_target",
        })
        self.append("FollowUpLogs", {
            "follow_up_id": "followup_other",
            "object_type": "creator",
            "object_id": "creator_other",
        })
        task = self.tasks_dir / "task_20260817T000000Z_aaaaaaaa"
        task.mkdir(parents=True)
        (task / "task.json").write_text(json.dumps({
            "id": task.name,
            "creator_library_creator_ids": ["creator_target"],
        }), encoding="utf-8")
        self.protection_file.write_text(json.dumps({
            "uid_target": {"email": {"value": "target@example.com"}},
            "uid_other": {"email": {"value": "other@example.com"}},
        }), encoding="utf-8")
        self.legacy_dir.mkdir()
        legacy_analysis = self.legacy_dir / "analysis_task_target.json"
        legacy_analysis.write_text(json.dumps({
            "creator_id": "creator_target", "account_uid": "uid_target",
        }), encoding="utf-8")
        self.legacy_file.write_text(json.dumps({"records": {
            "creator_target": {"creator_id": "creator_target", "account_uid": "uid_target"},
            "creator_other": {"creator_id": "creator_other", "account_uid": "uid_other"},
        }}), encoding="utf-8")

        impact = self.preview("creator_target")
        self.assertTrue(impact["can_delete"])
        result = self.delete("creator_target", impact["preview_fingerprint"])

        self.assertEqual({"creator_id": "creator_target", "deleted": True}, result)
        self.assertEqual([], [row for row in rows(self.workbook_path, "Creators") if row["creator_id"] == "creator_target"])
        self.assertEqual(1, len(rows(self.workbook_path, "Creators")))
        self.assertEqual("creator_other", rows(self.workbook_path, "Creators")[0]["creator_id"])
        self.assertEqual("account_other", rows(self.workbook_path, "CreatorAccounts")[0]["account_id"])
        self.assertEqual("other", rows(self.workbook_path, "Videos")[0]["video_url"])
        self.assertEqual("followup_other", rows(self.workbook_path, "FollowUpLogs")[0]["follow_up_id"])
        self.assertEqual("campaign_keep", rows(self.workbook_path, "Campaigns")[0]["campaign_id"])
        self.assertEqual("product_keep", rows(self.workbook_path, "Products")[0]["product_id"])
        self.assertEqual("agency_keep", rows(self.workbook_path, "Agencies")[0]["agency_id"])
        self.assertFalse(task.exists())
        self.assertFalse(legacy_analysis.exists())
        self.assertEqual(
            {"uid_other": {"email": {"value": "other@example.com"}}},
            json.loads(self.protection_file.read_text(encoding="utf-8")),
        )
        self.assertEqual(
            {"creator_other": {"creator_id": "creator_other", "account_uid": "uid_other"}},
            json.loads(self.legacy_file.read_text(encoding="utf-8"))["records"],
        )
        manifests = list((self.runtime_dir / "delete_transactions").glob("*/manifest.json"))
        self.assertEqual(1, len(manifests))
        self.assertEqual("CLEANED", json.loads(manifests[0].read_text(encoding="utf-8"))["phase"])

    def test_unarchived_and_archived_creators_share_hard_delete_semantics(self) -> None:
        for creator_id, archived in (("creator_active", False), ("creator_archived", True)):
            self.seed_creator(creator_id, archived=archived)
            impact = self.preview(creator_id)
            self.assertTrue(impact["can_delete"])
            self.assertNotIn(
                "CREATOR_NOT_ARCHIVED", {item["code"] for item in impact["blockers"]}
            )
            self.assertTrue(self.delete(creator_id, impact["preview_fingerprint"])["deleted"])

    def test_stale_preview_and_new_blocker_prevent_mutation(self) -> None:
        self.seed_creator("creator_stale")
        fingerprint = self.preview("creator_stale")["preview_fingerprint"]
        self.append("CreatorSnapshots", {
            "snapshot_id": "snapshot_new",
            "creator_id": "creator_stale",
        })
        before = self.workbook_path.read_bytes()
        self.assert_error(
            "DELETE_PREVIEW_STALE",
            lambda: self.delete("creator_stale", fingerprint),
        )
        self.assertEqual(before, self.workbook_path.read_bytes())

        current = self.preview("creator_stale")
        self.append("Cooperations", {
            "cooperation_id": "cooperation_new",
            "creator_id": "creator_stale",
        })
        blocked = self.assert_error(
            "DELETE_BLOCKED",
            lambda: self.delete("creator_stale", current["preview_fingerprint"]),
        )
        self.assertEqual("DELETE_BLOCKED", blocked.code)

    def test_active_campaign_and_video_ownership_conflict_block_delete(self) -> None:
        self.seed_creator("creator_conflict")
        self.seed_creator("creator_other")
        self.append("CreatorAccounts", {
            "account_id": "account_conflict",
            "creator_id": "creator_conflict",
            "account_uid": "uid_conflict",
        })
        self.append("CreatorSnapshots", {
            "snapshot_id": "snapshot_conflict",
            "creator_id": "creator_conflict",
        })
        self.append("VideoSnapshots", {
            "video_snapshot_id": "video_conflict",
            "snapshot_id": "snapshot_conflict",
            "creator_id": "creator_other",
        })
        self.append("Campaigns", {"campaign_id": "campaign_active", "name": "Active"})
        self.append("CampaignCreators", {
            "id": "relation_active",
            "campaign_id": "campaign_active",
            "creator_id": "creator_conflict",
            "account_id": "account_conflict",
            "archived_at": "",
        })
        impact = self.preview("creator_conflict")
        codes = {item["code"] for item in impact["blockers"]}
        self.assertIn("VIDEO_SNAPSHOT_OWNERSHIP_CONFLICT", codes)
        self.assertIn("ACTIVE_CAMPAIGN_RELATION", codes)
        before = self.workbook_path.read_bytes()
        self.assert_error(
            "DELETE_BLOCKED",
            lambda: self.delete("creator_conflict", impact["preview_fingerprint"]),
        )
        self.assertEqual(before, self.workbook_path.read_bytes())

    def test_pending_manifest_blocks_before_new_transaction_or_mutation(self) -> None:
        self.seed_creator("creator_pending")
        fingerprint = self.preview("creator_pending")["preview_fingerprint"]
        pending = StagedDeleteTransaction(
            self.runtime_dir,
            "creator_previous",
            transaction_id="delete_previous_pending",
        )
        pending.prepare()
        before = self.workbook_path.read_bytes()
        self.assert_error(
            "CREATOR_DELETE_FAILED",
            lambda: self.delete("creator_pending", fingerprint),
        )
        self.assertEqual(before, self.workbook_path.read_bytes())
        roots = list((self.runtime_dir / "delete_transactions").iterdir())
        self.assertEqual(["delete_previous_pending"], [item.name for item in roots])

    def test_incomplete_delete_plan_locator_coverage_fails_closed(self) -> None:
        self.seed_creator("creator_incomplete")
        assessment = self.impact_service.inspect_delete_impact("creator_incomplete")
        plan = dict(assessment["plan"])
        plan["delete_locators"] = []
        with self.assertRaisesRegex(RuntimeError, "no exact locators"):
            self.factory.creator_hard_delete().transaction_inputs(plan)
        self.assertEqual(1, len([
            row for row in rows(self.workbook_path, "Creators")
            if row["creator_id"] == "creator_incomplete"
        ]))

    def test_workbook_is_saved_once(self) -> None:
        self.seed_creator("creator_once")
        fingerprint = self.preview("creator_once")["preview_fingerprint"]
        with mock.patch.object(
            self.factory.store,
            "_save_now",
            wraps=self.factory.store._save_now,
        ) as save:
            self.delete("creator_once", fingerprint)
        self.assertEqual(1, save.call_count)

    def test_rescan_and_fingerprint_check_run_while_shared_lock_is_held(self) -> None:
        self.seed_creator("creator_locked_scan")
        fingerprint = self.preview("creator_locked_scan")["preview_fingerprint"]
        real_inspect = self.impact_service.inspect_delete_impact

        def inspect(creator_id):
            self.assertTrue(shared_storage_lock_held())
            return real_inspect(creator_id)

        with mock.patch.object(self.impact_service, "inspect_delete_impact", side_effect=inspect):
            self.delete("creator_locked_scan", fingerprint)

    def test_typed_shared_lock_timeout_is_a_conflict_without_mutation(self) -> None:
        self.seed_creator("creator_lock_timeout")
        before = self.workbook_path.read_bytes()

        @contextmanager
        def timed_out_lock(**_options):
            raise SharedStorageLockTimeout("busy")
            yield

        with mock.patch(
            "services.creator_hard_delete_service.shared_storage_lock",
            timed_out_lock,
        ):
            error = self.assert_error(
                "SHARED_STORAGE_LOCK_TIMEOUT",
                lambda: self.service.delete_creator(
                    "creator_lock_timeout",
                    confirm=True,
                    preview_fingerprint="f" * 64,
                ),
            )
        self.assertEqual(409, error.status)
        self.assertEqual(before, self.workbook_path.read_bytes())
        self.assertFalse((self.runtime_dir / "delete_transactions").exists())

    def test_json_failure_restores_staged_artifact_before_workbook_mutation(self) -> None:
        self.seed_creator("creator_json_failure")
        self.append("CreatorAccounts", {
            "account_id": "account_json_failure",
            "creator_id": "creator_json_failure",
            "account_uid": "uid_json_failure",
        })
        task = self.tasks_dir / "task_20260817T000003Z_dddddddd"
        task.mkdir(parents=True)
        (task / "task.json").write_text(json.dumps({
            "id": task.name,
            "creator_library_creator_ids": ["creator_json_failure"],
        }), encoding="utf-8")
        protection = {"uid_json_failure": {"email": {"value": "json@example.com"}}}
        self.protection_file.write_text(json.dumps(protection), encoding="utf-8")
        fingerprint = self.preview("creator_json_failure")["preview_fingerprint"]
        with mock.patch.object(
            StagedDeleteTransaction,
            "write_json",
            side_effect=OSError("json replace failed"),
        ):
            self.assert_error(
                "CREATOR_DELETE_FAILED",
                lambda: self.delete("creator_json_failure", fingerprint),
            )
        self.assertTrue(task.is_dir())
        self.assertEqual(protection, json.loads(self.protection_file.read_text(encoding="utf-8")))
        self.assertEqual(1, len([row for row in rows(self.workbook_path, "Creators") if row["creator_id"] == "creator_json_failure"]))
        manifest = next((self.runtime_dir / "delete_transactions").glob("*/manifest.json"))
        self.assertEqual("ROLLED_BACK", json.loads(manifest.read_text(encoding="utf-8"))["phase"])

    def test_precommit_failure_rolls_back_json_artifact_and_workbook(self) -> None:
        self.seed_creator("creator_rollback")
        self.append("CreatorAccounts", {
            "account_id": "account_rollback",
            "creator_id": "creator_rollback",
            "account_uid": "uid_rollback",
        })
        task = self.tasks_dir / "task_20260817T000001Z_bbbbbbbb"
        task.mkdir(parents=True)
        (task / "task.json").write_text(json.dumps({
            "id": task.name,
            "creator_library_creator_ids": ["creator_rollback"],
        }), encoding="utf-8")
        original_protection = {"uid_rollback": {"email": {"value": "rollback@example.com"}}}
        self.protection_file.write_text(json.dumps(original_protection), encoding="utf-8")
        fingerprint = self.preview("creator_rollback")["preview_fingerprint"]
        repository = self.factory.creator_hard_delete()
        with mock.patch.object(
            repository,
            "verify_delete",
            side_effect=RuntimeError("verification failed C:\\private\\path"),
        ):
            error = self.assert_error(
                "CREATOR_DELETE_FAILED",
                lambda: self.delete("creator_rollback", fingerprint),
            )
        self.assertEqual({"ok": False, "error": "CREATOR_DELETE_FAILED"}, error.to_response())
        self.assertTrue(task.is_dir())
        self.assertEqual(original_protection, json.loads(self.protection_file.read_text(encoding="utf-8")))
        self.assertEqual(1, len([row for row in rows(self.workbook_path, "Creators") if row["creator_id"] == "creator_rollback"]))
        manifests = list((self.runtime_dir / "delete_transactions").glob("*/manifest.json"))
        self.assertEqual("ROLLED_BACK", json.loads(manifests[0].read_text(encoding="utf-8"))["phase"])

    def test_error_logger_failure_cannot_prevent_precommit_rollback(self) -> None:
        self.seed_creator("creator_log_failure")
        fingerprint = self.preview("creator_log_failure")["preview_fingerprint"]
        repository = self.factory.creator_hard_delete()
        service = CreatorHardDeleteService(
            lambda: self.impact_service,
            lambda: repository,
            lambda: self.runtime_dir,
            lambda _exc: (_ for _ in ()).throw(OSError("log unavailable")),
            lock_timeout=2,
        )
        with mock.patch.object(
            repository,
            "verify_delete",
            side_effect=RuntimeError("verification failed"),
        ):
            self.assert_error(
                "CREATOR_DELETE_FAILED",
                lambda: service.delete_creator(
                    "creator_log_failure",
                    confirm=True,
                    preview_fingerprint=fingerprint,
                ),
            )
        self.assertEqual(1, len([
            row for row in rows(self.workbook_path, "Creators")
            if row["creator_id"] == "creator_log_failure"
        ]))
        manifest = next((self.runtime_dir / "delete_transactions").glob("*/manifest.json"))
        self.assertEqual("ROLLED_BACK", json.loads(manifest.read_text(encoding="utf-8"))["phase"])

    def test_cleanup_failure_keeps_committed_delete_and_reports_pending(self) -> None:
        self.seed_creator("creator_cleanup")
        task = self.tasks_dir / "task_20260817T000002Z_cccccccc"
        task.mkdir(parents=True)
        (task / "task.json").write_text(json.dumps({
            "id": task.name,
            "creator_library_creator_ids": ["creator_cleanup"],
        }), encoding="utf-8")
        fingerprint = self.preview("creator_cleanup")["preview_fingerprint"]
        with mock.patch.object(
            StagedDeleteTransaction,
            "_remove_path",
            side_effect=OSError("cleanup busy"),
        ):
            result = self.delete("creator_cleanup", fingerprint)
        self.assertTrue(result["deleted"])
        self.assertTrue(result["cleanup_pending"])
        self.assertFalse(task.exists())
        self.assertEqual([], [row for row in rows(self.workbook_path, "Creators") if row["creator_id"] == "creator_cleanup"])
        manifests = list((self.runtime_dir / "delete_transactions").glob("*/manifest.json"))
        self.assertEqual("CLEANUP_PENDING", json.loads(manifests[0].read_text(encoding="utf-8"))["phase"])

    def test_handler_uses_exact_delete_route_and_stable_error_shape(self) -> None:
        class StubService:
            def delete_creator(self, creator_id, **kwargs):
                self.call = (creator_id, kwargs)
                return {"creator_id": creator_id, "deleted": True}

        service = StubService()
        handler = FakeHandler()
        request = {
            "method": "DELETE",
            "path": "/api/creator-library/creator_route",
            "query": {},
            "get_payload": lambda: {"confirm": True, "preview_fingerprint": "f" * 64},
        }
        handled = creator_handler.handle(handler, request, {
            "services": {
                "creator": object(),
                "agency": object(),
                "creator_delete_impact": object(),
                "creator_hard_delete": service,
            },
            "config": {
                "legacy_cooperation_pattern": __import__("re").compile(r"$^"),
                "legacy_cooperation_read_only_message": "legacy",
            },
        })
        self.assertTrue(handled)
        self.assertEqual((200, {"ok": True, "data": {"creator_id": "creator_route", "deleted": True}}), handler.response)
        self.assertEqual(True, service.call[1]["confirm"])


if __name__ == "__main__":
    unittest.main()

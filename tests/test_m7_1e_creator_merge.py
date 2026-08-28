from __future__ import annotations

import json
import shutil
import sys
import unittest
from contextlib import nullcontext
from pathlib import Path
from unittest.mock import Mock, patch
from uuid import uuid4

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))
sys.path.insert(0, str(ROOT / "tests"))

from creator_repository import CreatorRepository, _WORKBOOK_SHEETS
from campaign_creator_repository import CampaignCreatorRepository
from repositories.creator_delete_impact_repository import CreatorDeleteImpactRepository
from repositories.creator_merge_repository import CreatorMergePlanError, CreatorMergeRepository
from services.creator_merge_service import CreatorMergeError, CreatorMergeService
from services.feishu_sync_service import FeishuSyncService
from http_handlers import creator_handler
from test_support.runtime_sandbox import test_artifact_path


class FailingMergeRepository(CreatorMergeRepository):
    def __init__(self, *args, fail_at: str, **kwargs):
        super().__init__(*args, **kwargs)
        self.fail_at = fail_at

    def _checkpoint(self, name: str) -> None:
        if name == self.fail_at:
            raise RuntimeError(f"injected:{name}")


class CreatorMergeTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime = test_artifact_path("m7_1e_merge")
        runtime.mkdir(exist_ok=True)
        self.root = runtime / uuid4().hex
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.tasks_dir = self.root / "tasks"
        self.tasks_dir.mkdir()
        self.mail_path = self.root / "mail_messages.json"
        self.mail_path.write_text(json.dumps({"messages": []}), encoding="utf-8")
        self.legacy_path = self.root / "legacy.json"
        self.legacy_path.write_text(json.dumps({"records": {}}), encoding="utf-8")
        self._create_workbook()
        self.lock_patches = [
            patch("services.creator_merge_service.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()),
            patch("excel_workbook_store.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()),
            patch("app_logging.get_logger", return_value=Mock()),
        ]
        for item in self.lock_patches:
            item.start()

    def tearDown(self) -> None:
        for item in reversed(self.lock_patches):
            item.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def _create_workbook(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in _WORKBOOK_SHEETS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(list(headers))
        self._append(workbook["Creators"], {"creator_id": "primary", "name": "Primary", "country": "Brazil", "archived_at": ""})
        self._append(workbook["Creators"], {"creator_id": "secondary", "name": "Secondary", "country": "Japan", "archived_at": "2025-01-01"})
        self._append(workbook["Creators"], {"creator_id": "unrelated", "name": "Unrelated"})
        self._append(workbook["CreatorAccounts"], {"account_id": "a-primary", "creator_id": "primary", "account_uid": "uid-primary", "platform": "YouTube", "profile_url": "https://youtube.com/@primary"})
        self._append(workbook["CreatorAccounts"], {"account_id": "a-secondary", "creator_id": "secondary", "account_uid": "uid-secondary", "platform": "TikTok", "profile_url": "https://tiktok.com/@secondary"})
        self._append(workbook["CreatorAccounts"], {"account_id": "a-unrelated", "creator_id": "unrelated", "account_uid": "uid-unrelated", "platform": "Instagram"})
        self._append(workbook["Videos"], {"creator_id": "secondary", "video_url": "https://video.example/1", "views": 10})
        self._append(workbook["CreatorSnapshots"], {"snapshot_id": "snapshot-secondary", "creator_id": "secondary", "account_uid": "uid-secondary", "platform": "TikTok"})
        self._append(workbook["VideoSnapshots"], {"video_snapshot_id": "video-snapshot-secondary", "snapshot_id": "snapshot-secondary", "creator_id": "secondary", "video_url": "https://video.example/1"})
        self._append(workbook["Cooperations"], {"cooperation_id": "coop-secondary", "creator_id": "secondary", "campaign": "Legacy"})
        self._append(workbook["Campaigns"], {"campaign_id": "campaign-secondary", "name": "Campaign"})
        self._append(workbook["CampaignCreators"], {"id": "relation-secondary", "campaign_id": "campaign-secondary", "creator_id": "secondary", "account_id": "a-secondary", "stage": "completed"})
        self._append(workbook["FollowUpLogs"], {"follow_up_id": "follow-secondary", "object_type": "creator", "object_id": "secondary"})
        self._append(workbook["_AnalysisData"], {"creator_id": "primary", "account_uid": "uid-primary", "analysis_json": json.dumps({"creator_id": "primary", "analysis_id": "primary"})})
        self._append(workbook["_AnalysisData"], {"creator_id": "secondary", "account_uid": "uid-secondary", "analysis_json": json.dumps({"creator_id": "secondary", "analysis_id": "secondary"})})
        self._append(workbook["_Metadata"], {"schema_version": "test", "last_update_time": "unchanged"})
        workbook.save(self.workbook_path)
        workbook.close()

    @staticmethod
    def _append(sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _repository(self, cls=CreatorMergeRepository, **kwargs):
        return cls(
            self.workbook_path,
            tasks_dir=self.tasks_dir,
            mail_messages_path=self.mail_path,
            legacy_library_file=self.legacy_path,
            **kwargs,
        )

    def _service(self, repository=None):
        repository = repository or self._repository()
        self.invalidations = []
        return CreatorMergeService(
            lambda: repository,
            cache_invalidators=(lambda: self.invalidations.append("creator"),),
        )

    def test_preview_is_read_only_and_success_migrates_all_workbook_references(self) -> None:
        repository = self._repository()
        before = self.workbook_path.read_bytes()
        preview = repository.preview("primary", "secondary")

        self.assertTrue(preview["safe_to_merge"])
        self.assertEqual(before, self.workbook_path.read_bytes())
        self.assertEqual("Primary", preview["primary"]["display_name"])
        self.assertEqual(["TikTok"], preview["secondary"]["platforms"])

        result = self._service(repository).execute(
            "primary", "secondary", confirm=True,
            preview_fingerprint=preview["preview_fingerprint"],
        )
        self.assertTrue(result["merged"])
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            creators = CreatorMergeRepository.rows(workbook["Creators"])
            self.assertEqual(["primary", "unrelated"], [row["creator_id"] for row in creators])
            primary = next(row for row in creators if row["creator_id"] == "primary")
            self.assertEqual("Primary", primary["name"])
            self.assertEqual("Brazil", primary["country"])
            accounts = CreatorMergeRepository.rows(workbook["CreatorAccounts"])
            moved = next(row for row in accounts if row["account_uid"] == "uid-secondary")
            self.assertEqual("primary", moved["creator_id"])
            self.assertEqual("a-secondary", moved["account_id"])
            self.assertEqual(2, sum(row["creator_id"] == "primary" for row in accounts))
            for sheet in ("Videos", "CreatorSnapshots", "VideoSnapshots", "Cooperations", "CampaignCreators"):
                self.assertNotIn("secondary", {str(row.get("creator_id") or "") for row in CreatorMergeRepository.rows(workbook[sheet])})
            followups = CreatorMergeRepository.rows(workbook["FollowUpLogs"])
            self.assertEqual("primary", followups[0]["object_id"])
            self.assertEqual(1, len(CreatorMergeRepository.rows(workbook["_AnalysisData"])))
            self.assertEqual("unchanged", CreatorMergeRepository.rows(workbook["_Metadata"])[0]["last_update_time"])
        finally:
            workbook.close()
        self.assertEqual(["creator"], self.invalidations)

    def test_multiple_secondary_accounts_and_existing_multi_account_primary(self) -> None:
        workbook = load_workbook(self.workbook_path)
        self._append(workbook["CreatorAccounts"], {"account_id": "a-primary-2", "creator_id": "primary", "account_uid": "uid-primary-2", "platform": "Instagram"})
        self._append(workbook["CreatorAccounts"], {"account_id": "a-secondary-2", "creator_id": "secondary", "account_uid": "uid-secondary-2", "platform": "YouTube"})
        self._append(workbook["Campaigns"], {"campaign_id": "campaign-multi", "name": "Multi", "platform": "Instagram"})
        workbook.save(self.workbook_path)
        workbook.close()
        repository = self._repository()
        preview = repository.preview("primary", "secondary")
        self.assertEqual(2, preview["migration_summary"]["accounts"])
        repository.execute("primary", "secondary", preview_fingerprint=preview["preview_fingerprint"])
        detail = CreatorRepository(self.workbook_path).getCreatorDetail("primary")
        self.assertEqual(4, len(detail["accounts"]))
        library = CreatorRepository(self.workbook_path).getCreatorLibrarySnapshot()
        self.assertEqual(1, sum(row["creator_id"] == "primary" for row in library["creators"]))
        self.assertFalse(any(row["creator_id"] == "secondary" for row in library["creators"]))
        campaign_result = CampaignCreatorRepository(self.workbook_path).batch_add_creators(
            "campaign-multi", ["primary"]
        )
        self.assertEqual("added", campaign_result[0]["status"])
        campaign_relation = CampaignCreatorRepository(self.workbook_path).getCampaignCreators(
            "campaign-multi"
        )[0]
        self.assertEqual("a-primary-2", campaign_relation["account_id"])
        inventory = FeishuSyncService(CreatorRepository(self.workbook_path), lambda: None)._local_inventory()
        self.assertEqual(2, inventory["creator_total"])
        self.assertEqual(5, inventory["account_total"])
        self.assertEqual(4, sum(row["creator_id"] == "primary" for row in inventory["accounts"].values()))

        delete_impact = CreatorDeleteImpactRepository(
            self.workbook_path,
            tasks_dir=self.tasks_dir,
            legacy_library_file=self.legacy_path,
        ).scan_creator_delete_impact("primary")
        account_locators = delete_impact["resource_locators"]["creator_accounts"]
        self.assertEqual(4, len(account_locators))
        self.assertEqual(4, len({item["stable_id"] for item in account_locators}))

    def test_same_missing_duplicate_account_campaign_insight_and_orphan_block(self) -> None:
        repository = self._repository()
        self.assertIn("SAME_CREATOR", {item["code"] for item in repository.preview("primary", "primary")["conflicts"]})
        self.assertIn("CREATOR_NOT_FOUND", {item["code"] for item in repository.preview("primary", "missing")["conflicts"]})

        workbook = load_workbook(self.workbook_path)
        self._append(workbook["CreatorAccounts"], {"account_id": "duplicate", "creator_id": "primary", "account_uid": "uid-secondary"})
        self._append(workbook["CampaignCreators"], {"id": "relation-primary", "campaign_id": "campaign-secondary", "creator_id": "primary", "account_id": "a-primary"})
        self._append(workbook["Insights"], {"creator_id": "primary", "average_views": 1})
        self._append(workbook["Insights"], {"creator_id": "secondary", "average_views": 2})
        self._append(workbook["FollowUpLogs"], {"follow_up_id": "unknown", "object_type": "mystery", "object_id": "secondary"})
        workbook.save(self.workbook_path)
        workbook.close()
        codes = {item["code"] for item in repository.preview("primary", "secondary")["conflicts"]}
        self.assertTrue({"DUPLICATE_ACCOUNT_UID", "CAMPAIGN_DUPLICATE", "INSIGHT_CONFLICT", "UNSUPPORTED_REFERENCE"}.issubset(codes))

    def test_external_creator_keyed_mail_task_and_legacy_references_block(self) -> None:
        (self.tasks_dir / "task.json").write_text(json.dumps({"creator_library_creator_ids": ["secondary"]}), encoding="utf-8")
        self.mail_path.write_text(json.dumps({"messages": [{"matched_creator_id": "secondary", "subject": "private"}]}), encoding="utf-8")
        self.legacy_path.write_text(json.dumps({"records": {"x": {"creator_id": "secondary"}}}), encoding="utf-8")
        preview = self._repository().preview("primary", "secondary")
        sources = {item.get("source") for item in preview["conflicts"]}
        self.assertTrue({"Tasks", "Mail", "Legacy"}.issubset(sources))
        self.assertNotIn("private", json.dumps(preview, ensure_ascii=False))

    def test_confirmation_and_stale_preview_fail_closed(self) -> None:
        service = self._service()
        preview = service.preview("primary", "secondary")
        with self.assertRaisesRegex(CreatorMergeError, "MERGE_CONFIRMATION_REQUIRED"):
            service.execute("primary", "secondary", confirm=False, preview_fingerprint=preview["preview_fingerprint"])
        with self.assertRaisesRegex(CreatorMergeError, "STALE_PREVIEW"):
            service.execute("primary", "secondary", confirm=True, preview_fingerprint="stale")
        broken_repository = Mock()
        broken_repository.execute.side_effect = RuntimeError("save failed")
        with self.assertRaisesRegex(CreatorMergeError, "MERGE_FAILED"):
            self._service(broken_repository).execute(
                "primary", "secondary", confirm=True, preview_fingerprint="current"
            )
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            self.assertTrue(CreatorMergeRepository.row_by_key(workbook["Creators"], "creator_id", "secondary"))
        finally:
            workbook.close()

    def test_failure_at_each_in_memory_stage_persists_nothing(self) -> None:
        original = self.workbook_path.read_bytes()
        for checkpoint in ("after_accounts", "after_snapshots", "after_campaigns", "before_save"):
            repository = self._repository(FailingMergeRepository, fail_at=checkpoint)
            preview = repository.preview("primary", "secondary")
            with self.assertRaisesRegex(RuntimeError, f"injected:{checkpoint}"):
                repository.execute("primary", "secondary", preview_fingerprint=preview["preview_fingerprint"])
            self.assertEqual(original, self.workbook_path.read_bytes())

    def test_handler_preview_and_execute_contract(self) -> None:
        service = self._service()

        class Handler:
            def __init__(self):
                self.responses = []

            def _json(self, payload, status=200):
                self.responses.append((status, payload))

        context = {
            "services": {
                "creator": object(),
                "agency": object(),
                "creator_delete_impact": object(),
                "creator_hard_delete": object(),
                "creator_merge": service,
            }
        }
        preview_handler = Handler()
        preview_request = {
            "method": "POST", "path": "/api/creator-library/merge/preview", "query": {},
            "get_payload": lambda: {"primary_creator_id": "primary", "secondary_creator_id": "secondary"},
        }
        self.assertTrue(creator_handler.handle(preview_handler, preview_request, context))
        self.assertTrue(preview_handler.responses[0][1]["safe_to_merge"])
        fingerprint = preview_handler.responses[0][1]["preview_fingerprint"]

        execute_handler = Handler()
        execute_request = {
            "method": "POST", "path": "/api/creator-library/merge/execute", "query": {},
            "get_payload": lambda: {
                "primary_creator_id": "primary", "secondary_creator_id": "secondary",
                "confirm": True, "preview_fingerprint": fingerprint,
            },
        }
        self.assertTrue(creator_handler.handle(execute_handler, execute_request, context))
        self.assertEqual(200, execute_handler.responses[0][0])
        self.assertTrue(execute_handler.responses[0][1]["merged"])


if __name__ == "__main__":
    unittest.main()

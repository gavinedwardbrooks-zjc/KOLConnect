from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
sys.path.insert(0, str(ROOT))

import creator_repository
from scripts.migrate_creator_crm_columns import migrate_workbook


class CreatorCrmColumnsTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        self.repository = creator_repository.CreatorRepository(self.workbook_path)
        with mock.patch.object(creator_repository, "log_event"):
            self.repository.getCreators()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    @staticmethod
    def _append(sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _seed_creator(
        self,
        *,
        creator_id: str = "creator_one",
        bio: str = "",
        archived_at: str = "",
        legacy_bio: str | None = None,
        legacy_archived_at: str | None = None,
        with_metadata: bool = True,
    ) -> None:
        workbook = load_workbook(self.workbook_path)
        self._append(workbook["Creators"], {
            "creator_id": creator_id,
            "name": "Creator One",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@creator-one",
            "created_at": "2026-08-01T00:00:00Z",
            "updated_at": "2026-08-01T00:00:00Z",
            "bio": bio,
            "archived_at": archived_at,
        })
        self._append(workbook["CreatorAccounts"], {
            "account_id": "account_one",
            "creator_id": creator_id,
            "account_uid": "tiktok|creator-one",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@creator-one",
        })
        if with_metadata:
            crm = {}
            if legacy_bio is not None:
                crm["bio"] = legacy_bio
            if legacy_archived_at is not None:
                crm["archived_at"] = legacy_archived_at
            self._append(workbook["_AnalysisData"], {
                "creator_id": creator_id,
                "account_uid": "tiktok|creator-one",
                "analysis_json": json.dumps({
                    "analysis_id": creator_id,
                    "creator": {"creator_name": "Creator One", "bio": "analysis bio"},
                    "_crm": crm,
                }),
            })
        workbook.save(self.workbook_path)
        workbook.close()

    def _creator_row(self) -> dict:
        workbook = load_workbook(self.workbook_path)
        try:
            return self.repository._creator_row(workbook["Creators"], "creator_one")
        finally:
            workbook.close()

    def _legacy_crm(self) -> dict:
        workbook = load_workbook(self.workbook_path)
        try:
            metadata = self.repository._metadata_row(workbook["_AnalysisData"], "creator_one")
            return self.repository._decode_analysis(metadata.get("analysis_json")).get("_crm", {})
        finally:
            workbook.close()

    def test_new_workbook_has_real_columns_at_end(self) -> None:
        workbook = load_workbook(self.workbook_path)
        headers = [cell.value for cell in workbook["Creators"][1]]
        workbook.close()
        self.assertEqual(["bio", "archived_at"], headers[-2:])

    def test_old_workbook_columns_are_appended_without_reordering(self) -> None:
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Creators"]
        original = [cell.value for cell in sheet[1]][:-2]
        sheet.delete_cols(sheet.max_column - 1, 2)
        workbook.save(self.workbook_path)
        workbook.close()
        with mock.patch.object(creator_repository, "log_event"):
            self.repository.getCreators()
        workbook = load_workbook(self.workbook_path)
        headers = [cell.value for cell in workbook["Creators"][1]]
        workbook.close()
        self.assertEqual(original, headers[:-2])
        self.assertEqual(["bio", "archived_at"], headers[-2:])

    def test_real_bio_is_authoritative(self) -> None:
        self._seed_creator(bio="real bio", legacy_bio="legacy bio")
        self.assertEqual("real bio", self.repository.getCreators()[0]["bio"])

    def test_legacy_bio_is_fallback(self) -> None:
        self._seed_creator(legacy_bio="legacy bio")
        self.assertEqual("legacy bio", self.repository.getCreators()[0]["bio"])

    def test_legacy_archive_remains_hidden_by_default(self) -> None:
        self._seed_creator(legacy_archived_at="2026-08-01T01:00:00Z")
        self.assertEqual([], self.repository.getCreators())
        self.assertEqual(1, len(self.repository.getCreators(include_archived=True)))

    def test_missing_metadata_is_safe(self) -> None:
        self._seed_creator(with_metadata=False)
        record = self.repository.getCreators()[0]
        self.assertEqual("", record["bio"])
        self.assertFalse(record["archived_at"])

    def test_patch_bio_writes_real_column_and_legacy_compatibility(self) -> None:
        self._seed_creator(legacy_bio="old")
        self.repository.updateCreator("creator_one", {"bio": "new bio"})
        self.assertEqual("new bio", self._creator_row()["bio"])
        self.assertEqual("new bio", self._legacy_crm()["bio"])

    def test_patch_archive_writes_real_column_and_legacy_compatibility(self) -> None:
        self._seed_creator()
        value = "2026-08-02T00:00:00Z"
        self.repository.updateCreator("creator_one", {"archived_at": value})
        self.assertEqual(value, self._creator_row()["archived_at"])
        self.assertEqual(value, self._legacy_crm()["archived_at"])

    def test_restore_clears_real_and_legacy_values(self) -> None:
        self._seed_creator(
            archived_at="2026-08-02T00:00:00Z",
            legacy_archived_at="2026-08-02T00:00:00Z",
        )
        self.repository.updateCreator("creator_one", {"archived_at": None})
        self.assertFalse(self._creator_row()["archived_at"])
        self.assertIsNone(self._legacy_crm()["archived_at"])
        self.assertEqual(1, len(self.repository.getCreators()))

    def test_set_creator_archived_is_idempotent(self) -> None:
        self._seed_creator()
        first = self.repository.set_creator_archived("creator_one", True)
        second = self.repository.set_creator_archived("creator_one", True)
        self.assertEqual(first["archived_at"], second["archived_at"])
        self.repository.set_creator_archived("creator_one", False)
        self.repository.set_creator_archived("creator_one", False)
        self.assertFalse(self._creator_row()["archived_at"])

    def test_task_import_preserves_real_bio_and_archive(self) -> None:
        archived_at = "2026-08-02T00:00:00Z"
        self._seed_creator(bio="curated", archived_at=archived_at)
        self.repository.importTaskResults(
            "task_20260807T100000Z_aaaaaaaa",
            [{
                "account_uid": "tiktok|creator-one",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@creator-one",
                "creator_name": "Creator One",
                "bio": "incoming",
                "scrape_status": "success",
            }],
            source="test",
        )
        row = self._creator_row()
        self.assertEqual("curated", row["bio"])
        self.assertEqual(archived_at, row["archived_at"])

    def test_task_import_uses_record_bio_before_legacy_fallback(self) -> None:
        self._seed_creator(legacy_bio="legacy")
        self.repository.importTaskResults(
            "task_20260807T100001Z_bbbbbbbb",
            [{
                "account_uid": "tiktok|creator-one",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@creator-one",
                "creator_name": "Creator One",
                "bio": "fresh record",
                "scrape_status": "success",
            }],
            source="test",
        )
        self.assertEqual("fresh record", self._creator_row()["bio"])

    def test_task_import_empty_bio_does_not_clear_real_value(self) -> None:
        self._seed_creator(bio="curated")
        self.repository.importTaskResults(
            "task_20260807T100002Z_cccccccc",
            [{
                "account_uid": "tiktok|creator-one",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@creator-one",
                "creator_name": "Creator One",
                "bio": "",
                "scrape_status": "partial_success",
            }],
            source="test",
        )
        self.assertEqual("curated", self._creator_row()["bio"])

    def test_migration_dry_run_does_not_modify_workbook(self) -> None:
        self._seed_creator(legacy_bio="legacy", legacy_archived_at="2026-08-02T00:00:00Z")
        before = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        report = migrate_workbook(self.workbook_path)
        after = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        self.assertEqual(before, after)
        self.assertEqual("dry-run", report["mode"])
        self.assertEqual(1, report["bio_migrated"])
        self.assertEqual(1, report["archived_at_migrated"])

    def test_migration_apply_creates_backup_and_copies_values(self) -> None:
        self._seed_creator(legacy_bio="legacy", legacy_archived_at="2026-08-02T00:00:00Z")
        report = migrate_workbook(self.workbook_path, dry_run=False)
        self.assertTrue(Path(report["backup_path"]).is_file())
        row = self._creator_row()
        self.assertEqual("legacy", row["bio"])
        self.assertEqual("2026-08-02T00:00:00Z", row["archived_at"])

    def test_migration_apply_appends_missing_columns(self) -> None:
        self._seed_creator(legacy_bio="legacy", legacy_archived_at="2026-08-02T00:00:00Z")
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Creators"]
        sheet.delete_cols(sheet.max_column - 1, 2)
        workbook.save(self.workbook_path)
        workbook.close()
        migrate_workbook(self.workbook_path, dry_run=False)
        workbook = load_workbook(self.workbook_path)
        headers = [cell.value for cell in workbook["Creators"][1]]
        row = self.repository._creator_row(workbook["Creators"], "creator_one")
        workbook.close()
        self.assertEqual(["bio", "archived_at"], headers[-2:])
        self.assertEqual("legacy", row["bio"])

    def test_migration_never_overwrites_conflicting_real_values(self) -> None:
        self._seed_creator(
            bio="real",
            archived_at="2026-08-03T00:00:00Z",
            legacy_bio="legacy",
            legacy_archived_at="2026-08-02T00:00:00Z",
        )
        report = migrate_workbook(self.workbook_path, dry_run=False)
        row = self._creator_row()
        self.assertEqual("real", row["bio"])
        self.assertEqual("2026-08-03T00:00:00Z", row["archived_at"])
        self.assertEqual(2, report["conflicts"])

    def test_migration_is_idempotent(self) -> None:
        self._seed_creator(legacy_bio="legacy", legacy_archived_at="2026-08-02T00:00:00Z")
        migrate_workbook(self.workbook_path, dry_run=False)
        second = migrate_workbook(self.workbook_path, dry_run=False)
        self.assertEqual(0, second["bio_migrated"])
        self.assertEqual(0, second["archived_at_migrated"])
        self.assertEqual(2, second["skipped_existing"])


if __name__ == "__main__":
    unittest.main()

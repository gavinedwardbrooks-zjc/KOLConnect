from __future__ import annotations

import hashlib
import json
import shutil
import sys
import unittest
from contextlib import nullcontext
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import Mock, patch
from uuid import uuid4

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))
sys.path.insert(0, str(ROOT / "tests"))

from creator_repository import CreatorRepository, _WORKBOOK_SHEETS
from dashboard_repository import DashboardRepository
from dashboard_service import DashboardService
from http_handlers import clean_reset_handler
from services.clean_reset_service import BUSINESS_SHEETS, CleanResetError, CleanResetService
from test_support.runtime_sandbox import test_artifact_path


class CleanResetServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime_root = test_artifact_path("m7_1_clean_reset")
        runtime_root.mkdir(parents=True, exist_ok=True)
        self.root = runtime_root / uuid4().hex
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.settings_path = self.root / "settings.json"
        self.data_protection_path = self.root / "data_protection.json"
        self.mail_messages_path = self.root / "mail_messages.json"
        self.tasks_dir = self.root / "tasks"
        self.tasks_dir.mkdir()
        self._create_workbook()
        self.settings_path.write_text(
            json.dumps({
                "profiles": {"selected": "profile-a"},
                "mail": {"accounts": [{"id": "mail-a"}]},
                "feishu": {"app_id": "configured", "app_secret": "secret-placeholder"},
            }),
            encoding="utf-8",
        )
        self.data_protection_path.write_text(
            json.dumps({"account-a": {"allowed": True}}), encoding="utf-8"
        )
        self.mail_messages_path.write_text(
            json.dumps({
                "version": 1,
                "accounts": {"mail-a": {"cursor": "1"}},
                "messages": [{"id": "message-a"}],
            }),
            encoding="utf-8",
        )
        (self.tasks_dir / "task.json").write_text("{}", encoding="utf-8")
        self.invalidations = []
        self.service = CleanResetService(
            self.workbook_path,
            settings_path=self.settings_path,
            data_protection_path=self.data_protection_path,
            mail_messages_path=self.mail_messages_path,
            tasks_dir=self.tasks_dir,
            cache_invalidators=(lambda: self.invalidations.append("creator"),),
            now_provider=lambda: datetime(2026, 8, 24, 9, 30, tzinfo=timezone.utc),
        )
        self.lock_patches = [
            patch("services.clean_reset_service.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()),
            patch("excel_workbook_store.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()),
            patch("runtime_paths.shared_storage_lock", side_effect=lambda *a, **k: nullcontext()),
            patch("app_logging.get_logger", return_value=Mock()),
        ]
        for lock_patch in self.lock_patches:
            lock_patch.start()

    def tearDown(self) -> None:
        for lock_patch in reversed(self.lock_patches):
            lock_patch.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def _create_workbook(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in _WORKBOOK_SHEETS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(list(headers))
            if name == "_Metadata":
                sheet.append(["schema_version", "7"])
            else:
                sheet.append([f"sample-{name}", *([""] * (len(headers) - 1))])
        workbook.save(self.workbook_path)
        workbook.close()

    @staticmethod
    def _digest(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def test_preview_is_read_only_and_execute_requires_explicit_confirmation(self) -> None:
        before = {
            path: self._digest(path)
            for path in (self.workbook_path, self.settings_path, self.data_protection_path, self.mail_messages_path)
        }

        preview = self.service.preview()

        self.assertEqual("success", preview["status"])
        self.assertTrue(all(preview["clear_sheets"][name] == 1 for name in BUSINESS_SHEETS))
        self.assertEqual(1, preview["preserve_sheets"]["_Metadata"])
        self.assertEqual(1, preview["external_business_data"]["task_files"])
        self.assertFalse((self.root / "backups").exists())
        self.assertEqual(before, {path: self._digest(path) for path in before})
        with self.assertRaisesRegex(ValueError, "CLEAN_RESET_CONFIRMATION_REQUIRED"):
            self.service.execute(confirm=False)
        self.assertFalse((self.root / "backups").exists())

    def test_execute_backs_up_clears_and_preserves_schema_configuration(self) -> None:
        settings_before = self._digest(self.settings_path)
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            headers_before = {name: [cell.value for cell in workbook[name][1]] for name in workbook.sheetnames}
            metadata_before = list(workbook["_Metadata"].values)
        finally:
            workbook.close()

        first = self.service.execute(confirm=True)

        backup = Path(first["backup"]["path"])
        self.assertTrue(backup.is_file())
        self.assertEqual(self.workbook_path.parent / "backups", backup.parent)
        self.assertEqual(settings_before, self._digest(self.settings_path))
        self.assertEqual({}, json.loads(self.data_protection_path.read_text(encoding="utf-8")))
        mail = json.loads(self.mail_messages_path.read_text(encoding="utf-8"))
        self.assertEqual([], mail["messages"])
        self.assertEqual({}, mail["accounts"])
        self.assertEqual([], list(self.tasks_dir.iterdir()))
        self.assertEqual(["creator"], self.invalidations)
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            self.assertEqual(headers_before, {name: [cell.value for cell in workbook[name][1]] for name in workbook.sheetnames})
            self.assertEqual(metadata_before, list(workbook["_Metadata"].values))
            for name in BUSINESS_SHEETS:
                self.assertEqual(1, workbook[name].max_row, name)
        finally:
            workbook.close()

        second = self.service.execute(confirm=True)
        self.assertNotEqual(first["backup"]["path"], second["backup"]["path"])
        self.assertTrue(Path(second["backup"]["path"]).is_file())
        self.assertTrue(all(value == 0 for value in second["after"].values()))

        repository = CreatorRepository(self.workbook_path)
        self.assertEqual([], repository.getCreators())
        dashboard = DashboardService(DashboardRepository(repository))
        self.assertEqual(0, dashboard.getOverview()["total_creators"])
        self.assertEqual([], dashboard.getPlatformDistribution())
        growth = dashboard.getCreatorGrowthTrend()
        self.assertEqual(30, len(growth))
        self.assertTrue(all(item["count"] == 0 for item in growth))

    def test_unknown_sheet_fails_closed_without_backup(self) -> None:
        workbook = load_workbook(self.workbook_path)
        workbook.create_sheet("UnexpectedBusinessData").append(["secret"])
        workbook.save(self.workbook_path)
        workbook.close()

        preview = self.service.preview()
        self.assertEqual("blocked", preview["status"])
        self.assertIn("UNKNOWN_SHEET:UnexpectedBusinessData", preview["review_items"])
        with self.assertRaisesRegex(CleanResetError, "CLEAN_RESET_SCHEMA_REVIEW_REQUIRED"):
            self.service.execute(confirm=True)
        self.assertFalse((self.root / "backups").exists())

    def test_json_failure_rolls_back_workbook_and_external_data(self) -> None:
        workbook_before = self._digest(self.workbook_path)
        protection_before = self._digest(self.data_protection_path)
        mail_before = self._digest(self.mail_messages_path)
        from services import clean_reset_service as module

        real_atomic_write = module.atomic_write_json
        calls = 0

        def fail_first_write(path, payload):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise PermissionError("simulated write failure")
            return real_atomic_write(path, payload)

        with patch("services.clean_reset_service.atomic_write_json", side_effect=fail_first_write):
            with self.assertRaises(PermissionError):
                self.service.execute(confirm=True)

        self.assertEqual(workbook_before, self._digest(self.workbook_path))
        self.assertEqual(protection_before, self._digest(self.data_protection_path))
        self.assertEqual(mail_before, self._digest(self.mail_messages_path))
        self.assertTrue((self.tasks_dir / "task.json").is_file())
        self.assertEqual([], self.invalidations)


class CleanResetHandlerTests(unittest.TestCase):
    class Handler:
        def __init__(self) -> None:
            self.responses = []

        def _json(self, payload, status=200):
            self.responses.append((status, payload))

        def _error(self, message):
            raise AssertionError(message)

    class Service:
        def __init__(self) -> None:
            self.confirmations = []

        def preview(self):
            return {"status": "success", "review_items": [], "summary": {"creators": 1}}

        def execute(self, *, confirm):
            self.confirmations.append(confirm)
            if confirm is not True:
                raise ValueError("CLEAN_RESET_CONFIRMATION_REQUIRED")
            return {"status": "success", "backup": {"filename": "backup.xlsx"}}

    def test_preview_route_is_read_only_and_execute_forwards_literal_confirmation(self) -> None:
        service = self.Service()
        handler = self.Handler()
        context = {"services": {"clean_reset": service}, "logging": {"error": lambda *args: None}}
        preview_request = {
            "method": "POST",
            "path": "/api/settings/clean-reset/preview",
            "get_payload": lambda: (_ for _ in ()).throw(AssertionError("preview must not parse mutation payload")),
        }
        self.assertTrue(clean_reset_handler.handle(handler, preview_request, context))
        self.assertEqual([], service.confirmations)
        self.assertTrue(handler.responses[-1][1]["ok"])

        execute_request = {
            "method": "POST",
            "path": "/api/settings/clean-reset/execute",
            "get_payload": lambda: {"confirm": True},
        }
        self.assertTrue(clean_reset_handler.handle(handler, execute_request, context))
        self.assertEqual([True], service.confirmations)

    def test_execute_without_confirmation_is_blocked(self) -> None:
        service = self.Service()
        handler = self.Handler()
        context = {"services": {"clean_reset": service}, "logging": {"error": lambda *args: None}}
        request = {
            "method": "POST",
            "path": "/api/settings/clean-reset/execute",
            "get_payload": lambda: {},
        }
        self.assertTrue(clean_reset_handler.handle(handler, request, context))
        self.assertEqual((400, "CLEAN_RESET_CONFIRMATION_REQUIRED"), (
            handler.responses[-1][0], handler.responses[-1][1]["error"]
        ))


if __name__ == "__main__":
    unittest.main()

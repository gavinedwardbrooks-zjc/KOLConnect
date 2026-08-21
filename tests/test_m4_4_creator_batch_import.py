from __future__ import annotations

import io
import json
import shutil
import sys
import threading
import unittest
import urllib.error
import urllib.request
import uuid
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

from creator_batch_import import (  # noqa: E402
    TEMPLATE_HEADERS,
    CreatorBatchImportError,
    build_creator_import_template,
)
from creator_repository import CreatorRepository  # noqa: E402
from services.creator_service import CreatorService  # noqa: E402
import app_logging  # noqa: E402

with (
    mock.patch.object(app_logging, "log_event"),
    mock.patch.object(app_logging, "log_error"),
):
    import server  # noqa: E402


class FakeAgencyPort:
    def __init__(self, agency_ids: set[str] | None = None) -> None:
        self.agency_ids = agency_ids or set()

    def get_agency(self, agency_id: str) -> dict:
        if agency_id not in self.agency_ids:
            raise ValueError("未找到 Agency。")
        return {"agency_id": agency_id, "name": agency_id}


def workbook_bytes(
    rows: list[list[object]] | None = None,
    headers: list[str] | tuple[str, ...] = TEMPLATE_HEADERS,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"
    sheet.append(list(headers))
    for row in rows or []:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class CreatorBatchImportContractTests(unittest.TestCase):
    def setUp(self) -> None:
        self.log_patcher = mock.patch("creator_repository.log_event")
        self.log_patcher.start()
        self.root = ROOT / f".m4_4_test_{uuid.uuid4().hex}"
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.repository = CreatorRepository(self.workbook_path)
        self.service = CreatorService(
            lambda: self.repository,
            lambda: None,
            agency_port_provider=lambda: FakeAgencyPort({"agency_known"}),
        )

    def tearDown(self) -> None:
        shutil.rmtree(self.root, ignore_errors=True)
        self.log_patcher.stop()

    def import_rows(self, rows: list[list[object]]) -> dict:
        return self.service.import_creator_batch(workbook_bytes(rows))

    def creators(self) -> list[dict]:
        return self.repository.getCreators(include_archived=True)

    def accounts(self) -> list[dict]:
        return self.repository.getCreatorAccounts()

    def error_response(self, rows: list[list[object]]) -> dict:
        with self.assertRaises(CreatorBatchImportError) as raised:
            self.import_rows(rows)
        return raised.exception.to_response()

    def test_template_headers_are_exact_and_has_no_example_rows(self) -> None:
        data = build_creator_import_template()
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        self.assertEqual(list(TEMPLATE_HEADERS), [cell.value for cell in sheet[1]])
        self.assertEqual(1, sheet.max_row)
        workbook.close()

    def test_valid_rows_create_once_with_optional_fields_and_blank_rows(self) -> None:
        result = self.import_rows([
            ["TikTok", "https://www.tiktok.com/@one", "One", "BR", "pt", "Games", "one@example.com", "+55", "agency_known", "Bio"],
            ["", "", "", "", "", "", "", "", "", ""],
            ["YouTube", "https://www.youtube.com/@two"],
        ])
        self.assertEqual(
            {"total_rows": 2, "created": 2, "skipped_existing": 0},
            result,
        )
        self.assertEqual(2, len(self.creators()))
        self.assertEqual(2, len(self.accounts()))
        one = next(item for item in self.creators() if item["creator_name"] == "One")
        self.assertEqual("agency_known", one["agency_id"])
        self.assertEqual("Bio", one["bio"])

    def test_file_level_errors_are_stable(self) -> None:
        cases = (
            (b"not an xlsx", "INVALID_FILE"),
            (workbook_bytes([], ["platform", "name"]), "MISSING_REQUIRED_COLUMN"),
            (workbook_bytes([]), "EMPTY_IMPORT"),
        )
        for payload, code in cases:
            with self.subTest(code=code), self.assertRaises(CreatorBatchImportError) as raised:
                self.service.import_creator_batch(payload)
            self.assertEqual(code, raised.exception.code)

    def test_required_and_normalized_field_validation_reports_excel_rows(self) -> None:
        response = self.error_response([
            ["", "https://www.tiktok.com/@missing-platform"],
            ["TikTok", ""],
            ["Other", "https://www.tiktok.com/@invalid-platform"],
            ["TikTok", "https://example.com/not-tiktok"],
        ])
        self.assertEqual("BATCH_IMPORT_VALIDATION_FAILED", response["error"])
        self.assertEqual(4, response["summary"]["invalid_rows"])
        self.assertEqual(
            {(2, "MISSING_REQUIRED_FIELD", "platform"),
             (3, "MISSING_REQUIRED_FIELD", "profile_url"),
             (4, "INVALID_PLATFORM", "platform"),
             (5, "INVALID_PROFILE_URL", "profile_url")},
            {(item["row"], item["code"], item.get("field")) for item in response["rows"]},
        )
        self.assertEqual([], self.creators())

    def test_duplicate_in_file_reports_every_conflicting_row_and_writes_nothing(self) -> None:
        response = self.error_response([
            ["TikTok", "https://www.tiktok.com/@same"],
            ["tiktok", "https://www.tiktok.com/@same?lang=en"],
        ])
        duplicates = [row for row in response["rows"] if row["code"] == "DUPLICATE_IN_FILE"]
        self.assertEqual([2, 3], [row["row"] for row in duplicates])
        self.assertEqual(2, response["summary"]["invalid_rows"])
        self.assertEqual([], self.creators())

    def test_existing_creator_is_non_fatal_skip_in_mixed_batch(self) -> None:
        self.import_rows([["TikTok", "https://www.tiktok.com/@existing", "Existing"]])
        result = self.import_rows([
            ["TikTok", "https://www.tiktok.com/@existing", "Must Not Overwrite"],
            ["Instagram", "https://www.instagram.com/new_creator/", "New"],
        ])
        self.assertEqual(
            {"total_rows": 2, "created": 1, "skipped_existing": 1},
            result,
        )
        records = self.creators()
        self.assertEqual(2, len(records))
        existing = next(item for item in records if item["platform"] == "TikTok")
        self.assertEqual("Existing", existing["creator_name"])

    def test_unknown_agency_is_row_error_and_any_invalid_row_means_zero_write(self) -> None:
        response = self.error_response([
            ["TikTok", "https://www.tiktok.com/@valid"],
            ["YouTube", "https://www.youtube.com/@unknown", "", "", "", "", "", "", "agency_missing"],
        ])
        self.assertEqual(2, response["summary"]["total_rows"])
        self.assertEqual(1, response["summary"]["valid_new_rows"])
        self.assertEqual(1, response["summary"]["invalid_rows"])
        self.assertEqual("UNKNOWN_AGENCY", response["rows"][0]["code"])
        self.assertEqual([], self.creators())

    def test_multiple_invalid_rows_are_reported_without_sensitive_values(self) -> None:
        secret_values = [
            "private@example.com",
            "+55 1234567",
            "private biography",
            "https://example.com/private-profile",
        ]
        response = self.error_response([
            ["", secret_values[3], "", "", "", "", secret_values[0], secret_values[1], "", secret_values[2]],
            ["YouTube", "", "", "", "", "", "second@example.com"],
        ])
        serialized = json.dumps(response, ensure_ascii=False)
        for secret in secret_values:
            self.assertNotIn(secret, serialized)
        self.assertNotIn(str(self.workbook_path), serialized)
        self.assertNotIn("Traceback", serialized)
        self.assertEqual(2, response["summary"]["invalid_rows"])

    def test_batch_write_opens_and_saves_once(self) -> None:
        with (
            mock.patch.object(self.repository, "_load_workbook", wraps=self.repository._load_workbook) as load,
            mock.patch.object(self.repository, "_save_workbook", wraps=self.repository._save_workbook) as save,
        ):
            self.import_rows([
                ["TikTok", "https://www.tiktok.com/@one"],
                ["YouTube", "https://www.youtube.com/@two"],
            ])
        self.assertEqual(1, load.call_count, "one batch write uses one workbook read")
        self.assertEqual(1, save.call_count)

    def test_save_failure_preserves_original_workbook(self) -> None:
        self.import_rows([["TikTok", "https://www.tiktok.com/@existing"]])
        original = self.workbook_path.read_bytes()
        with mock.patch.object(
            self.repository,
            "_save_workbook",
            side_effect=RuntimeError("simulated save failure at C:/private/path"),
        ):
            with self.assertRaises(CreatorBatchImportError) as raised:
                self.import_rows([["YouTube", "https://www.youtube.com/@new"]])
        self.assertEqual("BATCH_IMPORT_WRITE_FAILED", raised.exception.code)
        self.assertEqual(original, self.workbook_path.read_bytes())
        self.assertNotIn("private", json.dumps(raised.exception.to_response()))


class CreatorBatchImportHttpTests(unittest.TestCase):
    def setUp(self) -> None:
        self.root = ROOT / f".m4_4_http_{uuid.uuid4().hex}"
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.patchers = [
            mock.patch.object(
                server, "_creator_library_workbook_path", return_value=self.workbook_path
            ),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
            mock.patch("creator_repository.log_event"),
        ]
        for patcher in self.patchers:
            patcher.start()
        self.httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.httpd.server_port}"

    def tearDown(self) -> None:
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=5)
        for patcher in reversed(self.patchers):
            patcher.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def request(self, method: str, path: str, payload: bytes | None = None):
        request = urllib.request.Request(
            self.base_url + path,
            data=payload,
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            } if payload is not None else {},
            method=method,
        )
        try:
            return urllib.request.urlopen(request, timeout=10)
        except urllib.error.HTTPError as exc:
            return exc

    def test_template_download_http_contract(self) -> None:
        with self.request("GET", "/api/creator-library/import-template") as response:
            payload = response.read()
            self.assertEqual(200, response.status)
            self.assertEqual(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                response.headers.get_content_type(),
            )
            self.assertIn(
                "KOLConnect_Creator_Import_Template.xlsx",
                response.headers.get("Content-Disposition", ""),
            )
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        self.assertEqual(list(TEMPLATE_HEADERS), [cell.value for cell in workbook.active[1]])
        workbook.close()

    def test_raw_xlsx_success_and_validation_failure_http_contracts(self) -> None:
        with self.request(
            "POST",
            "/api/creator-library/import",
            workbook_bytes([["TikTok", "https://www.tiktok.com/@http-one"]]),
        ) as response:
            body = json.loads(response.read().decode("utf-8"))
            self.assertEqual(200, response.status)
            self.assertEqual(
                {"ok": True, "data": {"total_rows": 1, "created": 1, "skipped_existing": 0}},
                body,
            )

        with self.request(
            "POST",
            "/api/creator-library/import",
            workbook_bytes([["YouTube", ""]]),
        ) as response:
            body = json.loads(response.read().decode("utf-8"))
            self.assertEqual(400, response.status)
            self.assertEqual(False, body["ok"])
            self.assertEqual("BATCH_IMPORT_VALIDATION_FAILED", body["error"])
            self.assertEqual(2, body["rows"][0]["row"])

    def test_upload_rejects_wrong_content_type_without_consuming_json(self) -> None:
        request = urllib.request.Request(
            self.base_url + "/api/creator-library/import",
            data=workbook_bytes([["TikTok", "https://www.tiktok.com/@wrong-type"]]),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            urllib.request.urlopen(request, timeout=10)
        except urllib.error.HTTPError as response:
            with response:
                body = json.loads(response.read().decode("utf-8"))
                self.assertEqual(400, response.status)
                self.assertEqual("INVALID_FILE", body["error"])
        else:
            self.fail("wrong content type must be rejected")

    def test_json_and_raw_body_readers_are_mutually_exclusive(self) -> None:
        handler = object.__new__(server.Handler)
        handler.command = "POST"
        handler.headers = {"Content-Length": "2"}
        handler.rfile = io.BytesIO(b"{}")
        request = handler._request_context(SimpleNamespace(path="/api/test"), {})
        self.assertEqual(b"{}", request["get_raw_body"]())
        with self.assertRaises(RuntimeError):
            request["get_payload"]()

        handler = object.__new__(server.Handler)
        handler.command = "POST"
        handler.headers = {"Content-Length": "2"}
        handler.rfile = io.BytesIO(b"{}")
        request = handler._request_context(SimpleNamespace(path="/api/test"), {})
        self.assertEqual({}, request["get_payload"]())
        with self.assertRaises(RuntimeError):
            request["get_raw_body"]()


if __name__ == "__main__":
    unittest.main()

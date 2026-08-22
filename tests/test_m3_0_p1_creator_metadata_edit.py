from __future__ import annotations

import json
import shutil
import sys
import unittest
import uuid
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository


class CreatorMetadataEditTests(unittest.TestCase):
    def setUp(self) -> None:
        self.root = ROOT / f".m3_0_p1_metadata_{uuid.uuid4().hex}"
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.lock_patcher = mock.patch(
            "local_storage_lock.get_shared_storage_lock_path",
            return_value=self.root / "locks" / "shared_storage.lock",
        )
        self.lock_patcher.start()
        self.log_patcher = mock.patch.object(creator_repository, "log_event")
        self.log_patcher.start()
        self.repository = creator_repository.CreatorRepository(self.workbook_path)
        self.repository.getCreators()
        self._seed_creator()

    def tearDown(self) -> None:
        self.log_patcher.stop()
        self.lock_patcher.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def _seed_creator(self) -> None:
        workbook = load_workbook(self.workbook_path)
        try:
            creators = workbook["Creators"]
            headers = [str(cell.value or "") for cell in creators[1]]
            creators.append([
                {
                    "creator_id": "creator_one",
                    "name": "Creator One",
                    "platform": "TikTok",
                    "profile_url": "https://www.tiktok.com/@creator-one",
                    "followers": "10K",
                    "content_category": "Gaming",
                    "created_at": "2026-08-01T00:00:00Z",
                    "updated_at": "2026-08-01T00:00:00Z",
                }.get(header, "")
                for header in headers
            ])
            metadata = workbook["_AnalysisData"]
            metadata_headers = [str(cell.value or "") for cell in metadata[1]]
            metadata.append([
                {
                    "creator_id": "creator_one",
                    "account_uid": "tiktok|creator-one",
                    "analysis_json": json.dumps({
                        "analysis_id": "creator_one",
                        "creator": {"creator_name": "Creator One"},
                        "_crm": {},
                    }),
                }.get(header, "")
                for header in metadata_headers
            ])
            workbook.save(self.workbook_path)
        finally:
            workbook.close()

    def _creator_row(self) -> dict[str, object]:
        workbook = load_workbook(self.workbook_path)
        try:
            return self.repository._creator_row(workbook["Creators"], "creator_one")
        finally:
            workbook.close()

    def test_updates_country(self) -> None:
        result = self.repository.updateCreator("creator_one", {"country": "Brazil"})
        self.assertEqual("creator_one", result["creator_id"])
        row = self._creator_row()
        self.assertEqual("Brazil", row["country"])

    def test_updates_language(self) -> None:
        self.repository.updateCreator("creator_one", {"language": "Portuguese"})
        row = self._creator_row()
        self.assertEqual("Portuguese", row["language"])

    def test_metadata_can_be_cleared(self) -> None:
        self.repository.updateCreator("creator_one", {
            "country": "Brazil",
            "language": "Portuguese",
        })
        self.repository.updateCreator("creator_one", {"country": ""})

        record = self.repository.getCreatorDetail("creator_one")["record"]
        self.assertEqual("", record["country"])
        self.assertEqual("Portuguese", record["language"])

    def test_existing_profile_fields_remain_editable(self) -> None:
        self.repository.updateCreator("creator_one", {
            "creator_name": "Creator Updated",
            "followers": "25K",
            "content_category": "Lifestyle",
        })

        row = self._creator_row()
        self.assertEqual("Creator Updated", row["name"])
        self.assertEqual("25K", row["followers"])
        self.assertEqual("Lifestyle", row["content_category"])

    def test_country_is_trimmed_before_persistence(self) -> None:
        self.repository.updateCreator("creator_one", {"country": "  Brazil  "})
        self.assertEqual("Brazil", self._creator_row()["country"])


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import csv
import importlib
import json
import os
import sys
import tempfile
import threading
import unittest
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository


def close_app_logger() -> None:
    app_logging = sys.modules.get("app_logging")
    if app_logging is None:
        return
    logger = app_logging.logging.getLogger(app_logging.LOGGER_NAME)
    for handler in list(logger.handlers):
        handler.close()
        logger.removeHandler(handler)
    app_logging._CONFIGURED = False


def runtime_data_environment(temp_root: str) -> dict[str, str]:
    return {
        "APPDATA": temp_root,
        "HOME": temp_root,
        "XDG_DATA_HOME": temp_root,
    }


OLD_CREATORS_HEADERS = [
    "creator_id", "name", "platform", "profile_url", "country", "language",
    "content_category", "followers", "insight_level", "status", "created_at",
    "tags", "updated_at",
]
OLD_SHEETS = {
    "Videos": ["creator_id", "video_url", "views", "likes", "comments", "captured_at"],
    "Insights": ["creator_id", "average_views", "median_views", "stability", "risks", "recommendation"],
    "CreatorSnapshots": [
        "snapshot_id", "creator_id", "platform", "account_uid", "followers",
        "average_views", "median_views", "video_count", "creator_score",
        "insight_level", "captured_at", "source",
    ],
    "VideoSnapshots": [
        "video_snapshot_id", "snapshot_id", "creator_id", "video_id",
        "video_url", "platform", "views", "likes", "comments", "captured_at",
    ],
    "Cooperations": [
        "cooperation_id", "creator_id", "campaign", "platform", "contact_date",
        "price", "published_count", "total_views", "average_views", "roi",
        "result", "note", "created_at",
    ],
    "_AnalysisData": [
        "creator_id", "task_id", "account_uid", "status_updated_at",
        "analysis_json", "source",
    ],
    "_Metadata": ["schema_version", "last_update_time"],
}


def analysis(
    task_id: str,
    account_uid: str,
    profile_url: str,
    *,
    name: str = "Alex",
    platform: str = "TikTok",
    creator_id: str = "",
) -> dict:
    payload = {
        "schema_version": "1.0",
        "analysis_id": f"analysis_{task_id}",
        "task_id": task_id,
        "account_uid": account_uid,
        "imported_at": "2026-07-31T10:00:00Z",
        "source": "chrome_extension",
        "creator": {
            "creator_name": name,
            "platform": platform,
            "profile_url": profile_url,
            "followers": "10K",
        },
        "video_analysis": {},
        "videos": [],
        "creator_insight": {"level": "good", "risks": [], "recommendation": ""},
    }
    if creator_id:
        payload["creator_id"] = creator_id
    return payload


def create_v13_workbook(path: Path) -> bytes:
    workbook = Workbook()
    creators = workbook.active
    creators.title = "Creators"
    creators.append(OLD_CREATORS_HEADERS)
    creators.append([
        "analysis_task_20260731T090000Z_aaaaaaaa",
        "Legacy Creator",
        "TikTok",
        "https://www.tiktok.com/@legacy",
        "Brazil",
        "Portuguese",
        "Gaming",
        "25K",
        "good",
        "discovered",
        "2026-07-01T00:00:00Z",
        "",
        "2026-07-01T00:00:00Z",
    ])
    for sheet_name, headers in OLD_SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
    creator_id = "analysis_task_20260731T090000Z_aaaaaaaa"
    workbook["Videos"].append([
        creator_id, "https://www.tiktok.com/@legacy/video/1", 1000, 100, 10,
        "2026-07-01T00:00:00Z",
    ])
    workbook["Insights"].append([creator_id, 1000, 900, 0.9, "[]", "Review"])
    workbook["CreatorSnapshots"].append([
        "snapshot_task_20260731T090000Z_aaaaaaaa",
        creator_id,
        "TikTok",
        "tiktok|https://www.tiktok.com/@legacy",
        "25K",
        1000,
        900,
        1,
        80,
        "good",
        "2026-07-01T00:00:00Z",
        "chrome_extension",
    ])
    workbook["VideoSnapshots"].append([
        "snapshot_task_20260731T090000Z_aaaaaaaa:1",
        "snapshot_task_20260731T090000Z_aaaaaaaa",
        creator_id,
        "1",
        "https://www.tiktok.com/@legacy/video/1",
        "TikTok",
        1000,
        100,
        10,
        "2026-07-01T00:00:00Z",
    ])
    workbook["Cooperations"].append([
        "cooperation_legacy",
        creator_id,
        "Launch",
        "TikTok",
        "2026-07-01",
        100,
        1,
        1000,
        1000,
        2,
        "completed",
        "",
        "2026-07-01T00:00:00Z",
    ])
    workbook["_AnalysisData"].append([
        "analysis_task_20260731T090000Z_aaaaaaaa",
        "task_20260731T090000Z_aaaaaaaa",
        "tiktok|https://www.tiktok.com/@legacy",
        "",
        "",
        "chrome_extension",
    ])
    workbook["_Metadata"].append(["1.3", "2026-07-01T00:00:00Z"])
    workbook.save(path)
    return path.read_bytes()


class Phase1RepositoryMigrationTests(unittest.TestCase):
    def test_old_creator_is_preserved_and_account_is_created(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            create_v13_workbook(workbook_path)
            repository = creator_repository.CreatorRepository(workbook_path)

            records = repository.getCreators()
            accounts = repository.getCreatorAccounts()
            workbook = load_workbook(workbook_path, read_only=True)
            try:
                sheet_names = set(workbook.sheetnames)
                preserved_counts = {
                    name: workbook[name].max_row - 1
                    for name in (
                        "Videos", "Insights", "CreatorSnapshots",
                        "VideoSnapshots", "Cooperations",
                    )
                }
            finally:
                workbook.close()

            self.assertEqual(1, len(records))
            self.assertEqual("Legacy Creator", records[0]["creator_name"])
            self.assertEqual(1, len(accounts))
            self.assertEqual(records[0]["creator_id"], accounts[0]["creator_id"])
            self.assertEqual(
                {
                    "Videos": 1,
                    "Insights": 1,
                    "CreatorSnapshots": 1,
                    "VideoSnapshots": 1,
                    "Cooperations": 1,
                },
                preserved_counts,
            )
            self.assertTrue(
                {"CreatorAccounts", "Agencies", "AgencyContacts", "FollowUpLogs"}.issubset(sheet_names)
            )

    def test_migration_is_repeatable_and_creates_only_one_timestamped_backup(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            root = Path(temp_dir)
            workbook_path = root / "Creator_Library.xlsx"
            create_v13_workbook(workbook_path)
            repository = creator_repository.CreatorRepository(workbook_path)

            repository.getCreators()
            first_accounts = repository.getCreatorAccounts()
            first_backups = list(root.glob("Creator_Library.pre_v2_*.xlsx"))
            repository.getCreators()
            second_accounts = repository.getCreatorAccounts()
            second_backups = list(root.glob("Creator_Library.pre_v2_*.xlsx"))

            self.assertEqual(1, len(first_accounts))
            self.assertEqual(1, len(second_accounts))
            self.assertEqual(1, len(first_backups))
            self.assertEqual(first_backups, second_backups)

    def test_migration_failure_keeps_original_and_backup_can_restore(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            root = Path(temp_dir)
            workbook_path = root / "Creator_Library.xlsx"
            original = create_v13_workbook(workbook_path)
            repository = creator_repository.CreatorRepository(workbook_path)

            with mock.patch.object(creator_repository.os, "replace", side_effect=OSError("replace failed")):
                with self.assertRaisesRegex(OSError, "replace failed"):
                    repository.getCreators()

            backups = list(root.glob("Creator_Library.pre_v2_*.xlsx"))
            self.assertEqual(original, workbook_path.read_bytes())
            self.assertEqual(1, len(backups))
            self.assertEqual(original, backups[0].read_bytes())


class Phase1IdentityAndRelationshipTests(unittest.TestCase):
    def repository(self, temp_dir: str) -> creator_repository.CreatorRepository:
        return creator_repository.CreatorRepository(Path(temp_dir) / "Creator_Library.xlsx")

    def test_repeated_extension_import_keeps_one_creator_and_account(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            repository = self.repository(temp_dir)
            uid = "tiktok|https://www.tiktok.com/@alex"
            first = repository.saveCreator(
                analysis(
                    "task_20260731T100000Z_aaaaaaaa",
                    uid,
                    "https://www.tiktok.com/@alex",
                )
            )
            second = repository.saveCreator(
                analysis(
                    "task_20260731T110000Z_bbbbbbbb",
                    uid,
                    "https://www.tiktok.com/@alex",
                )
            )

            self.assertEqual(first["creator_id"], second["creator_id"])
            self.assertEqual(1, len(repository.getCreators()))
            self.assertEqual(1, len(repository.getCreatorAccounts()))
            self.assertEqual(2, len(repository.getCreatorSnapshots(first["creator_id"])))

    def test_one_creator_can_have_multiple_platform_accounts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            repository = self.repository(temp_dir)
            first = repository.saveCreator(
                analysis(
                    "task_20260731T120000Z_aaaaaaaa",
                    "tiktok|https://www.tiktok.com/@alex",
                    "https://www.tiktok.com/@alex",
                )
            )
            repository.saveCreator(
                analysis(
                    "task_20260731T130000Z_bbbbbbbb",
                    "youtube|https://www.youtube.com/@alex",
                    "https://www.youtube.com/@alex",
                    platform="YouTube",
                    creator_id=first["creator_id"],
                )
            )

            self.assertEqual(1, len(repository.getCreators()))
            self.assertEqual(2, len(repository.getCreatorAccounts(first["creator_id"])))

    def test_same_name_does_not_merge_different_accounts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            repository = self.repository(temp_dir)
            repository.saveCreator(
                analysis(
                    "task_20260731T140000Z_aaaaaaaa",
                    "tiktok|https://www.tiktok.com/@same-name-one",
                    "https://www.tiktok.com/@same-name-one",
                    name="Same Name",
                )
            )
            repository.saveCreator(
                analysis(
                    "task_20260731T150000Z_bbbbbbbb",
                    "tiktok|https://www.tiktok.com/@same-name-two",
                    "https://www.tiktok.com/@same-name-two",
                    name="Same Name",
                )
            )

            self.assertEqual(2, len(repository.getCreators()))
            self.assertEqual(2, len(repository.getCreatorAccounts()))

    def test_agency_contacts_and_creator_relations_are_independent(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            repository = self.repository(temp_dir)
            creator_one = repository.saveCreator(
                analysis(
                    "task_20260731T160000Z_aaaaaaaa",
                    "tiktok|https://www.tiktok.com/@one",
                    "https://www.tiktok.com/@one",
                )
            )
            creator_two = repository.saveCreator(
                analysis(
                    "task_20260731T170000Z_bbbbbbbb",
                    "youtube|https://www.youtube.com/@two",
                    "https://www.youtube.com/@two",
                    platform="YouTube",
                )
            )
            agency = repository.saveAgency({"name": "Example Agency"})
            source_contact = repository.saveAgencyContact({"name": "Maria"})
            current_contact = repository.saveAgencyContact(
                {"name": "John", "agency_id": agency["agency_id"]}
            )
            repository.updateCreatorRelations(
                creator_one["creator_id"],
                {
                    "agency_id": agency["agency_id"],
                    "source_contact_id": source_contact["contact_id"],
                    "current_contact_id": current_contact["contact_id"],
                },
            )
            repository.updateCreatorRelations(
                creator_two["creator_id"],
                {"agency_id": agency["agency_id"]},
            )
            detail = repository.getAgencyDetail(agency["agency_id"])
            creator_detail = repository.getCreatorDetail(creator_one["creator_id"])

            self.assertEqual("", source_contact["agency_id"])
            self.assertEqual(2, len(detail["creators"]))
            self.assertNotEqual(
                creator_detail["record"]["source_contact_id"],
                creator_detail["record"]["current_contact_id"],
            )

    def test_task_import_skips_failures_and_is_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            repository = self.repository(temp_dir)
            task_id = "task_20260731T180000Z_aaaaaaaa"
            records = [
                {
                    "account_uid": "tiktok|https://www.tiktok.com/@valid",
                    "platform": "TikTok",
                    "profile_url": "https://www.tiktok.com/@valid",
                    "creator_name": "Valid",
                    "scrape_status": "success",
                },
                {
                    "account_uid": "instagram|https://www.instagram.com/failed/",
                    "platform": "Instagram",
                    "profile_url": "https://www.instagram.com/failed/",
                    "creator_name": "",
                    "scrape_status": "failed",
                },
                {
                    "account_uid": "",
                    "platform": "",
                    "profile_url": "",
                    "creator_name": "Plain text",
                    "scrape_status": "success",
                },
            ]
            first = repository.importTaskResults(task_id, records, source="系统抓取")
            second = repository.importTaskResults(task_id, records, source="系统抓取")

            self.assertEqual(1, first["created_creators"])
            self.assertEqual(1, first["created_accounts"])
            self.assertEqual(1, first["skipped_failed"])
            self.assertEqual(1, first["skipped_invalid"])
            self.assertEqual(0, second["created_creators"])
            self.assertEqual(0, second["created_accounts"])
            self.assertEqual(1, len(repository.getCreators()))
            self.assertEqual(1, len(repository.getCreatorAccounts()))


class Phase1ServerCompatibilityTests(unittest.TestCase):
    def test_unconfigured_feishu_and_basic_apis_work(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.dict(
            os.environ,
            runtime_data_environment(temp_dir),
        ):
            for module_name in (
                "server", "runtime_paths", "mail_sync", "task_manager",
                "dashboard_repository", "dashboard_service",
            ):
                sys.modules.pop(module_name, None)
            server = importlib.import_module("server")
            httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
            thread = threading.Thread(target=httpd.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{httpd.server_port}"
            try:
                for path in (
                    "/api/state",
                    "/api/dashboard",
                    "/api/creator-library",
                    "/api/tasks",
                    "/api/mail/inbox/messages",
                    "/api/local/agencies",
                    "/api/local/agency-contacts",
                ):
                    with urllib.request.urlopen(base_url + path, timeout=10) as response:
                        payload = json.loads(response.read().decode("utf-8"))
                    self.assertTrue(payload.get("ok", True), path)
            finally:
                httpd.shutdown()
                httpd.server_close()
                thread.join(timeout=5)
                close_app_logger()

    def test_feishu_sync_failure_does_not_change_local_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.dict(
            os.environ,
            runtime_data_environment(temp_dir),
        ):
            for module_name in (
                "server", "runtime_paths", "mail_sync", "task_manager",
                "dashboard_repository", "dashboard_service",
            ):
                sys.modules.pop(module_name, None)
            server = importlib.import_module("server")
            task = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@sync-test"],
                [],
                1,
                name="sync-test",
                target_platform="TikTok",
            )
            result = server.scraper_module.build_result(
                url="https://www.tiktok.com/@sync-test",
                platform="TikTok",
                name="Sync Test",
                scrape_status="success",
            )
            row = server.scraper_module.result_to_row(result)
            _task, paths = server.task_manager.load_task(server.TASKS_DIR, task["id"])
            server.task_manager.atomic_write_files(
                {
                    paths["results"]: server._csv_content(server.scraper_module.OUTPUT_FIELDS, [row]),
                    paths["progress"]: server._csv_content(server.scraper_module.PROGRESS_FIELDS, [row]),
                }
            )
            server.task_manager.update_task(server.TASKS_DIR, task["id"], status="completed")
            server.import_task_results_to_creator_library(task["id"])
            workbook_path = server.get_creator_repository().workbook_path
            before = workbook_path.read_bytes()

            with mock.patch.object(
                server.scraper_module,
                "push_to_feishu_four_tables",
                side_effect=RuntimeError("Feishu unavailable"),
            ):
                sync_result = server.sync_task_results_to_four_tables(task["id"])

            self.assertEqual("failed", sync_result["sync_status"])
            self.assertEqual(before, workbook_path.read_bytes())
            close_app_logger()


if __name__ == "__main__":
    unittest.main()

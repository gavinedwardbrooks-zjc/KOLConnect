from __future__ import annotations

import json
import sys
import tempfile
import threading
import unittest
import urllib.error
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import server
from campaign_creator_repository import CampaignCreatorRepository
from campaign_repository import CampaignRepository
from product_repository import ProductRepository


def append_mapping(sheet, values: dict) -> None:
    headers = [cell.value for cell in sheet[1]]
    sheet.append([values.get(header, "") for header in headers])


class LegacyCooperationReadOnlyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        creator_repository.CreatorRepository(self.workbook_path).getCreators()
        self._seed_workbook()

        self.patchers = [
            mock.patch.object(
                server,
                "get_creator_repository",
                side_effect=lambda: creator_repository.CreatorRepository(self.workbook_path),
            ),
            mock.patch.object(
                server,
                "get_product_repository",
                side_effect=lambda: ProductRepository(self.workbook_path),
            ),
            mock.patch.object(
                server,
                "get_campaign_repository",
                side_effect=lambda: CampaignRepository(self.workbook_path),
            ),
            mock.patch.object(
                server,
                "get_campaign_creator_repository",
                side_effect=lambda: CampaignCreatorRepository(self.workbook_path),
            ),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in self.patchers:
            patcher.start()

        self.httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.httpd.server_port}"

    def tearDown(self) -> None:
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=5)
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp_dir.cleanup()

    def _seed_workbook(self) -> None:
        workbook = load_workbook(self.workbook_path)
        try:
            append_mapping(workbook["Creators"], {
                "creator_id": "creator_legacy",
                "name": "Legacy Creator",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@legacy-creator",
                "status": "discovered",
                "created_at": "2026-08-01T00:00:00Z",
                "updated_at": "2026-08-01T00:00:00Z",
            })
            append_mapping(workbook["CreatorAccounts"], {
                "account_id": "account_legacy",
                "creator_id": "creator_legacy",
                "account_uid": "tiktok|legacy-creator",
                "platform": "TikTok",
                "username": "legacy-creator",
                "profile_url": "https://www.tiktok.com/@legacy-creator",
                "created_at": "2026-08-01T00:00:00Z",
                "updated_at": "2026-08-01T00:00:00Z",
            })
            append_mapping(workbook["Cooperations"], {
                "cooperation_id": "legacy_cooperation_one",
                "creator_id": "creator_legacy",
                "campaign": "Legacy Campaign",
                "platform": "TikTok",
                "contact_date": "2026-07-01",
                "price": 999999,
                "total_views": 999999,
                "average_views": 999999,
                "roi": 999,
                "result": "Historical",
                "created_at": "2026-07-01T00:00:00Z",
            })
            workbook.save(self.workbook_path)
        finally:
            workbook.close()

    def request(
        self,
        method: str,
        path: str,
        payload: dict | None = None,
    ) -> tuple[int, dict]:
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            self.base_url + path,
            data=data,
            method=method,
            headers={"Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                return response.status, json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            try:
                return exc.code, json.loads(exc.read().decode("utf-8"))
            finally:
                exc.close()

    def workbook_state(self) -> tuple[list[list], str]:
        workbook = load_workbook(self.workbook_path, data_only=True)
        try:
            cooperation_rows = [list(row) for row in workbook["Cooperations"].iter_rows(values_only=True)]
            creators = workbook["Creators"]
            headers = [cell.value for cell in creators[1]]
            status_index = headers.index("status")
            creator_status = next(
                row[status_index]
                for row in creators.iter_rows(min_row=2, values_only=True)
                if row[0] == "creator_legacy"
            )
            return cooperation_rows, str(creator_status or "")
        finally:
            workbook.close()

    def test_legacy_history_is_read_only_and_v2_flow_remains_active(self) -> None:
        status, detail = self.request("GET", "/api/creator-library/creator_legacy")
        self.assertEqual(200, status)
        self.assertEqual("legacy_cooperation_one", detail["cooperations"][0]["cooperation_id"])
        self.assertEqual(1, detail["cooperation_statistics"]["cooperation_count"])

        before_rows, before_status = self.workbook_state()
        path = "/api/creator-library/creator_legacy/cooperations"
        for method in ("POST", "PUT", "PATCH", "DELETE"):
            status, body = self.request(method, path, {"campaign": "Forbidden"})
            self.assertEqual(403, status, (method, body))
            self.assertIn("请使用 Campaign 创建新的合作", body["error"])

        with self.assertRaises(PermissionError):
            creator_repository.CreatorRepository(self.workbook_path).saveCooperation(
                "creator_legacy", {"campaign": "Direct write"}
            )

        after_rows, after_status = self.workbook_state()
        self.assertEqual(before_rows, after_rows)
        self.assertEqual(before_status, after_status)

        status, product_body = self.request(
            "POST", "/api/products", {"name": "Product A", "company_name": "Studio"}
        )
        self.assertEqual(201, status)
        status, campaign_body = self.request(
            "POST",
            "/api/campaigns",
            {
                "product_id": product_body["product"]["product_id"],
                "name": "Campaign A",
                "status": "running",
            },
        )
        self.assertEqual(201, status)
        campaign_id = campaign_body["campaign"]["campaign_id"]
        status, relation_body = self.request(
            "POST",
            f"/api/campaigns/{campaign_id}/creators",
            {
                "creator_id": "creator_legacy",
                "account_id": "account_legacy",
                "stage": "completed",
                "cost": 100,
                "views": 1000,
                "roi": 2,
                "performance_note": "Reviewed",
            },
        )
        self.assertEqual(201, status, relation_body)

        status, dashboard = self.request("GET", "/api/dashboard")
        self.assertEqual(200, status)
        performance = dashboard["cooperation_performance"]
        self.assertEqual(1, performance["total_campaigns"])
        self.assertEqual(100, performance["total_cost"])
        self.assertEqual(1000, performance["total_views"])
        self.assertEqual(2, performance["average_roi"])


if __name__ == "__main__":
    unittest.main()

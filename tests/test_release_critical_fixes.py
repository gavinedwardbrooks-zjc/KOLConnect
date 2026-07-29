from __future__ import annotations

import csv
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import scraper


def extension_analysis(task_id: str, account_uid: str, followers: str) -> dict:
    return {
        "schema_version": "1.0",
        "analysis_id": f"analysis_{task_id}",
        "task_id": task_id,
        "account_uid": account_uid,
        "imported_at": "2026-07-29T10:00:00Z",
        "source": "chrome_extension",
        "creator": {
            "creator_name": "Maria",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@maria",
            "followers": followers,
        },
        "video_analysis": {
            "average_views": 1000,
            "median_views": 900,
            "view_stability": 0.9,
        },
        "videos": [],
        "creator_insight": {
            "level": "good",
            "risks": [],
            "recommendation": "Review",
        },
    }


def scraper_args(progress_file: Path, output_file: Path) -> SimpleNamespace:
    return SimpleNamespace(
        file="links.txt",
        excel=None,
        urls=None,
        column="url",
        chrome_dir=None,
        chrome_profile="Default",
        no_browser=False,
        progress_file=str(progress_file),
        task_file=None,
        reset=False,
        feishu_app_id=None,
        feishu_app_secret=None,
        feishu_app_token=None,
        feishu_creator_table_id=None,
        feishu_account_table_id=None,
        four_table_sync=False,
        no_feishu=True,
        output=str(output_file),
        sync_result_file=None,
    )


class CreatorRepositoryRegressionTests(unittest.TestCase):
    def test_repeated_account_import_keeps_one_creator_and_one_insight_owner(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository,
            "log_event",
        ):
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            repository = creator_repository.CreatorRepository(workbook_path)
            account_uid = "TikTok|https://www.tiktok.com/@maria"
            first = repository.saveCreator(
                extension_analysis("task_20260729T100000Z_aaaaaaaa", account_uid, "10K")
            )
            second = repository.saveCreator(
                extension_analysis("task_20260729T110000Z_bbbbbbbb", account_uid, "11K")
            )

            workbook = load_workbook(workbook_path, read_only=True)
            try:
                creators = list(workbook["Creators"].iter_rows(min_row=2, values_only=True))
                snapshots = list(
                    workbook["CreatorSnapshots"].iter_rows(min_row=2, values_only=True)
                )
                insights = list(workbook["Insights"].iter_rows(min_row=2, values_only=True))
            finally:
                workbook.close()

            self.assertEqual(first["creator_id"], second["creator_id"])
            self.assertTrue(first["is_new_creator"])
            self.assertFalse(second["is_new_creator"])
            self.assertEqual(1, len(creators))
            self.assertEqual(2, len(snapshots))
            self.assertEqual(1, len(insights))
            self.assertEqual(first["creator_id"], insights[0][0])


class ScraperPersistenceRegressionTests(unittest.TestCase):
    URL = "https://www.tiktok.com/@maria"

    def run_main(
        self,
        progress_file: Path,
        output_file: Path,
        driver: mock.Mock,
        scrape_side_effect,
    ) -> None:
        args = scraper_args(progress_file, output_file)
        driver._potato_temp_user_data_dir = None
        with (
            mock.patch.object(scraper.argparse.ArgumentParser, "parse_args", return_value=args),
            mock.patch.object(scraper, "read_from_file", return_value=[self.URL]),
            mock.patch.object(scraper, "normalize_urls", return_value=[self.URL]),
            mock.patch.object(scraper, "make_chrome_driver", return_value=driver),
            mock.patch.object(scraper, "scrape_all", side_effect=scrape_side_effect),
            mock.patch.object(scraper, "SELENIUM_AVAILABLE", True),
            mock.patch.object(scraper, "PANDAS_AVAILABLE", False),
            mock.patch.object(scraper, "setup_console_encoding"),
        ):
            scraper.main()

    def saved_result(self) -> dict:
        return scraper.build_result(
            url=self.URL,
            platform="TikTok",
            name="Maria",
            emails=["maria@example.com"],
            last_scrape_time="2026-07-29T10:00:00Z",
        )

    def assert_csv_files(self, progress_file: Path, output_file: Path) -> None:
        with progress_file.open(encoding="utf-8-sig", newline="") as handle:
            progress_rows = list(csv.DictReader(handle))
        with output_file.open(encoding="utf-8-sig", newline="") as handle:
            output_rows = list(csv.DictReader(handle))
        self.assertEqual(1, len(progress_rows))
        self.assertEqual(1, len(output_rows))
        self.assertEqual(self.URL, output_rows[0][scraper.FIELD_URL])

    def test_driver_quit_failure_does_not_break_saved_results(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            progress_file = root / "progress.csv"
            output_file = root / "results.csv"
            result = self.saved_result()
            driver = mock.Mock()
            driver.quit.side_effect = RuntimeError("Chrome already closed")

            def scrape_side_effect(urls, *, driver, progress_file, task_file):
                scraper.save_progress(result, progress_file)
                return [result]

            self.run_main(progress_file, output_file, driver, scrape_side_effect)
            self.assert_csv_files(progress_file, output_file)

    def test_unexpected_scrape_error_recovers_results_from_progress(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            progress_file = root / "progress.csv"
            output_file = root / "results.csv"
            result = self.saved_result()
            driver = mock.Mock()

            def scrape_side_effect(urls, *, driver, progress_file, task_file):
                scraper.save_progress(result, progress_file)
                raise RuntimeError("network failed")

            with self.assertRaisesRegex(RuntimeError, "network failed"):
                self.run_main(progress_file, output_file, driver, scrape_side_effect)
            self.assert_csv_files(progress_file, output_file)


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import hashlib
import json
import shutil
import sys
import unittest
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import Mock, patch
from uuid import uuid4

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
TESTS = ROOT / "tests"
for path in (APP, TESTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from repository_factory import RepositoryFactory
from storage.authority import resolve_runtime_authority
from storage.errors import SQLiteActivationError
from storage.migration import ExcelToSQLiteMigrator
from storage.paths import SQLiteStoragePaths
from storage.sqlite_workbook_store import SQLiteWorkbookStore
from storage.sqlite_creator_repository import SQLiteCreatorRepository
from services.clean_reset_service import CleanResetService
from services.creator_delete_impact_service import CreatorDeleteImpactService
from services.creator_hard_delete_service import CreatorHardDeleteService
from services.dashboard_response_cache import DashboardResponseCache
from services.feishu_delete_intent_service import FeishuDeleteIntentStore
from services.feishu_sync_service import FeishuSyncService
from services.assistant_provider import AssistantIntent, MockAssistantProvider
from services.assistant_service import AssistantService
from services.workbook_backup_service import WorkbookBackupService
from staged_delete_transaction import StagedDeleteTransaction
from test_pre_m8_excel_sqlite_migration import build_fixture
from test_m7_1_feishu_sync_foundation import FakeClient


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class SQLiteRuntimeCutoverTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime = ROOT / ".pre_m8_batch2_test_runtime"
        runtime.mkdir(exist_ok=True)
        self.root = runtime / f"cutover_{uuid4().hex}"
        self.root.mkdir()
        (self.root / "logs").mkdir()
        self.lock_patch = patch(
            "local_storage_lock.get_shared_storage_lock_path",
            return_value=self.root / "locks" / "shared_storage.lock",
        )
        self.lock_patch.start()
        self.logs_patch = patch(
            "app_logging.get_logs_dir", return_value=self.root / "logs"
        )
        self.logs_patch.start()
        self.paths = SQLiteStoragePaths.for_app_data(self.root / "appdata")
        self.workbook = self.root / "legacy.xlsx"
        build_fixture(self.workbook)
        result = ExcelToSQLiteMigrator(self.paths).migrate(self.workbook)
        ExcelToSQLiteMigrator(self.paths).activate_synthetic(result)

    def tearDown(self) -> None:
        self.logs_patch.stop()
        self.lock_patch.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def factory(self) -> RepositoryFactory:
        return RepositoryFactory.for_runtime(
            self.workbook,
            storage_paths=self.paths,
            tasks_dir=self.root / "tasks",
            data_protection_file=self.root / "data_protection.json",
        )

    def test_runtime_contract_and_no_dual_write(self) -> None:
        source_before = digest(self.workbook)
        factory = self.factory()
        self.assertIsInstance(factory.store, SQLiteWorkbookStore)
        creators = factory.creator().getCreators(include_archived=True)
        self.assertEqual(4, len(creators))
        detail = factory.creator().getCreatorDetail("creator_0000")
        self.assertEqual(3, len(detail["accounts"]))
        campaign = factory.campaign().getCampaign("campaign_1")
        self.assertEqual(["TikTok", "YouTube"], campaign["platforms"])
        relations = factory.campaign_creator().getCampaignCreators(campaign_id="campaign_1")
        self.assertEqual(3, len(relations))
        self.assertEqual(2, len(relations[-1]["account_ids"]))

        revision = factory.store.business_revision()
        factory.creator().updateCreator("creator_0000", {"country": "Portugal"})
        self.assertEqual(revision + 1, factory.store.business_revision())
        self.assertEqual(source_before, digest(self.workbook))
        self.assertEqual(
            "Portugal",
            next(
                row for row in self.factory().creator().getCreators(include_archived=True)
                if row["creator_id"] == "creator_0000"
            )["country"],
        )

        workbook = load_workbook(self.workbook)
        workbook["Creators"].cell(2, 2, "STALE EXCEL NAME")
        workbook.save(self.workbook)
        workbook.close()
        self.assertNotEqual(
            "STALE EXCEL NAME",
            next(
                row for row in self.factory().creator().getCreators(include_archived=True)
                if row["creator_id"] == "creator_0000"
            )["creator_name"],
        )

    def test_write_exception_rolls_back_and_does_not_increment_revision(self) -> None:
        store = self.factory().store
        revision = store.business_revision()
        with self.assertRaisesRegex(RuntimeError, "rollback"):
            with store.workbook(write=True) as workbook:
                workbook["Creators"].cell(2, 2, "NOT COMMITTED")
                raise RuntimeError("rollback")
        self.assertEqual(revision, store.business_revision())
        self.assertNotEqual(
            "NOT COMMITTED",
            next(
                row for row in self.factory().creator().getCreators(include_archived=True)
                if row["creator_id"] == "creator_0000"
            )["creator_name"],
        )

    def test_snapshot_append_does_not_rebuild_base_tables(self) -> None:
        store = self.factory().store
        with store.factory.read_connection() as connection:
            creator_rowid = connection.execute(
                "SELECT rowid FROM creators WHERE creator_id='creator_0000'"
            ).fetchone()[0]
            snapshot_rowid = connection.execute(
                "SELECT rowid FROM creator_snapshots WHERE snapshot_id='snapshot_00000'"
            ).fetchone()[0]
        with store.workbook(write=True) as workbook:
            sheet = workbook["CreatorSnapshots"]
            headers = [cell.value for cell in sheet[1]]
            record = {
                "snapshot_id": "snapshot_incremental",
                "creator_id": "creator_0000",
                "account_uid": "account_0000_0",
                "platform": "TikTok",
                "captured_at": "2026-08-25T00:00:00Z",
            }
            sheet.append([record.get(header, "") for header in headers])
        with store.factory.read_connection() as connection:
            self.assertEqual(
                creator_rowid,
                connection.execute(
                    "SELECT rowid FROM creators WHERE creator_id='creator_0000'"
                ).fetchone()[0],
            )
            self.assertEqual(
                snapshot_rowid,
                connection.execute(
                    "SELECT rowid FROM creator_snapshots WHERE snapshot_id='snapshot_00000'"
                ).fetchone()[0],
            )
            self.assertEqual(
                1,
                connection.execute(
                    "SELECT COUNT(*) FROM creator_snapshots WHERE snapshot_id='snapshot_incremental'"
                ).fetchone()[0],
            )

    def test_creator_detail_uses_indexed_creator_scoped_snapshot_reads(self) -> None:
        factory = self.factory()
        self.assertIsInstance(factory.creator(), SQLiteCreatorRepository)
        detail = factory.creator().getCreatorDetail("creator_0000")
        self.assertTrue(detail["snapshots"])
        self.assertEqual(
            {"creator_0000"},
            {str(row["creator_id"]) for row in detail["snapshots"]},
        )
        with factory.store.creator_read_scope("creator_0000"):
            workbook = factory.store.open()
        try:
            snapshot_creator_ids = {
                str(row[1].value)
                for row in workbook["CreatorSnapshots"].iter_rows(min_row=2)
            }
            self.assertEqual({"creator_0000"}, snapshot_creator_ids)
        finally:
            workbook.close()
        with factory.store.factory.read_connection() as connection:
            plan = " ".join(
                str(row[3])
                for row in connection.execute(
                    "EXPLAIN QUERY PLAN SELECT * FROM creator_snapshots "
                    "WHERE creator_id=? ORDER BY captured_at DESC",
                    ("creator_0000",),
                )
            )
        self.assertIn("idx_creator_snapshots_creator_time", plan)

    def test_common_sqlite_reads_do_not_materialize_unrelated_high_volume_tables(self) -> None:
        store = self.factory().store
        statements: list[str] = []
        original_connect = store.factory.connect

        def traced_connect():
            connection = original_connect()
            connection.set_trace_callback(statements.append)
            return connection

        with patch.object(store.factory, "connect", side_effect=traced_connect):
            self.factory().creator().getCreatorsPage(page=1, page_size=12)
        sql = "\n".join(statements).lower()
        self.assertNotIn(" from video_snapshots", sql)
        self.assertNotIn(" from videos", sql)
        self.assertLess(
            sum(statement.lstrip().upper().startswith("SELECT") for statement in statements),
            20,
        )

        statements.clear()
        with patch.object(store.factory, "connect", side_effect=traced_connect):
            self.factory().campaign().getCampaign("campaign_1")
        sql = "\n".join(statements).lower()
        self.assertNotIn(" from creator_snapshots", sql)
        self.assertNotIn(" from video_snapshots", sql)
        self.assertLess(
            sum(statement.lstrip().upper().startswith("SELECT") for statement in statements),
            20,
        )

    def test_dashboard_uses_set_based_sqlite_reads_without_workbook_projection(self) -> None:
        factory = self.factory()
        repository = factory.dashboard(
            factory.creator(), factory.campaign_creator(), factory.campaign()
        )
        with patch.object(
            factory.store,
            "open",
            side_effect=AssertionError("Dashboard opened a workbook projection"),
        ):
            creators = repository.get_creators()
            relations = repository.get_campaign_creator_records(creators)
        self.assertTrue(creators)
        self.assertTrue(relations)

    def test_authority_marker_controls_mixed_files_and_corruption_fails_closed(self) -> None:
        authority = resolve_runtime_authority(self.paths, self.workbook)
        self.assertEqual("sqlite", authority.kind)
        self.paths.database_path.write_bytes(b"corrupt")
        with self.assertRaises(SQLiteActivationError):
            resolve_runtime_authority(self.paths, self.workbook)

    def test_new_install_bootstraps_sqlite_but_legacy_install_stays_excel(self) -> None:
        new_paths = SQLiteStoragePaths.for_app_data(self.root / "new_appdata")
        missing_workbook = self.root / "missing.xlsx"
        authority = resolve_runtime_authority(
            new_paths, missing_workbook, bootstrap_new_install=True
        )
        self.assertEqual("sqlite", authority.kind)
        self.assertTrue(new_paths.database_path.is_file())

        legacy_paths = SQLiteStoragePaths.for_app_data(self.root / "legacy_appdata")
        legacy_workbook = self.root / "legacy_only.xlsx"
        build_fixture(legacy_workbook, creator_count=0, include_campaign=False)
        legacy = resolve_runtime_authority(
            legacy_paths, legacy_workbook, bootstrap_new_install=True
        )
        self.assertEqual("legacy_excel", legacy.kind)
        self.assertFalse(legacy_paths.database_path.exists())

    def test_dashboard_revision_cache_rebuilds_only_after_commit(self) -> None:
        store = self.factory().store
        cache = DashboardResponseCache(lambda: date(2026, 8, 25))
        loader = Mock(side_effect=[{"version": 1}, {"version": 2}])
        self.assertEqual(1, cache.get_response(store, loader)["version"])
        self.assertEqual(1, cache.get_response(store, loader)["version"])
        self.factory().creator().updateCreator("creator_0000", {"language": "pt"})
        self.assertEqual(2, cache.get_response(store, loader)["version"])
        self.assertEqual(2, loader.call_count)

    def test_sqlite_backup_restore_export_and_managed_retention(self) -> None:
        store = self.factory().store
        service = WorkbookBackupService(
            lambda: store.workbook_path,
            store_provider=lambda: store,
            now_provider=lambda: datetime(2026, 8, 25, 8, 0, tzinfo=timezone.utc),
            token_provider=iter(("one", "two", "three")).__next__,
            retention=2,
        )
        first = service.create_backup()
        self.factory().creator().updateCreator("creator_0000", {"country": "Portugal"})
        second = service.create_backup()
        self.factory().creator().updateCreator("creator_0000", {"country": "Spain"})
        third = service.create_backup()
        backup_dir = self.paths.database_backup_dir
        self.assertFalse((backup_dir / first["filename"]).exists())
        self.assertTrue((backup_dir / second["filename"]).is_file())
        newest = backup_dir / third["filename"]
        store.restore_transaction_backup(newest)
        creator = self.factory().creator().getCreatorDetail("creator_0000")
        self.assertEqual("Spain", creator["record"]["country"])

        export_path = self.root / "compatibility.xlsx"
        store.export_workbook(export_path)
        exported = load_workbook(export_path, read_only=True)
        try:
            self.assertEqual(5, exported["Creators"].max_row)
            campaign_headers = [cell.value for cell in exported["Campaigns"][1]]
            platforms = exported["Campaigns"].cell(
                2, campaign_headers.index("platforms") + 1
            ).value
            self.assertEqual(["TikTok", "YouTube"], json.loads(platforms))
        finally:
            exported.close()

    def test_sqlite_clean_reset_preserves_authority_and_independent_intent(self) -> None:
        store = self.factory().store
        settings = self.root / "settings.json"
        protection = self.root / "data_protection.json"
        mail = self.root / "mail.json"
        tasks = self.root / "tasks"
        intent = self.root / "runtime" / "feishu_delete_intents" / "intent.json"
        settings.write_text('{"feishu":{"configured":true}}', encoding="utf-8")
        protection.write_text('{"creator_0000":{}}', encoding="utf-8")
        mail.write_text('{"accounts":{},"messages":[]}', encoding="utf-8")
        tasks.mkdir(exist_ok=True)
        (tasks / "task.json").write_text("{}", encoding="utf-8")
        intent.parent.mkdir(parents=True)
        intent.write_text('{"status":"pending_remote"}', encoding="utf-8")
        service = CleanResetService(
            self.workbook,
            settings_path=settings,
            data_protection_path=protection,
            mail_messages_path=mail,
            tasks_dir=tasks,
            store_provider=lambda: store,
            now_provider=lambda: datetime(2026, 8, 25, 9, 0, tzinfo=timezone.utc),
        )
        result = service.execute(confirm=True)
        self.assertEqual("success", result["status"])
        self.assertTrue(Path(result["backup"]["path"]).is_file())
        self.assertEqual([], self.factory().creator().getCreators(include_archived=True))
        self.assertTrue(self.paths.authority_marker_path.is_file())
        self.assertTrue(intent.is_file())
        self.assertEqual(
            {"feishu": {"configured": True}},
            json.loads(settings.read_text(encoding="utf-8")),
        )

    def test_clean_reset_backup_and_transaction_failures_leave_business_data(self) -> None:
        store = self.factory().store
        settings = self.root / "failure_settings.json"
        protection = self.root / "failure_protection.json"
        mail = self.root / "failure_mail.json"
        tasks = self.root / "failure_tasks"
        settings.write_text("{}", encoding="utf-8")
        protection.write_text("{}", encoding="utf-8")
        mail.write_text('{"accounts":{},"messages":[]}', encoding="utf-8")
        tasks.mkdir()
        service = CleanResetService(
            self.workbook,
            settings_path=settings,
            data_protection_path=protection,
            mail_messages_path=mail,
            tasks_dir=tasks,
            store_provider=lambda: store,
        )
        with patch.object(
            store, "create_transaction_backup", side_effect=OSError("backup failure")
        ):
            with self.assertRaisesRegex(OSError, "backup failure"):
                service.execute(confirm=True)
        self.assertEqual(4, len(self.factory().creator().getCreators(include_archived=True)))

        def delete_then_fail(connection, _workbook):
            connection.execute("DELETE FROM creator_tags")
            raise RuntimeError("transaction failure")

        with patch.object(store, "_prepare_and_replace", side_effect=delete_then_fail):
            with self.assertRaisesRegex(RuntimeError, "transaction failure"):
                service.execute(confirm=True)
        self.assertEqual(4, len(self.factory().creator().getCreators(include_archived=True)))
        with store.factory.read_connection() as connection:
            self.assertEqual(8, connection.execute("SELECT COUNT(*) FROM creator_tags").fetchone()[0])

    def test_sqlite_delete_transaction_rollback_restores_database_backup(self) -> None:
        transaction = StagedDeleteTransaction(
            self.root / "runtime", "creator_0000"
        )
        transaction.prepare()
        transaction.backup_workbook(self.factory().store)
        self.factory().creator().updateCreator("creator_0000", {"country": "Changed"})
        transaction.rollback()
        restored = self.factory().creator().getCreatorDetail("creator_0000")
        self.assertEqual("Brazil", restored["record"]["country"])
        self.assertEqual(
            "sqlite", transaction.load_manifest().get("storage_kind")
        )

    def test_remaining_repository_contracts_use_sqlite_authority(self) -> None:
        factory = self.factory()
        self.assertEqual("Product", factory.product().getProduct("product_1")["name"])
        self.assertEqual("Agency", factory.agency().get_agency("agency_1")["name"])
        self.assertEqual(1, len(factory.agency().list_contacts("agency_1")))
        detail = factory.creator().getCreatorDetail("creator_0000")
        self.assertEqual(3, len(detail["accounts"]))
        self.assertTrue(detail["snapshots"])
        with factory.store.read_only_workbook() as workbook:
            for sheet_name in (
                "Videos", "VideoSnapshots", "Insights", "Cooperations",
                "FollowUpLogs", "_AnalysisData",
            ):
                self.assertGreater(workbook[sheet_name].max_row, 1, sheet_name)
        campaign = factory.campaign().updateCampaign(
            "campaign_1", {"name": "SQLite Campaign", "platforms": []}
        )
        self.assertEqual("SQLite Campaign", campaign["name"])
        self.assertEqual([], campaign["platforms"])
        relation = factory.campaign_creator().getCampaignCreator("relation_1")
        self.assertEqual("", relation["creator_quote"])
        self.assertEqual("", relation["cost"])
        with factory.store.factory.read_connection() as connection:
            stored = connection.execute(
                "SELECT creator_quote, cost FROM campaign_creators WHERE id='relation_1'"
            ).fetchone()
            self.assertIsNone(stored["creator_quote"])
            self.assertIsNone(stored["cost"])

    def test_explicit_creator_import_is_transactional(self) -> None:
        factory = self.factory()
        repository = factory.creator()
        imported = repository.createCreatorsBatch([{
            "account_uid": "account_imported_1",
            "name": "Imported Creator",
            "platform": "TikTok",
            "profile_url": "https://example.test/imported",
            "country": "Portugal",
        }])
        self.assertEqual(1, imported["created"])
        self.assertIn(
            "account_imported_1",
            {row["account_uid"] for row in self.factory().creator().getCreatorAccounts()},
        )

        before = factory.store.business_revision()
        with patch.object(
            factory.store, "_prepare_and_replace", side_effect=RuntimeError("inject")
        ):
            with self.assertRaises(Exception):
                repository.createCreatorsBatch([{
                    "account_uid": "account_not_committed",
                    "name": "Rollback",
                    "platform": "YouTube",
                    "profile_url": "https://example.test/rollback",
                }])
        self.assertEqual(before, factory.store.business_revision())
        self.assertNotIn(
            "account_not_committed",
            {row["account_uid"] for row in self.factory().creator().getCreatorAccounts()},
        )

    def test_creator_merge_is_one_sqlite_commit_and_preserves_account_uid(self) -> None:
        repository = self.factory().creator_merge()
        preview = repository.preview("creator_0000", "creator_0003")
        self.assertTrue(preview["safe_to_merge"])
        before = self.factory().store.business_revision()
        repository.execute(
            "creator_0000", "creator_0003",
            preview_fingerprint=preview["preview_fingerprint"],
        )
        self.assertEqual(before + 1, self.factory().store.business_revision())
        creators = self.factory().creator().getCreators(include_archived=True)
        self.assertNotIn("creator_0003", {row["creator_id"] for row in creators})
        account = next(
            row for row in self.factory().creator().getCreatorAccounts()
            if row["account_uid"] == "account_0003_0"
        )
        self.assertEqual("creator_0000", account["creator_id"])

    def test_hard_delete_commits_sqlite_then_promotes_independent_intent(self) -> None:
        runtime = self.root / "hard_delete_runtime"
        impact = CreatorDeleteImpactService(
            lambda: self.factory().creator_delete_impact()
        )
        preview = impact.get_delete_impact("creator_0003")
        self.assertTrue(preview["can_delete"])
        intent_store = FeishuDeleteIntentStore(runtime)
        service = CreatorHardDeleteService(
            lambda: impact,
            lambda: self.factory().creator_hard_delete(),
            lambda: runtime,
            feishu_delete_intent_store=intent_store,
        )
        result = service.delete_creator(
            "creator_0003",
            confirm=True,
            preview_fingerprint=preview["preview_fingerprint"],
        )
        self.assertTrue(result["deleted"])
        self.assertNotIn(
            "creator_0003",
            {row["creator_id"] for row in self.factory().creator().getCreators(include_archived=True)},
        )
        intents = [
            json.loads(path.read_text(encoding="utf-8"))
            for path in (runtime / "feishu_delete_intents").glob("*.json")
        ]
        self.assertEqual(1, len(intents))
        self.assertEqual("pending_remote", intents[0]["status"])

    def test_thread_affine_concurrent_reads_and_bounded_write(self) -> None:
        def read_creator(_index: int) -> int:
            return len(self.factory().creator().getCreators(include_archived=True))

        with ThreadPoolExecutor(max_workers=8) as pool:
            counts = list(pool.map(read_creator, range(16)))
        self.assertEqual([4] * 16, counts)
        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = [pool.submit(read_creator, index) for index in range(6)]
            futures.append(pool.submit(
                lambda: self.factory().creator().updateCreator(
                    "creator_0000", {"language": "Portuguese-BR"}
                )
            ))
            results = [future.result(timeout=10) for future in futures]
        self.assertTrue(all(value == 4 for value in results[:-1]))

    def test_feishu_dry_run_reads_sqlite_and_fake_sync_converges(self) -> None:
        client = FakeClient()
        service = FeishuSyncService(self.factory().creator(), lambda: client)
        first = service.dry_run()
        self.assertEqual(4, first["creator_create_count"])
        self.assertEqual(6, first["account_create_count"])
        self.assertEqual(0, first["creator_conflict_count"])
        self.assertEqual(0, first["account_conflict_count"])
        completed = service.full_sync(confirm=True)
        self.assertEqual("success", completed["status"])
        second = service.dry_run()
        self.assertEqual(0, second["creator_create_count"])
        self.assertEqual(0, second["creator_update_count"])
        self.assertEqual(0, second["account_create_count"])
        self.assertEqual(0, second["account_update_count"])

    def test_assistant_reads_sqlite_only_through_allowlisted_tools(self) -> None:
        factory = self.factory()

        def search(arguments):
            filters = {
                key: arguments.get(key)
                for key in (
                    "country", "platform", "language", "content_category",
                    "search", "followers_min", "followers_max", "ai_tag",
                )
                if str(arguments.get(key) or "").strip()
            }
            return factory.creator().getCreatorsPage(
                page=1, page_size=100, filters=filters
            )["creators"]

        tools = {
            "search_creators": search,
            "get_creator_detail": factory.creator().getCreatorDetail,
            "list_campaigns": lambda arguments: factory.campaign().getCampaigns(
                status=str(arguments.get("status") or "")
            ),
            "get_campaign_detail": lambda campaign_id: {
                "campaign": factory.campaign().getCampaign(campaign_id),
                "campaign_creators": factory.campaign_creator().getCampaignCreators(
                    campaign_id=campaign_id
                ),
            },
            "get_task_status": lambda task_id: {"task": {"id": task_id}},
            "feishu_sync_dry_run": lambda: {"status": "ready"},
            "daily_summary": lambda: {"creator_total": 4},
            "create_capture_task": lambda _arguments: {},
            "feishu_full_sync": lambda: {},
        }
        service = AssistantService(
            MockAssistantProvider(AssistantIntent("search_creators", {"country": "Brazil"})),
            tools,
        )
        response = service.message("search", "sqlite-session", "trace-sqlite")
        self.assertTrue(response["ok"], response)
        self.assertEqual(2, response["data"]["total"])


if __name__ == "__main__":
    unittest.main()

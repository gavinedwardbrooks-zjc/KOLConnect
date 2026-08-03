from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import campaign_repository
from campaign_creator_repository import CAMPAIGN_CREATORS_HEADERS, CampaignCreatorRepository
from campaign_repository import CAMPAIGNS_HEADERS, CampaignRepository
from product_repository import PRODUCTS_HEADERS, ProductRepository


TARGET_SHEETS = {
    "Products": PRODUCTS_HEADERS,
    "Campaigns": CAMPAIGNS_HEADERS,
    "CampaignCreators": CAMPAIGN_CREATORS_HEADERS,
}


def append_mapping(sheet, headers: list[str], values: dict) -> None:
    sheet.append([values.get(header, "") for header in headers])


class ProductCampaignDataFoundationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        self.log_patch = mock.patch.object(creator_repository, "log_event")
        self.log_patch.start()

    def tearDown(self) -> None:
        self.log_patch.stop()
        self.temp_dir.cleanup()

    def initialize_workbook(self) -> None:
        creator_repository.CreatorRepository(self.workbook_path).getCreators()

    def seed_creator(self, *, creator_id: str = "creator_one", account_id: str = "account_one") -> None:
        self.initialize_workbook()
        workbook = load_workbook(self.workbook_path)
        append_mapping(
            workbook["Creators"],
            creator_repository._CREATORS_HEADERS,
            {
                "creator_id": creator_id,
                "name": "Original Creator",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@original",
                "status": "discovered",
                "created_at": "2026-07-01T00:00:00Z",
                "updated_at": "2026-07-01T00:00:00Z",
            },
        )
        append_mapping(
            workbook["CreatorAccounts"],
            creator_repository._CREATOR_ACCOUNTS_HEADERS,
            {
                "account_id": account_id,
                "creator_id": creator_id,
                "account_uid": "tiktok|https://www.tiktok.com/@original",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@original",
                "created_at": "2026-07-01T00:00:00Z",
                "updated_at": "2026-07-01T00:00:00Z",
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

    def create_product_and_campaign(self) -> tuple[dict, dict]:
        product = ProductRepository(self.workbook_path).createProduct(
            {"name": "Block Game", "company_name": "Example Studio"}
        )
        campaign = CampaignRepository(self.workbook_path).createCampaign(
            {
                "product_id": product["product_id"],
                "name": "Brazil TikTok Launch",
                "platform": "TikTok",
                "budget": 1000,
            }
        )
        return product, campaign

    def test_empty_workbook_is_created_with_product_campaign_sheets(self) -> None:
        self.assertEqual([], ProductRepository(self.workbook_path).getProducts())
        self.assertEqual([], CampaignRepository(self.workbook_path).getCampaigns())
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            self.assertTrue(set(TARGET_SHEETS).issubset(workbook.sheetnames))
            for sheet_name, headers in TARGET_SHEETS.items():
                self.assertEqual(headers, [cell.value for cell in workbook[sheet_name][1]])
        finally:
            workbook.close()

    def test_campaign_with_missing_product_remains_readable_and_logs_warning(self) -> None:
        self.initialize_workbook()
        workbook = load_workbook(self.workbook_path)
        append_mapping(
            workbook["Campaigns"],
            CAMPAIGNS_HEADERS,
            {
                "campaign_id": "campaign_orphan",
                "product_id": "product_missing",
                "name": "Orphan Campaign",
                "status": "draft",
                "created_at": "2026-08-03T00:00:00Z",
                "updated_at": "2026-08-03T00:00:00Z",
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

        with mock.patch.object(campaign_repository, "log_event") as warning_log:
            campaigns = CampaignRepository(self.workbook_path).getCampaigns()

        self.assertEqual(1, len(campaigns))
        self.assertEqual("", campaigns[0]["product_name"])
        warning_log.assert_called_once()
        self.assertIn("Campaign 关联 Product 缺失", warning_log.call_args.args[1])

    def test_existing_workbook_upgrade_preserves_creator_account_and_legacy_cooperation(self) -> None:
        self.seed_creator()
        workbook = load_workbook(self.workbook_path)
        for sheet_name in TARGET_SHEETS:
            del workbook[sheet_name]
        append_mapping(
            workbook["Cooperations"],
            creator_repository._COOPERATIONS_HEADERS,
            {
                "cooperation_id": "cooperation_legacy",
                "creator_id": "creator_one",
                "campaign": "Legacy Campaign",
                "price": 100,
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

        baseline = load_workbook(self.workbook_path, read_only=True)
        try:
            creator_before = list(baseline["Creators"].values)
            account_before = list(baseline["CreatorAccounts"].values)
            cooperation_before = list(baseline["Cooperations"].values)
        finally:
            baseline.close()

        ProductRepository(self.workbook_path).getProducts()

        upgraded = load_workbook(self.workbook_path, read_only=True)
        try:
            self.assertTrue(set(TARGET_SHEETS).issubset(upgraded.sheetnames))
            self.assertEqual(creator_before, list(upgraded["Creators"].values))
            self.assertEqual(account_before, list(upgraded["CreatorAccounts"].values))
            self.assertEqual(cooperation_before, list(upgraded["Cooperations"].values))
        finally:
            upgraded.close()

    def test_repeated_startup_is_idempotent(self) -> None:
        self.initialize_workbook()
        first_bytes = self.workbook_path.read_bytes()
        ProductRepository(self.workbook_path).getProducts()
        CampaignRepository(self.workbook_path).getCampaigns()
        CampaignCreatorRepository(self.workbook_path).getCampaignCreators()
        second_bytes = self.workbook_path.read_bytes()
        self.assertEqual(first_bytes, second_bytes)

    def test_product_crud(self) -> None:
        repository = ProductRepository(self.workbook_path)
        product = repository.createProduct({"name": "Product A", "company_name": "Company A"})
        self.assertEqual(product, repository.getProduct(product["product_id"]))
        updated = repository.updateProduct(product["product_id"], {"name": "Product B", "note": "Updated"})
        self.assertEqual("Product B", updated["name"])
        self.assertEqual("Updated", repository.getProduct(product["product_id"])["note"])
        repository.deleteProduct(product["product_id"])
        self.assertEqual([], repository.getProducts())

    def test_campaign_crud_and_delete_protection(self) -> None:
        product, campaign = self.create_product_and_campaign()
        repository = CampaignRepository(self.workbook_path)
        updated = repository.updateCampaign(campaign["campaign_id"], {"status": "sourcing", "budget": 1500})
        self.assertEqual("sourcing", updated["status"])
        self.assertEqual(1500, repository.getCampaign(campaign["campaign_id"])["budget"])
        with self.assertRaisesRegex(ValueError, "不能归档"):
            ProductRepository(self.workbook_path).deleteProduct(product["product_id"])
        repository.deleteCampaign(campaign["campaign_id"])
        ProductRepository(self.workbook_path).deleteProduct(product["product_id"])

    def test_campaign_creator_crud_uniqueness_and_parent_delete_protection(self) -> None:
        self.seed_creator()
        _product, campaign = self.create_product_and_campaign()
        repository = CampaignCreatorRepository(self.workbook_path)
        record = repository.createCampaignCreator(
            {
                "campaign_id": campaign["campaign_id"],
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "contacted",
                "publish_links": ["https://example.com/post/1"],
            }
        )
        with self.assertRaisesRegex(ValueError, "已加入"):
            repository.createCampaignCreator(
                {
                    "campaign_id": campaign["campaign_id"],
                    "creator_id": "creator_one",
                    "account_id": "account_one",
                }
            )
        updated = repository.updateCampaignCreator(record["id"], {"stage": "quoted", "creator_quote": 200})
        self.assertEqual("quoted", updated["stage"])
        CampaignRepository(self.workbook_path).deleteCampaign(campaign["campaign_id"])
        archived_campaign = CampaignRepository(self.workbook_path).getCampaign(campaign["campaign_id"])
        self.assertEqual("draft", archived_campaign["status"])
        self.assertTrue(archived_campaign["archived_at"])
        repository.deleteCampaignCreator(record["id"])
        self.assertEqual([], repository.getCampaignCreators(campaign_id=campaign["campaign_id"]))

    def test_legacy_archived_status_is_preserved_for_manual_confirmation(self) -> None:
        _product, campaign = self.create_product_and_campaign()
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Campaigns"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        archived_at_column = headers.index("archived_at") + 1
        status_column = headers.index("status") + 1
        campaign_id_column = headers.index("campaign_id") + 1
        row_index = next(
            row
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, campaign_id_column).value == campaign["campaign_id"]
        )
        sheet.cell(row_index, status_column, "archived")
        sheet.delete_cols(archived_at_column)
        workbook.save(self.workbook_path)
        workbook.close()

        storage = creator_repository.CreatorRepository(self.workbook_path)
        storage.getCreators()
        self.assertEqual(1, storage.last_campaign_lifecycle_report["count"])
        review_item = storage.last_campaign_lifecycle_report["manual_review_required"][0]
        self.assertEqual(campaign["campaign_id"], review_item["campaign_id"])
        self.assertEqual("archived", review_item["status"])

        repository = CampaignRepository(self.workbook_path)
        self.assertEqual([], repository.getCampaigns())
        legacy_campaign = repository.getCampaigns(include_archived=True)[0]
        self.assertEqual("archived", legacy_campaign["status"])
        self.assertTrue(legacy_campaign["archived_at"])
        with self.assertRaisesRegex(ValueError, "人工确认"):
            repository.setCampaignArchivedAt(campaign["campaign_id"], None)

    def test_invalid_foreign_keys_are_rejected_without_auto_creation(self) -> None:
        self.seed_creator()
        with self.assertRaisesRegex(ValueError, "产品不存在"):
            CampaignRepository(self.workbook_path).createCampaign(
                {"product_id": "product_missing", "name": "Invalid"}
            )
        _product, campaign = self.create_product_and_campaign()
        repository = CampaignCreatorRepository(self.workbook_path)
        invalid_payloads = [
            {"campaign_id": "campaign_missing", "creator_id": "creator_one", "account_id": "account_one"},
            {"campaign_id": campaign["campaign_id"], "creator_id": "creator_missing", "account_id": "account_one"},
            {"campaign_id": campaign["campaign_id"], "creator_id": "creator_one", "account_id": "account_missing"},
        ]
        for payload in invalid_payloads:
            with self.subTest(payload=payload), self.assertRaises(ValueError):
                repository.createCampaignCreator(payload)
        self.assertEqual([], repository.getCampaignCreators())

    def test_product_campaign_operations_do_not_change_existing_creator_data(self) -> None:
        self.seed_creator()
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            creator_before = list(workbook["Creators"].values)
            account_before = list(workbook["CreatorAccounts"].values)
        finally:
            workbook.close()
        _product, campaign = self.create_product_and_campaign()
        CampaignCreatorRepository(self.workbook_path).createCampaignCreator(
            {
                "campaign_id": campaign["campaign_id"],
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "contacted",
            }
        )
        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            self.assertEqual(creator_before, list(workbook["Creators"].values))
            self.assertEqual(account_before, list(workbook["CreatorAccounts"].values))
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()

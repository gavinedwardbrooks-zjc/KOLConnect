"""Focused coverage for M3.4 task-result propagation and T3 batch writes."""

from __future__ import annotations

from contextlib import nullcontext
from pathlib import Path
from unittest import mock
import sys
import unittest

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository as creator_repository_module  # noqa: E402
import scraper  # noqa: E402
from creator_repository import CreatorRepository  # noqa: E402
from services.creator_service import CreatorService  # noqa: E402


class BatchFieldPropagationTests(unittest.TestCase):
    def test_platform_results_share_the_existing_task_import_normalization(self) -> None:
        task = {"task_type": "scrape", "finished_at": "2026-08-21T00:00:00Z"}
        rows = []
        for platform, url, name in (
            ("Instagram", "https://www.instagram.com/creator/", "Instagram Creator"),
            ("YouTube", "https://www.youtube.com/@creator", "YouTube Creator"),
            ("TikTok", "https://www.tiktok.com/@creator", "TikTok Creator"),
        ):
            rows.append(
                scraper.result_to_row(
                    scraper.build_result(
                        url=url,
                        platform=platform,
                        name=name,
                        emails=[f"{platform.lower()}@example.com"],
                        whatsapp="+5511999999999",
                        follower_count="12.5K",
                        latest_publish_date="2026-08-20",
                        last_scrape_time="2026-08-21T00:00:00Z",
                    )
                )
            )

        records = CreatorService._task_rows_for_creator_library(task, tuple(rows))

        self.assertEqual(["Instagram", "YouTube", "TikTok"], [row["platform"] for row in records])
        for record, expected in zip(records, ("Instagram Creator", "YouTube Creator", "TikTok Creator")):
            self.assertTrue(record["account_uid"])
            self.assertEqual(expected, record["creator_name"])
            self.assertEqual("12.5K", record["followers"])
            self.assertEqual("+5511999999999", record["whatsapp"])
            self.assertEqual("2026-08-20", record["latest_post_date"])
            self.assertEqual("2026-08-21T00:00:00Z", record["last_scrape_time"])

        tiktok_record = records[-1]
        self.assertNotIn("views", tiktok_record)
        self.assertNotIn("likes", tiktok_record)
        self.assertNotIn("comments", tiktok_record)


class BatchWorkbookWriteTests(unittest.TestCase):
    def setUp(self) -> None:
        self.lock_patches = [
            mock.patch.object(
                creator_repository_module,
                "shared_storage_lock",
                side_effect=lambda *args, **kwargs: nullcontext(),
            ),
        ]
        for patcher in self.lock_patches:
            patcher.start()
        self.log_patcher = mock.patch.object(creator_repository_module, "log_event")
        self.log_patcher.start()
        self.repository = CreatorRepository(ROOT / "unused_m3_4_t3.xlsx")
        self.workbook = Workbook()
        self.repository._apply_schema_migrations(self.workbook, created=True)

    def tearDown(self) -> None:
        self.log_patcher.stop()
        for patcher in reversed(self.lock_patches):
            patcher.stop()

    def test_fifty_new_rows_use_one_load_one_save_and_append_rows(self) -> None:
        records = [
            {
                "account_uid": f"TikTok|https://www.tiktok.com/@creator_{index}",
                "platform": "TikTok",
                "profile_url": f"https://www.tiktok.com/@creator_{index}",
                "name": f"Creator {index}",
            }
            for index in range(50)
        ]
        with (
            mock.patch.object(self.repository, "_load_workbook", return_value=self.workbook) as load,
            mock.patch.object(self.repository, "_save_workbook") as save,
            mock.patch.object(self.repository, "_append_row", wraps=self.repository._append_row) as append,
        ):
            result = self.repository.createCreatorsBatch(records)

        self.assertEqual({"created": 50, "skipped_existing": 0}, result)
        self.assertEqual(1, load.call_count)
        self.assertEqual(1, save.call_count)
        self.assertEqual(100, append.call_count)

    def test_existing_identity_is_skipped_without_duplicate_append(self) -> None:
        existing = {
            "account_uid": "Instagram|https://www.instagram.com/existing/",
            "platform": "Instagram",
            "profile_url": "https://www.instagram.com/existing/",
            "name": "Existing",
        }
        with (
            mock.patch.object(self.repository, "_load_workbook", return_value=self.workbook),
            mock.patch.object(self.repository, "_save_workbook"),
        ):
            self.repository.createCreatorsBatch([existing])
            result = self.repository.createCreatorsBatch(
                [
                    existing,
                    {
                        "account_uid": "YouTube|https://www.youtube.com/@new",
                        "platform": "YouTube",
                        "profile_url": "https://www.youtube.com/@new",
                        "name": "New",
                    },
                ]
            )

        self.assertEqual({"created": 1, "skipped_existing": 1}, result)
        self.assertEqual(2, len(self.repository._rows(self.workbook["Creators"])))

    def test_orphaned_existing_account_is_updated_not_appended(self) -> None:
        record = {
            "account_uid": "TikTok|https://www.tiktok.com/@recovered",
            "platform": "TikTok",
            "profile_url": "https://www.tiktok.com/@recovered",
            "name": "Recovered",
        }
        self.repository._append_row(
            self.workbook["CreatorAccounts"],
            {"account_uid": record["account_uid"], "account_id": "legacy_orphan"},
        )
        with (
            mock.patch.object(self.repository, "_load_workbook", return_value=self.workbook),
            mock.patch.object(self.repository, "_save_workbook"),
        ):
            result = self.repository.createCreatorsBatch([record])

        self.assertEqual({"created": 1, "skipped_existing": 0}, result)
        accounts = self.repository._rows(self.workbook["CreatorAccounts"])
        self.assertEqual(1, len(accounts))
        self.assertTrue(accounts[0]["creator_id"])


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import hashlib
import json
import shutil
import sys
import time
import unittest
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))

from creator_repository import _WORKBOOK_SHEETS
from runtime_paths import atomic_write_json, load_json_with_backup
from storage.connection import SQLiteConnectionFactory
from storage.errors import (
    SQLiteActivationError,
    SQLiteMigrationAmbiguousIdentityError,
    SQLiteMigrationError,
)
from storage.migration import (
    ExcelToSQLiteMigrator,
    resolve_authority,
    semantic_digest,
    validate_source_workbook,
)
from storage.paths import SQLiteStoragePaths
from storage.schema import apply_schema_migrations


def append_row(sheet, values: dict[str, object]) -> None:
    headers = [str(cell.value or "") for cell in sheet[1]]
    sheet.append([values.get(header, "") for header in headers])


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build_fixture(
    path: Path,
    *,
    creator_count: int = 4,
    include_campaign: bool = True,
    snapshot_count: int = 1,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers in _WORKBOOK_SHEETS.items():
        sheet = workbook.create_sheet(name)
        sheet.append(list(headers))
    accounts: list[dict[str, object]] = []
    for index in range(creator_count):
        creator_id = f"creator_{index:04d}"
        append_row(workbook["Creators"], {
            "creator_id": creator_id,
            "name": f"Creator {index}",
            "platform": "TikTok",
            "country": "Brazil" if index % 2 == 0 else "",
            "language": "Portuguese",
            "content_category": "Gaming",
            "followers": 1000 + index,
            "status": "discovered",
            "tags": json.dumps(["priority", f"group-{index % 2}"]),
            "created_at": "2026-01-01T00:00:00Z",
            "updated_at": "2026-01-02T00:00:00Z",
            "archived_at": "2026-02-01T00:00:00Z" if index == creator_count - 1 else "",
        })
        account_platforms = ("TikTok", "YouTube", "Instagram") if index == 0 else ("TikTok",)
        for position, platform in enumerate(account_platforms):
            account_uid = f"account_{index:04d}_{position}"
            account = {
                "account_id": f"legacy_{index}_{position}",
                "creator_id": creator_id,
                "account_uid": account_uid,
                "platform": platform,
                "username": f"creator{index}_{position}",
                "profile_url": f"https://example.test/{index}/{position}",
                "followers": 1000 + index + position,
                "created_at": "2026-01-01T00:00:00Z",
                "updated_at": "2026-01-02T00:00:00Z",
            }
            accounts.append(account)
            append_row(workbook["CreatorAccounts"], account)
    if creator_count:
        append_row(workbook["Videos"], {
            "creator_id": "creator_0000", "video_url": "https://video.test/1",
            "views": 5000, "likes": 500, "comments": 50, "captured_at": "2026-01-03T00:00:00Z",
        })
        append_row(workbook["Insights"], {
            "creator_id": "creator_0000", "average_views": 5000, "median_views": 4900,
            "stability": 0.8, "risks": "", "recommendation": "fit",
        })
        for index in range(snapshot_count):
            snapshot_id = f"snapshot_{index:05d}"
            append_row(workbook["CreatorSnapshots"], {
                "snapshot_id": snapshot_id, "creator_id": "creator_0000", "platform": "TikTok",
                "account_uid": "account_0000_0", "followers": 1000 + index,
                "average_views": 5000, "median_views": 4900, "video_count": 1,
                "creator_score": 80.5, "captured_at": f"2026-01-{(index % 28) + 1:02d}T00:00:00Z",
                "source": "test",
            })
            append_row(workbook["VideoSnapshots"], {
                "video_snapshot_id": f"video_snapshot_{index:05d}", "snapshot_id": snapshot_id,
                "creator_id": "creator_0000", "video_id": "video_1",
                "video_url": "https://video.test/1", "platform": "TikTok", "views": 5000 + index,
                "likes": 500, "comments": 50, "captured_at": f"2026-01-{(index % 28) + 1:02d}T00:00:00Z",
            })
        append_row(workbook["Cooperations"], {
            "cooperation_id": "coop_1", "creator_id": "creator_0000", "campaign": "Legacy",
            "platform": "TikTok", "price": 100, "created_at": "2026-01-01T00:00:00Z",
        })
        append_row(workbook["Agencies"], {
            "agency_id": "agency_1", "name": "Agency", "country": "Brazil",
            "resource_files": "[]", "created_at": "2026-01-01T00:00:00Z",
        })
        append_row(workbook["AgencyContacts"], {
            "contact_id": "contact_1", "name": "Contact", "agency_id": "agency_1",
            "email": "public@example.test", "created_at": "2026-01-01T00:00:00Z",
        })
        append_row(workbook["FollowUpLogs"], {
            "follow_up_id": "follow_1", "object_type": "creator", "object_id": "creator_0000",
            "content": "Follow up", "created_at": "2026-01-01T00:00:00Z",
        })
        append_row(workbook["_AnalysisData"], {
            "creator_id": "creator_0000", "task_id": "task_1", "account_uid": "account_0000_0",
            "status_updated_at": "2026-01-01T00:00:00Z", "analysis_json": "{}", "source": "test",
        })
    if include_campaign and creator_count:
        append_row(workbook["Products"], {
            "product_id": "product_1", "name": "Product", "created_at": "2026-01-01T00:00:00Z",
        })
        append_row(workbook["Campaigns"], {
            "campaign_id": "campaign_1", "product_id": "product_1", "name": "Campaign",
            "country": "Brazil", "platform": "TikTok",
            "platforms": json.dumps(["TikTok", "YouTube"]), "status": "running",
            "budget": 5000, "created_at": "2026-01-01T00:00:00Z",
        })
        for index in range(min(3, creator_count)):
            selected = [f"account_{index:04d}_0"]
            if index == 0:
                selected.append("account_0000_1")
            append_row(workbook["CampaignCreators"], {
                "id": f"relation_{index}", "campaign_id": "campaign_1",
                "creator_id": f"creator_{index:04d}", "account_id": selected[0],
                "account_ids": json.dumps(selected), "stage": "completed",
                "creator_quote": "" if index else 500, "cost": "" if index else 450,
                "publish_links": json.dumps([f"https://publish.test/{index}"]),
                "publish_date": "2026-02-01",
                "planned_publish_dates": json.dumps(["2026-02-01", "2026-02-05"]),
                "views": 1000, "likes": 100, "comments": 10, "roi": 1.5,
                "created_at": "2026-01-01T00:00:00Z",
            })
    append_row(workbook["_Metadata"], {
        "schema_version": "2.0-product-campaign-phase2-api",
        "last_update_time": "2026-01-01T00:00:00Z",
    })
    workbook.save(path)
    workbook.close()


class ExcelSQLiteMigrationTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime = ROOT / ".pre_m8_c0_c3_test_runtime"
        runtime.mkdir(exist_ok=True)
        self.root = runtime / f"migration_{uuid4().hex}"
        self.root.mkdir()
        self.lock_patch = patch(
            "local_storage_lock.get_shared_storage_lock_path",
            return_value=self.root / "locks" / "shared_storage.lock",
        )
        self.lock_patch.start()
        self.paths = SQLiteStoragePaths.for_app_data(self.root / "appdata")
        self.workbook = self.root / "source.xlsx"
        build_fixture(self.workbook)

    def tearDown(self) -> None:
        self.lock_patch.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    def test_current_shape_migrates_all_entities_and_preserves_source(self) -> None:
        before = file_hash(self.workbook)
        result = ExcelToSQLiteMigrator(self.paths).migrate(self.workbook)
        self.assertEqual(before, result.source_sha256_before)
        self.assertEqual(before, result.source_sha256_after)
        self.assertEqual(before, file_hash(self.workbook))
        self.assertEqual(before, file_hash(result.backup_path))
        self.assertEqual(4, result.counts["creators"])
        self.assertEqual(6, result.counts["creator_accounts"])
        self.assertEqual(3, result.counts["campaign_creators"])
        self.assertEqual(4, result.counts["campaign_creator_accounts"])
        self.assertEqual(6, result.counts["campaign_creator_planned_dates"])
        factory = SQLiteConnectionFactory(result.staged_database_path)
        with factory.read_connection() as connection:
            self.assertEqual(result.semantic_digest, semantic_digest(connection))
            self.assertEqual(1, connection.execute(
                "SELECT COUNT(*) FROM creators WHERE archived_at IS NOT NULL"
            ).fetchone()[0])
            self.assertIsNone(connection.execute(
                "SELECT creator_quote FROM campaign_creators WHERE id='relation_1'"
            ).fetchone()[0])

    def test_empty_and_older_supported_shape_migrate(self) -> None:
        empty = self.root / "empty.xlsx"
        build_fixture(empty, creator_count=0, include_campaign=False, snapshot_count=0)
        result = ExcelToSQLiteMigrator(
            SQLiteStoragePaths.for_app_data(self.root / "empty_appdata")
        ).migrate(empty)
        self.assertEqual(0, result.counts["creators"])
        older = self.root / "older.xlsx"
        build_fixture(older, creator_count=1, include_campaign=False)
        workbook = load_workbook(older)
        sheet = workbook["Creators"]
        headers = [cell.value for cell in sheet[1]]
        for field in ("archived_at", "bio"):
            column = headers.index(field) + 1
            sheet.delete_cols(column)
            headers.pop(column - 1)
        workbook.save(older)
        workbook.close()
        older_result = ExcelToSQLiteMigrator(
            SQLiteStoragePaths.for_app_data(self.root / "older_appdata")
        ).migrate(older)
        self.assertEqual(1, older_result.counts["creators"])

    def test_duplicate_creator_and_account_identity_fail_closed(self) -> None:
        for sheet_name, key in (("Creators", "creator_id"), ("CreatorAccounts", "account_uid")):
            with self.subTest(sheet=sheet_name):
                path = self.root / f"duplicate_{sheet_name}.xlsx"
                build_fixture(path)
                workbook = load_workbook(path)
                sheet = workbook[sheet_name]
                headers = [str(cell.value or "") for cell in sheet[1]]
                values = [sheet.cell(2, index + 1).value for index in range(len(headers))]
                self.assertTrue(values[headers.index(key)])
                sheet.append(values)
                workbook.save(path)
                workbook.close()
                with self.assertRaises(SQLiteMigrationAmbiguousIdentityError):
                    ExcelToSQLiteMigrator(
                        SQLiteStoragePaths.for_app_data(self.root / f"app_{sheet_name}")
                    ).migrate(path)

    def test_orphan_references_fail_closed(self) -> None:
        cases = (
            ("CreatorAccounts", "creator_id", "missing_creator"),
            ("CampaignCreators", "creator_id", "missing_creator"),
            ("CreatorSnapshots", "creator_id", "missing_creator"),
            ("VideoSnapshots", "snapshot_id", "missing_snapshot"),
        )
        for index, (sheet_name, field, value) in enumerate(cases):
            with self.subTest(sheet=sheet_name, field=field):
                path = self.root / f"orphan_{index}.xlsx"
                build_fixture(path)
                workbook = load_workbook(path)
                sheet = workbook[sheet_name]
                headers = [str(cell.value or "") for cell in sheet[1]]
                sheet.cell(2, headers.index(field) + 1, value)
                workbook.save(path)
                workbook.close()
                with self.assertRaises(SQLiteMigrationError):
                    ExcelToSQLiteMigrator(
                        SQLiteStoragePaths.for_app_data(self.root / f"orphan_app_{index}")
                    ).migrate(path)

    def test_malformed_identity_relation_fails_closed(self) -> None:
        workbook = load_workbook(self.workbook)
        sheet = workbook["CampaignCreators"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.cell(2, headers.index("account_ids") + 1, "[invalid")
        workbook.save(self.workbook)
        workbook.close()
        with self.assertRaises(SQLiteMigrationError):
            ExcelToSQLiteMigrator(self.paths).migrate(self.workbook)

    def test_unknown_empty_placeholders_are_allowed_but_nonempty_data_is_rejected(self) -> None:
        workbook = load_workbook(self.workbook)
        creators = workbook["Creators"]
        creators.cell(1, creators.max_column + 1, "legacy_empty_column")
        workbook.create_sheet("LegacyEmpty").append(["legacy_header"])
        workbook.save(self.workbook)
        workbook.close()
        validate_source_workbook(self.workbook)

        workbook = load_workbook(self.workbook)
        creators = workbook["Creators"]
        headers = [str(cell.value or "") for cell in creators[1]]
        creators.cell(2, headers.index("legacy_empty_column") + 1, "unsupported-value")
        workbook.save(self.workbook)
        workbook.close()
        with self.assertRaisesRegex(SQLiteMigrationError, "Unsupported columns"):
            validate_source_workbook(self.workbook)

        workbook = load_workbook(self.workbook)
        creators = workbook["Creators"]
        creators.cell(2, headers.index("legacy_empty_column") + 1, None)
        workbook["LegacyEmpty"].append(["unsupported-value"])
        workbook.save(self.workbook)
        workbook.close()
        with self.assertRaisesRegex(SQLiteMigrationError, "Unsupported non-empty sheet"):
            validate_source_workbook(self.workbook)

    def test_interrupted_phases_never_activate_authority(self) -> None:
        phases = (
            "source_validated", "backup_created", "schema_created",
            "entity_imported:creators", "data_imported", "validated", "ready_for_activation",
        )
        for index, phase in enumerate(phases):
            with self.subTest(phase=phase):
                paths = SQLiteStoragePaths.for_app_data(self.root / f"interrupt_{index}")

                def inject(current: str) -> None:
                    if current == phase:
                        raise RuntimeError(f"stop:{phase}")

                with self.assertRaises(RuntimeError):
                    ExcelToSQLiteMigrator(paths, failure_injector=inject).migrate(self.workbook)
                self.assertFalse(paths.authority_marker_path.exists())
                self.assertEqual("legacy_excel", resolve_authority(paths))
                self.assertEqual(file_hash(self.workbook), validate_source_workbook(self.workbook)["source_sha256"])

    def test_synthetic_activation_interruption_recovers_deterministically(self) -> None:
        result = ExcelToSQLiteMigrator(self.paths).migrate(self.workbook)
        migrator = ExcelToSQLiteMigrator(self.paths)
        with self.assertRaises(SQLiteActivationError):
            migrator.activate_synthetic(result, inject_after_database_activation=True)
        self.assertEqual("legacy_excel", resolve_authority(self.paths))
        final_path = migrator.recover_synthetic_activation(result.migration_id)
        self.assertEqual(self.paths.database_path, final_path)
        self.assertEqual("sqlite_active", resolve_authority(self.paths))
        manifest, _source = load_json_with_backup(result.manifest_path)
        self.assertEqual("completed", manifest["phase"])

    def test_both_files_do_not_override_explicit_legacy_authority(self) -> None:
        self.paths.ensure_migration_directories()
        self.paths.database_path.write_bytes(b"not selected")
        atomic_write_json(self.paths.authority_marker_path, {
            "authority": "legacy_excel", "schema_version": 0,
        })
        self.assertEqual("legacy_excel", resolve_authority(self.paths))

    def test_newer_marker_is_unsupported_and_not_downgraded(self) -> None:
        self.paths.ensure_migration_directories()
        self.paths.database_path.write_bytes(b"newer")
        atomic_write_json(self.paths.authority_marker_path, {
            "authority": "sqlite", "schema_version": 999,
        })
        self.assertEqual("unsupported_schema", resolve_authority(self.paths))
        self.assertEqual(b"newer", self.paths.database_path.read_bytes())

    def test_sqlite_marker_never_activates_a_corrupt_database(self) -> None:
        self.paths.ensure_migration_directories()
        self.paths.database_path.write_bytes(b"not-a-sqlite-database")
        atomic_write_json(self.paths.authority_marker_path, {
            "authority": "sqlite", "schema_version": 1,
        })
        self.assertEqual("migration_error", resolve_authority(self.paths))

    def test_medium_fixture_migration_and_indexed_lookup_smoke(self) -> None:
        medium = self.root / "medium.xlsx"
        build_fixture(medium, creator_count=500, snapshot_count=2500)
        paths = SQLiteStoragePaths.for_app_data(self.root / "medium_appdata")
        started = time.perf_counter()
        result = ExcelToSQLiteMigrator(paths).migrate(medium)
        elapsed = time.perf_counter() - started
        factory = SQLiteConnectionFactory(result.staged_database_path)
        with factory.read_connection() as connection:
            creator = connection.execute(
                "SELECT creator_id FROM creators WHERE creator_id=?", ("creator_0499",)
            ).fetchone()
            campaign = connection.execute(
                "SELECT campaign_id FROM campaigns WHERE campaign_id=?", ("campaign_1",)
            ).fetchone()
            creator_plan = connection.execute(
                "EXPLAIN QUERY PLAN SELECT * FROM creator_accounts WHERE creator_id=?",
                ("creator_0000",),
            ).fetchall()
        self.assertEqual("creator_0499", creator[0])
        self.assertEqual("campaign_1", campaign[0])
        self.assertTrue(
            any("idx_creator_accounts_creator" in str(row[3]) for row in creator_plan)
        )
        self.assertLess(elapsed, 30.0)

    def test_source_validation_rejects_missing_file(self) -> None:
        with self.assertRaises(SQLiteMigrationError):
            validate_source_workbook(self.root / "missing.xlsx")


if __name__ == "__main__":
    unittest.main()

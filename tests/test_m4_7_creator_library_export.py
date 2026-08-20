from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

from creator_batch_import import parse_creator_import_workbook  # noqa: E402
from http_handlers import creator_handler  # noqa: E402
from services.creator_service import CreatorService  # noqa: E402


class SnapshotRepository:
    workbook_path = ROOT / "Creator_Library.xlsx"

    def __init__(self) -> None:
        self.records = {
            "creator_one": {
                "creator_id": "creator_one",
                "creator_name": "Creator One",
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@one",
                "country": "US",
                "language": "English",
                "content_category": "Gaming",
                "agency_id": "agency_one",
                "agency_name": "North Studio",
                "followers": "1200",
                "insight_level": "A",
                "status": "contacted",
                "created_at": "2026-08-20T00:00:00Z",
                "email": "private@example.com",
                "whatsapp": "+15551234567",
                "note": "private note",
            },
            "creator_two": {
                "creator_id": "creator_two",
                "creator_name": "Creator Two",
                "platform": "Instagram",
                "profile_url": "https://www.instagram.com/two",
                "country": "BR",
                "language": "Portuguese",
                "content_category": "Lifestyle",
                "agency_id": "agency_two",
                "agency_name": "South Studio",
                "followers": "3400",
                "insight_level": "B",
                "status": "discovered",
                "created_at": "2026-08-19T00:00:00Z",
            },
        }

    def getCreatorLibrarySnapshot(self) -> dict:
        return {"creator_id_index": self.records}


class FakeBinaryHandler:
    def __init__(self) -> None:
        self.binary = None
        self.json = None

    def _binary(self, data, content_type, filename):
        self.binary = (data, content_type, filename)

    def _json(self, payload, status=200):
        self.json = (payload, status)


class CreatorLibraryExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.repository = SnapshotRepository()
        self.service = CreatorService(lambda: self.repository, lambda: None)

    def _rows(self, creator_ids: list[str]) -> list[tuple]:
        payload = self.service.export_creators(creator_ids)
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        try:
            return list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()

    def test_single_creator_export_has_frozen_headers_and_no_private_fields(self) -> None:
        rows = self._rows(["creator_one"])
        self.assertEqual(
            (
                "platform", "profile_url", "name", "country", "language", "content_category",
                "agency_id", "followers", "insight_level", "status", "created_at",
            ),
            rows[0],
        )
        self.assertEqual(
            (
                "TikTok", "https://www.tiktok.com/@one", "Creator One", "US", "English", "Gaming",
                "agency_one", "1200", "A", "contacted", "2026-08-20T00:00:00Z",
            ),
            rows[1],
        )
        self.assertNotIn("private@example.com", rows[1])
        self.assertNotIn("+15551234567", rows[1])
        self.assertNotIn("private note", rows[1])

    def test_multiple_creator_export_preserves_requested_selection_order(self) -> None:
        rows = self._rows(["creator_two", "creator_one"])
        self.assertEqual(["Creator Two", "Creator One"], [row[2] for row in rows[1:]])

    def test_empty_or_missing_selection_is_rejected(self) -> None:
        for creator_ids in (None, [], "creator_one"):
            with self.subTest(creator_ids=creator_ids):
                with self.assertRaisesRegex(ValueError, "CREATOR_IDS_REQUIRED"):
                    self.service.export_creators(creator_ids)

    def test_unknown_creator_rejects_the_whole_export(self) -> None:
        with self.assertRaisesRegex(LookupError, "CREATOR_NOT_FOUND"):
            self.service.export_creators(["creator_one", "missing"])

    def test_exported_workbook_is_accepted_by_the_existing_import_parser(self) -> None:
        parsed = parse_creator_import_workbook(self.service.export_creators(["creator_one"]))
        self.assertEqual(1, len(parsed))
        self.assertEqual(
            {
                "platform": "TikTok",
                "profile_url": "https://www.tiktok.com/@one",
                "name": "Creator One",
                "country": "US",
                "language": "English",
                "content_category": "Gaming",
                "agency_id": "agency_one",
            },
            {key: parsed[0].values[key] for key in (
                "platform", "profile_url", "name", "country", "language", "content_category", "agency_id",
            )},
        )
        self.assertNotIn("followers", parsed[0].values)
        self.assertNotIn("status", parsed[0].values)

    def test_handler_returns_xlsx_and_contract_errors(self) -> None:
        context = {
            "services": {
                "creator": self.service,
                "agency": object(),
                "creator_delete_impact": object(),
                "creator_hard_delete": object(),
            },
            "config": {"legacy_cooperation_pattern": __import__("re").compile("$" )},
        }
        handler = FakeBinaryHandler()
        request = {
            "method": "POST",
            "path": "/api/creator-library/export",
            "query": {},
            "get_payload": lambda: {"creator_ids": ["creator_one"]},
        }
        self.assertTrue(creator_handler.handle(handler, request, context))
        self.assertEqual(creator_handler.XLSX_CONTENT_TYPE, handler.binary[1])
        self.assertEqual("KOLConnect_Creator_Export.xlsx", handler.binary[2])

        missing = FakeBinaryHandler()
        request["get_payload"] = lambda: {"creator_ids": []}
        creator_handler.handle(missing, request, context)
        self.assertEqual(({"ok": False, "error": "CREATOR_IDS_REQUIRED"}, 400), missing.json)

        absent = FakeBinaryHandler()
        request["get_payload"] = lambda: {}
        creator_handler.handle(absent, request, context)
        self.assertEqual(({"ok": False, "error": "CREATOR_IDS_REQUIRED"}, 400), absent.json)

        invalid_type = FakeBinaryHandler()
        request["get_payload"] = lambda: {"creator_ids": "creator_one"}
        creator_handler.handle(invalid_type, request, context)
        self.assertEqual(({"ok": False, "error": "CREATOR_IDS_REQUIRED"}, 400), invalid_type.json)

        unknown = FakeBinaryHandler()
        request["get_payload"] = lambda: {"creator_ids": ["missing"]}
        creator_handler.handle(unknown, request, context)
        self.assertEqual(({"ok": False, "error": "CREATOR_NOT_FOUND"}, 404), unknown.json)


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import json
import sys
import tempfile
import threading
import unittest
import urllib.error
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import server
from campaign_creator_repository import CampaignCreatorRepository
from campaign_repository import CampaignRepository
from product_repository import ProductRepository


class ProductCampaignApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        creator_repository.CreatorRepository(self.workbook_path).getCreators()
        self._seed_creator_accounts()

        self.patchers = [
            mock.patch.object(
                server,
                "get_creator_repository",
                side_effect=lambda: creator_repository.CreatorRepository(self.workbook_path),
            ),
            mock.patch.object(server, "get_product_repository", side_effect=lambda: ProductRepository(self.workbook_path)),
            mock.patch.object(server, "get_campaign_repository", side_effect=lambda: CampaignRepository(self.workbook_path)),
            mock.patch.object(
                server,
                "get_campaign_creator_repository",
                side_effect=lambda: CampaignCreatorRepository(self.workbook_path),
            ),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in self.patchers:
            patcher.start()

        self.httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.httpd.server_port}"

    def tearDown(self) -> None:
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=5)
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp_dir.cleanup()

    def _seed_creator_accounts(self) -> None:
        workbook = load_workbook(self.workbook_path)
        creators = workbook["Creators"]
        creator_accounts = workbook["CreatorAccounts"]
        creator_headers = [cell.value for cell in creators[1]]
        account_headers = [cell.value for cell in creator_accounts[1]]
        creators.append(
            [
                {
                    "creator_id": "creator_one",
                    "name": "Creator One",
                    "platform": "TikTok",
                    "profile_url": "https://www.tiktok.com/@creator-one",
                    "status": "discovered",
                    "created_at": "2026-07-01T00:00:00Z",
                    "updated_at": "2026-07-01T00:00:00Z",
                }.get(header, "")
                for header in creator_headers
            ]
        )
        for account_id, platform, url in (
            ("account_one", "TikTok", "https://www.tiktok.com/@creator-one"),
            ("account_two", "Instagram", "https://www.instagram.com/creator-one/"),
        ):
            creator_accounts.append(
                [
                    {
                        "account_id": account_id,
                        "creator_id": "creator_one",
                        "account_uid": f"{platform.lower()}|{url}",
                        "platform": platform,
                        "profile_url": url,
                        "created_at": "2026-07-01T00:00:00Z",
                        "updated_at": "2026-07-01T00:00:00Z",
                    }.get(header, "")
                    for header in account_headers
                ]
            )
        workbook.save(self.workbook_path)
        workbook.close()

    def request(self, method: str, path: str, payload: dict | None = None) -> tuple[int, dict]:
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            self.base_url + path,
            data=data,
            method=method,
            headers={"Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                return response.status, json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            try:
                return exc.code, json.loads(exc.read().decode("utf-8"))
            finally:
                exc.close()

    def create_product(self, name: str = "Product A") -> dict:
        status, body = self.request("POST", "/api/products", {"name": name, "company_name": "Studio"})
        self.assertEqual(201, status)
        return body["product"]

    def create_campaign(self, product_id: str, *, name: str = "Campaign A", status_value: str = "draft") -> dict:
        status, body = self.request(
            "POST",
            "/api/campaigns",
            {"product_id": product_id, "name": name, "status": status_value, "budget": 1000},
        )
        self.assertEqual(201, status)
        return body["campaign"]

    def test_product_crud_and_archive(self) -> None:
        product = self.create_product()
        status, body = self.request("GET", "/api/products")
        self.assertEqual(200, status)
        self.assertEqual([product["product_id"]], [item["product_id"] for item in body["products"]])

        status, body = self.request("PATCH", f"/api/products/{product['product_id']}", {"note": "Updated"})
        self.assertEqual(200, status)
        self.assertEqual("Updated", body["product"]["note"])

        status, body = self.request("PATCH", f"/api/products/{product['product_id']}", {"archived": True})
        self.assertEqual(200, status, body)
        self.assertTrue(body["product"]["archived_at"])
        self.assertEqual([], self.request("GET", "/api/products")[1]["products"])
        self.assertEqual(1, len(self.request("GET", "/api/products?include_archived=true")[1]["products"]))

    def test_product_list_counts_only_active_campaigns(self) -> None:
        product = self.create_product()
        active_campaign = self.create_campaign(product["product_id"], name="Active Campaign")
        archived_campaign = self.create_campaign(product["product_id"], name="Archived Campaign")
        self.assertEqual(
            200,
            self.request(
                "PATCH",
                f"/api/campaigns/{archived_campaign['campaign_id']}",
                {"archived": True},
            )[0],
        )

        status, body = self.request("GET", "/api/products")
        self.assertEqual(200, status)
        self.assertEqual(1, len(body["products"]))
        self.assertEqual(1, body["products"][0]["campaigns_count"])
        self.assertEqual(active_campaign["campaign_id"], self.request(
            "GET", f"/api/campaigns?product_id={product['product_id']}"
        )[1]["campaigns"][0]["campaign_id"])

    def test_product_patch_restore_preserves_campaign_relations(self) -> None:
        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "agreed",
                "creator_quote": 300,
            },
        )
        self.assertEqual(201, status)
        campaign_creator_id = body["campaign_creator"]["id"]

        self.assertEqual(
            200,
            self.request("PATCH", f"/api/campaigns/{campaign['campaign_id']}", {"archived": True})[0],
        )
        campaign_before = self.request("GET", f"/api/campaigns/{campaign['campaign_id']}")[1]["campaign"]
        relations_before = self.request(
            "GET",
            f"/api/campaigns/{campaign['campaign_id']}/creators?include_archived=true",
        )[1]["campaign_creators"]

        status, body = self.request(
            "PATCH",
            f"/api/products/{product['product_id']}",
            {"archived_at": "2026-07-31T12:00:00Z"},
        )
        self.assertEqual(200, status, body)
        self.assertTrue(body["product"]["archived_at"])
        self.assertEqual([], self.request("GET", "/api/products")[1]["products"])
        archived_products = self.request("GET", "/api/products?include_archived=true")[1]["products"]
        self.assertEqual(1, len(archived_products))
        self.assertEqual(0, archived_products[0]["campaigns_count"])

        status, body = self.request(
            "PATCH",
            f"/api/products/{product['product_id']}",
            {"archived_at": None},
        )
        self.assertEqual(200, status, body)
        self.assertIsNone(body["product"]["archived_at"])
        self.assertEqual(product["product_id"], self.request("GET", "/api/products")[1]["products"][0]["product_id"])

        campaign_after = self.request("GET", f"/api/campaigns/{campaign['campaign_id']}")[1]["campaign"]
        relations_after = self.request(
            "GET",
            f"/api/campaigns/{campaign['campaign_id']}/creators?include_archived=true",
        )[1]["campaign_creators"]
        self.assertEqual(campaign_before, campaign_after)
        self.assertEqual(relations_before, relations_after)
        self.assertEqual(campaign_creator_id, relations_after[0]["id"])

    def test_campaign_crud_filters_and_archive(self) -> None:
        product_a = self.create_product("Product A")
        product_b = self.create_product("Product B")
        campaign_a = self.create_campaign(
            product_a["product_id"], name="Campaign A", status_value="completed"
        )
        self.create_campaign(product_b["product_id"], name="Campaign B", status_value="running")

        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign_a['campaign_id']}/creators",
            {"creator_id": "creator_one", "account_id": "account_one", "stage": "completed"},
        )
        self.assertEqual(201, status, body)
        relation_id = body["campaign_creator"]["id"]

        status, body = self.request(
            "PATCH",
            f"/api/campaigns/{campaign_a['campaign_id']}",
            {"budget": 1500},
        )
        self.assertEqual(200, status)
        self.assertEqual("completed", body["campaign"]["status"])

        by_product = self.request("GET", f"/api/campaigns?product_id={product_a['product_id']}")[1]["campaigns"]
        self.assertEqual([campaign_a["campaign_id"]], [item["campaign_id"] for item in by_product])
        self.assertEqual(1, len(self.request("GET", "/api/campaigns?status=running")[1]["campaigns"]))

        archive_time = "2026-08-01T08:00:00Z"
        status, body = self.request(
            "PATCH",
            f"/api/campaigns/{campaign_a['campaign_id']}",
            {"archived_at": archive_time},
        )
        self.assertEqual(200, status, body)
        self.assertEqual("completed", body["campaign"]["status"])
        self.assertEqual(archive_time, body["campaign"]["archived_at"])
        self.assertEqual([], self.request("GET", f"/api/campaigns?product_id={product_a['product_id']}")[1]["campaigns"])

        archived = self.request("GET", "/api/campaigns?include_archived=true")[1]["campaigns"]
        archived_campaign = next(item for item in archived if item["campaign_id"] == campaign_a["campaign_id"])
        self.assertEqual("completed", archived_campaign["status"])
        product_rows = self.request("GET", "/api/products")[1]["products"]
        self.assertEqual(0, next(item for item in product_rows if item["product_id"] == product_a["product_id"])["campaigns_count"])
        relations = self.request(
            "GET", f"/api/campaigns/{campaign_a['campaign_id']}/creators"
        )[1]["campaign_creators"]
        self.assertEqual([relation_id], [item["id"] for item in relations])

        status, body = self.request(
            "PATCH",
            f"/api/campaigns/{campaign_a['campaign_id']}",
            {"archived_at": None},
        )
        self.assertEqual(200, status, body)
        self.assertEqual("completed", body["campaign"]["status"])
        self.assertIsNone(body["campaign"]["archived_at"])
        self.assertEqual(1, len(self.request("GET", f"/api/campaigns?product_id={product_a['product_id']}")[1]["campaigns"]))

        status, body = self.request(
            "PATCH",
            f"/api/campaigns/{campaign_a['campaign_id']}",
            {"status": "archived"},
        )
        self.assertEqual(400, status, body)
        self.assertEqual(
            "completed",
            self.request("GET", f"/api/campaigns/{campaign_a['campaign_id']}")[1]["campaign"]["status"],
        )

    def test_campaign_list_aggregates_product_name_and_active_creator_count(self) -> None:
        product = self.create_product("BlockBlast")
        campaign = self.create_campaign(product["product_id"], name="Brazil Launch")
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "contacted",
            },
        )
        self.assertEqual(201, status, body)
        relation_id = body["campaign_creator"]["id"]

        status, body = self.request("GET", "/api/campaigns")
        self.assertEqual(200, status, body)
        self.assertEqual("BlockBlast", body["campaigns"][0]["product_name"])
        self.assertEqual(1, body["campaigns"][0]["creators_count"])
        by_creator = self.request("GET", "/api/campaigns?creator_id=creator_one")[1]["campaigns"]
        self.assertEqual([campaign["campaign_id"]], [item["campaign_id"] for item in by_creator])
        self.assertEqual([], self.request("GET", "/api/campaigns?creator_id=creator_missing")[1]["campaigns"])

        status, body = self.request(
            "PATCH", f"/api/campaign-creators/{relation_id}", {"archived": True}
        )
        self.assertEqual(200, status, body)
        campaigns = self.request("GET", "/api/campaigns")[1]["campaigns"]
        self.assertEqual(0, campaigns[0]["creators_count"])
        self.assertEqual([], self.request("GET", "/api/campaigns?creator_id=creator_one")[1]["campaigns"])

    def test_campaign_list_returns_empty_array_without_products_or_campaigns(self) -> None:
        status, body = self.request("GET", "/api/campaigns")
        self.assertEqual(200, status, body)
        self.assertEqual([], body["campaigns"])

    def test_campaign_list_keeps_orphaned_product_reference_readable(self) -> None:
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Campaigns"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([
            {
                "campaign_id": "campaign_orphan",
                "product_id": "product_missing",
                "name": "Orphan Campaign",
                "status": "draft",
                "created_at": "2026-08-03T00:00:00Z",
                "updated_at": "2026-08-03T00:00:00Z",
            }.get(header, "")
            for header in headers
        ])
        workbook.save(self.workbook_path)
        workbook.close()

        with mock.patch("campaign_repository.log_event") as warning_log:
            status, body = self.request("GET", "/api/campaigns")

        self.assertEqual(200, status, body)
        self.assertEqual("", body["campaigns"][0]["product_name"])
        warning_log.assert_called_once()

    def test_campaign_detail_returns_campaign_and_enriched_creator_account_data(self) -> None:
        product = self.create_product("BlockBlast")
        campaign = self.create_campaign(product["product_id"], name="Creator Campaign")
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "executing",
                "cost": 200,
                "roi": 2.5,
                "publish_links": ["https://www.tiktok.com/@creator-one/video/1"],
            },
        )
        self.assertEqual(201, status, body)

        status, body = self.request("GET", f"/api/campaigns/{campaign['campaign_id']}")
        self.assertEqual(200, status, body)
        self.assertEqual("BlockBlast", body["campaign"]["product_name"])
        self.assertEqual(1, body["campaign"]["creators_count"])

        status, body = self.request("GET", f"/api/campaigns/{campaign['campaign_id']}/creators")
        self.assertEqual(200, status, body)
        self.assertEqual(1, len(body["campaign_creators"]))
        relation = body["campaign_creators"][0]
        self.assertEqual("Creator One", relation["creator_name"])
        self.assertIsNone(relation["agency_id"])
        self.assertIsNone(relation["agency_name"])
        self.assertEqual("account_one", relation["account_id"])
        self.assertEqual("TikTok", relation["account_platform"])
        self.assertEqual("https://www.tiktok.com/@creator-one", relation["account_url"])
        self.assertEqual("executing", relation["stage"])
        self.assertEqual(200, relation["cost"])
        self.assertEqual(2.5, relation["roi"])
        self.assertIn("video/1", relation["publish_links"])

    def test_campaign_creator_aggregates_agency_without_extra_api_requests(self) -> None:
        workbook = load_workbook(self.workbook_path)
        creators = workbook["Creators"]
        agencies = workbook["Agencies"]
        creator_headers = [str(cell.value or "") for cell in creators[1]]
        agency_headers = [str(cell.value or "") for cell in agencies[1]]
        agency_id_column = creator_headers.index("agency_id") + 1
        creator_id_column = creator_headers.index("creator_id") + 1
        creator_row = next(
            row
            for row in range(2, creators.max_row + 1)
            if creators.cell(row, creator_id_column).value == "creator_one"
        )
        creators.cell(creator_row, agency_id_column, "agency_one")
        agencies.append([
            {
                "agency_id": "agency_one",
                "name": "Agency One",
                "created_at": "2026-07-01T00:00:00Z",
                "updated_at": "2026-07-01T00:00:00Z",
            }.get(header, "")
            for header in agency_headers
        ])
        workbook.save(self.workbook_path)
        workbook.close()

        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {"creator_id": "creator_one", "account_id": "account_one"},
        )
        self.assertEqual(201, status, body)
        self.assertEqual("agency_one", body["campaign_creator"]["agency_id"])
        self.assertEqual("Agency One", body["campaign_creator"]["agency_name"])

        status, body = self.request("GET", f"/api/campaigns/{campaign['campaign_id']}/creators")
        self.assertEqual(200, status, body)
        self.assertEqual("agency_one", body["campaign_creators"][0]["agency_id"])
        self.assertEqual("Agency One", body["campaign_creators"][0]["agency_name"])

        status, body = self.request("GET", "/api/creator-library")
        self.assertEqual(200, status, body)
        creator_record = next(item for item in body["records"] if item["creator_id"] == "creator_one")
        self.assertEqual("agency_one", creator_record["agency_id"])
        self.assertEqual("Agency One", creator_record["agency_name"])

        status, body = self.request("GET", "/api/creator-library/creator_one")
        self.assertEqual(200, status, body)
        self.assertEqual("agency_one", body["record"]["agency_id"])
        self.assertEqual("Agency One", body["record"]["agency_name"])

    def test_missing_agency_name_does_not_break_legacy_campaign_creator(self) -> None:
        workbook = load_workbook(self.workbook_path)
        creators = workbook["Creators"]
        headers = [str(cell.value or "") for cell in creators[1]]
        creator_id_column = headers.index("creator_id") + 1
        agency_id_column = headers.index("agency_id") + 1
        creator_row = next(
            row
            for row in range(2, creators.max_row + 1)
            if creators.cell(row, creator_id_column).value == "creator_one"
        )
        creators.cell(creator_row, agency_id_column, "agency_missing")
        workbook.save(self.workbook_path)
        workbook.close()

        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {"creator_id": "creator_one", "account_id": "account_one"},
        )
        self.assertEqual(201, status, body)
        self.assertEqual("agency_missing", body["campaign_creator"]["agency_id"])
        self.assertIsNone(body["campaign_creator"]["agency_name"])

    def test_campaign_creator_crud_and_default_account_update(self) -> None:
        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "contacted",
                "creator_quote": 250,
            },
        )
        self.assertEqual(201, status)
        record = body["campaign_creator"]

        status, body = self.request(
            "PATCH",
            f"/api/campaign-creators/{record['id']}",
            {
                "account_id": "account_two",
                "stage": "executing",
                "cost": 200,
                "views": 10000,
                "likes": 500,
                "comments": 20,
                "roi": 2.5,
            },
        )
        self.assertEqual(200, status)
        self.assertEqual("account_two", body["campaign_creator"]["account_id"])
        self.assertEqual(2.5, body["campaign_creator"]["roi"])
        self.assertEqual(1, len(self.request("GET", f"/api/campaigns/{campaign['campaign_id']}/creators")[1]["campaign_creators"]))

    def test_foreign_keys_duplicates_and_enums_are_rejected(self) -> None:
        status, body = self.request(
            "POST", "/api/campaigns", {"product_id": "product_missing", "name": "Invalid"}
        )
        self.assertEqual(404, status)
        self.assertIn("产品不存在", body["error"])

        product = self.create_product()
        status, _body = self.request(
            "POST",
            "/api/campaigns",
            {"product_id": product["product_id"], "name": "Invalid", "status": "unknown"},
        )
        self.assertEqual(400, status)
        campaign = self.create_campaign(product["product_id"])

        invalid_payload = {"creator_id": "creator_missing", "account_id": "account_one"}
        status, body = self.request(
            "POST", f"/api/campaigns/{campaign['campaign_id']}/creators", invalid_payload
        )
        self.assertEqual(404, status)
        self.assertIn("达人不存在", body["error"])

        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {"creator_id": "creator_one", "account_id": "account_missing"},
        )
        self.assertEqual(404, status)
        self.assertIn("达人账号不存在", body["error"])

        valid_payload = {"creator_id": "creator_one", "account_id": "account_one", "stage": "pending_contact"}
        self.assertEqual(201, self.request("POST", f"/api/campaigns/{campaign['campaign_id']}/creators", valid_payload)[0])
        self.assertEqual(409, self.request("POST", f"/api/campaigns/{campaign['campaign_id']}/creators", valid_payload)[0])

    def test_archive_keeps_excel_rows_and_protects_active_product(self) -> None:
        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        status, body = self.request("PATCH", f"/api/products/{product['product_id']}", {"archived": True})
        self.assertEqual(409, status)
        self.assertIn("不能归档", body["error"])

        status, body = self.request(
            "POST",
            f"/api/campaigns/{campaign['campaign_id']}/creators",
            {"creator_id": "creator_one", "account_id": "account_one"},
        )
        record_id = body["campaign_creator"]["id"]
        self.assertEqual(200, self.request("PATCH", f"/api/campaign-creators/{record_id}", {"archived": True})[0])
        self.assertEqual([], self.request("GET", f"/api/campaigns/{campaign['campaign_id']}/creators")[1]["campaign_creators"])

        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            rows = list(workbook["CampaignCreators"].values)
            self.assertEqual(2, len(rows))
            self.assertEqual(record_id, rows[1][0])
        finally:
            workbook.close()

        self.assertEqual(200, self.request("PATCH", f"/api/campaigns/{campaign['campaign_id']}", {"archived": True})[0])
        self.assertEqual(200, self.request("PATCH", f"/api/products/{product['product_id']}", {"archived": True})[0])

    def test_readding_archived_campaign_creator_restores_same_relation(self) -> None:
        product = self.create_product()
        campaign = self.create_campaign(product["product_id"])
        path = f"/api/campaigns/{campaign['campaign_id']}/creators"
        status, body = self.request(
            "POST",
            path,
            {
                "creator_id": "creator_one",
                "account_id": "account_one",
                "stage": "contacted",
                "creator_quote": 250,
            },
        )
        self.assertEqual(201, status, body)
        relation_id = body["campaign_creator"]["id"]
        created_at = body["campaign_creator"]["created_at"]

        self.assertEqual(
            200,
            self.request("PATCH", f"/api/campaign-creators/{relation_id}", {"archived": True})[0],
        )
        status, body = self.request(
            "POST",
            path,
            {
                "creator_id": "creator_one",
                "account_id": "account_two",
                "stage": "executing",
                "cost": 200,
            },
        )
        self.assertEqual(201, status, body)
        restored = body["campaign_creator"]
        self.assertEqual(relation_id, restored["id"])
        self.assertEqual(created_at, restored["created_at"])
        self.assertIsNone(restored["archived_at"])
        self.assertEqual("account_two", restored["account_id"])
        self.assertEqual("executing", restored["stage"])
        self.assertEqual(250, restored["creator_quote"])
        self.assertEqual(200, restored["cost"])

        status, body = self.request("GET", path)
        self.assertEqual(200, status, body)
        self.assertEqual([relation_id], [item["id"] for item in body["campaign_creators"]])
        self.assertEqual(409, self.request("POST", path, {
            "creator_id": "creator_one",
            "account_id": "account_two",
        })[0])

        workbook = load_workbook(self.workbook_path, read_only=True)
        try:
            rows = list(workbook["CampaignCreators"].values)
            self.assertEqual(2, len(rows))
            self.assertEqual(relation_id, rows[1][0])
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()

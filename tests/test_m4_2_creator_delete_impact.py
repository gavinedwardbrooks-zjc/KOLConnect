from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import threading
import unittest
import urllib.error
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import server
from repository_factory import RepositoryFactory


def append_row(workbook, sheet_name: str, values: dict) -> None:
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    sheet.append([values.get(header, "") for header in headers])


class CreatorDeleteImpactHttpTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.tasks_dir = self.root / "tasks"
        self.protection_file = self.root / "data_protection.json"
        self.legacy_analysis_dir = self.root / "creator_analysis"
        self.legacy_library_file = self.root / "creator_library.json"
        RepositoryFactory.for_path(self.workbook_path).creator().getCreators()
        self._seed_creator("creator_archived", archived=True)
        self._seed_creator("creator_active", archived=False)

        self.patchers = [
            mock.patch.object(
                server,
                "_creator_library_workbook_path",
                return_value=self.workbook_path,
            ),
            mock.patch.object(server, "TASKS_DIR", self.tasks_dir),
            mock.patch.object(server, "DATA_PROTECTION_FILE", self.protection_file),
            mock.patch.object(server, "CREATOR_ANALYSIS_DIR", self.legacy_analysis_dir),
            mock.patch.object(server, "CREATOR_LIBRARY_FILE", self.legacy_library_file),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in self.patchers:
            patcher.start()
        self.httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.httpd.server_port}"

    def tearDown(self) -> None:
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=5)
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp_dir.cleanup()

    def _seed_creator(self, creator_id: str, *, archived: bool) -> None:
        workbook = load_workbook(self.workbook_path)
        append_row(
            workbook,
            "Creators",
            {
                "creator_id": creator_id,
                "name": f"Name {creator_id}",
                "email": f"{creator_id}@private.example",
                "whatsapp": "+1-private",
                "profile_url": f"https://private.example/{creator_id}",
                "bio": "private biography",
                "archived_at": "2026-08-14T00:00:00Z" if archived else "",
                "updated_at": "2026-08-14T00:00:00Z",
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

    def request(self, method: str, creator_id: str) -> tuple[int, dict]:
        request = urllib.request.Request(
            self.base_url
            + f"/api/creator-library/{creator_id}/delete-impact",
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                return response.status, json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            with exc:
                return exc.code, json.loads(exc.read().decode("utf-8"))

    def test_success_is_read_only_minimal_and_deterministic(self) -> None:
        before = self.workbook_path.read_bytes()
        before_hash = hashlib.sha256(before).hexdigest()
        before_mtime = self.workbook_path.stat().st_mtime_ns

        status, first = self.request("GET", "creator_archived")
        second_status, second = self.request("GET", "creator_archived")

        self.assertEqual(200, status)
        self.assertEqual(200, second_status)
        self.assertTrue(first["ok"])
        self.assertTrue(first["can_delete"])
        self.assertEqual([], first["blockers"])
        self.assertEqual(1, first["impact"]["creators"])
        self.assertRegex(first["preview_fingerprint"], r"^[0-9a-f]{64}$")
        self.assertEqual(
            first["preview_fingerprint"], second["preview_fingerprint"]
        )
        serialized = json.dumps(first, ensure_ascii=False)
        for private_value in (
            "creator_archived@private.example",
            "+1-private",
            "https://private.example/creator_archived",
            "private biography",
        ):
            self.assertNotIn(private_value, serialized)
        self.assertEqual(before_hash, hashlib.sha256(self.workbook_path.read_bytes()).hexdigest())
        self.assertEqual(before_mtime, self.workbook_path.stat().st_mtime_ns)

    def test_not_found_not_archived_and_no_mutation_endpoint(self) -> None:
        status, missing = self.request("GET", "missing")
        self.assertEqual(404, status)
        self.assertEqual("未找到达人分析记录。", missing["error"])

        status, active = self.request("GET", "creator_active")
        self.assertEqual(200, status)
        self.assertFalse(active["can_delete"])
        self.assertIn(
            "CREATOR_NOT_ARCHIVED",
            {item["code"] for item in active["blockers"]},
        )

        status, _body = self.request("POST", "creator_archived")
        self.assertEqual(404, status)

    def test_scanner_failure_fails_closed(self) -> None:
        task_id = "task_20260814T000000Z_deadbeef"
        task_root = self.tasks_dir / task_id
        task_root.mkdir(parents=True)
        (task_root / "task.json").write_text("{invalid", encoding="utf-8")
        status, body = self.request("GET", "creator_archived")
        self.assertEqual(500, status)
        self.assertEqual("删除影响扫描未完成。", body["error"])
        self.assertNotIn("can_delete", body)

    def test_active_campaign_remains_an_http_blocker(self) -> None:
        workbook = load_workbook(self.workbook_path)
        append_row(
            workbook,
            "CreatorAccounts",
            {
                "account_id": "active_account",
                "creator_id": "creator_archived",
                "account_uid": "active_uid",
            },
        )
        append_row(
            workbook,
            "CreatorSnapshots",
            {
                "snapshot_id": "active_snapshot",
                "creator_id": "creator_archived",
            },
        )
        append_row(
            workbook,
            "Campaigns",
            {"campaign_id": "active_campaign", "name": "Active"},
        )
        append_row(
            workbook,
            "CampaignCreators",
            {
                "id": "active_relation",
                "campaign_id": "active_campaign",
                "creator_id": "creator_archived",
                "account_id": "active_account",
                "archived_at": "",
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

        status, body = self.request("GET", "creator_archived")
        self.assertEqual(200, status)
        self.assertFalse(body["can_delete"])
        codes = {item["code"] for item in body["blockers"]}
        self.assertIn("ACTIVE_CAMPAIGN_RELATION", codes)
        self.assertNotIn("UNRESOLVED_SNAPSHOT_RETENTION", codes)
        self.assertEqual(1, body["impact"]["campaign_creators"]["active"])


class CreatorDeleteImpactRepositoryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.tasks_dir = self.root / "tasks"
        self.protection_file = self.root / "data_protection.json"
        self.legacy_analysis_dir = self.root / "creator_analysis"
        self.legacy_library_file = self.root / "creator_library.json"
        self.factory = RepositoryFactory.for_path(
            self.workbook_path,
            tasks_dir=self.tasks_dir,
            data_protection_file=self.protection_file,
            legacy_analysis_dir=self.legacy_analysis_dir,
            legacy_library_file=self.legacy_library_file,
        )
        self.factory.creator().getCreators()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _append(self, sheet_name: str, values: dict) -> None:
        workbook = load_workbook(self.workbook_path)
        append_row(workbook, sheet_name, values)
        workbook.save(self.workbook_path)
        workbook.close()

    def _seed_full_impact(self) -> None:
        rows = {
            "Creators": {
                "creator_id": "creator_target",
                "name": "Target",
                "archived_at": "2026-08-14T00:00:00Z",
                "updated_at": "2026-08-14T00:00:00Z",
                "agency_id": "agency_1",
                "current_contact_id": "contact_1",
            },
            "CreatorAccounts": {
                "account_id": "account_1",
                "creator_id": "creator_target",
                "account_uid": "uid_1",
            },
            "Videos": {"creator_id": "creator_target", "video_url": "private"},
            "Insights": {"creator_id": "creator_target", "average_views": 1},
            "_AnalysisData": {
                "creator_id": "creator_target",
                "account_uid": "uid_1",
                "analysis_json": json.dumps(
                    {"analysis_id": "creator_target", "account_uid": "uid_1"}
                ),
            },
            "CreatorSnapshots": {
                "snapshot_id": "snapshot_1",
                "creator_id": "creator_target",
                "account_uid": "uid_1",
            },
            "VideoSnapshots": {
                "video_snapshot_id": "video_snapshot_1",
                "snapshot_id": "snapshot_1",
                "creator_id": "creator_target",
            },
            "Cooperations": {
                "cooperation_id": "cooperation_1",
                "creator_id": "creator_target",
            },
            "Products": {"product_id": "product_1", "name": "Product"},
            "Campaigns": {
                "campaign_id": "campaign_1",
                "product_id": "product_1",
                "name": "Campaign",
            },
            "CampaignCreators": {
                "id": "relation_1",
                "campaign_id": "campaign_1",
                "creator_id": "creator_target",
                "account_id": "account_1",
                "archived_at": "",
            },
            "FollowUpLogs": {
                "follow_up_id": "followup_1",
                "object_type": "creator",
                "object_id": "creator_target",
                "content": "private",
            },
            "Agencies": {"agency_id": "agency_1", "name": "Agency"},
            "AgencyContacts": {
                "contact_id": "contact_1",
                "agency_id": "agency_1",
                "name": "Contact",
            },
        }
        workbook = load_workbook(self.workbook_path)
        for sheet_name, values in rows.items():
            append_row(workbook, sheet_name, values)
        append_row(
            workbook,
            "VideoSnapshots",
            {
                "video_snapshot_id": "video_snapshot_indirect",
                "snapshot_id": "snapshot_1",
                "creator_id": "",
            },
        )
        append_row(
            workbook,
            "CampaignCreators",
            {
                "id": "relation_archived",
                "campaign_id": "campaign_1",
                "creator_id": "creator_target",
                "account_id": "account_1",
                "archived_at": "2026-08-13T00:00:00Z",
            },
        )
        append_row(
            workbook,
            "Creators",
            {
                "creator_id": "creator_other",
                "name": "Other",
                "archived_at": "2026-08-14T00:00:00Z",
            },
        )
        append_row(
            workbook,
            "Videos",
            {"creator_id": "creator_other", "video_url": "other"},
        )
        append_row(
            workbook,
            "_AnalysisData",
            {
                "creator_id": "creator_other",
                "analysis_json": json.dumps({"creator_id": "creator_target"}),
            },
        )
        workbook.save(self.workbook_path)
        workbook.close()

        task_id = "task_20260814T000000Z_1234abcd"
        task_root = self.tasks_dir / task_id
        task_root.mkdir(parents=True)
        (task_root / "task.json").write_text(
            json.dumps(
                {
                    "id": task_id,
                    "creator_library_creator_ids": ["creator_target"],
                    "creator_library_account_ids": ["account_1"],
                    "creator_snapshot_id": "snapshot_1",
                }
            ),
            encoding="utf-8",
        )
        (task_root / "results.csv").write_text("private,data\n", encoding="utf-8")
        self.protection_file.write_text(
            json.dumps({"uid_1": {"email": {"value": "private"}}}),
            encoding="utf-8",
        )
        self.legacy_analysis_dir.mkdir()
        (self.legacy_analysis_dir / "analysis_task_one.json").write_text(
            json.dumps({"analysis_id": "creator_target", "account_uid": "uid_1"}),
            encoding="utf-8",
        )
        self.legacy_library_file.write_text(
            json.dumps({"records": {"creator_target": {"status": "archived"}}}),
            encoding="utf-8",
        )

    def test_all_structured_references_and_retention_are_counted(self) -> None:
        self._seed_full_impact()
        service = server.CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        )
        before_workbook = self.workbook_path.read_bytes()
        before_files = {
            path: (path.read_bytes(), path.stat().st_mtime_ns)
            for path in self.root.rglob("*") if path.is_file()
        }

        result = service.get_delete_impact("creator_target")

        self.assertEqual(1, result["impact"]["creator_accounts"])
        self.assertEqual(1, result["impact"]["videos"])
        self.assertEqual(1, result["impact"]["insights"])
        self.assertEqual(1, result["impact"]["analysis_data"])
        self.assertEqual(1, result["impact"]["creator_snapshots"])
        self.assertEqual(2, result["impact"]["video_snapshots"]["total"])
        self.assertEqual(1, result["impact"]["video_snapshots"]["direct"])
        self.assertEqual(1, result["impact"]["video_snapshots"]["indirect"])
        self.assertEqual(1, result["impact"]["cooperations"])
        self.assertEqual(1, result["impact"]["campaign_creators"]["active"])
        self.assertEqual(1, result["impact"]["campaign_creators"]["archived"])
        self.assertEqual(1, result["impact"]["follow_up_logs"])
        self.assertEqual(1, result["impact"]["embedded_analysis_references"])
        self.assertEqual(1, result["impact"]["task_artifacts"])
        self.assertEqual(1, result["impact"]["data_protection"])
        self.assertEqual(2, result["impact"]["legacy_sources"])
        self.assertEqual(1, result["retained"]["agencies"])
        self.assertEqual(1, result["retained"]["agency_contacts"])
        self.assertEqual(1, result["retained"]["campaigns"])
        self.assertEqual(1, result["retained"]["products"])
        codes = {item["code"] for item in result["blockers"]}
        self.assertIn("ACTIVE_CAMPAIGN_RELATION", codes)
        self.assertIn("COOPERATION_RETENTION_ANONYMIZATION_GAP", codes)
        self.assertIn("EMBEDDED_ANALYSIS_REFERENCE", codes)
        self.assertNotIn("UNRESOLVED_SNAPSHOT_RETENTION", codes)
        self.assertNotIn("UNRESOLVED_TASK_ARTIFACT", codes)
        self.assertFalse(result["can_delete"])
        self.assertEqual(before_workbook, self.workbook_path.read_bytes())
        for path, (contents, mtime) in before_files.items():
            self.assertEqual(contents, path.read_bytes())
            self.assertEqual(mtime, path.stat().st_mtime_ns)

        other = service.get_delete_impact("creator_other")
        self.assertEqual(1, other["impact"]["videos"])
        self.assertEqual(0, other["impact"]["creator_accounts"])
        self.assertEqual(0, other["impact"]["task_artifacts"])

    def test_broken_references_and_unknown_followup_type_block(self) -> None:
        self._append(
            "Creators",
            {
                "creator_id": "creator_broken",
                "name": "Broken",
                "archived_at": "2026-08-14T00:00:00Z",
            },
        )
        self._append(
            "CampaignCreators",
            {
                "id": "relation_broken",
                "campaign_id": "missing_campaign",
                "creator_id": "creator_broken",
                "account_id": "missing_account",
            },
        )
        self._append(
            "FollowUpLogs",
            {
                "follow_up_id": "followup_unknown",
                "object_type": "unknown",
                "object_id": "creator_broken",
            },
        )
        result = server.CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        ).get_delete_impact("creator_broken")
        codes = {item["code"] for item in result["blockers"]}
        self.assertIn("BROKEN_REFERENCE", codes)
        self.assertIn("UNKNOWN_FOLLOWUP_OBJECT_TYPE", codes)
        self.assertFalse(result["can_delete"])

    def test_fingerprint_changes_only_after_relevant_structural_change(self) -> None:
        self._append(
            "Creators",
            {
                "creator_id": "creator_fingerprint",
                "name": "Fingerprint",
                "archived_at": "2026-08-14T00:00:00Z",
                "updated_at": "2026-08-14T00:00:00Z",
            },
        )
        service = server.CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        )
        first = service.get_delete_impact("creator_fingerprint")
        second = service.get_delete_impact("creator_fingerprint")
        self.assertEqual(first["preview_fingerprint"], second["preview_fingerprint"])

        self._append(
            "CreatorAccounts",
            {
                "account_id": "account_new",
                "creator_id": "creator_fingerprint",
                "account_uid": "uid_new",
            },
        )
        changed = service.get_delete_impact("creator_fingerprint")
        self.assertNotEqual(first["preview_fingerprint"], changed["preview_fingerprint"])

    def test_unmapped_task_artifact_fails_closed_without_identity_guessing(self) -> None:
        self._append(
            "Creators",
            {
                "creator_id": "creator_unmapped",
                "name": "Unmapped",
                "archived_at": "2026-08-14T00:00:00Z",
            },
        )
        task_id = "task_20260814T000000Z_cafebabe"
        task_root = self.tasks_dir / task_id
        task_root.mkdir(parents=True)
        (task_root / "task.json").write_text(
            json.dumps({"id": task_id}), encoding="utf-8"
        )
        (task_root / "links.txt").write_text(
            "https://example.invalid/identity-must-not-be-matched\n",
            encoding="utf-8",
        )
        result = server.CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        ).get_delete_impact("creator_unmapped")
        self.assertEqual(0, result["impact"]["task_artifacts"])
        self.assertEqual(1, result["impact"]["unmapped_task_artifacts"])
        self.assertIn(
            "UNRESOLVED_TASK_OWNERSHIP",
            {item["code"] for item in result["blockers"]},
        )
        self.assertFalse(result["can_delete"])


if __name__ == "__main__":
    unittest.main()

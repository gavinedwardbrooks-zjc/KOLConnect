from __future__ import annotations

import os
import sys
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

from http_handlers import settings_handler
from local_storage_lock import SharedStorageLockTimeout
from services.workbook_backup_service import (
    WorkbookBackupError,
    WorkbookBackupNotFoundError,
    WorkbookBackupService,
)


class _Handler:
    def __init__(self) -> None:
        self.payload = None
        self.status = None

    def _json(self, payload, status=200) -> None:
        self.payload = payload
        self.status = status


class WorkbookBackupTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(dir=ROOT, prefix=".m5_4_e3_")
        self.root = Path(self.temp.name)
        self.workbook_path = self.root / "Creator_Library.xlsx"
        workbook = Workbook()
        workbook.active.title = "Creators"
        workbook.active.append(["creator_id", "creator_name"])
        workbook.active.append(["creator_one", "Creator One"])
        workbook.save(self.workbook_path)
        workbook.close()
        self.environment = mock.patch.dict(
            os.environ,
            {
                "APPDATA": str(self.root / "runtime"),
                "HOME": str(self.root),
                "XDG_DATA_HOME": str(self.root / "runtime"),
            },
        )
        self.environment.start()

    def tearDown(self) -> None:
        self.environment.stop()
        self.temp.cleanup()

    def service(self, tokens=None) -> WorkbookBackupService:
        token_values = iter(tokens or ["token-one"])
        return WorkbookBackupService(
            lambda: self.workbook_path,
            now_provider=lambda: datetime(2026, 8, 22, 8, 30, tzinfo=timezone.utc),
            token_provider=lambda: next(token_values),
        )

    def test_backup_is_valid_and_api_response_uses_backups_directory(self) -> None:
        service = self.service()
        handler = _Handler()
        handled = settings_handler.handle(
            handler,
            {
                "method": "POST",
                "path": "/api/settings/creator-library/backup",
                "get_payload": lambda: {},
            },
            {
                "state": {"get": lambda: {}},
                "services": {"workbook_backup": service},
                "modules": {},
            },
        )

        self.assertTrue(handled)
        self.assertEqual(200, handler.status)
        self.assertTrue(handler.payload["ok"])
        backup = handler.payload["backup"]
        backup_path = self.workbook_path.parent / "backups" / backup["filename"]
        self.assertTrue(backup_path.is_file())
        self.assertGreater(backup["size"], 0)
        self.assertEqual("2026-08-22T08:30:00Z", backup["created_at"])
        validated = load_workbook(backup_path, read_only=True)
        validated.close()

    def test_missing_workbook_fails_without_creating_backup(self) -> None:
        missing = self.root / "missing.xlsx"
        service = WorkbookBackupService(lambda: missing)
        with self.assertRaises(WorkbookBackupNotFoundError):
            service.create_backup()
        self.assertFalse((self.root / "backups").exists())

    def test_api_reports_missing_workbook(self) -> None:
        handler = _Handler()
        service = WorkbookBackupService(lambda: self.root / "missing.xlsx")
        settings_handler.handle(
            handler,
            {
                "method": "POST",
                "path": "/api/settings/creator-library/backup",
                "get_payload": lambda: {},
            },
            {
                "state": {"get": lambda: {}},
                "services": {"workbook_backup": service},
                "modules": {},
            },
        )
        self.assertEqual(404, handler.status)
        self.assertFalse(handler.payload["ok"])

    def test_lock_failure_is_reported_and_not_claimed_as_success(self) -> None:
        with mock.patch(
            "services.workbook_backup_service.ExcelWorkbookStore.create_transaction_backup",
            side_effect=SharedStorageLockTimeout("busy"),
        ):
            with self.assertRaisesRegex(WorkbookBackupError, "其他操作"):
                self.service().create_backup()

    def test_same_timestamp_still_creates_unique_backup_names(self) -> None:
        service = self.service(["first", "second"])
        first = service.create_backup()
        second = service.create_backup()
        self.assertNotEqual(first["filename"], second["filename"])
        backups = list((self.root / "backups").glob("*.xlsx"))
        self.assertEqual(2, len(backups))


if __name__ == "__main__":
    unittest.main()

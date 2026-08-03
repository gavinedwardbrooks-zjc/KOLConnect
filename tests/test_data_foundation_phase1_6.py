from __future__ import annotations

import csv
import importlib
import json
import os
import sys
import tempfile
import time
import unittest
from collections import Counter
from pathlib import Path
from unittest import mock

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
from campaign_creator_repository import CAMPAIGN_CREATORS_HEADERS
from campaign_repository import CAMPAIGNS_HEADERS
from dashboard_repository import DashboardRepository
from dashboard_service import DashboardService
from product_repository import PRODUCTS_HEADERS


def close_app_logger() -> None:
    app_logging = sys.modules.get("app_logging")
    if app_logging is None:
        return
    logger = app_logging.logging.getLogger(app_logging.LOGGER_NAME)
    for handler in list(logger.handlers):
        handler.close()
        logger.removeHandler(handler)
    app_logging._CONFIGURED = False


def append_mapping(sheet, headers: list[str], values: dict) -> None:
    sheet.append([values.get(header, "") for header in headers])


def create_schema2_workbook(
    path: Path,
    *,
    creator_count: int,
    snapshots_per_creator: int,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_headers = {
        "Creators": creator_repository._CREATORS_HEADERS,
        "CreatorAccounts": creator_repository._CREATOR_ACCOUNTS_HEADERS,
        "Videos": creator_repository._VIDEOS_HEADERS,
        "Insights": creator_repository._INSIGHTS_HEADERS,
        "CreatorSnapshots": creator_repository._CREATOR_SNAPSHOTS_HEADERS,
        "VideoSnapshots": creator_repository._VIDEO_SNAPSHOTS_HEADERS,
        "Cooperations": creator_repository._COOPERATIONS_HEADERS,
        "Agencies": creator_repository._AGENCIES_HEADERS,
        "AgencyContacts": creator_repository._AGENCY_CONTACTS_HEADERS,
        "FollowUpLogs": creator_repository._FOLLOW_UP_LOGS_HEADERS,
        "Products": PRODUCTS_HEADERS,
        "Campaigns": CAMPAIGNS_HEADERS,
        "CampaignCreators": CAMPAIGN_CREATORS_HEADERS,
        "_AnalysisData": creator_repository._ANALYSIS_METADATA_HEADERS,
        "_Metadata": creator_repository._WORKBOOK_METADATA_HEADERS,
    }
    for sheet_name, headers in sheet_headers.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)

    for creator_index in range(creator_count):
        creator_id = f"creator_{creator_index:05d}"
        account_uid = f"tiktok|https://www.tiktok.com/@creator{creator_index:05d}"
        profile_url = f"https://www.tiktok.com/@creator{creator_index:05d}"
        append_mapping(
            workbook["Creators"],
            creator_repository._CREATORS_HEADERS,
            {
                "creator_id": creator_id,
                "name": f"Creator {creator_index}",
                "platform": "TikTok",
                "profile_url": profile_url,
                "followers": str(creator_index),
                "insight_level": "good",
                "status": "discovered",
                "created_at": "2026-07-01T00:00:00Z",
                "updated_at": "2026-07-01T00:00:00Z",
            },
        )
        append_mapping(
            workbook["CreatorAccounts"],
            creator_repository._CREATOR_ACCOUNTS_HEADERS,
            {
                "account_id": f"account_{creator_index:05d}",
                "creator_id": creator_id,
                "account_uid": account_uid,
                "platform": "TikTok",
                "username": f"creator{creator_index:05d}",
                "profile_url": profile_url,
                "followers": str(creator_index),
                "data_source": "test",
                "scrape_status": "success",
                "created_at": "2026-07-01T00:00:00Z",
                "updated_at": "2026-07-01T00:00:00Z",
            },
        )
        append_mapping(
            workbook["Insights"],
            creator_repository._INSIGHTS_HEADERS,
            {"creator_id": creator_id, "average_views": 1000, "median_views": 900},
        )
        append_mapping(
            workbook["_AnalysisData"],
            creator_repository._ANALYSIS_METADATA_HEADERS,
            {
                "creator_id": creator_id,
                "account_uid": account_uid,
                "source": "test",
            },
        )
        for snapshot_index in range(snapshots_per_creator):
            append_mapping(
                workbook["CreatorSnapshots"],
                creator_repository._CREATOR_SNAPSHOTS_HEADERS,
                {
                    "snapshot_id": f"snapshot_{creator_index:05d}_{snapshot_index:02d}",
                    "creator_id": creator_id,
                    "platform": "TikTok",
                    "account_uid": account_uid,
                    "followers": str(snapshot_index),
                    "average_views": snapshot_index * 100,
                    "median_views": snapshot_index * 90,
                    "video_count": 10,
                    "creator_score": snapshot_index,
                    "insight_level": "good",
                    "captured_at": f"2026-07-{snapshot_index + 1:02d}T00:00:00Z",
                    "source": "test",
                },
            )
    append_mapping(
        workbook["_Metadata"],
        creator_repository._WORKBOOK_METADATA_HEADERS,
        {
            "schema_version": creator_repository.CREATOR_LIBRARY_SCHEMA_VERSION,
            "last_update_time": "2026-07-31T00:00:00Z",
        },
    )
    workbook["_AnalysisData"].sheet_state = "hidden"
    workbook["_Metadata"].sheet_state = "hidden"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


class CountingRepository(creator_repository.CreatorRepository):
    def __init__(self, workbook_path: Path) -> None:
        super().__init__(workbook_path)
        self.load_count = 0
        self.row_read_counts: Counter[str] = Counter()

    def _load_workbook(self):
        self.load_count += 1
        return super()._load_workbook()

    def _rows(self, sheet):
        self.row_read_counts[sheet.title] += 1
        return creator_repository.CreatorRepository._rows(sheet)


def reload_server(temp_appdata: str):
    close_app_logger()
    for module_name in (
        "server",
        "runtime_paths",
        "mail_sync",
        "task_manager",
        "dashboard_repository",
        "dashboard_service",
        "app_logging",
    ):
        sys.modules.pop(module_name, None)
    with mock.patch.dict(os.environ, {"APPDATA": temp_appdata}):
        return importlib.import_module("server")


class CreatorLibraryPerformanceTests(unittest.TestCase):
    def test_get_creators_reads_snapshots_once_for_large_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            create_schema2_workbook(
                workbook_path,
                creator_count=1500,
                snapshots_per_creator=10,
            )
            repository = CountingRepository(workbook_path)

            started = time.perf_counter()
            records = repository.getCreators()
            elapsed = time.perf_counter() - started

            self.assertEqual(1500, len(records))
            self.assertEqual(1, repository.load_count)
            self.assertEqual(1, repository.row_read_counts["Creators"])
            self.assertEqual(1, repository.row_read_counts["CreatorAccounts"])
            self.assertEqual(1, repository.row_read_counts["CreatorSnapshots"])
            self.assertEqual("9", records[0]["followers"])
            self.assertLess(elapsed, 45)

    def test_detail_reuses_one_workbook_and_request_indexes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            create_schema2_workbook(
                workbook_path,
                creator_count=2,
                snapshots_per_creator=3,
            )
            repository = CountingRepository(workbook_path)

            detail = repository.getCreatorDetail("creator_00000")

            self.assertEqual("creator_00000", detail["record"]["creator_id"])
            self.assertEqual(1, repository.load_count)
            self.assertEqual(1, repository.row_read_counts["CreatorSnapshots"])
            self.assertEqual(3, len(detail["snapshots"]))

    def test_dashboard_request_reuses_creator_and_campaign_creator_reads(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, mock.patch.object(
            creator_repository, "log_event"
        ):
            workbook_path = Path(temp_dir) / "Creator_Library.xlsx"
            create_schema2_workbook(
                workbook_path,
                creator_count=2,
                snapshots_per_creator=3,
            )
            repository = CountingRepository(workbook_path)
            campaign_creator_repository = mock.Mock()
            campaign_creator_repository.getCampaignCreators.return_value = []
            campaign_repository = mock.Mock()
            campaign_repository.getCampaigns.return_value = []
            service = DashboardService(DashboardRepository(
                repository,
                campaign_creator_repository,
                campaign_repository,
            ))

            service.getOverview()
            service.getCreatorHealth()
            service.getCooperationPerformance()
            service.getActionItems()

            self.assertEqual(1, repository.load_count)
            self.assertEqual(1, repository.row_read_counts["Creators"])
            self.assertEqual(1, repository.row_read_counts["CreatorSnapshots"])
            self.assertEqual(0, repository.row_read_counts["Cooperations"])
            campaign_creator_repository.getCampaignCreators.assert_called_once_with(
                include_archived=False
            )
            campaign_repository.getCampaigns.assert_called_once_with(include_archived=True)


class HistoricalTaskBoundaryTests(unittest.TestCase):
    def test_page_reads_do_not_backfill_large_historical_task(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            data_dir = Path(temp_dir) / "KOLConnect"
            workbook_path = data_dir / "Creator_Library.xlsx"
            create_schema2_workbook(
                workbook_path,
                creator_count=2,
                snapshots_per_creator=1,
            )
            server = reload_server(temp_dir)
            task = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@historical"],
                [],
                1330,
                name="historical",
                target_platform="TikTok",
            )
            _task, paths = server.task_manager.load_task(server.TASKS_DIR, task["id"])
            metadata = json.loads(paths["metadata"].read_text(encoding="utf-8"))
            metadata.pop("creator_library_import_eligible", None)
            metadata["status"] = "completed"
            paths["metadata"].write_text(
                json.dumps(metadata, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            with paths["results"].open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=server.scraper_module.OUTPUT_FIELDS)
                writer.writeheader()
                for index in range(1330):
                    result = server.scraper_module.build_result(
                        url=f"https://www.tiktok.com/@historical{index}",
                        platform="TikTok",
                        name=f"Historical {index}",
                        scrape_status="success",
                    )
                    writer.writerow(server.scraper_module.result_to_row(result))

            workbook_before = workbook_path.read_bytes()
            task_before = paths["metadata"].read_bytes()
            library = server.get_creator_library()
            dashboard = server.get_dashboard_data()

            self.assertEqual(2, len(library["records"]))
            self.assertEqual(2, dashboard["overview"]["total_creators"])
            self.assertEqual(workbook_before, workbook_path.read_bytes())
            self.assertEqual(task_before, paths["metadata"].read_bytes())
            close_app_logger()

    def test_new_completed_task_import_is_incremental_and_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            server = reload_server(temp_dir)
            task = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@new-task"],
                [],
                1,
                name="new-task",
                target_platform="TikTok",
            )
            self.assertTrue(task["creator_library_import_eligible"])
            result = server.scraper_module.build_result(
                url="https://www.tiktok.com/@new-task",
                platform="TikTok",
                name="New Task",
                scrape_status="success",
            )
            row = server.scraper_module.result_to_row(result)
            _task, paths = server.task_manager.load_task(server.TASKS_DIR, task["id"])
            server.task_manager.atomic_write_files(
                {
                    paths["results"]: server._csv_content(
                        server.scraper_module.OUTPUT_FIELDS,
                        [row],
                    ),
                    paths["progress"]: server._csv_content(
                        server.scraper_module.PROGRESS_FIELDS,
                        [row],
                    ),
                }
            )
            server.task_manager.update_task(
                server.TASKS_DIR,
                task["id"],
                status="completed",
            )

            first = server.import_task_results_to_creator_library(task["id"])
            server = reload_server(temp_dir)
            second = server.import_task_results_to_creator_library(task["id"])
            repository = server.get_creator_repository()
            creators = repository.getCreators()
            accounts = repository.getCreatorAccounts()

            self.assertEqual("success", first["status"])
            self.assertEqual("success", second["status"])
            self.assertEqual(1, len(creators))
            self.assertEqual(1, len(accounts))
            self.assertEqual(task["id"], accounts[0]["source_task_id"])
            self.assertEqual(1, len(repository.getCreatorSnapshots(creators[0]["creator_id"])))
            close_app_logger()

    def test_historical_unfinished_and_email_recheck_tasks_are_skipped(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            server = reload_server(temp_dir)
            unfinished = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@unfinished"],
                [],
                1,
                target_platform="TikTok",
            )
            email_recheck = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@email-recheck"],
                [],
                1,
                target_platform="TikTok",
                task_type="email_recheck",
            )
            historical = server.task_manager.create_task(
                server.TASKS_DIR,
                ["https://www.tiktok.com/@historical"],
                [],
                1,
                target_platform="TikTok",
            )
            _task, historical_paths = server.task_manager.load_task(
                server.TASKS_DIR,
                historical["id"],
            )
            historical_metadata = json.loads(
                historical_paths["metadata"].read_text(encoding="utf-8")
            )
            historical_metadata.pop("creator_library_import_eligible", None)
            historical_metadata["status"] = "completed"
            historical_paths["metadata"].write_text(
                json.dumps(historical_metadata, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            server.task_manager.update_task(
                server.TASKS_DIR,
                email_recheck["id"],
                status="completed",
            )

            self.assertEqual(
                "task_not_completed",
                server.import_task_results_to_creator_library(unfinished["id"])["reason"],
            )
            self.assertEqual(
                "email_recheck_task",
                server.import_task_results_to_creator_library(email_recheck["id"])["reason"],
            )
            self.assertEqual(
                "historical_task_requires_manual_import",
                server.import_task_results_to_creator_library(historical["id"])["reason"],
            )
            close_app_logger()


if __name__ == "__main__":
    unittest.main()

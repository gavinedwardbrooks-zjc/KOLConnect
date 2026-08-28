from __future__ import annotations

import json
import shutil
import sys
import unittest
import uuid
from contextlib import nullcontext
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
sys.path.insert(0, str(ROOT / "tests"))

from excel_workbook_store import ExcelWorkbookStore  # noqa: E402
from excel_workbook_store import WORKBOOK_LOCK  # noqa: E402
import creator_repository  # noqa: E402
from local_storage_lock import LOCAL_STORAGE_MUTATION_LOCK  # noqa: E402
from repository_factory import RepositoryFactory  # noqa: E402
from repositories.task_repository import TASK_STORAGE_LOCK  # noqa: E402
from services.creator_delete_impact_service import CreatorDeleteImpactService  # noqa: E402
from test_support.runtime_sandbox import test_artifact_path  # noqa: E402
from services.creator_delete_plan import build_creator_delete_plan  # noqa: E402
from staged_delete_transaction import (  # noqa: E402
    StagedDeleteTransaction,
    recover_pending_delete_transactions,
)
import runtime_paths  # noqa: E402


def append_row(workbook, sheet_name: str, values: dict) -> None:
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    sheet.append([values.get(header, "") for header in headers])


class WorkspaceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.root = test_artifact_path("m4_6b_test", uuid.uuid4().hex)
        self.root.mkdir()

    def tearDown(self) -> None:
        shutil.rmtree(self.root, ignore_errors=True)


class AtomicWriteJsonTests(WorkspaceTestCase):
    @staticmethod
    def _windows_permission_error(winerror: int) -> PermissionError:
        error = PermissionError(13, "Access is denied")
        error.winerror = winerror
        return error

    def test_atomic_write_uses_unique_sibling_temp_files(self) -> None:
        target = self.root / "manifest.json"
        real_replace = runtime_paths.os.replace
        sources: list[Path] = []

        def record_replace(source, destination):
            sources.append(Path(source))
            return real_replace(source, destination)

        with (
            mock.patch.object(runtime_paths, "shared_storage_lock", side_effect=nullcontext),
            mock.patch.object(runtime_paths.os, "replace", side_effect=record_replace),
        ):
            runtime_paths.atomic_write_json(target, {"sequence": 1})
            runtime_paths.atomic_write_json(target, {"sequence": 2})

        self.assertEqual(2, len(sources))
        self.assertNotEqual(sources[0], sources[1])
        self.assertEqual(target.parent, sources[0].parent)
        self.assertEqual(target.parent, sources[1].parent)
        self.assertTrue(sources[0].name.startswith("manifest.json."))
        self.assertTrue(sources[0].name.endswith(".tmp"))
        self.assertEqual({"sequence": 2}, json.loads(target.read_text(encoding="utf-8")))

    def test_windows_transient_replace_error_classifier(self) -> None:
        with mock.patch.object(runtime_paths.os, "name", "nt"):
            for winerror in (5, 32, 33):
                with self.subTest(winerror=winerror):
                    self.assertTrue(
                        runtime_paths._is_windows_transient_replace_error(
                            self._windows_permission_error(winerror)
                        )
                    )
            self.assertFalse(
                runtime_paths._is_windows_transient_replace_error(
                    self._windows_permission_error(13)
                )
            )

    def test_atomic_write_retries_transient_windows_replace_errors(self) -> None:
        for winerror in (5, 32, 33):
            with self.subTest(winerror=winerror):
                target = self.root / f"manifest_{winerror}.json"
                real_replace = runtime_paths.os.replace
                calls = 0

                def fail_once(source, destination):
                    nonlocal calls
                    calls += 1
                    if calls == 1:
                        raise self._windows_permission_error(winerror)
                    return real_replace(source, destination)

                with (
                    mock.patch.object(runtime_paths, "shared_storage_lock", side_effect=nullcontext),
                    mock.patch.object(
                        runtime_paths,
                        "_is_windows_transient_replace_error",
                        return_value=True,
                    ),
                    mock.patch.object(runtime_paths.os, "replace", side_effect=fail_once),
                    mock.patch.object(runtime_paths.time, "sleep") as sleep,
                ):
                    runtime_paths.atomic_write_json(target, {"saved": True})

                self.assertEqual(2, calls)
                sleep.assert_called_once_with(0.05)
                self.assertEqual({"saved": True}, json.loads(target.read_text(encoding="utf-8")))

    def test_atomic_write_does_not_retry_other_windows_permission_errors(self) -> None:
        target = self.root / "manifest.json"

        def fail_with_non_transient_error(*_args):
            raise self._windows_permission_error(13)

        with (
            mock.patch.object(runtime_paths, "shared_storage_lock", side_effect=nullcontext),
            mock.patch.object(
                runtime_paths,
                "_is_windows_transient_replace_error",
                return_value=False,
            ),
            mock.patch.object(
                runtime_paths.os,
                "replace",
                side_effect=fail_with_non_transient_error,
            ) as replace,
            mock.patch.object(runtime_paths.time, "sleep") as sleep,
        ):
            with self.assertRaises(PermissionError):
                runtime_paths.atomic_write_json(target, {"saved": False})

        replace.assert_called_once()
        sleep.assert_not_called()

    def test_atomic_write_reraises_persistent_windows_access_denied(self) -> None:
        target = self.root / "manifest.json"

        def always_fail(*_args):
            raise self._windows_permission_error(32)

        with (
            mock.patch.object(runtime_paths, "shared_storage_lock", side_effect=nullcontext),
            mock.patch.object(
                runtime_paths,
                "_is_windows_transient_replace_error",
                return_value=True,
            ),
            mock.patch.object(
                runtime_paths.os,
                "replace",
                side_effect=always_fail,
            ) as replace,
            mock.patch.object(runtime_paths.time, "sleep") as sleep,
        ):
            with self.assertRaises(PermissionError):
                runtime_paths.atomic_write_json(target, {"saved": False})

        self.assertEqual(runtime_paths.WINDOWS_REPLACE_MAX_RETRIES + 1, replace.call_count)
        self.assertEqual(
            [mock.call(delay) for delay in runtime_paths.WINDOWS_REPLACE_RETRY_DELAYS],
            sleep.call_args_list,
        )
        self.assertFalse(target.exists())


class DeletePlanClassificationTests(WorkspaceTestCase):
    def setUp(self) -> None:
        super().setUp()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        self.tasks_dir = self.root / "tasks"
        self.protection_file = self.root / "data_protection.json"
        self.legacy_dir = self.root / "creator_analysis"
        self.legacy_file = self.root / "creator_library.json"
        self.factory = RepositoryFactory.for_path(
            self.workbook_path,
            tasks_dir=self.tasks_dir,
            data_protection_file=self.protection_file,
            legacy_analysis_dir=self.legacy_dir,
            legacy_library_file=self.legacy_file,
        )
        with mock.patch.object(creator_repository, "log_event"):
            self.factory.creator().getCreators()

    def append(self, sheet_name: str, values: dict) -> None:
        workbook = load_workbook(self.workbook_path)
        append_row(workbook, sheet_name, values)
        workbook.save(self.workbook_path)
        workbook.close()

    def seed_creator(self, creator_id: str, account_id: str, account_uid: str) -> None:
        self.append("Creators", {
            "creator_id": creator_id,
            "name": creator_id,
            "archived_at": "2026-08-17T00:00:00Z",
            "updated_at": "2026-08-17T00:00:00Z",
        })
        self.append("CreatorAccounts", {
            "account_id": account_id,
            "creator_id": creator_id,
            "account_uid": account_uid,
        })

    def create_task(self, task_id: str, metadata: dict) -> Path:
        root = self.tasks_dir / task_id
        root.mkdir(parents=True)
        (root / "task.json").write_text(
            json.dumps({"id": task_id, **metadata}), encoding="utf-8"
        )
        (root / "links.txt").write_text("https://example.invalid/\n", encoding="utf-8")
        return root

    def test_exact_resources_receive_delete_locators_without_campaign_root(self) -> None:
        self.seed_creator("creator_target", "account_target", "uid_target")
        self.append("CreatorSnapshots", {
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_target",
            "account_uid": "uid_target",
        })
        self.append("VideoSnapshots", {
            "video_snapshot_id": "video_snapshot_target",
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_target",
        })
        self.append("Campaigns", {"campaign_id": "campaign_keep", "name": "Keep"})
        self.append("CampaignCreators", {
            "id": "relation_delete",
            "campaign_id": "campaign_keep",
            "creator_id": "creator_target",
            "account_id": "account_target",
            "archived_at": "2026-08-16T00:00:00Z",
        })
        self.append("FollowUpLogs", {
            "follow_up_id": "followup_delete",
            "object_type": "creator",
            "object_id": "creator_target",
        })
        self.create_task(
            "task_20260817T000000Z_aaaaaaaa",
            {"creator_library_creator_ids": ["creator_target"]},
        )
        self.protection_file.write_text(
            json.dumps({"uid_target": {"email": {"value": "private"}}}),
            encoding="utf-8",
        )
        self.legacy_dir.mkdir()
        (self.legacy_dir / "analysis_task_target.json").write_text(
            json.dumps({"creator_id": "creator_target", "account_uid": "uid_target"}),
            encoding="utf-8",
        )

        repository = self.factory.creator_delete_impact()
        snapshot = repository.scan_creator_delete_impact("creator_target")
        result = CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        ).get_delete_impact("creator_target")
        snapshot_plan = build_creator_delete_plan(snapshot)
        decisions = {
            item["source"]: item["classification"]
            for item in snapshot_plan["decisions"]
        }

        self.assertEqual("DELETE", decisions["creator_snapshots"])
        self.assertEqual("DELETE", decisions["video_snapshots"])
        self.assertEqual("DELETE", decisions["campaign_creators"])
        self.assertEqual("DELETE", decisions["follow_up_logs"])
        self.assertEqual("DELETE", decisions["task_artifacts"])
        self.assertEqual("DELETE", decisions["data_protection"])
        self.assertEqual("DELETE", decisions["legacy_sources"])
        self.assertTrue(result["can_delete"])
        self.assertTrue(
            all(set(item) == {"source", "count"} for item in result["unresolved"])
        )
        locator_ids = {
            item["stable_id"]
            for items in snapshot["resource_locators"].values()
            for item in items
        }
        self.assertIn("snapshot_target", locator_ids)
        self.assertIn("video_snapshot_target", locator_ids)
        self.assertIn("relation_delete", locator_ids)
        self.assertIn("creator_target", locator_ids)
        self.assertIn("account_target", locator_ids)
        self.assertNotIn("campaign_keep", locator_ids)

    def test_conflicting_and_unresolvable_resources_fail_closed(self) -> None:
        self.seed_creator("creator_target", "account_target", "uid_shared")
        self.seed_creator("creator_other", "account_other", "uid_shared")
        self.append("CreatorSnapshots", {
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_target",
        })
        self.append("VideoSnapshots", {
            "video_snapshot_id": "video_conflict",
            "snapshot_id": "snapshot_target",
            "creator_id": "creator_other",
        })
        self.append("Cooperations", {
            "cooperation_id": "cooperation_block",
            "creator_id": "creator_target",
            "note": "private",
        })
        self.append("FollowUpLogs", {
            "follow_up_id": "followup_unknown",
            "object_type": "unknown",
            "object_id": "creator_target",
        })
        self.append("_AnalysisData", {
            "creator_id": "creator_other",
            "analysis_json": json.dumps({"creator_id": "creator_target"}),
        })
        self.create_task(
            "task_20260817T000001Z_bbbbbbbb",
            {"creator_library_creator_ids": ["creator_target", "creator_other"]},
        )
        self.create_task("task_20260817T000002Z_cccccccc", {})
        self.protection_file.write_text(
            json.dumps({"uid_shared": {"email": {"value": "private"}}}),
            encoding="utf-8",
        )
        self.legacy_dir.mkdir()
        (self.legacy_dir / "analysis_task_unknown.json").write_text(
            json.dumps({"status": "archived"}), encoding="utf-8"
        )

        repository = self.factory.creator_delete_impact()
        snapshot = repository.scan_creator_delete_impact("creator_target")
        result = CreatorDeleteImpactService(
            self.factory.creator_delete_impact
        ).get_delete_impact("creator_target")
        snapshot_plan = build_creator_delete_plan(snapshot)
        decisions = {
            item["source"]: item["classification"]
            for item in snapshot_plan["decisions"]
        }
        codes = {item["code"] for item in result["blockers"]}

        self.assertEqual("BLOCK", decisions["video_snapshots"])
        self.assertEqual("BLOCK", decisions["cooperations"])
        self.assertEqual("BLOCK", decisions["task_artifacts"])
        self.assertEqual("BLOCK", decisions["unmapped_task_artifacts"])
        self.assertEqual("BLOCK", decisions["data_protection"])
        self.assertEqual("BLOCK", decisions["legacy_sources"])
        self.assertEqual("BLOCK", decisions["embedded_analysis_references"])
        self.assertIn("VIDEO_SNAPSHOT_OWNERSHIP_CONFLICT", codes)
        self.assertIn("COOPERATION_RETENTION_ANONYMIZATION_GAP", codes)
        self.assertIn("SHARED_TASK_CREATOR_REFERENCE", codes)
        self.assertIn("SHARED_DATA_PROTECTION_UID", codes)
        self.assertIn("UNRELIABLE_LEGACY_IDENTITY", codes)
        self.assertIn("EMBEDDED_ANALYSIS_REFERENCE", codes)
        self.assertIn("UNKNOWN_FOLLOWUP_OBJECT_TYPE", codes)
        self.assertFalse(result["can_delete"])

    def test_policy_version_participates_in_deterministic_fingerprint(self) -> None:
        self.seed_creator("creator_target", "account_target", "uid_target")
        service = CreatorDeleteImpactService(self.factory.creator_delete_impact)
        first = service.get_delete_impact("creator_target")["preview_fingerprint"]
        second = service.get_delete_impact("creator_target")["preview_fingerprint"]
        self.assertEqual(first, second)
        with mock.patch(
            "services.creator_delete_impact_service.POLICY_VERSION",
            "m4.6b-delete-safety-v2",
        ):
            changed = service.get_delete_impact("creator_target")[
                "preview_fingerprint"
            ]
        self.assertNotEqual(first, changed)


class StagedDeleteTransactionTests(WorkspaceTestCase):
    def create_workbook(self, path: Path, value: str) -> None:
        workbook = Workbook()
        workbook.active["A1"] = value
        workbook.save(path)
        workbook.close()

    def transaction(self) -> StagedDeleteTransaction:
        transaction = StagedDeleteTransaction(
            self.root / "runtime",
            "creator_safe",
            transaction_id="delete_test_transaction",
        )
        transaction.prepare()
        return transaction

    def test_manifest_is_atomic_minimal_and_rejects_invalid_state(self) -> None:
        transaction = self.transaction()
        manifest = transaction.transition("STAGED")
        raw = transaction.manifest_path.read_text(encoding="utf-8")
        self.assertEqual("STAGED", manifest["phase"])
        for sensitive in ("email", "whatsapp", "bio", "notes", "analysis_json"):
            self.assertNotIn(sensitive, raw.casefold())
        transaction.manifest_path.write_text("{}", encoding="utf-8")
        with self.assertRaisesRegex(RuntimeError, "manifest"):
            transaction.load_manifest()

        other = StagedDeleteTransaction(
            self.root / "runtime",
            "creator_safe",
            transaction_id="delete_transition_transaction",
        )
        other.prepare()
        other.transition("STAGED")
        other.transition("COMMITTED")
        other.transition("CLEANED")
        with self.assertRaisesRegex(ValueError, "phase transition"):
            other.transition("MUTATING")

    def test_local_storage_writers_share_one_process_lock(self) -> None:
        self.assertIs(WORKBOOK_LOCK, LOCAL_STORAGE_MUTATION_LOCK)
        self.assertIs(TASK_STORAGE_LOCK, LOCAL_STORAGE_MUTATION_LOCK)

    def test_file_and_directory_quarantine_rollback_and_cleanup(self) -> None:
        transaction = self.transaction()
        single = self.root / "artifact.json"
        single.write_text("data", encoding="utf-8")
        directory = self.root / "task_dir"
        directory.mkdir()
        (directory / "task.json").write_text("{}", encoding="utf-8")
        quarantined_file = transaction.stage_path(single)
        quarantined_dir = transaction.stage_path(directory)
        self.assertFalse(single.exists())
        self.assertFalse(directory.exists())
        transaction.transition("STAGED")
        transaction.recover()
        self.assertTrue(single.is_file())
        self.assertTrue(directory.is_dir())
        self.assertFalse(quarantined_file.exists())
        self.assertFalse(quarantined_dir.exists())

        committed = StagedDeleteTransaction(
            self.root / "runtime",
            "creator_safe",
            transaction_id="delete_committed_transaction",
        )
        committed.prepare()
        disposable = self.root / "disposable"
        disposable.mkdir()
        quarantined = committed.stage_path(disposable)
        committed.transition("STAGED")
        committed.transition("COMMITTED")
        self.assertEqual("CLEANED", committed.recover()["phase"])
        self.assertFalse(quarantined.exists())
        self.assertFalse(disposable.exists())

    def test_recovery_uses_planned_move_after_crash_during_stage(self) -> None:
        transaction = self.transaction()
        artifact = self.root / "crash_window_artifact"
        artifact.write_text("recoverable", encoding="utf-8")
        real_replace = __import__("os").replace

        def replace_then_crash(source, destination):
            real_replace(source, destination)
            if Path(source) == artifact:
                raise OSError("simulated crash after rename")

        with mock.patch("staged_delete_transaction.os.replace", replace_then_crash):
            with self.assertRaisesRegex(OSError, "simulated crash"):
                transaction.stage_path(artifact)

        self.assertFalse(artifact.exists())
        manifest = transaction.load_manifest()
        self.assertEqual("planned", manifest["quarantine_moves"][0]["state"])
        recovered = recover_pending_delete_transactions(self.root / "runtime")
        self.assertEqual("ROLLED_BACK", recovered[0]["phase"])
        self.assertEqual("recoverable", artifact.read_text(encoding="utf-8"))

    def test_json_and_workbook_restore_use_exact_transaction_backups(self) -> None:
        transaction = self.transaction()
        workbook_path = self.root / "Creator_Library.xlsx"
        self.create_workbook(workbook_path, "before")
        store = ExcelWorkbookStore(workbook_path)
        exact_backup = transaction.backup_workbook(store)
        unrelated_backup = self.root / "Creator_Library.xlsx.bak"
        self.create_workbook(unrelated_backup, "wrong")
        self.create_workbook(workbook_path, "after")

        protection = self.root / "data_protection.json"
        protection.write_text(
            json.dumps({"uid_target": {"value": 1}, "uid_other": {"value": 2}}),
            encoding="utf-8",
        )
        transaction.backup_json(protection, label="data_protection")
        transaction.transition("STAGED")
        transaction.transition("MUTATING")
        transaction.write_json(protection, {"uid_other": {"value": 2}})
        transaction.recover()

        restored = load_workbook(workbook_path, read_only=True)
        self.assertEqual("before", restored.active["A1"].value)
        restored.close()
        self.assertTrue(exact_backup.is_file())
        self.assertEqual(
            {"uid_target": {"value": 1}, "uid_other": {"value": 2}},
            json.loads(protection.read_text(encoding="utf-8")),
        )

    def test_cleanup_failure_is_pending_and_never_rolls_back_commit(self) -> None:
        transaction = self.transaction()
        artifact = self.root / "committed_artifact"
        artifact.mkdir()
        quarantine = transaction.stage_path(artifact)
        transaction.transition("STAGED")
        transaction.transition("COMMITTED")
        with mock.patch.object(transaction, "_remove_path", side_effect=OSError("busy")):
            result = transaction.finalize_cleanup()
        self.assertEqual("CLEANUP_PENDING", result["phase"])
        self.assertTrue(result["commit_marker"])
        self.assertTrue(quarantine.exists())
        self.assertFalse(artifact.exists())
        self.assertEqual("CLEANED", transaction.recover()["phase"])

    def test_same_volume_and_plan_preflight_fail_closed(self) -> None:
        transaction = self.transaction()
        artifact = self.root / "artifact"
        artifact.write_text("x", encoding="utf-8")
        with mock.patch.object(transaction, "_same_volume", return_value=False):
            with self.assertRaisesRegex(ValueError, "same volume"):
                transaction.stage_path(artifact)

        workbook_path = self.root / "Creator_Library.xlsx"
        self.create_workbook(workbook_path, "safe")
        blocked = transaction.preflight(
            {"blocked": [{"code": "BLOCK"}], "delete_locators": []},
            workbook_path=workbook_path,
        )
        self.assertEqual("BLOCKED", blocked["status"])
        self.assertIn("DELETE_PLAN_BLOCKED", blocked["reasons"])


if __name__ == "__main__":
    unittest.main()

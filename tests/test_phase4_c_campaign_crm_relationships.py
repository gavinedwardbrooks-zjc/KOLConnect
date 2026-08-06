from __future__ import annotations

import json
import sys
import tempfile
import threading
import unittest
import urllib.request
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

import creator_repository
import server
from campaign_creator_repository import CampaignCreatorRepository
from campaign_repository import CampaignRepository
from dashboard_repository import DashboardRepository
from product_repository import ProductRepository


class CampaignCrmRelationshipTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Creator_Library.xlsx"
        self.creator_repository = creator_repository.CreatorRepository(self.workbook_path)
        self.creator_repository.getCreators()
        self._seed_creator()
        self.product_repository = ProductRepository(self.workbook_path)
        self.campaign_repository = CampaignRepository(self.workbook_path)
        self.relation_repository = CampaignCreatorRepository(self.workbook_path)
        self.product = self.product_repository.createProduct({
            "name": "Product One",
            "company_name": "Company",
        })

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    @staticmethod
    def _append(sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _seed_creator(self) -> None:
        workbook = load_workbook(self.workbook_path)
        self._append(workbook["Creators"], {
            "creator_id": "creator_one",
            "name": "Creator One",
            "platform": "TikTok",
            "status": "discovered",
            "created_at": "2026-08-01T00:00:00Z",
            "updated_at": "2026-08-01T00:00:00Z",
        })
        self._append(workbook["CreatorAccounts"], {
            "account_id": "account_one",
            "creator_id": "creator_one",
            "account_uid": "tiktok|creator-one",
            "platform": "TikTok",
        })
        self._append(workbook["CreatorSnapshots"], {
            "snapshot_id": "snapshot_one",
            "creator_id": "creator_one",
            "followers": 1000,
            "captured_at": "2026-08-01T00:00:00Z",
        })
        workbook.save(self.workbook_path)
        workbook.close()

    def _create_campaign(self, name: str) -> dict:
        return self.campaign_repository.createCampaign({
            "product_id": self.product["product_id"],
            "name": name,
        })

    def _add_creator(self, campaign_id: str) -> dict:
        return self.relation_repository.createCampaignCreator({
            "campaign_id": campaign_id,
            "creator_id": "creator_one",
            "account_id": "account_one",
        })

    def _business_counts(self) -> tuple[int, int, int]:
        workbook = load_workbook(self.workbook_path)
        counts = tuple(
            len(creator_repository.CreatorRepository._rows(workbook[sheet]))
            for sheet in ("Creators", "CreatorAccounts", "CreatorSnapshots")
        )
        workbook.close()
        return counts

    def test_remove_creator_deletes_only_relation_and_is_idempotent(self) -> None:
        campaign_one = self._create_campaign("Campaign One")
        campaign_two = self._create_campaign("Campaign Two")
        relation_one = self._add_creator(campaign_one["campaign_id"])
        relation_two = self._add_creator(campaign_two["campaign_id"])
        before = self._business_counts()

        self.relation_repository.archiveCampaignCreator(relation_one["id"])
        removed = self.relation_repository.remove_creator_from_campaign(relation_one["id"])
        self.assertTrue(removed["deleted"])
        self.assertEqual(before, self._business_counts())
        self.assertEqual([], self.relation_repository.getCampaignCreators(
            campaign_id=campaign_one["campaign_id"], include_archived=True,
        ))
        self.assertEqual(
            relation_two["id"],
            self.relation_repository.getCampaignCreators(campaign_id=campaign_two["campaign_id"])[0]["id"],
        )
        self.assertFalse(self.relation_repository.remove_creator_from_campaign(relation_one["id"])["deleted"])

    def test_delete_campaign_removes_its_relations_and_preserves_other_business_data(self) -> None:
        deleted_campaign = self._create_campaign("Delete Me")
        kept_campaign = self._create_campaign("Keep Me")
        self._add_creator(deleted_campaign["campaign_id"])
        kept_relation = self._add_creator(kept_campaign["campaign_id"])
        before = self._business_counts()

        self.campaign_repository.archiveCampaign(deleted_campaign["campaign_id"])
        result = self.campaign_repository.delete_campaign(deleted_campaign["campaign_id"])
        self.assertTrue(result["deleted"])
        self.assertEqual(1, result["removed_campaign_creators"])
        self.assertEqual(before, self._business_counts())
        with self.assertRaisesRegex(ValueError, "不存在"):
            self.campaign_repository.getCampaign(deleted_campaign["campaign_id"])
        self.assertEqual(kept_campaign["campaign_id"], self.campaign_repository.getCampaign(
            kept_campaign["campaign_id"]
        )["campaign_id"])
        self.assertEqual(kept_relation["id"], self.relation_repository.getCampaignCreators(
            campaign_id=kept_campaign["campaign_id"]
        )[0]["id"])

        repeated = self.campaign_repository.delete_campaign(deleted_campaign["campaign_id"])
        self.assertFalse(repeated["deleted"])
        self.assertEqual(0, repeated["removed_campaign_creators"])

        dashboard_records = DashboardRepository(self.creator_repository).get_campaign_creator_records()
        self.assertEqual({kept_campaign["campaign_id"]}, {
            record["campaign_id"] for record in dashboard_records
        })
        self.assertTrue(all(record["campaign_name"] for record in dashboard_records))

    def test_delete_api_is_idempotent(self) -> None:
        campaign = self._create_campaign("API Campaign")
        relation = self._add_creator(campaign["campaign_id"])
        patchers = [
            mock.patch.object(server, "get_campaign_repository", return_value=self.campaign_repository),
            mock.patch.object(server, "get_campaign_creator_repository", return_value=self.relation_repository),
            mock.patch.object(server, "log_event"),
            mock.patch.object(server, "log_error"),
            mock.patch.object(server, "_record_last_error"),
        ]
        for patcher in patchers:
            patcher.start()
        httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.Handler)
        thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        thread.start()
        base_url = f"http://127.0.0.1:{httpd.server_port}"
        try:
            request = urllib.request.Request(
                f"{base_url}/api/campaign-creators/{relation['id']}", method="DELETE"
            )
            with urllib.request.urlopen(request) as response:
                removed = json.loads(response.read().decode("utf-8"))
            self.assertTrue(removed["deleted"])

            second_request = urllib.request.Request(
                f"{base_url}/api/campaign-creators/{relation['id']}", method="DELETE"
            )
            with urllib.request.urlopen(second_request) as response:
                repeated = json.loads(response.read().decode("utf-8"))
            self.assertFalse(repeated["deleted"])

            self._add_creator(campaign["campaign_id"])
            campaign_request = urllib.request.Request(
                f"{base_url}/api/campaigns/{campaign['campaign_id']}", method="DELETE"
            )
            with urllib.request.urlopen(campaign_request) as response:
                deleted = json.loads(response.read().decode("utf-8"))
            self.assertTrue(deleted["deleted"])
            self.assertEqual(1, deleted["removed_campaign_creators"])

            repeated_campaign_request = urllib.request.Request(
                f"{base_url}/api/campaigns/{campaign['campaign_id']}", method="DELETE"
            )
            with urllib.request.urlopen(repeated_campaign_request) as response:
                repeated_campaign = json.loads(response.read().decode("utf-8"))
            self.assertFalse(repeated_campaign["deleted"])
            self.assertEqual(0, repeated_campaign["removed_campaign_creators"])
        finally:
            httpd.shutdown()
            httpd.server_close()
            thread.join(timeout=2)
            for patcher in reversed(patchers):
                patcher.stop()


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import shutil
import sys
import unittest
from contextlib import nullcontext
from pathlib import Path
from unittest.mock import Mock, patch
from uuid import uuid4

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))
sys.path.insert(0, str(ROOT / "tests"))

from campaign_creator_repository import CampaignCreatorRepository
from campaign_repository import CampaignRepository
from creator_repository import _WORKBOOK_SHEETS
from http_handlers import campaign_handler, task_handler
from product_repository import ProductRepository
from test_support.runtime_sandbox import test_artifact_path


class Handler:
    def __init__(self) -> None:
        self.payload = None
        self.status = None

    def _json(self, payload, status=200):
        self.payload = payload
        self.status = status

    def _repository_error(self, exc):
        raise exc


class CampaignDetailCompatibilityTests(unittest.TestCase):
    def setUp(self) -> None:
        runtime = test_artifact_path("m7_1f")
        runtime.mkdir(exist_ok=True)
        self.root = runtime / uuid4().hex
        self.root.mkdir()
        self.workbook_path = self.root / "Creator_Library.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in _WORKBOOK_SHEETS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(list(headers))
        workbook.save(self.workbook_path)
        workbook.close()
        self.lock_patch = patch(
            "excel_workbook_store.shared_storage_lock",
            side_effect=lambda *args, **kwargs: nullcontext(),
        )
        self.lock_patch.start()

    def tearDown(self) -> None:
        self.lock_patch.stop()
        shutil.rmtree(self.root, ignore_errors=True)

    @staticmethod
    def _append(sheet, values: dict) -> None:
        headers = [str(cell.value or "") for cell in sheet[1]]
        sheet.append([values.get(header, "") for header in headers])

    def _create_campaign(self, **changes):
        product = ProductRepository(self.workbook_path).createProduct({"name": "BLOCK BLAST!"})
        payload = {
            "name": "father day",
            "product_id": product["product_id"],
            "country": "",
            "platform": "",
            "start_date": "",
            "end_date": "",
        }
        payload.update(changes)
        return CampaignRepository(self.workbook_path).createCampaign(payload), product

    def test_clean_baseline_and_create_detail_roundtrip_with_zero_creators(self) -> None:
        repository = CampaignRepository(self.workbook_path)
        self.assertEqual([], repository.getCampaigns())
        campaign, product = self._create_campaign()

        detail = repository.getCampaign(campaign["campaign_id"])
        self.assertEqual("father day", detail["name"])
        self.assertEqual(product["name"], detail["product_name"])
        self.assertEqual(0, detail["creators_count"])
        for field in ("country", "platform", "start_date", "end_date"):
            self.assertEqual("", detail[field])
        self.assertEqual(
            [],
            CampaignCreatorRepository(self.workbook_path).getCampaignCreators(
                campaign_id=campaign["campaign_id"]
            ),
        )

        handler = Handler()
        request = {
            "method": "GET",
            "path": f"/api/campaigns/{campaign['campaign_id']}",
            "query": {},
        }
        context = {
            "repositories": {
                "campaign": lambda: repository,
                "campaign_creator": lambda: CampaignCreatorRepository(self.workbook_path),
            },
            "services": {},
        }
        self.assertTrue(campaign_handler.handle(handler, request, context))
        self.assertEqual(200, handler.status)
        self.assertEqual("father day", handler.payload["campaign"]["name"])

    def test_missing_campaign_is_explicit_404_and_missing_product_is_optional(self) -> None:
        handler = Handler()
        request = {"method": "GET", "path": "/api/campaigns/missing", "query": {}}
        context = {
            "repositories": {
                "campaign": lambda: CampaignRepository(self.workbook_path),
                "campaign_creator": lambda: CampaignCreatorRepository(self.workbook_path),
            },
            "services": {},
        }
        self.assertTrue(campaign_handler.handle(handler, request, context))
        self.assertEqual(404, handler.status)
        self.assertEqual("CAMPAIGN_NOT_FOUND", handler.payload["error"])

        workbook = load_workbook(self.workbook_path)
        self._append(workbook["Campaigns"], {
            "campaign_id": "legacy-campaign",
            "product_id": "missing-product",
            "name": "Legacy",
            "status": "draft",
        })
        workbook.save(self.workbook_path)
        workbook.close()
        with patch("campaign_repository.log_event"):
            detail = CampaignRepository(self.workbook_path).getCampaign("legacy-campaign")
        self.assertEqual("", detail["product_name"])
        self.assertEqual(0, detail["creators_count"])

    def test_stale_creator_is_safe_and_multi_account_campaign_selects_platform_once(self) -> None:
        campaign, _ = self._create_campaign(platform="TikTok")
        workbook = load_workbook(self.workbook_path)
        self._append(workbook["Creators"], {"creator_id": "creator-one", "name": "Creator One"})
        self._append(workbook["CreatorAccounts"], {
            "account_id": "youtube-account",
            "creator_id": "creator-one",
            "account_uid": "youtube-uid",
            "platform": "YouTube",
        })
        self._append(workbook["CreatorAccounts"], {
            "account_id": "tiktok-account",
            "creator_id": "creator-one",
            "account_uid": "tiktok-uid",
            "platform": "TikTok",
        })
        self._append(workbook["CampaignCreators"], {
            "id": "stale-relation",
            "campaign_id": campaign["campaign_id"],
            "creator_id": "missing-creator",
            "account_id": "missing-account",
            "stage": "pending_contact",
        })
        workbook.save(self.workbook_path)
        workbook.close()

        repository = CampaignCreatorRepository(self.workbook_path)
        result = repository.batch_add_creators(campaign["campaign_id"], ["creator-one"])
        self.assertEqual("added", result[0]["status"])
        relations = repository.getCampaignCreators(campaign_id=campaign["campaign_id"])
        self.assertEqual(2, len(relations))
        selected = next(row for row in relations if row["creator_id"] == "creator-one")
        self.assertEqual("tiktok-account", selected["account_id"])
        self.assertEqual("TikTok", selected["account_platform"])
        stale = next(row for row in relations if row["id"] == "stale-relation")
        self.assertEqual("", stale["creator_name"])
        self.assertEqual("", stale["account_platform"])

    def test_legacy_task_sync_route_is_not_registered(self) -> None:
        handler = Handler()
        request = {
            "method": "POST",
            "path": "/api/tasks/task-one/sync-four-tables",
            "query": {},
            "get_payload": lambda: {},
        }
        service = Mock()
        context = {"services": {"task": service}}
        self.assertFalse(task_handler.handle(handler, request, context))
        service.sync_four_tables.assert_not_called()


if __name__ == "__main__":
    unittest.main()
